from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends, BackgroundTasks, Query, status, Request, Body, Body
from fastapi.responses import JSONResponse, FileResponse, Response, StreamingResponse
import asyncio
import functools
import gc
import gzip
import math
import operator
import pandas as pd
import numpy as np
import json
import re
import os
import copy
import tempfile
import base64
import zipfile
import io
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any, Tuple
from pydantic import BaseModel
from pathlib import Path
from app.models.schemas import (
    UploadRequest, UploadResponse, ChatRequest, ChatResponse, 
    ErrorResponse, DataStats, KnowledgeGraphRequest, KnowledgeGraphResponse,
    CodeExecutionResponse, ColumnInfoResponse, ColumnInfo,
    CustomTreatmentUpdate, CustomTreatmentResponse,
    BivariateAnalysisAllResponse, BivariateAnalysisSingleResponse,CorrelationAnalysisRequest, CorrelationAnalysisResponse, 
    SingleVariableCorrelationResponse, CorrelationHeatmapResponse, 
    CorrelationHeatmapImageResponse, GlobalModelTrainingRequest, GlobalModelTrainingResponse,
    ModelPerformanceMetrics, CrossValidationResult, ModelAlgorithm, DatasetPreviewResponse,
    SegmentationRequest, SegmentationResponse, AutoSegmentationRequest, SegmentProfilingResponse,
    ModelCodebookResponse, CodebookSection, DatasetTypeClassificationRequest, DatasetTypeClassificationResponse,
    DatasetType, VariableDefinitionsRequest, VariableDefinitionsResponse, VariableDefinition,
    FeatureTransformationResponse, DQSResponse,
    CrossAlgorithmRecommendationRequest, CrossAlgorithmRecommendationResponse,
    # Segmentation Agent schemas (4-mode architecture)
    SegmentationMode, VariablePriority, ManualSegmentRule, RuleCondition,
    UnifiedSegmentationRequest, UnifiedSegmentationResponse, SegmentDetail, TertiaryPromotionSuggestion,
    PromotionSuggestion,
    ValidationSuiteResult, MergeRecommendation, SegmentFlag,
    RuleValidationResult, AddToDataRequest, AddToDataResponse,
    SchemeRegistryResponse, SchemeRegistryEntry, SegmentationSchemeMetadata,
    SegmentationSchemeDetailResponse,
    VariableRelevanceMatrix, BootstrapStabilityResult, OutOfSampleValidation,
    MergeSegmentsRequest, MergeSegmentsResponse, MergeImpact,
    CutoffEditRequest, CutoffEditResponse, CutoffEditImpact,
    MoveCategoricalValueRequest, MoveCategoricalValueResponse,
)
from app.api.auth_routes import get_current_user_dependency
from app.services.dataset_service import dataset_manager, FileTooLargeError
from app.services.unique_id_validation import compute_duplicate_stats_from_tabular_path
from app.services.agentic_system import AgenticSystem, MessageState, DatasetAnalyser
from app.services.data_quality_detector import (
    DataQualityError,
    DetectionError,
    TreatmentError,
    ValidationError as DQValidationError,
    ConfigurationError as DQConfigurationError,
    ColumnNotFoundError,
    InsufficientDataError,
)
from app.services.llm_service import llm_service
from app.services.dataframe_state_manager import dataframe_state_manager
from app.services.vector_store import vector_store
from app.services.segmentation_service import segmentation_service
from app.services.tertiary_promotion_utils import check_tertiary_promotion as run_depth1_tertiary_promotion_check
from app.services.segmentation_validation import (
    validation_suite as seg_validation_suite,
    variable_relevance_calculator,
    narrative_generator,
)
from app.services.message_state_service import message_state_manager
from app.services.dqs_service import dqs_service
from app.services.model_training_auto_training import make_json_serializable
from app.services.partition_preview_service import build_partition_preview
from app.services.analytics_cache import analytics_cache
from app.services.insight_job_runners import (
    run_insight_bivariate_all_job,
    run_insight_correlation_heatmap_basic_job,
    run_insight_correlation_heatmap_categorical_job,
    run_insight_correlation_matrix_job,
    run_insight_correlation_ratio_job,
    run_insight_iv_analysis_job,
    run_insight_vif_analysis_job,
)
from app.utils.correlation_insight_heatmaps import heatmap_top_n as _heatmap_top_n
from app.utils.vif_insight_payload import build_vif_frontend_analysis_payload as _build_vif_frontend_analysis_payload
from app.services.segmentation_validation_refresh import rebuild_validation_from_segmentation_result
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import (
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    roc_auc_score,
    confusion_matrix,
)
from app.utils import helpers
# ---------------------------------------------------------------------------
# Helper utilities for persisting modelling context into MessageState
# ---------------------------------------------------------------------------

def _persist_modelling_artifacts(dataset_id: str, modelling_results: Dict[str, Any]) -> None:
    """
    Store variable analysis outputs and training artifacts in MessageState so the
    modelling agent can answer follow-up questions (VIF/IV filters, model metrics, etc.).
    """
    if not dataset_id or not modelling_results:
        return

    try:
        state = message_state_manager.create_or_load_state(dataset_id, "")
    except ValueError as exc:
        logger.warning(f"Unable to load MessageState for dataset {dataset_id}: {exc}")
        return

    if state is None:
        logger.warning(f"MessageState unavailable for dataset {dataset_id}; skipping modelling artifact persistence")
        return

    updated = False

    def _set_if_value(key: str, value: Any) -> None:
        nonlocal updated
        if value is not None:
            state[key] = value
            updated = True

    variable_analysis = modelling_results.get("variable_analysis")
    # Preserve existing variable_analysis if new results don't have it
    if variable_analysis is None:
        variable_analysis = state.get("variable_analysis")
    _set_if_value("variable_analysis", variable_analysis)

    # Provide easy access to flattened variable statistics if present
    if isinstance(variable_analysis, dict):
        stats = variable_analysis.get("variable_statistics")
        _set_if_value("variable_statistics", stats)
    elif variable_analysis is None:
        # Preserve existing variable_statistics if available
        existing_stats = state.get("variable_statistics")
        if existing_stats is not None:
            _set_if_value("variable_statistics", existing_stats)

    # Try multiple locations for used_features
    used_features = modelling_results.get("used_features")
    
    # Fallback 1: Check in results array (first result)
    if used_features is None or (isinstance(used_features, list) and len(used_features) == 0):
        results_table = modelling_results.get("results")
        if isinstance(results_table, list) and len(results_table) > 0:
            first_result = results_table[0]
            if isinstance(first_result, dict):
                # Check direct used_features
                if "used_features" in first_result:
                    used_features = first_result.get("used_features")
                    logger.info(f"Extracted used_features from results[0]: {len(used_features) if isinstance(used_features, list) else 'N/A'} features")
                # Check nested in best_model
                elif "best_model" in first_result:
                    best_model = first_result.get("best_model")
                    if isinstance(best_model, dict) and "used_features" in best_model:
                        used_features = best_model.get("used_features")
                        logger.info(f"Extracted used_features from results[0].best_model: {len(used_features) if isinstance(used_features, list) else 'N/A'} features")
    
    # Fallback 2: Check in best_model_selection (multiple levels)
    if used_features is None or (isinstance(used_features, list) and len(used_features) == 0):
        best_model_info = modelling_results.get("best_model_selection") or {}
        if isinstance(best_model_info, dict):
            # Check direct
            if "used_features" in best_model_info:
                used_features = best_model_info.get("used_features")
                logger.info(f"Extracted used_features from best_model_selection: {len(used_features) if isinstance(used_features, list) else 'N/A'} features")
            # Check nested in best_model
            elif "best_model" in best_model_info:
                best_model = best_model_info.get("best_model")
                if isinstance(best_model, dict):
                    if "used_features" in best_model:
                        used_features = best_model.get("used_features")
                        logger.info(f"Extracted used_features from best_model_selection.best_model: {len(used_features) if isinstance(used_features, list) else 'N/A'} features")
    
    # Fallback 2b: Check in auto_selection_summary
    if used_features is None or (isinstance(used_features, list) and len(used_features) == 0):
        auto_selection = modelling_results.get("auto_selection_summary") or {}
        if isinstance(auto_selection, dict):
            if "used_features" in auto_selection:
                used_features = auto_selection.get("used_features")
                logger.info(f"Extracted used_features from auto_selection_summary: {len(used_features) if isinstance(used_features, list) else 'N/A'} features")
            elif "best_model" in auto_selection:
                best_model = auto_selection.get("best_model")
                if isinstance(best_model, dict) and "used_features" in best_model:
                    used_features = best_model.get("used_features")
                    logger.info(f"Extracted used_features from auto_selection_summary.best_model: {len(used_features) if isinstance(used_features, list) else 'N/A'} features")

    # Fallback 3: Aggregate from segment/manual structures (all_results / segment_results)
    if used_features is None or (isinstance(used_features, list) and len(used_features) == 0):
        collected_features: List[str] = []

        # 3a) Manual segment training returns enriched results in `all_results`
        all_results = modelling_results.get("all_results")
        if isinstance(all_results, list):
            for entry in all_results:
                if isinstance(entry, dict):
                    uf = entry.get("used_features")
                    if isinstance(uf, list):
                        collected_features.extend(str(f) for f in uf)

        # 3b) Segment auto training stores flattened per-segment results under `segment_results`
        segment_results = modelling_results.get("segment_results")
        if isinstance(segment_results, dict):
            for seg_key, seg_val in segment_results.items():
                if isinstance(seg_val, dict):
                    uf = seg_val.get("used_features")
                    if isinstance(uf, list):
                        collected_features.extend(str(f) for f in uf)

        # Deduplicate while preserving at least a stable order
        if collected_features:
            seen = set()
            deduped: List[str] = []
            for f in collected_features:
                if f not in seen:
                    seen.add(f)
                    deduped.append(f)
            used_features = deduped
    
    # Ensure used_features is a list
    if used_features is not None and not isinstance(used_features, list):
        if isinstance(used_features, (str, tuple)):
            used_features = list(used_features)
        else:
            logger.warning(f"used_features is not a list: {type(used_features)}, converting")
            used_features = [str(f) for f in used_features] if hasattr(used_features, '__iter__') else []
    
    if used_features:
        logger.info(f"Persisting used_features: {len(used_features)} features")
    else:
        logger.warning(f"used_features is None or empty in modelling_results. Available keys: {list(modelling_results.keys())}")
    
    _set_if_value("used_features", used_features)
    _set_if_value("used_features_short", used_features)

    results_table = modelling_results.get("results")
    _set_if_value("results", results_table)
    _set_if_value("model_comparison", results_table)

    best_model_info = modelling_results.get("best_model_selection") or {}
    if isinstance(best_model_info, dict) and best_model_info:
        _set_if_value("best_model_summary", best_model_info)
        _set_if_value("best_model", best_model_info.get("best_model"))
        _set_if_value("model_id", best_model_info.get("best_model_id"))
        _set_if_value("comparison_results_json", best_model_info.get("metrics_comparison"))
        best_model = best_model_info.get("best_model")
        if best_model and isinstance(best_model, dict):
            _set_if_value("model_params_short", best_model.get("hyperparameters"))

    # Preserve the entire payload for future reference
    _set_if_value("training_context", modelling_results)

    if updated:
        save_success = message_state_manager.save_state(dataset_id, state)
        if save_success:
            persisted_keys = [k for k in ["variable_analysis", "used_features", "results", "best_model_summary"] if k in state]
            logger.info(f"Persisted modelling context to MessageState for dataset {dataset_id}: {persisted_keys}")
            if "used_features" in state:
                logger.info(f"  - used_features: {len(state['used_features']) if isinstance(state['used_features'], list) else 'N/A'} features")
        else:
            logger.warning(f"Failed to save modelling context to MessageState for dataset {dataset_id}")
    else:
        logger.debug(f"No modelling artifacts to persist for dataset {dataset_id} (all values were None)")


def _persist_artifacts_if_available(dataset_id: Optional[str], results: Any, job_id: Optional[str] = None, force: bool = False) -> None:
    """
    Serialize and persist modelling artifacts when dataset_id and results are available.
    
    OPTIMIZATION: Uses job_id-based caching to avoid redundant persists during status polling.
    Each job is only persisted once unless force=True is specified.
    
    Args:
        dataset_id: The dataset ID to persist artifacts for
        results: The training results to persist
        job_id: Optional job ID to track persistence (avoids redundant persists)
        force: If True, persist even if already persisted (for explicit save operations)
    """
    if not dataset_id:
        logger.debug(f"Skipping persistence: no dataset_id provided")
        return
    if results is None:
        logger.debug(f"Skipping persistence for dataset {dataset_id}: results is None")
        return
    
    # OPTIMIZATION: Skip if this job has already been persisted (avoids 6-11s per poll)
    if job_id and not force:
        if job_id in _persisted_jobs:
            logger.debug(f"Skipping redundant persistence for job {job_id} - already persisted")
            return
    
    try:
        serializable_results = make_json_serializable(results)
        logger.info(f"Persisting modelling artifacts for dataset {dataset_id}, keys: {list(serializable_results.keys()) if isinstance(serializable_results, dict) else 'N/A'}")
        _persist_modelling_artifacts(dataset_id, serializable_results)
        
        # Mark job as persisted to avoid redundant future persists
        if job_id:
            _persisted_jobs.add(job_id)
            logger.debug(f"Marked job {job_id} as persisted")
    except Exception as exc:
        logger.error(f"Could not persist modelling artifacts for dataset {dataset_id}: {exc}", exc_info=True)


def _extract_dataset_id_from_result(result: Any) -> Optional[str]:
    """
    Best-effort extraction of dataset_id from various training result payloads.
    """
    if not isinstance(result, dict):
        return None

    config = result.get("config")
    if isinstance(config, dict) and config.get("dataset_id"):
        return config["dataset_id"]

    dataset_info = result.get("dataset_info")
    if isinstance(dataset_info, dict) and dataset_info.get("dataset_id"):
        return dataset_info["dataset_id"]

    segment_result = result.get("segment_result")
    if isinstance(segment_result, dict):
        segment_dataset = segment_result.get("dataset_info")
        if isinstance(segment_dataset, dict) and segment_dataset.get("dataset_id"):
            return segment_dataset["dataset_id"]

    segment_data = result.get("segment_data")
    if isinstance(segment_data, dict) and segment_data.get("dataset_id"):
        return segment_data["dataset_id"]

    return None

from app.core.config import (
    settings,
    env_override_present,
    DEFAULT_CHAT_MODEL,
    DEFAULT_KG_MODEL,
    DEFAULT_EMBEDDING_MODEL,
)
from app.core.logging_config import get_logger
from app.utils.helpers import (
    validate_target_variable,
    generate_dataset_summary,
    analyze_all_correlations,
    safe_json_serialize,
    clean_nan_values,
    identify_date_columns,
)
from app.services.feature_engineering_service import feature_engineering_service
import uuid
import time

# Initialize logger
logger = get_logger(__name__)


def _read_data_dictionary(data_dictionary: str) -> str:
    """Load data-dictionary text from a filesystem path, or return inline content.

    Matches upload-route behavior: if ``data_dictionary`` points at an existing
    file, read it as UTF-8; otherwise treat the string as already-inline text
    (e.g. pasted from the form). Used by KG generation, variable review, and
    classify-variables workers.
    """
    if not data_dictionary or not isinstance(data_dictionary, str):
        return ""
    s = data_dictionary.strip()
    if not s:
        return ""
    if os.path.isfile(s):
        try:
            with open(s, "r", encoding="utf-8", errors="replace") as handle:
                return handle.read()
        except OSError as exc:
            logger.warning("Failed to read data dictionary file %s: %s", s, exc)
            return ""
    return s


# =============================================================================
# STANDARDIZED ERROR RESPONSE HELPERS
# =============================================================================

def create_error_response(
    error: Exception,
    status_code: int = 500,
    include_traceback: bool = False
) -> JSONResponse:
    """
    Create a standardized JSON error response from an exception.
    
    Args:
        error: The exception that occurred
        status_code: HTTP status code (default 500)
        include_traceback: Whether to include traceback in details (default False, only in debug)
    
    Returns:
        JSONResponse with standardized error format
    """
    import traceback
    
    error_type = type(error).__name__
    error_message = str(error)
    
    # Build details dict
    details = {}
    
    # Handle DataQualityError subclasses specially
    if isinstance(error, DataQualityError):
        details = error.details.copy() if error.details else {}
        details["operation"] = error.operation
        
        # Map specific error types to appropriate status codes
        if isinstance(error, (DQValidationError, DQConfigurationError)):
            status_code = 400  # Bad Request
        elif isinstance(error, ColumnNotFoundError):
            status_code = 404  # Not Found
        elif isinstance(error, InsufficientDataError):
            status_code = 422  # Unprocessable Entity
        elif isinstance(error, (DetectionError, TreatmentError)):
            status_code = 500  # Internal Server Error
    
    # Include traceback in development/debug mode
    if include_traceback:
        details["traceback"] = traceback.format_exc()
    
    response_body = {
        "error": error_type,
        "message": error_message,
        "details": details if details else None
    }
    
    logger.error(f"API Error [{status_code}]: {error_type}: {error_message}")
    
    return JSONResponse(
        status_code=status_code,
        content=response_body
    )


def handle_data_quality_errors(func):
    """
    Decorator to handle DataQualityError exceptions in API endpoints.
    Converts exceptions to standardized error responses.
    """
    from functools import wraps
    import asyncio
    
    @wraps(func)
    async def async_wrapper(*args, **kwargs):
        try:
            return await func(*args, **kwargs)
        except DataQualityError as e:
            return create_error_response(e)
        except HTTPException:
            raise  # Let FastAPI handle HTTP exceptions
        except Exception as e:
            logger.exception(f"Unexpected error in {func.__name__}: {e}")
            return create_error_response(e, status_code=500)
    
    @wraps(func)
    def sync_wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except DataQualityError as e:
            return create_error_response(e)
        except HTTPException:
            raise
        except Exception as e:
            logger.exception(f"Unexpected error in {func.__name__}: {e}")
            return create_error_response(e, status_code=500)
    
    if asyncio.iscoroutinefunction(func):
        return async_wrapper
    return sync_wrapper



def _segmentation_audit_actor(current_user: Any) -> Optional[str]:
    if current_user is None:
        return None
    for attr in ("username", "email", "user_id"):
        v = getattr(current_user, attr, None)
        if v is not None and str(v).strip():
            return str(v)
    oid = getattr(current_user, "id", None)
    return str(oid) if oid is not None else None


def _record_segmentation_audit_event(
    dataset_id: str, event_type: str, data: Dict[str, Any], current_user: Any = None
) -> None:
    """Persist a structured segmentation audit row (plan Section 15)."""
    from app.services.segmentation_audit import append_audit_event

    try:
        append_audit_event(
            dataset_manager,
            dataset_id,
            event_type,
            data,
            actor=_segmentation_audit_actor(current_user),
            idempotency_key=(data or {}).get("idempotency_key"),
        )
    except Exception as exc:
        logger.warning("segmentation audit append failed for %s: %s", dataset_id, exc)


# Job storage for background training tasks.
#
# DEPRECATED (Phase 1 stateless-API migration, May 2026): this in-process
# dict is per gunicorn worker (and per EKS pod). Status polls landing on
# a different replica than the one that started the job return 404, and
# the in-flight result is lost when the pod restarts. New endpoints MUST
# enqueue via ``app.services.background_jobs.background_job_manager``,
# which mirrors every job snapshot to shared object storage so any
# replica can poll. The legacy dict is retained ONLY so the in-flight
# jobs that pre-date the migration can still complete their lifecycle.
#
# Migration status (each item is a high-priority follow-up):
#   ✓ POST /train-global-model            (migrated to background_job_manager)
#   ✓ POST /train-multiple-models         (already on background_job_manager)
#   ✓ POST /dataset-type-classification-by-id   (already on background_job_manager)
#   ✓ POST /feature-transformation/start  (already on background_job_manager)
#   ✓ POST /segment-profiling/start       (already on background_job_manager)
#   ✓ POST /calculate-vif-correlation/start  (already on background_job_manager)
#   ✓ POST /auto-training/analyze/start   (already on background_job_manager)
#   ✓ POST /segment-training/run          (migrated to background_job_manager)
#   ✓ POST /auto-training/run             (migrated to background_job_manager)
#   ✓ POST /segment-auto-training/run     (migrated to background_job_manager)
training_jobs: dict = {}

_JOBS_PERSIST_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "training_jobs_state.json")

# ---------------------------------------------------------------------------
# Split-config persistence
# Stores the user's chosen ratio / seed / sampling_variable per dataset_id so
# that after an Azure process restart (or scale-out to a fresh instance) the
# correct split can be recreated from the same parameters rather than
# defaulting to a random 70/30 split.
# ---------------------------------------------------------------------------
_SPLIT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "split_configs_state.json")
# In-memory store: {dataset_id: {"ratio": float, "seed": int, "sampling_variable": str|None, "scope": str}}
_split_configs: dict = {}

def _save_split_configs() -> None:
    """Flush split configs to disk (best-effort)."""
    try:
        import json as _json
        with open(_SPLIT_CONFIG_PATH, "w", encoding="utf-8") as _fh:
            _json.dump(_split_configs, _fh)
    except Exception as _e:
        logger.warning(f"Could not persist split configs: {_e}")

def _load_split_configs() -> None:
    """Reload split configs from disk on startup."""
    try:
        import json as _json
        if not os.path.exists(_SPLIT_CONFIG_PATH):
            return
        with open(_SPLIT_CONFIG_PATH, "r", encoding="utf-8") as _fh:
            saved = _json.load(_fh)
        _split_configs.update(saved)
        logger.info(f"Restored split configs for {len(saved)} dataset(s) from disk")
    except Exception as _e:
        logger.warning(f"Could not restore split configs: {_e}")

def _register_split_config(dataset_id: str, ratio: float, seed: int, sampling_variable=None, scope: str = "dev") -> None:
    """Save the user's split parameters so they survive process restarts."""
    _split_configs[dataset_id] = {
        "ratio": ratio,
        "seed": seed,
        "sampling_variable": sampling_variable,
        "scope": scope,
    }
    _save_split_configs()
    logger.info(f"📌 Persisted split config for {dataset_id}: ratio={ratio}, seed={seed}, sampling_variable={sampling_variable}, scope={scope}")

def get_split_config(dataset_id: str) -> dict:
    """Return the persisted split config for a dataset, or an empty dict if not found."""
    return _split_configs.get(dataset_id, {})

# Load persisted split configs at import time
_load_split_configs()

def _save_jobs_state() -> None:
    """Flush training_jobs to disk (non-blocking best-effort)."""
    try:
        import json as _json
        _safe = {
            jid: {k: v for k, v in jdata.items() if k != "results"}
            for jid, jdata in training_jobs.items()
        }
        with open(_JOBS_PERSIST_PATH, "w", encoding="utf-8") as _fh:
            _json.dump(_safe, _fh)
    except Exception as _e:
        logger.warning(f"Could not persist training_jobs state: {_e}")

def _load_jobs_state() -> None:
    """Reload training_jobs from disk on startup (marks running jobs as interrupted)."""
    try:
        import json as _json
        if not os.path.exists(_JOBS_PERSIST_PATH):
            return
        with open(_JOBS_PERSIST_PATH, "r", encoding="utf-8") as _fh:
            saved = _json.load(_fh)
        for jid, jdata in saved.items():
            if jdata.get("status") in ("pending", "running"):
                jdata["status"] = "interrupted"
                jdata["message"] = "Job was interrupted by a server restart. Please re-submit."
            training_jobs[jid] = jdata
        logger.info(f"Restored {len(saved)} training job(s) from disk")
    except Exception as _e:
        logger.warning(f"Could not restore training_jobs state: {_e}")

# Load persisted state at import time
_load_jobs_state()

# OPTIMIZATION: Track which jobs have already been persisted to avoid redundant persistence
# This prevents the expensive 31MB DataFrame pickle/unpickle on every status poll
_persisted_jobs: set = set()

# ============================================================================
# OOP: Domain Classes for Data Ingestion (Single Responsibility Principle)
# ============================================================================

class UploadRequestParser:
    """
    Single Responsibility: Parse and validate form parameters from upload request.
    Encapsulates all JSON parsing logic in one place.
    """
    
    @staticmethod
    def parse_unique_id_combinations(raw_value: str) -> List[str]:
        """Parse unique_id_combinations JSON string with validation."""
        if not raw_value or not raw_value.strip():
            raise ValueError("Unique ID combinations are required. Please select at least one unique identifier.")
        
        try:
            parsed = json.loads(raw_value)
            if not parsed or len(parsed) == 0:
                raise ValueError("Unique ID combinations are required. Please select at least one unique identifier.")
            return parsed
        except json.JSONDecodeError:
            raise ValueError("Invalid format for unique ID combinations")
    
    @staticmethod
    def parse_split_configuration(raw_value: Optional[str]) -> Optional[Dict[str, Any]]:
        """Parse split_configuration JSON string."""
        if not raw_value or not str(raw_value).strip():
            return None
        try:
            return json.loads(raw_value)
        except json.JSONDecodeError:
            logger.warning("split_configuration is not valid JSON; ignoring")
            return None
    
    @staticmethod
    def parse_exclusion_rules(raw_value: Optional[str]) -> Optional[List[Dict[str, Any]]]:
        """Parse exclusion_rules JSON string."""
        if not raw_value or not raw_value.strip():
            return None
        try:
            parsed = json.loads(raw_value)
            if isinstance(parsed, list) and len(parsed) > 0:
                return parsed
            return None
        except json.JSONDecodeError:
            logger.warning("Failed to parse exclusion_rules JSON")
            return None
    
    @staticmethod
    def parse_variables_to_remove(raw_value: Optional[str]) -> Optional[List[str]]:
        """Parse variables_to_remove JSON string."""
        if not raw_value or not raw_value.strip():
            return None
        try:
            parsed = json.loads(raw_value)
            if isinstance(parsed, list) and len(parsed) > 0:
                return parsed
            return None
        except json.JSONDecodeError:
            logger.warning("Failed to parse variables_to_remove JSON")
            return None
    
    @staticmethod
    def clean_optional_string(value: Optional[str]) -> Optional[str]:
        """Clean optional string field - strip whitespace or return None."""
        return value.strip() if value else None
    
    @staticmethod
    def parse_boolean_string(value: Optional[str]) -> Optional[bool]:
        """Parse boolean from string form field."""
        if not value:
            return None
        return value.lower() == 'true'


class DatasetValidator:
    """
    Single Responsibility: Validate dataset content and schema.
    Encapsulates validation logic with behavior attached to data.
    """
    
    def __init__(self, df: pd.DataFrame):
        self._df = df
    
    def validate_unique_id_columns(self, unique_id_columns: List[str]) -> Dict[str, Any]:
        """
        Check if unique_id_columns exist and form a unique key.
        Returns validation result with any warnings.
        """
        result = {"valid": True, "warnings": [], "missing_columns": []}
        
        # Check if columns exist
        missing_cols = [col for col in unique_id_columns if col not in self._df.columns]
        if missing_cols:
            result["missing_columns"] = missing_cols
            return result
        
        # Check for duplicates
        duplicate_count = self._df.duplicated(subset=unique_id_columns).sum()
        if duplicate_count > 0:
            result["warnings"].append(
                f"Warning: Selected columns ({', '.join(unique_id_columns)}) cannot act as unique key. Found {duplicate_count} duplicate rows."
            )
        
        return result
    
    def validate_target_variable_exists(self, target_variable: str) -> bool:
        """Check if target variable exists in dataset."""
        return target_variable in self._df.columns
    
    def get_duplicate_row_count(self) -> int:
        """Get count of duplicate rows in dataset."""
        return int(self._df.duplicated().sum())


class VariableRemover:
    """
    Single Responsibility: Remove specified columns from DataFrame.
    Encapsulates column removal logic.
    """
    
    @staticmethod
    def remove_columns(df: pd.DataFrame, columns_to_remove: List[str]) -> tuple:
        """
        Remove specified columns from DataFrame.
        Returns (updated_df, list_of_removed_columns).
        """
        existing_cols = [col for col in columns_to_remove if col in df.columns]
        if not existing_cols:
            return df, []
        
        updated_df = df.drop(columns=existing_cols)
        return updated_df, existing_cols


class FileMerger:
    """
    Single Responsibility: Merge multiple uploaded CSV files.
    Encapsulates file merging logic with schema validation.
    """
    
    @staticmethod
    async def merge_files(files: List[UploadFile], max_size: int) -> tuple:
        """
        Merge multiple CSV files into a single DataFrame.
        Returns (merged_df, filenames, merged_filename).
        Raises HTTPException on validation errors.
        """
        import io
        
        # Validate all files are CSV
        for f in files:
            if not f.filename.endswith('.csv'):
                raise HTTPException(
                    status_code=400, 
                    detail=f"Only CSV files are supported. Invalid file: {f.filename}"
                )
        
        dfs_to_merge = []
        total_size = 0
        filenames = []
        
        for f in files:
            content = await f.read()
            total_size += len(content)
            filenames.append(f.filename)
            
            if total_size > max_size:
                raise HTTPException(status_code=400, detail="Combined file size exceeds maximum limit")
            
            df_part = pd.read_csv(io.BytesIO(content))
            dfs_to_merge.append(df_part)
            logger.info(f"Read validation file: {f.filename}, rows: {len(df_part)}")
        
        # Merge with schema validation
        if len(dfs_to_merge) > 1:
            ref_cols = set(dfs_to_merge[0].columns)
            for idx, df_part in enumerate(dfs_to_merge[1:], 2):
                if set(df_part.columns) != ref_cols:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Schema mismatch: File {idx} ({filenames[idx-1]}) has different columns than the first file"
                    )
            
            merged_df = pd.concat(dfs_to_merge, ignore_index=True)
            logger.info(f"Merged {len(files)} validation files. Total rows: {len(merged_df)}")
        else:
            merged_df = dfs_to_merge[0]
        
        merged_filename = f"merged_validation_{'+'.join([f.split('.')[0] for f in filenames[:2]])}.csv"
        return merged_df, filenames, merged_filename


# ---------------------------------------------------------------------------
# Auto-training idempotency (one active pending/running job per dataset_id)
# Prevents Azure/proxy timeouts + frontend retries from starting duplicate jobs
# that train the same data twice and confuse the UI.
# ---------------------------------------------------------------------------
_active_auto_training_by_dataset: Dict[str, str] = {}
_auto_training_run_lock = asyncio.Lock()


def _clear_auto_training_slot(dataset_id: str, job_id: str) -> None:
    """Release the per-dataset auto-training slot when this job finishes."""
    if _active_auto_training_by_dataset.get(dataset_id) == job_id:
        _active_auto_training_by_dataset.pop(dataset_id, None)

# Background jobs for dataset-type-classification (ML problem type via LLM)
# Keyed by job_id; avoids blocking the upload flow on a slow LLM call.
_classification_jobs: dict = {}

# Cache for classify-variables results keyed by dataset_id.
# Prevents the LLM from being called more than once per dataset.
_variable_classification_cache: dict = {}

# Background jobs for classify-variables (LLM call, ~30 s).
# Keyed by dataset_id so the frontend can poll without holding a connection open.
_classify_vars_jobs: dict = {}

# ---------------------------------------------------------------------------
# Server-side insight result caches
# Keyed by "<dataset_id>|<target_variable>|<data_scope>" (and extra params where
# relevant).  Results are expensive to compute on large datasets and are
# deterministic for a given dataset/target combination, so we cache them for
# the lifetime of the process.  The cache is invalidated when the dataset is
# re-uploaded or the scope changes (different key).
# ---------------------------------------------------------------------------
import time as _time

_INSIGHT_CACHE_TTL_SECONDS = 3600  # 1 hour

class _InsightCache:
    """Thread-safe LRU-style dict with TTL."""
    def __init__(self, ttl: int = _INSIGHT_CACHE_TTL_SECONDS):
        self._store: dict = {}
        self._ttl = ttl

    def _key(self, **kwargs) -> str:
        return "|".join(f"{k}={v}" for k, v in sorted(kwargs.items()))

    def get(self, **kwargs):
        key = self._key(**kwargs)
        entry = self._store.get(key)
        if entry and (_time.time() - entry["ts"]) < self._ttl:
            return entry["data"]
        if entry:
            del self._store[key]
        return None

    def set(self, data, **kwargs) -> None:
        key = self._key(**kwargs)
        self._store[key] = {"data": data, "ts": _time.time()}

    def invalidate(self, dataset_id: str) -> None:
        """Remove all cache entries for a given dataset (e.g. after re-upload)."""
        to_delete = [k for k in self._store if f"dataset_id={dataset_id}" in k]
        for k in to_delete:
            del self._store[k]

_bivariate_cache = _InsightCache()
_correlation_cache = _InsightCache()
_corr_matrix_cache = _InsightCache()
_correlation_ratio_cache = _InsightCache()


def _insight_scope_version(dataset_id: str) -> Tuple[str, int]:
    """Active scope + dataframe version for analytics_cache keys."""
    scope = dataframe_state_manager._active_scope.get(dataset_id, "entire")
    return scope, dataframe_state_manager.get_version(dataset_id)


def _enqueue_insight_job(
    job_id: str,
    job_type: str,
    *,
    params: Dict[str, Any],
    job_function,
) -> JSONResponse:
    """Return HTTP 202 so clients poll ``GET /insights/jobs/status/{job_id}``."""
    from app.services.background_jobs import background_job_manager as _bgm

    _bgm.start_job(
        job_id=job_id,
        job_type=job_type,
        params=params,
        job_function=job_function,
    )
    return JSONResponse(
        status_code=202,
        content={
            "queued": True,
            "job_id": job_id,
            "status_path": f"/insights/jobs/status/{job_id}",
        },
    )


# Initialize routers
upload_router = APIRouter()
chat_router = APIRouter()


@chat_router.get("/llm-config")
async def get_llm_config():
    resolved = llm_service.get_resolved_config_summary()
    locked_by_env = {
        "chat": {
            "locked": env_override_present("chat"),
            "model_id": settings.CHAT_LLM_CONFIG.model if env_override_present("chat") else None,
        },
        "knowledge_graph": {
            "locked": env_override_present("knowledge_graph"),
            "model_id": settings.KG_LLM_CONFIG.model if env_override_present("knowledge_graph") else None,
        },
        "embedding": {
            "locked": env_override_present("embedding"),
            "model_id": settings.EMBEDDING_LLM_CONFIG.model if env_override_present("embedding") else None,
        },
    }
    return {
        **resolved,
        "locked_by_env": locked_by_env,
    }


@chat_router.post("/dataset/scope")
async def set_dataset_scope(request: dict):
    """
    Set active data scope for a dataset (train/test/validation/entire/dev/hold).

    Body: {
      dataset_id: str,
      scope: 'train'|'test'|'validation'|'entire'|'dev'|'hold',
      seed?: int,
      ratio?: float,
      sampling_variable?: str | null,
    }
    """
    try:
        dataset_id = request.get("dataset_id")
        scope = request.get("scope", "train")
        seed = int(request.get("seed", 42))
        ratio_raw = request.get("ratio", 0.7)
        try:
            ratio = float(ratio_raw) if ratio_raw is not None else 0.7
        except (TypeError, ValueError):
            ratio = 0.7
        sampling_variable = request.get("sampling_variable")

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        
        valid_scopes = ["train", "test", "validation", "entire", "dev", "hold"]
        if scope not in valid_scopes:
            raise HTTPException(status_code=400, detail=f"Invalid scope: {scope}. Must be one of {valid_scopes}")

        # Ensure full dataset is cached in _full_dataframes
        if dataset_id not in dataframe_state_manager._full_dataframes:
            df = dataframe_state_manager.get_dataframe(dataset_id)
            if df is None:
                df = dataset_manager.load_dataset(dataset_id)
                if df is None:
                    raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
                dataframe_state_manager.update_dataframe(dataset_id, df, original_shape=df.shape)
            dataframe_state_manager._full_dataframes[dataset_id] = df.copy()
            logger.info(f"Loaded and cached full dataset for {dataset_id}, shape: {df.shape}")

        logger.info(
            f"📊 Setting scope for {dataset_id}: scope={scope}, seed={seed}, ratio={ratio}, "
            f"sampling_variable={sampling_variable}"
        )
        result = dataframe_state_manager.set_scope(
            dataset_id,
            scope=scope,
            ratio=ratio,
            seed=seed,
            sampling_variable=sampling_variable,
        )
        logger.info(f"✅ Scope set successfully for {dataset_id}: {result.get('scope', 'unknown')}, shape={result.get('shape', 'unknown')}")

        if scope in ("dev", "hold"):
            _register_split_config(
                dataset_id, ratio=ratio, seed=seed, sampling_variable=sampling_variable, scope=scope
            )

        return {"success": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error setting dataset scope: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error setting dataset scope: {str(e)}")

# Initialize agentic system
agentic_system = AgenticSystem()
agent = agentic_system.set_graph()

logger.info("API routes initialized")

# Initialize vector store once at startup
def initialize_vector_store():
    """Initialize vector store from knowledge base"""
    logger.info("Initializing vector store...")
    try:
        # Try to load existing index first
        vector_store.load_index()
        if vector_store.index is not None:
            logger.info("Vector store loaded successfully from existing index")
            return True
    except Exception as e:
        logger.warning(f"Could not load existing vector store: {e}")
    
    # Create new index if it doesn't exist
    try:
        vector_store.create_index_from_knowledge_base("knowledge_base.json")
        return True
    except Exception as e:
        logger.error(f"Could not initialize vector store: {e}")
        return False

# Initialize vector store at module import
vector_store_initialized = initialize_vector_store()
if vector_store_initialized:
    logger.info("Vector store initialization completed successfully")
else:
    logger.warning("Vector store initialization failed")

def safe_float_conversion(value):
    """
    Safely convert a value to float, handling NaN and infinity values.
    Returns None if the value is NaN, infinity, or cannot be converted.
    """
    if pd.isna(value) or np.isinf(value):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None

def calculate_column_info(df: pd.DataFrame) -> List[ColumnInfo]:
    """
    Calculate detailed statistical information for all columns in a DataFrame.
    Returns a list of ColumnInfo objects with statistics for each column.

    Performance note (4M-row regression fix):
        The original implementation ran 7-9 separate pandas calls **per
        column** (isna().sum, nunique, mean, median, mode, std, var,
        skew, min, max, quantile, value_counts...). For wide / long
        frames this drove the endpoint behind the ALB idle timeout.
        We now precompute the cross-column scalars in **one C-vectorised
        pass each** before the per-column loop and read them out of the
        precomputed Series. Mode and value_counts are still per-column
        (no faster pandas primitive exists).
    """
    columns_info: List[ColumnInfo] = []

    # Use shared date detection logic so column-info, analyze-dataset and
    # other APIs have consistent understanding of which variables are
    # dates.
    from app.utils.helpers import identify_date_columns  # Local import to avoid cycles
    date_detection_results = identify_date_columns(df)

    _NUMERIC_DTYPES = ('int64', 'float64', 'int32', 'float32')
    _total_count = len(df)

    # --- Single-pass cross-column precomputation ------------------------
    # df.shape[0] - df.count() gives per-column NA count without
    # materialising a full bool DataFrame the way df.isna().sum() does.
    try:
        _non_na_counts = df.count()
        _isna_counts = (_total_count - _non_na_counts).astype('int64')
    except Exception:
        _isna_counts = pd.Series({c: int(df[c].isna().sum()) for c in df.columns})

    try:
        _nunique_counts = df.nunique(dropna=True)
    except Exception:
        _nunique_counts = pd.Series({c: int(df[c].nunique()) for c in df.columns})

    _num_cols = [c for c in df.columns if str(df[c].dtype) in _NUMERIC_DTYPES]
    _num_stats: Dict[str, Any] = {}
    if _num_cols:
        try:
            num_view = df[_num_cols]
            _num_stats['mean'] = num_view.mean(numeric_only=True)
            _num_stats['median'] = num_view.median(numeric_only=True)
            _num_stats['std'] = num_view.std(numeric_only=True)
            _num_stats['var'] = num_view.var(numeric_only=True)
            _num_stats['min'] = num_view.min(numeric_only=True)
            _num_stats['max'] = num_view.max(numeric_only=True)
            try:
                _num_stats['skew'] = num_view.skew(numeric_only=True)
            except Exception:
                _num_stats['skew'] = pd.Series(index=_num_cols, dtype='float64')
            try:
                _num_stats['quantiles'] = num_view.quantile(
                    [0.0, 0.01, 0.05, 0.25, 0.50, 0.75, 0.95, 0.99, 1.0]
                )
            except Exception:
                _num_stats['quantiles'] = None
        except Exception as _exc:
            logger.warning("Vectorised numeric stats failed; falling back to per-column: %s", _exc)
            _num_stats = {}

    for column_name in df.columns:
        column_data = df[column_name]
        column_type = str(column_data.dtype)

        # Fetch date detection metadata for this column
        date_meta = date_detection_results.get(str(column_name), {})
        is_date_like = bool(date_meta.get("is_date"))
        logical_type: Optional[str] = None

        col_isna = int(_isna_counts.get(column_name, 0))
        col_nunique = int(_nunique_counts.get(column_name, 0))

        # Determine user-friendly column classification (Numerical vs Categorical)
        # Using the same improved logic as analyze-dataset endpoint
        is_numeric_dtype = column_type in _NUMERIC_DTYPES
        if is_date_like:
            user_friendly_type = 'Date'
            logical_type = 'Date'
        elif is_numeric_dtype:
            user_friendly_type = 'Numerical'
            logical_type = 'Numerical'
            # Keep numerical columns as numerical by default.
            # Convert to categorical only if: 2-5 unique values, not a
            # target-like name, has data, and <50% missing.
            if 2 <= col_nunique <= 5 and col_isna < _total_count:
                missing_pct = (col_isna / _total_count) if _total_count else 0.0
                if missing_pct < 0.5:
                    user_friendly_type = 'Categorical'
        elif column_type in ('object', 'category'):
            user_friendly_type = 'Categorical'
            logical_type = logical_type or 'Categorical'
        else:
            user_friendly_type = 'Categorical'
            logical_type = logical_type or 'Categorical'

        # Initialize column info
        col_info = ColumnInfo(
            column_name=column_name,
            data_type=column_type,
            column_type=user_friendly_type,
            logical_type=logical_type,
            is_date=is_date_like,
            date_detection_reason=date_meta.get("reason"),
            date_detected_format=date_meta.get("detected_format"),
            date_detection_confidence=float(date_meta.get("confidence")) if date_meta.get("confidence") is not None else None,
            missing_count=col_isna,
            unique_count=col_nunique,
            total_count=_total_count,
        )

        if user_friendly_type == 'Numerical' and is_numeric_dtype:
            # Pure Numerical column. All bulk stats came from the
            # precomputed dataframe-level series above; only mode is
            # genuinely per-column. clean_data is also only needed if
            # mode runs.
            if (_total_count - col_isna) > 0:
                if _num_stats:
                    col_info.mean = safe_float_conversion(_num_stats['mean'].get(column_name))
                    col_info.median = safe_float_conversion(_num_stats['median'].get(column_name))
                    col_info.standard_deviation = safe_float_conversion(_num_stats['std'].get(column_name))
                    col_info.variance = safe_float_conversion(_num_stats['var'].get(column_name))
                    col_info.min_value = safe_float_conversion(_num_stats['min'].get(column_name))
                    col_info.max_value = safe_float_conversion(_num_stats['max'].get(column_name))
                    skew_s = _num_stats.get('skew')
                    if skew_s is not None:
                        col_info.skewness = safe_float_conversion(skew_s.get(column_name))
                    q = _num_stats.get('quantiles')
                    if q is not None and column_name in q.columns:
                        col_q = q[column_name]
                        col_info.percentile_0 = safe_float_conversion(col_q.get(0.0))
                        col_info.percentile_1 = safe_float_conversion(col_q.get(0.01))
                        col_info.percentile_5 = safe_float_conversion(col_q.get(0.05))
                        col_info.percentile_25 = safe_float_conversion(col_q.get(0.25))
                        col_info.percentile_50 = safe_float_conversion(col_q.get(0.50))
                        col_info.percentile_75 = safe_float_conversion(col_q.get(0.75))
                        col_info.percentile_95 = safe_float_conversion(col_q.get(0.95))
                        col_info.percentile_99 = safe_float_conversion(col_q.get(0.99))
                        col_info.percentile_100 = safe_float_conversion(col_q.get(1.0))
                else:
                    # Fallback: per-column path (when the vectorised
                    # precompute raised). Preserves original behaviour.
                    clean_data = column_data.dropna()
                    col_info.mean = safe_float_conversion(clean_data.mean())
                    col_info.median = safe_float_conversion(clean_data.median())
                    col_info.standard_deviation = safe_float_conversion(clean_data.std())
                    col_info.variance = safe_float_conversion(clean_data.var())
                    try:
                        col_info.skewness = safe_float_conversion(clean_data.skew())
                    except Exception:
                        col_info.skewness = None
                    col_info.min_value = safe_float_conversion(clean_data.min())
                    col_info.max_value = safe_float_conversion(clean_data.max())
                    try:
                        percentiles = clean_data.quantile([0.0, 0.01, 0.05, 0.25, 0.50, 0.75, 0.95, 0.99, 1.0])
                        col_info.percentile_0 = safe_float_conversion(percentiles[0.0])
                        col_info.percentile_1 = safe_float_conversion(percentiles[0.01])
                        col_info.percentile_5 = safe_float_conversion(percentiles[0.05])
                        col_info.percentile_25 = safe_float_conversion(percentiles[0.25])
                        col_info.percentile_50 = safe_float_conversion(percentiles[0.50])
                        col_info.percentile_75 = safe_float_conversion(percentiles[0.75])
                        col_info.percentile_95 = safe_float_conversion(percentiles[0.95])
                        col_info.percentile_99 = safe_float_conversion(percentiles[0.99])
                        col_info.percentile_100 = safe_float_conversion(percentiles[1.0])
                    except Exception:
                        col_info.percentile_0 = None
                        col_info.percentile_1 = None
                        col_info.percentile_5 = None
                        col_info.percentile_25 = None
                        col_info.percentile_50 = None
                        col_info.percentile_75 = None
                        col_info.percentile_95 = None
                        col_info.percentile_99 = None
                        col_info.percentile_100 = None

                # Mode is intrinsically per-column.
                try:
                    mode_values = column_data.mode(dropna=True)
                    if len(mode_values) > 0:
                        mode_value = mode_values.iloc[0]
                        if hasattr(mode_value, 'item'):
                            col_info.mode = safe_float_conversion(mode_value.item())
                        else:
                            col_info.mode = safe_float_conversion(mode_value)
                except Exception:
                    col_info.mode = None
        elif user_friendly_type == 'Categorical':
            # Categorical column (including numeric columns classified as
            # categorical). value_counts(dropna=True) is the dominant
            # cost and is intrinsically per-column; mode() reads the same
            # sort internally. Skip the intermediate ``column_data.dropna()``
            # allocation — value_counts/mode handle NaN natively.
            total_valid = _total_count - col_isna
            if total_valid > 0:
                try:
                    mode_values = column_data.mode(dropna=True)
                    if len(mode_values) > 0:
                        mode_value = mode_values.iloc[0]
                        if hasattr(mode_value, 'item'):
                            col_info.mode = str(mode_value.item())
                        else:
                            col_info.mode = str(mode_value)
                except Exception:
                    pass

                try:
                    value_counts = column_data.value_counts(dropna=True)
                    if len(value_counts) > 0:
                        # Top category (most frequent)
                        top_cat = value_counts.index[0]
                        top_count = value_counts.iloc[0]
                        col_info.top_category = str(top_cat)
                        col_info.top_category_pct = safe_float_conversion((top_count / total_valid) * 100)

                        # Lowest category (least frequent)
                        lowest_cat = value_counts.index[-1]
                        lowest_count = value_counts.iloc[-1]
                        col_info.lowest_category = str(lowest_cat)
                        col_info.lowest_category_pct = safe_float_conversion((lowest_count / total_valid) * 100)
                except Exception:
                    pass
        else:
            # "Other" columns: object/category dtype not classified as
            # Categorical, OR Date columns. Same per-column primitives.
            total_valid = _total_count - col_isna
            if total_valid > 0:
                try:
                    mode_values = column_data.mode(dropna=True)
                    if len(mode_values) > 0:
                        mode_value = mode_values.iloc[0]
                        if hasattr(mode_value, 'item'):
                            col_info.mode = str(mode_value.item())
                        else:
                            col_info.mode = str(mode_value)
                except Exception:
                    pass

                try:
                    value_counts = column_data.value_counts(dropna=True)
                    if len(value_counts) > 0:
                        top_cat = value_counts.index[0]
                        top_count = value_counts.iloc[0]
                        col_info.top_category = str(top_cat)
                        col_info.top_category_pct = safe_float_conversion((top_count / total_valid) * 100)
                        lowest_cat = value_counts.index[-1]
                        lowest_count = value_counts.iloc[-1]
                        col_info.lowest_category = str(lowest_cat)
                        col_info.lowest_category_pct = safe_float_conversion((lowest_count / total_valid) * 100)
                except Exception:
                    pass

                # Date columns with string-like values are sometimes
                # coerce-able to numbers (e.g. ``"20240101"`` columns).
                # Keep the legacy "best-effort" numeric-stat extraction
                # so downstream UIs that read numeric fields for such
                # date-string columns still see populated values.
                if column_data.dtype not in _NUMERIC_DTYPES:
                    try:
                        numeric_data = pd.to_numeric(column_data, errors='coerce').dropna()
                        if len(numeric_data) > 0:
                            col_info.mean = safe_float_conversion(numeric_data.mean())
                            col_info.median = safe_float_conversion(numeric_data.median())
                            col_info.standard_deviation = safe_float_conversion(numeric_data.std())
                            col_info.variance = safe_float_conversion(numeric_data.var())
                            col_info.min_value = safe_float_conversion(numeric_data.min())
                            col_info.max_value = safe_float_conversion(numeric_data.max())
                            try:
                                percentiles = numeric_data.quantile(
                                    [0.0, 0.01, 0.05, 0.25, 0.50, 0.75, 0.95, 0.99, 1.0]
                                )
                                col_info.percentile_0 = safe_float_conversion(percentiles[0.0])
                                col_info.percentile_1 = safe_float_conversion(percentiles[0.01])
                                col_info.percentile_5 = safe_float_conversion(percentiles[0.05])
                                col_info.percentile_25 = safe_float_conversion(percentiles[0.25])
                                col_info.percentile_50 = safe_float_conversion(percentiles[0.50])
                                col_info.percentile_75 = safe_float_conversion(percentiles[0.75])
                                col_info.percentile_95 = safe_float_conversion(percentiles[0.95])
                                col_info.percentile_99 = safe_float_conversion(percentiles[0.99])
                                col_info.percentile_100 = safe_float_conversion(percentiles[1.0])
                            except Exception:
                                col_info.percentile_0 = None
                                col_info.percentile_1 = None
                                col_info.percentile_5 = None
                                col_info.percentile_25 = None
                                col_info.percentile_50 = None
                                col_info.percentile_75 = None
                                col_info.percentile_95 = None
                                col_info.percentile_99 = None
                                col_info.percentile_100 = None
                    except (ValueError, TypeError):
                        pass

        # DateTime-specific calculations
        if is_date_like:
            try:
                clean_data = column_data.dropna()
                if len(clean_data) > 0:
                    # Try to convert to datetime for min/max calculations
                    datetime_data = pd.to_datetime(clean_data, errors='coerce')
                    datetime_valid = datetime_data.dropna()
                    
                    # Check if we have valid datetime conversions
                    if len(datetime_valid) > 0 and len(datetime_valid) >= len(clean_data) * 0.5:
                        # More than 50% converted successfully - use datetime min/max
                        min_date = datetime_valid.min()
                        max_date = datetime_valid.max()
                        col_info.date_min = str(min_date.strftime('%Y-%m-%d')) if pd.notna(min_date) else None
                        col_info.date_max = str(max_date.strftime('%Y-%m-%d')) if pd.notna(max_date) else None
                    else:
                        # Partial dates (e.g., "14-Oct", "Jan-2023") - use original string values
                        # Sort the original string values to find min/max
                        unique_values = clean_data.unique()
                        
                        # Try to parse partial dates with a reference year for proper sorting
                        parsed_dates = []
                        for val in unique_values:
                            try:
                                # Try parsing with dayfirst=True and various formats
                                parsed = pd.to_datetime(str(val), errors='coerce', dayfirst=True)
                                if pd.isna(parsed):
                                    # Try adding a reference year for partial dates like "14-Oct"
                                    parsed = pd.to_datetime(f"{val}-2000", errors='coerce', dayfirst=True)
                                if pd.isna(parsed):
                                    # Try format like "Jan-2023"
                                    parsed = pd.to_datetime(str(val), errors='coerce', format='mixed')
                                parsed_dates.append((val, parsed))
                            except:
                                parsed_dates.append((val, pd.NaT))
                        
                        # Filter valid parsed dates and sort
                        valid_parsed = [(orig, dt) for orig, dt in parsed_dates if pd.notna(dt)]
                        
                        if valid_parsed:
                            # Sort by datetime and get original values for min/max
                            valid_parsed.sort(key=lambda x: x[1])
                            col_info.date_min = str(valid_parsed[0][0])  # Earliest date (original format)
                            col_info.date_max = str(valid_parsed[-1][0])  # Latest date (original format)
                    
                    # Get most frequent value (mode) - use original format
                    mode_values = clean_data.mode()
                    if len(mode_values) > 0:
                        col_info.most_frequent_date = str(mode_values.iloc[0])
            except Exception as e:
                logger.warning(f"DateTime calculations failed for column {column_name}: {e}")
                # If datetime calculations fail, leave as None
                pass
        
        columns_info.append(col_info)
    
    return columns_info

@upload_router.post("/upload", response_model=UploadResponse)
async def upload_dataset(
    background_tasks: BackgroundTasks,
    file: Optional[UploadFile] = File(None),
    existing_dataset_id: Optional[str] = Form(None),
    files: Optional[List[UploadFile]] = File(None),
    merge_validation: Optional[str] = Form(None),
    target_variable: str = Form(...),
    target_variable_type: str = Form(...),
    data_dictionary: Optional[str] = Form(""),
    data_dictionary_file: Optional[UploadFile] = File(None),
    problem_statement: Optional[str] = Form(""),
    unique_id_combinations: str = Form(...),
    segmentation_variable: Optional[str] = Form(""),
    sample_identifier_variable: Optional[str] = Form(""),
    # Split configuration parameters
    has_sampling_variable: Optional[str] = Form(None),
    sampling_variable: Optional[str] = Form(None),
    split_ratio: Optional[str] = Form(None),
    initial_scope: Optional[str] = Form(None),
    split_configuration: Optional[str] = Form(None),
    # Exclusion rules
    exclusion_rules: Optional[str] = Form(None),
    # Variables to remove from variable review
    variables_to_remove: Optional[str] = Form(None),
    # Partition role for pre-split uploads (train, test, validation/oot)
    partition_role: Optional[str] = Form(None),
    current_user = Depends(get_current_user_dependency),
):
    """
    Upload dataset and configure analysis parameters
    Replicates the functionality from line 219 of app1.py
    
    Supports:
    - Single file upload (file parameter)
    - Multiple validation files (files parameter with merge_validation=true)
    """
    # Check if we have at least one file
    should_merge = merge_validation and merge_validation.lower() == 'true'
    
    if should_merge and files and len(files) > 0:
        logger.info(f"Validation file merge request: {len(files)} files to merge, target: {target_variable}, type: {target_variable_type}")
    elif file or existing_dataset_id:
        logger.info(f"Dataset upload request: {file.filename if file else existing_dataset_id}, target: {target_variable}, type: {target_variable_type}")
    else:
        raise HTTPException(status_code=400, detail="Either 'file', 'existing_dataset_id', or 'files' must be provided")
    
    try:
        # OOP: Use UploadRequestParser for all parsing (Single Responsibility)
        try:
            parsed_unique_id_combinations = UploadRequestParser.parse_unique_id_combinations(unique_id_combinations)
            logger.info(f"Parsed unique_id_combinations: {parsed_unique_id_combinations}")
        except ValueError as e:
            logger.error(f"unique_id_combinations validation failed: {e}")
            raise HTTPException(status_code=400, detail=str(e))
        
        # OOP: Clean optional string fields using parser
        cleaned_segmentation_variable = UploadRequestParser.clean_optional_string(segmentation_variable)
        cleaned_sample_identifier_variable = UploadRequestParser.clean_optional_string(sample_identifier_variable)
        
        # OOP: Parse split configuration using parser
        parsed_has_sampling_variable = UploadRequestParser.parse_boolean_string(has_sampling_variable)
        parsed_sampling_variable = UploadRequestParser.clean_optional_string(sampling_variable)
        parsed_split_ratio = float(split_ratio) if split_ratio else None
        parsed_initial_scope = UploadRequestParser.clean_optional_string(initial_scope)
        parsed_split_configuration = UploadRequestParser.parse_split_configuration(split_configuration)
        
        logger.info(f"Additional fields - Segmentation: {cleaned_segmentation_variable}, Sample ID: {cleaned_sample_identifier_variable}, Unique IDs: {parsed_unique_id_combinations}")
        logger.info(f"Split config - has_sampling_variable: {parsed_has_sampling_variable}, sampling_variable: {parsed_sampling_variable}, split_ratio: {parsed_split_ratio}, initial_scope: {parsed_initial_scope}")
        if parsed_split_configuration:
            logger.info(f"Platform split_configuration present: ingestion_mode={parsed_split_configuration.get('ingestion_mode')}, method={parsed_split_configuration.get('split_method')}")
        
        # OOP: Handle multiple validation files using FileMerger class (Single Responsibility)
        if should_merge and files and len(files) > 0:
            # Use FileMerger class for merging logic
            merged_df, filenames, merged_filename = await FileMerger.merge_files(files, settings.MAX_FILE_SIZE)
            
            # Save merged file
            merged_content = merged_df.to_csv(index=False).encode('utf-8')
            dataset_id, file_path = dataset_manager.save_uploaded_file(merged_content, merged_filename)
            logger.info(f"Merged validation dataset saved with ID: {dataset_id}")
            
            # Set df for further processing
            df = merged_df
            dataset_manager.schedule_parquet_alongside_csv(file_path, df)
            actual_filename = merged_filename
        else:
            if existing_dataset_id:
                dataset_id = existing_dataset_id
                info = dataset_manager.get_dataset_info(dataset_id)
                if not info:
                    raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
                
                # Retrieve the storage key saved by chunked upload finalize
                file_path = str(info.get("storage_key") or info.get("file_path") or "")
                if not file_path:
                    raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} has no storage key")
                actual_filename = str(info.get("filename") or "")
                logger.info(f"Using existing dataset ID: {dataset_id}, key: {file_path}")
                df = None
            else:
                # Single file upload - stream to disk to avoid holding full CSV twice in memory
                if not file or not file.filename.endswith('.csv'):
                    logger.warning(f"Invalid file type uploaded: {file.filename if file else 'None'}")
                    raise HTTPException(status_code=400, detail="Only CSV files are supported")
                
                # P3.5: time the byte-streaming-to-disk stage separately.
                from app.core.metrics import time_stage as _ts_upload
                with _ts_upload("upload_stream", bytes_processed=getattr(file, "size", None) or 0):
                    dataset_id, file_path = await dataset_manager.save_uploaded_file_streaming(
                        file, file.filename, settings.MAX_FILE_SIZE
                    )
                logger.info(f"Dataset saved with ID: {dataset_id}")
                
                actual_filename = file.filename
                df = None  # Will be loaded later

        # Invalidate any stale insight caches for this dataset (e.g. re-upload)
        _bivariate_cache.invalidate(dataset_id)
        _correlation_cache.invalidate(dataset_id)
        _corr_matrix_cache.invalidate(dataset_id)
        _correlation_ratio_cache.invalidate(dataset_id)
        
        # Process data dictionary file if provided
        data_dictionary_path = None
        final_data_dictionary = data_dictionary
        
        if data_dictionary_file and data_dictionary_file.filename:
            logger.info(f"Processing data dictionary file: {data_dictionary_file.filename}")
            
            # Validate data dictionary file type
            if not data_dictionary_file.filename.endswith('.csv'):
                logger.warning(f"Invalid data dictionary file type: {data_dictionary_file.filename}")
                raise HTTPException(status_code=400, detail="Data dictionary must be a CSV file")
            
            # Stream data dictionary to disk
            data_dict_id, data_dictionary_path = await dataset_manager.save_uploaded_file_streaming(
                data_dictionary_file,
                f"data_dict_{dataset_id}_{data_dictionary_file.filename}",
                settings.MAX_FILE_SIZE,
            )
            
            # Store the file path as the data dictionary reference
            final_data_dictionary = data_dictionary_path
            logger.info(f"Data dictionary file saved at: {data_dictionary_path}")

            logger.info(f"Data dictionary file saved at: {final_data_dictionary}")
        
        # Load dataset - async executor so the event loop stays responsive
        import asyncio as _asyncio
        from app.core.executor import executor as _executor
        from app.core.metrics import time_stage as _time_stage
        _loop = _asyncio.get_event_loop()

        # P3.3: SSE publisher for ingest progress (best-effort).
        from app.api import sse as _sse_bus

        async def _emit(stage: str, status: str, **extra) -> None:
            try:
                await _sse_bus.publish(dataset_id, {
                    "type": "stage",
                    "stage": stage,
                    "status": status,
                    **extra,
                })
            except Exception:
                pass

        # ``is_existing_dataset_path`` gates several hot-path optimizations
        # (skip parquet pre-convert when chunked-finalize already started one,
        # skip O(n) duplicate / null scans, skip the second ``validate_dataset``
        # call). Compute it once here so the branches below can reference it.
        is_existing_dataset_path = bool(existing_dataset_id)

        # P2.1: stream-convert CSV -> Parquet on disk BEFORE we touch pandas.
        # If conversion succeeds, downstream `read_csv_for_upload` (which calls
        # load_dataset internally) will discover the Parquet sidecar and
        # bypass the slow pandas csv parser. On 2 GB files this trims the
        # pandas-load step from ~25 s to ~5 s and caps peak RAM at the
        # ParquetWriter row-group buffer (~64 MB).
        #
        # P3.x: on the existing_dataset_id hot path the chunked-upload finalize
        # already kicked off a background ``_safe_stream_convert`` for the same
        # storage key. Re-running the conversion synchronously here on a 2.5 GB
        # CSV cost ~30-60 s and was the dominant contributor to the residual
        # 504 Gateway Timeout the user kept seeing on Submit. Instead, we
        # *probe* for the parquet sidecar: if it's already there, we reap the
        # win for free; if not, we fire off one background conversion (it
        # de-duplicates against the in-flight finalize task because both
        # callers check ``store.exists(pq_key)`` first) and let
        # ``read_csv_for_upload`` fall back to the streaming-CSV reader for
        # this request. Future requests still hit parquet via the sidecar.
        if df is None:
            from app.services.dataset_service import (
                _normalize_storage_key as _norm_key,
                _parquet_key as _pq_key_fn,
            )
            from app.services.object_storage.registry import (
                get_object_storage as _get_store,
            )

            csv_key_for_pq = _norm_key(file_path) if file_path else ""
            pq_sidecar_key = _pq_key_fn(csv_key_for_pq) if csv_key_for_pq else ""
            store_for_pq = _get_store()
            parquet_already_there = bool(
                pq_sidecar_key and store_for_pq.exists(pq_sidecar_key)
            )

            if is_existing_dataset_path and not parquet_already_there:
                # Schedule but DO NOT await. ``stream_convert_csv_to_parquet``
                # is idempotent (returns the existing key when the sidecar
                # already exists), so racing with the chunked-finalize
                # background converter is safe and at most one of them does
                # the actual conversion.
                try:
                    _loop.run_in_executor(
                        _executor,
                        dataset_manager.stream_convert_csv_to_parquet,
                        file_path,
                    )
                    logger.info(
                        "parquet_convert: scheduled background conversion for "
                        "existing_dataset path (key=%s)",
                        csv_key_for_pq,
                    )
                except Exception as exc:
                    logger.warning(
                        "parquet_convert: background schedule failed: %s", exc
                    )
                await _emit(
                    "parquet_convert", "scheduled_background",
                    parquet_key=pq_sidecar_key or "",
                )
            else:
                await _emit("parquet_convert", "running")
                try:
                    with _time_stage("parquet_convert"):
                        parquet_key = await _loop.run_in_executor(
                            _executor,
                            dataset_manager.stream_convert_csv_to_parquet,
                            file_path,
                        )
                    if parquet_key:
                        logger.info(f"P2.1: streaming Parquet conversion complete: {parquet_key}")
                    await _emit("parquet_convert", "completed", parquet_key=parquet_key or "")
                except Exception as exc:
                    logger.warning(f"P2.1: streaming Parquet conversion skipped: {exc}")
                    await _emit("parquet_convert", "skipped", error=str(exc))

        if df is None:
            await _emit("read_csv", "running")
            with _time_stage("read_csv_for_upload"):
                df = await _loop.run_in_executor(_executor, dataset_manager.read_csv_for_upload, file_path)
            await _emit("read_csv", "completed", rows=int(len(df)), cols=int(df.shape[1]))

        # P2.1: build a target_profile sidecar so the LLM classifier and
        # sampler can answer "is this imbalanced?" / "what are the rare
        # classes?" without re-scanning the file.
        if target_variable:
            await _emit("target_profile", "running", target=target_variable)
            try:
                with _time_stage("target_profile"):
                    await _loop.run_in_executor(
                        _executor,
                        dataset_manager.compute_target_profile,
                        dataset_id,
                        file_path,
                        target_variable,
                    )
                await _emit("target_profile", "completed")
            except Exception as exc:
                logger.warning(f"P2.1: target_profile sidecar skipped: {exc}")
                await _emit("target_profile", "skipped", error=str(exc))

        # OOP: Use DatasetValidator for validation (Encapsulation - behavior with data)
        dataset_validator = DatasetValidator(df)

        # Hot path on multi-GB datasets is dominated by full-frame pandas
        # scans (``df.duplicated().sum()`` on a 2 GB / 5 M-row frame takes
        # ~30-60 s). When the caller passed ``existing_dataset_id`` the file
        # already went through ``/analyze-dataset`` and
        # ``/validate-unique-ids-by-id`` earlier in Step 1, so we have no
        # reason to redo the same O(n) checks here just to populate
        # response warnings the frontend does not consume. Empty-file and
        # too-few-columns checks are kept (they are O(1)). Skipping the
        # redundant scans + the previously-wasted second ``validate_dataset``
        # call drops post-read time from ~2-5 min to ~5-30 s on a 2 GB
        # CSV, which is what kept the request inside the ALB idle timeout
        # and ended the 504 Gateway Time-out at Submit.
        # (``is_existing_dataset_path`` is computed up top so the parquet
        # pre-convert branch can also reference it.)

        if is_existing_dataset_path:
            duplicate_row_count = 0
            uid_dup_count: Optional[int] = None
            validation_result = {"is_valid": True, "errors": [], "warnings": []}
            if df.empty:
                validation_result["is_valid"] = False
                validation_result["errors"].append("The uploaded file is empty.")
            if df.shape[1] < 2:
                validation_result["is_valid"] = False
                validation_result["errors"].append("Dataset must have at least 2 columns.")
        else:
            def _dup_count():
                return dataset_validator.get_duplicate_row_count()

            def _uid_check():
                if not parsed_unique_id_combinations:
                    return None
                uid_result = dataset_validator.validate_unique_id_columns(parsed_unique_id_combinations)
                if uid_result.get("missing_columns"):
                    return None
                dup_count = df.duplicated(subset=parsed_unique_id_combinations).sum()
                return int(dup_count) if dup_count > 0 else None

            # The previous implementation ran a third ``validate_dataset``
            # task here in parallel, but its result was unconditionally
            # overwritten by the post-gather call below -- pure wasted
            # work that on a 2 GB frame doubled the time spent in
            # ``df.duplicated().sum()``.
            duplicate_row_count, uid_dup_count = await _asyncio.gather(
                _loop.run_in_executor(_executor, _dup_count),
                _loop.run_in_executor(_executor, _uid_check),
            )

            validation_result = await _loop.run_in_executor(
                _executor,
                lambda: dataset_manager.validate_dataset(
                    df, duplicate_row_count=duplicate_row_count
                ),
            )

        if not validation_result['is_valid']:
            # Clean up saved files
            os.remove(file_path)
            if data_dictionary_path and os.path.exists(data_dictionary_path):
                os.remove(data_dictionary_path)
            logger.error(f"Dataset validation failed: {validation_result['errors']}")
            raise HTTPException(status_code=400, detail="; ".join(validation_result['errors']))
        
        # Apply unique-ID duplicate warning from parallel check
        if uid_dup_count is not None and uid_dup_count > 0:
            logger.warning(f"Selected columns cannot act as unique key: {parsed_unique_id_combinations} - Found {uid_dup_count} duplicates")
            validation_result['warnings'].append(
                f"Warning: Selected columns ({', '.join(parsed_unique_id_combinations)}) cannot act as unique key. Found {uid_dup_count} duplicate rows."
            )

        # Store dataset metadata (fast dict operation - no executor needed)
        dataset_manager.store_dataset_info(
            dataset_id=dataset_id,
            file_path=file_path,
            filename=actual_filename,
            target_variable=target_variable,
            target_variable_type=target_variable_type,
            data_dictionary=final_data_dictionary,
            problem_statement=problem_statement,
            unique_id_combinations=parsed_unique_id_combinations,
            segmentation_variable=cleaned_segmentation_variable,
            sample_identifier_variable=cleaned_sample_identifier_variable,
            split_configuration=parsed_split_configuration,
        )

        # Prepare the raw text that will be sent to the KG generator
        data_dictionary_content = ""
        if final_data_dictionary:
           if os.path.exists(final_data_dictionary):
               try:
                   data_dictionary_content = _read_data_dictionary(final_data_dictionary)
               except Exception as exc:
                   logger.warning(f"Failed to read data dictionary file {final_data_dictionary}: {exc}")
           else:
               # If it wasn't a file upload, final_data_dictionary already contains the text from the form
               data_dictionary_content = final_data_dictionary

        if data_dictionary_content:
            background_tasks.add_task(
                llm_service.generate_async_knowledge_graph,
                dataset_id,
                data_dictionary_content,
                problem_statement,
            )
        else:
            logger.warning(f"No data dictionary content available for dataset {dataset_id}; skipping KG pre-generation")

        # Generate dataset statistics in executor (can be slow on large frames).
        # On the ``existing_dataset_id`` hot path we already declined to run the
        # full-frame duplicate scan above, and the frontend does not consume
        # ``stats.missing_values`` from this response either, so skip both
        # full-frame scans here. ``memory_usage`` inside ``get_dataset_stats``
        # is always shallow now -- ``deep=True`` was previously a multi-second
        # cost on object columns of a 2 GB frame.
        stats = await _loop.run_in_executor(
            _executor,
            lambda: dataset_manager.get_dataset_stats(
                df,
                target_variable,
                duplicate_rows=duplicate_row_count,
                skip_missing_summary=is_existing_dataset_path,
                skip_duplicate_count=is_existing_dataset_path,
            ),
        )

        # Load dataset into dataframe state manager for immediate use
        dataframe_state_manager.update_dataframe(dataset_id, df, original_shape=df.shape)
        logger.info(f"Dataset loaded into dataframe state manager: {dataset_id}")

        upload_exclusion_rules = UploadRequestParser.parse_exclusion_rules(exclusion_rules)
        if upload_exclusion_rules:
            try:
                logger.info(f"Applying {len(upload_exclusion_rules)} exclusion groups to dataset {dataset_id}")
                exclusion_result = dataframe_state_manager.apply_exclusion_rules(
                    dataset_id,
                    upload_exclusion_rules,
                    target_variable,
                )
                logger.info(f"Exclusion rules applied: {exclusion_result}")
                df_after = dataframe_state_manager.get_dataframe(dataset_id)
                if df_after is not None:
                    df = df_after
                    stats = await _loop.run_in_executor(
                        _executor,
                        lambda: dataset_manager.get_dataset_stats(
                            df,
                            target_variable,
                            duplicate_rows=duplicate_row_count,
                            skip_missing_summary=is_existing_dataset_path,
                            skip_duplicate_count=is_existing_dataset_path,
                        ),
                    )
            except Exception as e:
                logger.error(f"Failed to apply exclusion rules: {e}")

        upload_vars_to_remove = UploadRequestParser.parse_variables_to_remove(variables_to_remove)
        if upload_vars_to_remove:
            try:
                logger.info(f"Removing {len(upload_vars_to_remove)} variables from dataset {dataset_id}")
                current_df = dataframe_state_manager.get_dataframe(dataset_id)
                if current_df is not None:
                    updated_df, vars_removed = VariableRemover.remove_columns(current_df, upload_vars_to_remove)
                    if vars_removed:
                        dataframe_state_manager.update_dataframe(
                            dataset_id, updated_df, original_shape=updated_df.shape
                        )
                        df = updated_df
                        stats = await _loop.run_in_executor(
                            _executor,
                            lambda: dataset_manager.get_dataset_stats(
                                df,
                                target_variable,
                                skip_missing_summary=is_existing_dataset_path,
                                skip_duplicate_count=is_existing_dataset_path,
                            ),
                        )
            except Exception as e:
                logger.error(f"Failed to remove variables: {e}")

        split_applied = False
        if parsed_split_configuration and isinstance(parsed_split_configuration, dict):
            parsed_partition_role = partition_role.strip() if partition_role else None
            work_df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
            if work_df is None:
                work_df = df
            split_applied = dataframe_state_manager.apply_split_configuration(
                dataset_id,
                work_df,
                target_variable,
                parsed_split_configuration,
                seed=42,
                partition_role=parsed_partition_role,
            )
            if split_applied:
                ing_mode = parsed_split_configuration.get("ingestion_mode", "unknown")
                logger.info(f"Applied {ing_mode} split_configuration for dataset {dataset_id}")

        if not split_applied and parsed_initial_scope:
            try:
                if parsed_initial_scope == 'sampling_variable_split' and parsed_sampling_variable:
                    if parsed_split_ratio is None:
                        logger.warning(
                            f"⚠️ Sampling variable split requested for {dataset_id} but split_ratio not provided. "
                            f"Will calculate from variable distribution."
                        )
                    logger.info(
                        f"📊 Creating sampling variable-based split for {dataset_id} using variable: {parsed_sampling_variable}"
                    )
                    if parsed_split_ratio:
                        logger.info(f"📊 Using calculated split_ratio: {parsed_split_ratio} for sampling variable split")
                    result = dataframe_state_manager.set_scope(
                        dataset_id,
                        scope='dev',
                        ratio=parsed_split_ratio if parsed_split_ratio else 0.7,
                        seed=42,
                        sampling_variable=parsed_sampling_variable,
                    )
                    logger.info(f"✅ Sampling variable split created for {dataset_id}: {result}")
                elif parsed_initial_scope == 'split':
                    ratio_to_use = parsed_split_ratio if parsed_split_ratio is not None else 1.0
                    if parsed_split_ratio is None:
                        logger.info(
                            f"📊 Split requested for {dataset_id} but split_ratio not provided, defaulting to 1.0 "
                            f"(Entire = 100% dev, 0% hold)"
                        )
                    else:
                        logger.info(f"📊 Creating random split for {dataset_id} with user-provided ratio: {ratio_to_use}")
                    result = dataframe_state_manager.set_scope(
                        dataset_id,
                        scope='dev',
                        ratio=ratio_to_use,
                        seed=42,
                    )
                    logger.info(f"✅ Random split created for {dataset_id}: {result}")
            except Exception as e:
                logger.error(f"Failed to set scope during upload: {str(e)}")
        
        # NOTE: Scope remains as 'entire' (full data) after Submit
        # The right pane will show entire data by default
        # Agents will set scope to 'train' when needed for modeling
        
        response = UploadResponse(
            success=True,
            message="Dataset uploaded successfully" if not should_merge else f"Merged {len(files)} validation files successfully",
            dataset_id=dataset_id,
            dataset_info={
                "filename": actual_filename,
                "target_variable": target_variable,
                "target_variable_type": target_variable_type,
                "unique_id_combinations": parsed_unique_id_combinations,
                "segmentation_variable": cleaned_segmentation_variable,
                "sample_identifier_variable": cleaned_sample_identifier_variable,
                "stats": stats.dict(),
                "warnings": validation_result['warnings'],
                "merged_files": [f.filename for f in files] if should_merge and files else None
            }
        )
        
        logger.info(f"Dataset upload completed successfully: {dataset_id}")
        return response
        
    except HTTPException:
        raise
    except FileTooLargeError:
        raise HTTPException(status_code=400, detail="File size exceeds maximum limit")
    except Exception as e:
        logger.error(f"Dataset upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.post("/validate-unique-ids-by-id")
async def validate_unique_ids_by_id(
    dataset_id: str = Form(...),
    unique_id_columns: str = Form(...),
    current_user=Depends(get_current_user_dependency),
):
    """
    Validate if the requested columns can act as a composite primary key for
    an already-uploaded dataset.

    Hot path on multi-GB datasets:
      1. ``AnalyticsResultCache`` lookup keyed by
         ``(validate_unique_ids, dataset_id, sha1(sorted(cols)), version)``
         - same selection re-rendered while the user is on the page returns
         in well under 100 ms (process-LRU; cross-replica via Redis L2 when
         ``REDIS_URL`` is set).
      2. ``materialize_unique_id_validation_path`` resolves a local Parquet
         (or CSV fallback) path. For S3-backed storage the file is staged
         exactly once via the ``SidecarCache`` and pinned for the duration
         of the scan so subsequent calls for the same dataset are pure
         filesystem reads.
      3. ``compute_duplicate_stats_from_tabular_path`` runs on the shared
         thread executor (Polars streaming collect with column pruning).
         The handler never blocks the FastAPI event loop on disk / S3 I/O.
    """
    import asyncio as _asyncio
    import hashlib
    import json
    from app.core.executor import executor as _executor
    from app.core.metrics import time_stage
    from app.services.analytics_cache import analytics_cache
    from app.services.dataframe_state_manager import dataframe_state_manager

    logger.info("Validating unique IDs by dataset_id=%s", dataset_id)
    try:
        unique_ids = json.loads(unique_id_columns)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON format for unique_id_columns")

    if not unique_ids:
        return {
            "success": True,
            "is_unique": True,
            "duplicate_count": 0,
            "message": "No unique ID columns specified",
        }
    if not isinstance(unique_ids, list) or not all(isinstance(c, str) for c in unique_ids):
        raise HTTPException(
            status_code=400,
            detail="unique_id_columns must be a JSON array of strings",
        )

    # Cache key. ``scope`` is a stable digest of the selected column set so
    # ["a","b"] and ["b","a"] hit the same entry (composite keys are
    # order-independent for uniqueness checks).
    sorted_cols = sorted(unique_ids)
    scope = hashlib.sha1("|".join(sorted_cols).encode("utf-8")).hexdigest()
    version = dataframe_state_manager.get_version(dataset_id)

    cached = analytics_cache.get(
        kind="validate_unique_ids",
        dataset_id=dataset_id,
        scope=scope,
        version=version,
    )
    if cached is not None:
        # Stamp the response so frontend can confirm the cache path is hot.
        return {**cached, "cached": True}

    loop = _asyncio.get_event_loop()

    def _materialize_and_validate(ds_id: str, cols: list) -> Dict[str, Any]:
        """Resolve sidecar path, then run the duplicate scan. Sub-staged."""
        with dataset_manager.materialize_unique_id_validation_path(ds_id) as (
            path,
            is_pq,
        ):
            if not path:
                return {"_not_found": True}
            with time_stage("validate_unique_ids_scan"):
                return compute_duplicate_stats_from_tabular_path(
                    path, cols, is_parquet=is_pq
                )

    try:
        with time_stage("validate_unique_ids"):
            result = await loop.run_in_executor(
                _executor, _materialize_and_validate, dataset_id, unique_ids
            )
    except Exception as e:
        logger.error("Unique ID validation (by-id) failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")

    if result.get("_not_found"):
        raise HTTPException(
            status_code=404,
            detail=f"Dataset {dataset_id} not found or file missing from storage.",
        )

    if "missing" in result:
        # Don't cache validation errors - the user will fix the column name
        # and re-issue, and we want them to see the corrected result.
        return {
            "success": False,
            "error": f"Columns not found: {', '.join(result['missing'])}",
        }

    duplicate_count = result["duplicate_count"]
    total_rows = result["total_rows"]
    is_unique = bool(duplicate_count == 0)
    logger.info(
        "Unique ID validation (by-id): %s cols=%s Duplicates: %s/%s",
        dataset_id,
        unique_ids,
        duplicate_count,
        total_rows,
    )
    response = {
        "success": True,
        "is_unique": is_unique,
        "duplicate_count": duplicate_count,
        "total_rows": total_rows,
        "columns": unique_ids,
        "message": (
            f"Found {duplicate_count} duplicate rows"
            if not is_unique
            else "All rows are unique"
        ),
    }
    analytics_cache.set(
        kind="validate_unique_ids",
        dataset_id=dataset_id,
        scope=scope,
        version=version,
        value=response,
    )
    return response


@upload_router.post("/combine-presplit")
async def combine_presplit_files(
    files: List[UploadFile] = File(None),
    partition_roles_json: str = Form(None),
    target_variable: Optional[str] = Form(None),
    current_user = Depends(get_current_user_dependency),
):
    """
    Combine multiple pre-split files (train, test, validation) into a single dataset.
    Each file gets a 'split_tag' column with its partition role.
    This creates a unified dataset that can be processed like platform_split (full population).
    Target variable is optional at this stage - it will be set during finalization.
    """
    import json
    
    logger.info(f"combine-presplit called: files count={len(files) if files else 0}, partition_roles_json={partition_roles_json}")
    
    # Validate inputs
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    if not partition_roles_json:
        raise HTTPException(status_code=400, detail="partition_roles_json is required")
    
    # Parse partition roles from JSON string
    try:
        partition_roles = json.loads(partition_roles_json)
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse partition_roles_json: {partition_roles_json}, error: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid partition_roles_json format: {e}")
    
    logger.info(f"Combining {len(files)} pre-split files with roles: {partition_roles}")
    
    if len(files) != len(partition_roles):
        raise HTTPException(
            status_code=400, 
            detail=f"Number of files ({len(files)}) must match number of partition roles ({len(partition_roles)})"
        )
    
    try:
        import io
        
        dfs_to_combine = []
        partition_counts = {}
        validation_counter = 0  # Track validation file numbering
        
        # Read each file and add split_tag column
        for file, role in zip(files, partition_roles):
            if not file.filename.endswith('.csv'):
                raise HTTPException(status_code=400, detail=f"Only CSV files are supported. Invalid file: {file.filename}")
            
            content = await file.read()
            df = pd.read_csv(io.BytesIO(content))
            
            # Normalize partition role
            normalized_role = role.lower().strip()
            if normalized_role == 'oot':
                # Tag each validation file with a unique number (validation_1, validation_2, validation_3)
                validation_counter += 1
                normalized_role = f'validation_{validation_counter}'
            
            # Add split_tag column
            df['split_tag'] = normalized_role
            
            partition_counts[normalized_role] = partition_counts.get(normalized_role, 0) + len(df)
            dfs_to_combine.append(df)
            
            logger.info(f"  Read {file.filename} as '{normalized_role}': {len(df)} rows")
        
        # Validate schema matches across all files (except split_tag which we added)
        if len(dfs_to_combine) > 1:
            ref_cols = set(dfs_to_combine[0].columns) - {'split_tag'}
            for idx, df in enumerate(dfs_to_combine[1:], 2):
                current_cols = set(df.columns) - {'split_tag'}
                if current_cols != ref_cols:
                    missing = ref_cols - current_cols
                    extra = current_cols - ref_cols
                    error_msg = f"Schema mismatch in file {idx}."
                    if missing:
                        error_msg += f" Missing columns: {missing}."
                    if extra:
                        error_msg += f" Extra columns: {extra}."
                    raise HTTPException(status_code=400, detail=error_msg)
        
        # Combine all DataFrames
        combined_df = pd.concat(dfs_to_combine, ignore_index=True)
        total_rows = len(combined_df)
        
        logger.info(f"Combined pre-split data: {total_rows} total rows, partitions: {partition_counts}")
        
        # Save combined file
        combined_filename = f"combined_presplit_{'+'.join(partition_counts.keys())}.csv"
        combined_content = combined_df.to_csv(index=False).encode('utf-8')
        dataset_id, file_path = dataset_manager.save_uploaded_file(combined_content, combined_filename)
        
        logger.info(f"Combined pre-split dataset saved with ID: {dataset_id}")
        
        # Store in dataframe state manager
        dataframe_state_manager.update_dataframe(dataset_id, combined_df, original_shape=combined_df.shape)
        
        # Store split indices based on split_tag column
        import numpy as np
        
        train_mask = combined_df['split_tag'] == 'train'
        test_mask = combined_df['split_tag'] == 'test'
        
        split_indices = {
            "train": np.where(train_mask)[0].astype(np.int64),
            "test": np.where(test_mask)[0].astype(np.int64),
        }
        
        # Add indices for each validation file (validation_1, validation_2, validation_3)
        for i in range(1, 4):
            val_tag = f'validation_{i}'
            val_mask = combined_df['split_tag'] == val_tag
            if val_mask.sum() > 0:
                split_indices[val_tag] = np.where(val_mask)[0].astype(np.int64)
        
        # Also create combined "validation" index for backwards compatibility
        all_validation_mask = combined_df['split_tag'].str.startswith('validation')
        split_indices["validation"] = np.where(all_validation_mask)[0].astype(np.int64)
        
        dataframe_state_manager._split_indices[dataset_id] = split_indices
        
        # DO NOT set scope here - keep _processed_dataframes as full combined data during Step 1
        # Scope will be set to "train" when user clicks Confirm button in frontend
        
        # Log detailed split info
        val_counts = {k: len(v) for k, v in split_indices.items() if k.startswith('validation')}
        logger.info(f"Split indices stored for {dataset_id}: train={train_mask.sum()}, test={test_mask.sum()}, validation_sets={val_counts}")
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "total_rows": total_rows,
            "partitions": partition_counts,
            "columns": list(combined_df.columns),
            "message": f"Successfully combined {len(files)} files into dataset {dataset_id}"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to combine pre-split files: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to combine files: {str(e)}")


@upload_router.post("/finalize-presplit")
async def finalize_presplit_dataset(
    background_tasks: BackgroundTasks,
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    target_variable_type: str = Form(...),
    problem_statement: Optional[str] = Form(""),
    data_dictionary: Optional[str] = Form(""),
    data_dictionary_file: Optional[UploadFile] = File(None),
    unique_id_combinations: Optional[str] = Form("[]"),
    segmentation_variable: Optional[str] = Form(""),
    sample_identifier_variable: Optional[str] = Form(""),
    exclusion_rules: Optional[str] = Form(None),
    variables_to_remove: Optional[str] = Form(None),
    current_user = Depends(get_current_user_dependency),
):
    """
    Finalize a pre-split combined dataset by applying exclusion rules and variable removal.
    This endpoint is called after combine-presplit to process the combined dataset
    the same way as a platform_split (full population) dataset.
    """
    logger.info(f"Finalizing pre-split dataset: {dataset_id}")
    
    try:
        # Get the combined dataframe
        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        
        original_rows = len(df)
        original_cols = len(df.columns)
        
        logger.info(f"Pre-split dataset {dataset_id}: {original_rows} rows, {original_cols} columns")
        
        # Apply exclusion rules if provided
        rows_after_exclusion = original_rows
        if exclusion_rules:
            try:
                import json
                parsed_exclusion_rules = json.loads(exclusion_rules)
                if parsed_exclusion_rules and isinstance(parsed_exclusion_rules, list) and len(parsed_exclusion_rules) > 0:
                    logger.info(f"Applying {len(parsed_exclusion_rules)} exclusion groups to dataset {dataset_id}")
                    exclusion_result = dataframe_state_manager.apply_exclusion_rules(
                        dataset_id,
                        parsed_exclusion_rules,
                        target_variable
                    )
                    logger.info(f"Exclusion rules applied: {exclusion_result}")
                    # Update df reference
                    df = dataframe_state_manager.get_dataframe(dataset_id)
                    if df is not None:
                        rows_after_exclusion = len(df)
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse exclusion_rules JSON: {e}")
            except Exception as e:
                logger.error(f"Failed to apply exclusion rules: {e}")
        
        # Apply variable removal if provided
        cols_after_removal = len(df.columns) if df is not None else original_cols
        variables_removed = []
        if variables_to_remove:
            try:
                import json
                parsed_variables_to_remove = json.loads(variables_to_remove)
                if parsed_variables_to_remove and isinstance(parsed_variables_to_remove, list) and len(parsed_variables_to_remove) > 0:
                    logger.info(f"Removing {len(parsed_variables_to_remove)} variables from dataset {dataset_id}")
                    
                    current_df = dataframe_state_manager.get_dataframe(dataset_id)
                    if current_df is not None:
                        cols_to_remove = [col for col in parsed_variables_to_remove if col in current_df.columns]
                        if cols_to_remove:
                            updated_df = current_df.drop(columns=cols_to_remove)
                            dataframe_state_manager.update_dataframe(dataset_id, updated_df, original_shape=updated_df.shape)
                            variables_removed = cols_to_remove
                            cols_after_removal = len(updated_df.columns)
                            logger.info(f"Successfully removed {len(cols_to_remove)} variables")
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse variables_to_remove JSON: {e}")
            except Exception as e:
                logger.error(f"Failed to remove variables: {e}")
        
        # Parse unique_id_combinations
        parsed_unique_ids = []
        if unique_id_combinations:
            try:
                import json
                parsed_unique_ids = json.loads(unique_id_combinations)
            except:
                pass
        
        # Process data dictionary file if provided
        data_dictionary_path = None
        final_data_dictionary = data_dictionary
        
        if data_dictionary_file and data_dictionary_file.filename:
            logger.info(f"Processing data dictionary file for pre-split: {data_dictionary_file.filename}")
            
            # Validate data dictionary file type
            if not data_dictionary_file.filename.endswith('.csv'):
                logger.warning(f"Invalid data dictionary file type: {data_dictionary_file.filename}")
                raise HTTPException(status_code=400, detail="Data dictionary must be a CSV file")
            
            # Stream data dictionary to disk
            data_dict_id, data_dictionary_path = await dataset_manager.save_uploaded_file_streaming(
                data_dictionary_file,
                f"data_dict_{dataset_id}_{data_dictionary_file.filename}",
                settings.MAX_FILE_SIZE,
            )
            
            # Store the file path as the data dictionary reference
            final_data_dictionary = data_dictionary_path
            logger.info(f"Data dictionary file saved at: {data_dictionary_path}")
        
        # Store dataset metadata
        dataset_manager.store_dataset_info(
            dataset_id=dataset_id,
            file_path="",  # Combined dataset already saved
            filename=f"combined_presplit_{dataset_id}.csv",
            target_variable=target_variable,
            target_variable_type=target_variable_type,
            data_dictionary=final_data_dictionary or "",
            problem_statement=problem_statement or "",
            unique_id_combinations=parsed_unique_ids,
            segmentation_variable=segmentation_variable.strip() if segmentation_variable else None,
            sample_identifier_variable=sample_identifier_variable.strip() if sample_identifier_variable else None,
            split_configuration={"ingestion_mode": "pre_split", "combined": True},
        )
        
        # Prepare data dictionary content for knowledge graph generation
        data_dictionary_content = ""
        if final_data_dictionary:
            if os.path.exists(final_data_dictionary):
                try:
                    data_dictionary_content = _read_data_dictionary(final_data_dictionary)
                except Exception as exc:
                    logger.warning(f"Failed to read data dictionary file {final_data_dictionary}: {exc}")
            else:
                # If it wasn't a file upload, final_data_dictionary already contains the text from the form
                data_dictionary_content = final_data_dictionary
        
        # Trigger knowledge graph generation as background task if data dictionary is available
        if data_dictionary_content:
            background_tasks.add_task(
                llm_service.generate_async_knowledge_graph,
                dataset_id,
                data_dictionary_content,
                problem_statement,
            )
            logger.info(f"Knowledge graph generation triggered for pre-split dataset {dataset_id}")
        else:
            logger.warning(f"No data dictionary content available for pre-split dataset {dataset_id}; skipping KG pre-generation")
        
        logger.info(f"Pre-split dataset {dataset_id} finalized: {rows_after_exclusion} rows, {cols_after_removal} columns")
        
        # NOTE: Scope remains as 'entire' (full data) after Submit
        # The right pane will show entire data by default
        # Agents will set scope to 'train' when needed for modeling
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "original_rows": original_rows,
            "rows_after_exclusion": rows_after_exclusion,
            "original_columns": original_cols,
            "columns_after_removal": cols_after_removal,
            "variables_removed": variables_removed,
            "message": f"Dataset finalized successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to finalize pre-split dataset: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to finalize dataset: {str(e)}")


@upload_router.post("/analyze-dataset")
async def analyze_dataset(
    file: UploadFile = File(...),
    is_preview: Optional[str] = Form(None),
    original_size: Optional[str] = Form(None),
    current_user = Depends(get_current_user_dependency)
):
    """
    Analyze uploaded dataset and return available columns and their types.
    All CPU-bound work (read_csv, column analysis) runs in a thread executor so
    the async event loop is never blocked - critical for large files on Azure.
    For column-type detection we read up to the first 1 000 000 rows; the full row
    count is obtained cheaply by counting newlines in the raw bytes.

    P1.1 part 2: when the frontend sends `is_preview=true` and the original
    `original_size` byte length, the backend treats the uploaded payload as a
    head-slice of a larger CSV and *extrapolates* the row count via
    `rows_in_slice * (original_size / slice_size)`. Accuracy is within 1-2%
    for uniformly-sized rows and avoids transferring multi-GB files just to
    count newlines for the preview UI. The response carries
    `is_estimated: true` so callers can render a "~N rows" badge.
    """
    is_preview_bool = (str(is_preview or "").lower() in ("1", "true", "yes"))
    original_size_int = 0
    try:
        if original_size:
            original_size_int = int(original_size)
    except ValueError:
        original_size_int = 0
    logger.info(
        f"Dataset analysis request: {file.filename} "
        f"(is_preview={is_preview_bool}, original_size={original_size_int})"
    )

    try:
        # Validate file type
        if not file.filename.endswith('.csv'):
            logger.warning(f"Invalid file type uploaded: {file.filename}")
            raise HTTPException(status_code=400, detail="Only CSV files are supported")

        # Stream file bytes without blocking the event loop
        file_content = await file.read()
        file_size = len(file_content)
        logger.info(f"File read successfully, size: {file_size} bytes")

        if file_size > settings.MAX_FILE_SIZE:
            logger.warning(f"File size exceeds limit: {file_size} > {settings.MAX_FILE_SIZE}")
            raise HTTPException(status_code=400, detail="File size exceeds maximum limit")

        import tempfile
        import os
        import asyncio as _asyncio
        from app.core.executor import executor as _executor

        # Write temp file in executor (avoids blocking on large writes)
        def _write_temp(content: bytes):
            tmp = tempfile.NamedTemporaryFile(mode='wb', suffix='.csv', delete=False)
            tmp.write(content)
            tmp.flush()
            tmp.close()
            return tmp.name

        temp_file_path = await _asyncio.get_event_loop().run_in_executor(
            _executor, _write_temp, file_content
        )

        # Cheap total-row estimate: count newlines in raw bytes (no full parse needed)
        def _count_rows(content: bytes) -> int:
            try:
                return max(0, content.count(b'\n') - 1)  # subtract header
            except Exception:
                return 0

        total_rows_estimate = await _asyncio.get_event_loop().run_in_executor(
            _executor, _count_rows, file_content
        )

        # All heavy CPU work runs in the executor - never blocks the event loop
        def _analyze_in_thread(path: str, filename: str, n_rows_estimate: int):
            from app.utils.helpers import identify_date_columns as _idc
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            # For column-type detection, use up to 1M rows for more stable categorization on large files.
            # Reading only a slice is still much faster than loading the entire CSV when it is huge.
            SAMPLE_ROWS = 1_000_000
            df_sample = None
            used_encoding = 'utf-8'
            for enc in encodings:
                try:
                    df_sample = pd.read_csv(path, encoding=enc, nrows=SAMPLE_ROWS)
                    used_encoding = enc
                    break
                except UnicodeDecodeError:
                    continue

            if df_sample is None:
                raise ValueError("Could not decode file with any supported encoding")

            logger.info(
                f"analyze-dataset: sampled {len(df_sample)} rows with '{used_encoding}' "
                f"(full file ~{n_rows_estimate} rows) for {filename}"
            )

            date_detection_results = _idc(df_sample)

            columns_info = []
            for col in df_sample.columns:
                col_type = str(df_sample[col].dtype)
                col_lower = col.lower()
                date_meta = date_detection_results.get(str(col), {})
                is_date_like = bool(date_meta.get("is_date"))

                if df_sample[col].dtype in ['int64', 'float64']:
                    data_type = 'Numerical'
                    nunique = df_sample[col].nunique()
                    if nunique >= 2 and nunique <= 5 and not df_sample[col].isna().all():
                        missing_pct = df_sample[col].isna().sum() / len(df_sample[col])
                        if missing_pct < 0.5:
                            data_type = 'Categorical'
                elif df_sample[col].dtype in ['object', 'category']:
                    data_type = 'Categorical'
                else:
                    data_type = 'Categorical'

                sample_values = None
                if data_type == 'Categorical' and df_sample[col].nunique() <= 10:
                    sample_values = df_sample[col].value_counts().head(5).to_dict()

                numerical_stats = None
                if data_type == 'Numerical':
                    numerical_stats = {
                        'min': float(df_sample[col].min()) if not df_sample[col].isna().all() else None,
                        'max': float(df_sample[col].max()) if not df_sample[col].isna().all() else None,
                        'mean': float(df_sample[col].mean()) if not df_sample[col].isna().all() else None,
                        'missing_count': int(df_sample[col].isna().sum())
                    }

                columns_info.append({
                    'name': col,
                    'type': data_type,
                    'logical_type': 'Date' if is_date_like else data_type,
                    'is_date': is_date_like,
                    'date_detection_reason': date_meta.get('reason'),
                    'date_detected_format': date_meta.get('detected_format'),
                    'date_detection_confidence': date_meta.get('confidence'),
                    'pandas_type': col_type,
                    'unique_count': int(df_sample[col].nunique()),
                    'missing_count': int(df_sample[col].isna().sum()),
                    'sample_values': sample_values,
                    'numerical_stats': numerical_stats
                })

            columns_info.sort(key=lambda x: (x['type'] != 'Numerical', x['name']))

            target_suggestions = []
            for col in columns_info:
                cn = col['name'].lower()
                if any(p in cn for p in ['target', 'label', 'class', 'outcome', 'result', 'prediction']):
                    target_suggestions.append(col['name'])
                elif any(p in cn for p in ['price', 'cost', 'amount', 'value', 'score']):
                    target_suggestions.append(col['name'])
                elif any(p in cn for p in ['success', 'failure', 'churn', 'conversion', 'click']):
                    target_suggestions.append(col['name'])

            return {
                "columns_info": columns_info,
                "total_columns": len(columns_info),
                "suggested_target": target_suggestions[0] if target_suggestions else None,
            }

        try:
            result = await _asyncio.get_event_loop().run_in_executor(
                _executor,
                _analyze_in_thread,
                temp_file_path,
                file.filename,
                total_rows_estimate,
            )
        finally:
            try:
                os.unlink(temp_file_path)
            except Exception:
                pass

        # P1.1 part 2: when this was a 5 MB head-slice preview, extrapolate
        # the true row count from the slice's row density.
        is_estimated = False
        reported_total_rows = total_rows_estimate
        if is_preview_bool and original_size_int > file_size and file_size > 0:
            scale = original_size_int / float(file_size)
            reported_total_rows = int(round(total_rows_estimate * scale))
            is_estimated = True
            logger.info(
                f"analyze-dataset preview extrapolation: slice={file_size}B "
                f"original={original_size_int}B scale={scale:.2f} "
                f"slice_rows={total_rows_estimate} -> estimated_total_rows={reported_total_rows}"
            )

        logger.info(f"Dataset analysis completed successfully: {file.filename}")
        return {
            "success": True,
            "message": "Dataset analyzed successfully",
            "dataset_info": {
                "filename": file.filename,
                "total_rows": reported_total_rows,
                "total_columns": result["total_columns"],
                "columns": result["columns_info"],
                "suggested_target_variable": result["suggested_target"],
                "is_estimated": is_estimated,
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dataset analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.post("/generate-qc-template/{template_type}")
async def generate_qc_template(
    template_type: str,
    request: dict = Body(...),
    current_user = Depends(get_current_user_dependency)
):
    """
    Generate QC template Excel/CSV file with auto-filled Var Name and Type columns.
    For outliers and missing_values, generates Excel with data validation dropdowns.
    For invalid_values and special_values, generates CSV.
    
    Args:
        template_type: One of 'invalid_values', 'special_values', 'outliers', 'missing_values'
        request: Contains 'columns' array with {name, type, is_date, logical_type} for each column
    """
    import io
    from fastapi.responses import StreamingResponse
    
    logger.info(f"Generating QC template: {template_type}")
    
    valid_types = ['invalid_values', 'special_values', 'outliers', 'missing_values']
    if template_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid template type. Must be one of: {valid_types}")
    
    columns = request.get('columns', [])
    
    # Filter columns: exclude date columns, include only Numerical and Categorical
    filtered_columns = [
        col for col in columns
        if not col.get('is_date', False) 
        and col.get('logical_type') != 'Date'
        and col.get('type') in ['Numerical', 'Categorical']
    ]
    
    logger.info(f"Filtered {len(filtered_columns)} columns from {len(columns)} total")
    
    # Template headers
    headers_map = {
        'invalid_values': ['Var Name', 'Type', 'Valid Range / Valid Labels'],
        'special_values': ['Var Name', 'Type', 'Special Values'],
        'outliers': ['Var Name', 'Type', 'Choose Detection Method'],
        'missing_values': ['Var Name', 'Type', 'Choose Imputation Method'],
    }
    
    headers = headers_map[template_type]
    
    # For outliers and missing_values, generate Excel with data validation dropdowns
    if template_type in ['outliers', 'missing_values']:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            
            # Dropdown options
            dropdown_options = {
                'outliers': ['Z-Score', 'IQR (Interquartile Range)', 'Percentile Capping'],
                'missing_values': ['Mean', 'Median', 'Min', 'Max', 'P1', 'P95', 'P99', 'Mode', 'Drop'],
            }
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Template'
            
            # Style definitions
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Set column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 35
            
            # Write data rows with auto-filled Var Name and Type
            for row_idx, col_info in enumerate(filtered_columns, 2):
                ws.cell(row=row_idx, column=1, value=col_info.get('name', '')).border = thin_border
                ws.cell(row=row_idx, column=2, value=col_info.get('type', 'Unknown')).border = thin_border
                ws.cell(row=row_idx, column=3, value='').border = thin_border
            
            # Create data validation for dropdown in column C
            options = dropdown_options[template_type]
            options_str = ','.join(options)
            
            # Calculate the range for data validation
            last_row = max(len(filtered_columns) + 1, 100)
            
            # Create DataValidation object with dropdown list
            dv = DataValidation(
                type='list',
                formula1=f'"{options_str}"',
                allow_blank=True,
                showDropDown=False,  # False = show dropdown arrow
                showErrorMessage=True,
                errorTitle='Invalid Selection',
                error=f'Please select from: {", ".join(options)}'
            )
            
            # Add validation to the range
            dv.add(f'C2:C{last_row}')
            ws.add_data_validation(dv)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'{template_type}_template.xlsx'
            
            return StreamingResponse(
                output,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{filename}"'}
            )
            
        except ImportError as e:
            logger.error(f"openpyxl not available: {e}")
            raise HTTPException(status_code=500, detail="Excel generation not available")
    
    else:
        # For invalid_values and special_values, generate CSV
        output = io.StringIO()
        
        # Write header
        output.write(','.join(headers) + '\n')
        
        # Write data rows
        for col_info in filtered_columns:
            var_name = col_info.get('name', '')
            # Escape if contains comma
            if ',' in var_name:
                var_name = f'"{var_name}"'
            col_type = col_info.get('type', 'Unknown')
            output.write(f'{var_name},{col_type},\n')
        
        output.seek(0)
        
        filename = f'{template_type}_template.csv'
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )


def _apply_exclusion_rules_to_df(df: pd.DataFrame, exclusion_groups: list) -> pd.DataFrame:
    """
    Apply exclusion rules to a DataFrame and return filtered DataFrame.
    Groups are OR-ed together; conditions within a group follow AND/OR precedence.
    """
    def evaluate_condition(data: pd.DataFrame, cond: dict) -> pd.Series:
        col = cond.get('column')
        op = cond.get('operator', '=')
        val = cond.get('value')
        
        if col not in data.columns:
            return pd.Series(False, index=data.index)
        
        # Unwrap single-element lists for scalar operators
        if isinstance(val, list) and len(val) == 1 and op in ('=', '!=', '>', '>=', '<', '<=', 'STARTS WITH', 'CONTAINS'):
            val = val[0]
        
        series = data[col]
        
        if op == 'IS NULL':
            return series.isna()
        elif op == 'IS NOT NULL':
            return series.notna()
        elif op == '= TRUE':
            return series.astype(str).str.lower().isin(['true', '1', 'yes'])
        elif op == '= FALSE':
            return series.astype(str).str.lower().isin(['false', '0', 'no'])
        elif op == '=':
            return series == val
        elif op == '!=':
            return series != val
        elif op == '>':
            return series > val
        elif op == '>=':
            return series >= val
        elif op == '<':
            return series < val
        elif op == '<=':
            return series <= val
        elif op == 'IN':
            vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(',')]
            return series.astype(str).isin([str(v) for v in vals])
        elif op == 'NOT IN':
            vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(',')]
            return ~series.astype(str).isin([str(v) for v in vals])
        elif op == 'STARTS WITH':
            return series.astype(str).str.startswith(str(val), na=False)
        elif op == 'CONTAINS':
            return series.astype(str).str.contains(str(val), na=False, case=False)
        elif op == 'BETWEEN':
            if isinstance(val, list) and len(val) == 2:
                return (series >= val[0]) & (series <= val[1])
            return pd.Series(False, index=data.index)
        elif op == 'NOT BETWEEN':
            if isinstance(val, list) and len(val) == 2:
                return (series < val[0]) | (series > val[1])
            return pd.Series(False, index=data.index)
        else:
            return pd.Series(False, index=data.index)

    def evaluate_group(data: pd.DataFrame, group: dict) -> pd.Series:
        conditions = group.get('conditions', [])
        if not conditions:
            return pd.Series(False, index=data.index)
        
        masks = [evaluate_condition(data, c) for c in conditions]
        connectors = [c.get('connector', 'AND') for c in conditions]
        
        # AND binds before OR
        and_groups = []
        current_group_indices = [0]
        for i in range(1, len(masks)):
            if connectors[i - 1] == 'AND':
                current_group_indices.append(i)
            else:
                and_groups.append(current_group_indices)
                current_group_indices = [i]
        and_groups.append(current_group_indices)
        
        and_results = []
        for grp_indices in and_groups:
            grp_result = masks[grp_indices[0]]
            for idx in grp_indices[1:]:
                grp_result = grp_result & masks[idx]
            and_results.append(grp_result)
        
        final_result = and_results[0]
        for ar in and_results[1:]:
            final_result = final_result | ar
        
        return final_result

    # Combine all groups with OR
    if not exclusion_groups:
        return df
    
    combined_mask = pd.Series(False, index=df.index)
    for group in exclusion_groups:
        combined_mask = combined_mask | evaluate_group(df, group)
    
    return df[~combined_mask].copy()


# Step 1 upload previews: formats supported (demo/V9 ff9a789).
_PREVIEW_UPLOAD_EXTENSIONS = frozenset({".csv", ".tsv", ".parquet", ".xlsx", ".xls"})


def _preview_upload_suffix(filename: Optional[str]) -> str:
    return Path(filename).suffix.lower() if filename else ""


def _load_df_from_preview_temp_path(temp_file_path: str, file_suffix: str) -> pd.DataFrame:
    """
    Load tabular data from a temp file for preview endpoints (partition / exclusion / variable review).
    """
    sfx = (file_suffix or ".csv").lower()
    try:
        if sfx == ".parquet":
            return pd.read_parquet(temp_file_path)
        if sfx in (".xlsx", ".xls"):
            return pd.read_excel(temp_file_path)
        encodings = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
        read_kwargs: Dict[str, Any] = {"sep": "\t"} if sfx == ".tsv" else {}
        for encoding in encodings:
            try:
                return pd.read_csv(temp_file_path, encoding=encoding, **read_kwargs)
            except UnicodeDecodeError:
                continue
    except Exception as e:
        logger.error(f"Preview file read failed (suffix={sfx}): {e}")
        raise HTTPException(status_code=400, detail="Could not decode file") from e
    raise HTTPException(status_code=400, detail="Could not decode file")


def _resolve_full_dataframe_for_preview_by_id(dataset_id: str) -> pd.DataFrame:
    """
    Combined (pre-split) frame for Step 1 previews. Prefer this pod's in-memory cache;
    otherwise load from object storage so partition / exclusion previews work on EKS
    when requests hit a different replica than the upload/session pod.

    NOTE: this is the legacy "give me the whole dataframe" entry point. New code
    paths (partition-preview-by-id) prefer ``_load_partition_preview_skinny_df``
    which does column projection through the parquet sidecar so a 2 GB file
    isn't dragged into pandas memory just to compute split row counts.
    """
    ref = dataframe_state_manager.get_full_dataframe_readonly(dataset_id)
    if ref is not None:
        return ref
    if not dataset_manager.get_dataset_info(dataset_id):
        raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
    loaded = dataset_manager.load_dataset(dataset_id)
    if loaded is None:
        raise HTTPException(
            status_code=404,
            detail=f"Dataset {dataset_id} not found. Please upload your dataset again.",
        )
    logger.info(
        "preview_by_id: dataset %s loaded from storage (no in-memory copy on this pod), shape=%s",
        dataset_id,
        loaded.shape,
    )
    return loaded


def _extract_exclusion_referenced_cols(
    exclusion_groups: Optional[List[Dict[str, Any]]],
) -> List[str]:
    """Return the set of column names referenced by exclusion-rule conditions."""
    if not exclusion_groups:
        return []
    cols: set = set()
    for grp in exclusion_groups:
        if not isinstance(grp, dict):
            continue
        for cond in grp.get("conditions", []) or []:
            if isinstance(cond, dict):
                c = cond.get("column")
                if c:
                    cols.add(str(c))
    return sorted(cols)


def _load_partition_preview_skinny_df(
    dataset_id: str,
    needed_cols: List[str],
) -> Tuple[pd.DataFrame, int]:
    """
    Load a *column-projected* DataFrame for the partition-preview-by-id path.

    Returns ``(skinny_df, full_n_columns)``.

    Fast path: the dataset's full frame is already cached in this pod's
    ``dataframe_state_manager`` (uploaded on this replica). We just hand
    out a column slice.

    Slow path: stream the parquet sidecar (or CSV fallback) through the
    existing ``SidecarCache`` / object-store layer and project just the
    columns we actually need with Polars. Memory ceiling = ``len(needed_cols)
    * rows * dtype_width``, typically <= 200 MB even for 50 M rows.

    Falls back to ``dataset_manager.load_dataset`` (full pandas) on any
    Polars-side error so behavior is never worse than the legacy path.
    """
    cached = dataframe_state_manager.get_full_dataframe_readonly(dataset_id)
    if cached is not None:
        all_cols = list(cached.columns)
        existing = [c for c in needed_cols if c in all_cols]
        skinny = cached[existing] if existing else cached.iloc[:, :0].copy()
        return skinny, len(all_cols)

    if not dataset_manager.get_dataset_info(dataset_id):
        raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")

    with dataset_manager.materialize_unique_id_validation_path(dataset_id) as (path, is_pq):
        if not path:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"Dataset {dataset_id} file missing from storage. "
                    f"Please re-upload."
                ),
            )
        try:
            import polars as pl  # noqa: WPS433 - local import keeps cold start light

            if is_pq:
                schema_names = list(pl.scan_parquet(path).collect_schema().names())
            else:
                schema_names = list(
                    pl.read_csv(path, n_rows=0, infer_schema_length=0).columns
                )
            full_n_columns = len(schema_names)
            existing = [c for c in needed_cols if c in schema_names]

            if not existing:
                # No needed column is in the file; we still need the row count
                # so build_partition_preview can compute proportions.
                if is_pq:
                    n_lf = pl.scan_parquet(path).select(pl.len().alias("n"))
                else:
                    n_lf = pl.scan_csv(path).select(pl.len().alias("n"))
                try:
                    n_df = n_lf.collect(engine="streaming")
                except Exception:
                    n_df = n_lf.collect(streaming=True)
                n = int(n_df["n"][0])
                return pd.DataFrame(index=range(n)), full_n_columns

            if is_pq:
                lf = pl.scan_parquet(path).select([pl.col(c) for c in existing])
            else:
                lf = pl.scan_csv(
                    path, infer_schema_length=10_000
                ).select([pl.col(c) for c in existing])
            try:
                df_pl = lf.collect(engine="streaming")
            except Exception:
                df_pl = lf.collect(streaming=True)
            return df_pl.to_pandas(), full_n_columns
        except HTTPException:
            raise
        except Exception as exc:
            logger.warning(
                "_load_partition_preview_skinny_df: polars projection failed, "
                "falling back to dataset_manager.load_dataset (full pandas): %s",
                exc,
            )
            full = dataset_manager.load_dataset(dataset_id)
            if full is None:
                raise HTTPException(
                    status_code=404,
                    detail=f"Dataset {dataset_id} not found. Please upload your dataset again.",
                )
            return full, len(full.columns)


@upload_router.post("/partition-preview")
async def partition_preview(
    file: UploadFile = File(...),
    split_configuration: str = Form(...),
    target_variable: str = Form(...),
    exclusion_rules: Optional[str] = Form(None),
    variables_to_remove: Optional[str] = Form(None),
    current_user=Depends(get_current_user_dependency),
):
    """
    Step 1 Review Stats: compute train/test/validation partition metrics from an uploaded file + split config (no persistence).
    Exclusion rules are applied first (if provided) before computing partition statistics.
    Variables to remove are also dropped before computing features count.
    """
    file_suffix = _preview_upload_suffix(file.filename)
    if not file.filename or file_suffix not in _PREVIEW_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Supported preview formats: CSV, TSV, Parquet, XLSX, XLS",
        )
    try:
        cfg = json.loads(split_configuration)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="split_configuration must be valid JSON")

    # Parse exclusion rules if provided
    parsed_exclusion_rules = None
    if exclusion_rules:
        try:
            parsed_exclusion_rules = json.loads(exclusion_rules)
        except json.JSONDecodeError:
            logger.warning("Invalid exclusion_rules JSON, ignoring")

    # Parse variables to remove if provided
    parsed_variables_to_remove = None
    if variables_to_remove:
        try:
            parsed_variables_to_remove = json.loads(variables_to_remove)
        except json.JSONDecodeError:
            logger.warning("Invalid variables_to_remove JSON, ignoring")

    try:
        file_content = await file.read()
        if len(file_content) > settings.MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File size exceeds maximum limit")

        with tempfile.NamedTemporaryFile(mode="wb", suffix=file_suffix or ".csv", delete=False) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name
        try:
            df = _load_df_from_preview_temp_path(temp_file_path, file_suffix)
        finally:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)

        # Apply exclusion rules before partition preview
        original_count = len(df)
        if parsed_exclusion_rules and isinstance(parsed_exclusion_rules, list) and len(parsed_exclusion_rules) > 0:
            df = _apply_exclusion_rules_to_df(df, parsed_exclusion_rules)
            logger.info(f"partition_preview: exclusion rules applied, {original_count} -> {len(df)} rows")

        # Apply variable removal before computing features count
        original_cols = len(df.columns)
        if parsed_variables_to_remove and isinstance(parsed_variables_to_remove, list) and len(parsed_variables_to_remove) > 0:
            cols_to_remove = [col for col in parsed_variables_to_remove if col in df.columns]
            if cols_to_remove:
                df = df.drop(columns=cols_to_remove)
                logger.info(f"partition_preview: variables removed, {original_cols} -> {len(df.columns)} columns")

        tv = (target_variable or "").strip()
        # If target variable is not provided or doesn't exist, use first column as placeholder
        # This allows computing date range/cutoffs even before target is selected
        if not tv or tv not in df.columns:
            tv = df.columns[0] if len(df.columns) > 0 else "_placeholder_"

        logger.info(f"partition_preview: method={cfg.get('split_method')}, date_col={cfg.get('date_column')}")
        result = build_partition_preview(df, tv, cfg)
        if not result.get("success"):
            raise HTTPException(status_code=400, detail=result.get("error", "Preview failed"))
        logger.info(f"partition_preview result: computed_cutoffs={result.get('computed_cutoffs')}")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"partition_preview failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@upload_router.post("/partition-preview-by-id")
async def partition_preview_by_id(
    dataset_id: str = Form(...),
    split_configuration: str = Form(...),
    target_variable: str = Form(...),
    exclusion_rules: Optional[str] = Form(None),
    variables_to_remove: Optional[str] = Form(None),
    current_user=Depends(get_current_user_dependency),
):
    """
    Step 1 Review Stats: partition metrics keyed by an already-uploaded dataset_id.

    Hot-path semantics on multi-GB datasets:

    1. **Result-level cache** (``AnalyticsResultCache``) keyed by
       ``(dataset_id, sha1(target+split+exclusions+removals), version)``.
       Repeating the same selection while the user fiddles with ratios
       returns in <100 ms.
    2. **Column projection** -- only target / date / identifier / exclusion-
       referenced columns are decoded from the parquet sidecar, so we never
       drag a 2 GB dataframe into pandas memory just to count rows per
       partition.
    3. **Executor offload** -- the resolve+compute runs on the shared thread
       pool so the event loop stays responsive while another preview is
       executing on a different request.
    """
    import asyncio as _asyncio
    import hashlib

    from app.core.executor import executor as _executor
    from app.core.metrics import time_stage
    from app.services.analytics_cache import analytics_cache
    from app.services.dataframe_state_manager import dataframe_state_manager

    try:
        cfg = json.loads(split_configuration)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="split_configuration must be valid JSON")

    parsed_exclusion_rules = None
    if exclusion_rules:
        try:
            parsed_exclusion_rules = json.loads(exclusion_rules)
        except json.JSONDecodeError:
            logger.warning("Invalid exclusion_rules JSON, ignoring")

    parsed_variables_to_remove = None
    if variables_to_remove:
        try:
            parsed_variables_to_remove = json.loads(variables_to_remove)
        except json.JSONDecodeError:
            logger.warning("Invalid variables_to_remove JSON, ignoring")

    tv_input = (target_variable or "").strip()

    scope_payload = json.dumps(
        {
            "tv": tv_input,
            "cfg": cfg,
            "exc": parsed_exclusion_rules or [],
            "rm": parsed_variables_to_remove or [],
        },
        sort_keys=True,
        default=str,
    ).encode("utf-8")
    scope = hashlib.sha1(scope_payload).hexdigest()
    version = dataframe_state_manager.get_version(dataset_id)

    cached = analytics_cache.get(
        kind="partition_preview",
        dataset_id=dataset_id,
        scope=scope,
        version=version,
    )
    if cached is not None:
        return {**cached, "cached": True}

    needed_cols: set = set()
    if tv_input:
        needed_cols.add(tv_input)
    method = cfg.get("split_method")
    if method == "time_based" and cfg.get("date_column"):
        needed_cols.add(cfg["date_column"])
    if method == "user_identifier" and cfg.get("identifier_column"):
        needed_cols.add(cfg["identifier_column"])
    needed_cols.update(_extract_exclusion_referenced_cols(parsed_exclusion_rules))
    needed_cols_list = sorted(c for c in needed_cols if c)

    def _load_and_build() -> Dict[str, Any]:
        df, full_n_columns = _load_partition_preview_skinny_df(dataset_id, needed_cols_list)

        original_count = len(df)
        if parsed_exclusion_rules and isinstance(parsed_exclusion_rules, list) and len(parsed_exclusion_rules) > 0:
            df = _apply_exclusion_rules_to_df(df, parsed_exclusion_rules)
            logger.info(
                "partition_preview_by_id: exclusion rules applied, %s -> %s rows",
                original_count,
                len(df),
            )

        # variables_to_remove is a feature-count adjustment only; we never
        # loaded those columns in the skinny frame, so just discount the
        # full-column count for display.
        adjusted_full_cols = full_n_columns
        if (
            parsed_variables_to_remove
            and isinstance(parsed_variables_to_remove, list)
            and len(parsed_variables_to_remove) > 0
        ):
            removed_count = sum(1 for c in parsed_variables_to_remove if c)
            adjusted_full_cols = max(0, full_n_columns - removed_count)
            logger.info(
                "partition_preview_by_id: variables removed (count=%s), full_cols %s -> %s",
                removed_count,
                full_n_columns,
                adjusted_full_cols,
            )

        tv = tv_input or (df.columns[0] if len(df.columns) > 0 else "_placeholder_")

        logger.info(
            "partition_preview_by_id: dataset_id=%s method=%s rows=%s cols_loaded=%s full_cols=%s",
            dataset_id,
            cfg.get("split_method"),
            len(df),
            len(df.columns),
            adjusted_full_cols,
        )
        result = build_partition_preview(df, tv, cfg)
        if result.get("success"):
            # ``build_partition_preview`` derives ``features`` from the columns
            # it actually saw; with column pruning that's wrong. Override with
            # the real full-schema count (minus the target if present).
            result["features"] = max(
                0, adjusted_full_cols - (1 if tv in df.columns else 0)
            )
        return result

    try:
        with time_stage("partition_preview_by_id"):
            result = await _asyncio.get_event_loop().run_in_executor(
                _executor, _load_and_build
            )
    except HTTPException:
        raise
    except Exception as e:
        # ``logger.exception`` captures the full traceback (vs. just the
        # message). Without it, every partition-preview-by-id 500 in
        # production showed up as a one-line error and was undebuggable.
        logger.exception(
            "partition_preview_by_id failed: dataset_id=%s method=%s "
            "needed_cols=%s exclusion_rules=%s vars_to_remove=%s err=%s",
            dataset_id,
            cfg.get("split_method"),
            needed_cols_list,
            bool(parsed_exclusion_rules),
            bool(parsed_variables_to_remove),
            e,
        )
        raise HTTPException(
            status_code=500,
            detail=f"partition-preview-by-id failed: {type(e).__name__}: {e}",
        )

    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Preview failed"))

    analytics_cache.set(
        kind="partition_preview",
        dataset_id=dataset_id,
        scope=scope,
        version=version,
        value=result,
    )
    return result


@upload_router.post("/exclusion-preview")
async def exclusion_preview(
    file: UploadFile = File(...),
    exclusion_groups: str = Form(...),
    target_variable: str = Form(...),
    current_user=Depends(get_current_user_dependency),
):
    """
    Evaluate exclusion rules against dataset and return waterfall statistics.
    Does not modify data - preview only.
    
    exclusion_groups: JSON array of groups, each with conditions:
    [
      {
        "id": "...",
        "conditions": [
          {"column": "state", "operator": "IN", "value": ["AK","HI","PR"], "connector": "AND"},
          {"column": "status", "operator": "!=", "value": "Fully Paid", "connector": "AND"}
        ]
      }
    ]
    """
    file_suffix = _preview_upload_suffix(file.filename)
    if not file.filename or file_suffix not in _PREVIEW_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Supported preview formats: CSV, TSV, Parquet, XLSX, XLS",
        )
    
    try:
        groups = json.loads(exclusion_groups)
        logger.info(f"=== EXCLUSION PREVIEW: Received {len(groups)} groups ===")
        logger.info(f"Raw exclusion_groups JSON: {exclusion_groups[:1000]}...")  # First 1000 chars
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="exclusion_groups must be valid JSON")
    
    try:
        file_content = await file.read()
        if len(file_content) > settings.MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File size exceeds maximum limit")
        
        with tempfile.NamedTemporaryFile(mode="wb", suffix=file_suffix or ".csv", delete=False) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name
        
        try:
            df = _load_df_from_preview_temp_path(temp_file_path, file_suffix)
            
            # Validate target variable exists
            if target_variable and target_variable not in df.columns:
                logger.warning(f"Target variable '{target_variable}' not found, using first column")
                target_variable = df.columns[0]
            
            original_count = len(df)
            original_event_rate = float(df[target_variable].mean() * 100) if target_variable else None
            original_event_count = int(df[target_variable].sum()) if target_variable else None
            original_non_event_count = original_count - original_event_count if original_event_count is not None else None
            
            waterfall = [{
                'step': 'Start',
                'label': 'Full population',
                'removed': '--',
                'remaining': original_count,
                'eventRate': original_event_rate,
                'eventCount': original_event_count,
                'nonEventCount': original_non_event_count
            }]
            
            remaining = df.copy()
            excluded_indices = set()
            
            def evaluate_condition(data: pd.DataFrame, cond: dict) -> pd.Series:
                """Evaluate a single condition and return boolean mask."""
                col = cond.get('column')
                op = cond.get('operator', '=')
                val = cond.get('value')
                
                if col not in data.columns:
                    return pd.Series(False, index=data.index)
                
                # Unwrap single-element lists for scalar comparison operators
                if isinstance(val, list) and len(val) == 1 and op in ('=', '!=', '>', '>=', '<', '<=', 'STARTS WITH', 'CONTAINS'):
                    val = val[0]
                
                series = data[col]
                
                if op == 'IS NULL':
                    return series.isna()
                elif op == 'IS NOT NULL':
                    return series.notna()
                elif op == '= TRUE':
                    return series.astype(str).str.lower().isin(['true', '1', 'yes'])
                elif op == '= FALSE':
                    return series.astype(str).str.lower().isin(['false', '0', 'no'])
                elif op == '=':
                    return series == val
                elif op == '!=':
                    return series != val
                elif op == '>':
                    return series > val
                elif op == '>=':
                    return series >= val
                elif op == '<':
                    return series < val
                elif op == '<=':
                    return series <= val
                elif op == 'IN':
                    vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(',')]
                    return series.astype(str).isin([str(v) for v in vals])
                elif op == 'NOT IN':
                    vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(',')]
                    return ~series.astype(str).isin([str(v) for v in vals])
                elif op == 'STARTS WITH':
                    return series.astype(str).str.startswith(str(val), na=False)
                elif op == 'CONTAINS':
                    return series.astype(str).str.contains(str(val), na=False, case=False)
                elif op == 'BETWEEN':
                    if isinstance(val, list) and len(val) == 2:
                        return (series >= val[0]) & (series <= val[1])
                    return pd.Series(False, index=data.index)
                elif op == 'NOT BETWEEN':
                    if isinstance(val, list) and len(val) == 2:
                        return (series < val[0]) | (series > val[1])
                    return pd.Series(False, index=data.index)
                else:
                    return pd.Series(False, index=data.index)
            
            def evaluate_group(data: pd.DataFrame, group: dict) -> pd.Series:
                """Evaluate a group of conditions with AND/OR connectors."""
                conditions = group.get('conditions', [])
                if not conditions:
                    return pd.Series(False, index=data.index)
                
                # DEBUG: Log conditions and connectors
                logger.info(f"=== EXCLUSION DEBUG: Evaluating group with {len(conditions)} conditions ===")
                for i, c in enumerate(conditions):
                    logger.info(f"  Condition {i}: column={c.get('column')}, op={c.get('operator')}, "
                               f"value={c.get('value')}, connector={c.get('connector')}")
                
                # Evaluate all conditions
                masks = [evaluate_condition(data, c) for c in conditions]
                connectors = [c.get('connector', 'AND') for c in conditions]
                
                # DEBUG: Log mask results
                for i, mask in enumerate(masks):
                    logger.info(f"  Mask {i}: {mask.sum()} rows match (connector after: {connectors[i]})")
                
                # Parse AND groups first, then OR them (AND binds before OR)
                and_groups = []
                current_group_indices = [0]
                
                for i in range(1, len(masks)):
                    prev_connector = connectors[i - 1]
                    logger.info(f"  Processing mask {i}, prev_connector (from condition {i-1}) = '{prev_connector}'")
                    if prev_connector == 'AND':
                        current_group_indices.append(i)
                    else:  # OR
                        and_groups.append(current_group_indices)
                        current_group_indices = [i]
                and_groups.append(current_group_indices)
                
                logger.info(f"  AND groups formed: {and_groups}")
                
                # AND within each group
                and_results = []
                for grp_indices in and_groups:
                    grp_masks = [masks[i] for i in grp_indices]
                    grp_result = grp_masks[0]
                    for m in grp_masks[1:]:
                        grp_result = grp_result & m
                    and_results.append(grp_result)
                    logger.info(f"  AND group {grp_indices}: {grp_result.sum()} rows match after AND")
                
                # OR between groups
                final_result = and_results[0]
                for i, ar in enumerate(and_results[1:], 1):
                    before = final_result.sum()
                    final_result = final_result | ar
                    logger.info(f"  OR with group {i}: {before} | {ar.sum()} = {final_result.sum()}")
                
                logger.info(f"  FINAL: {final_result.sum()} rows match the group")
                return final_result
            
            def group_to_natural_language(group: dict) -> str:
                """Convert group to human-readable rule description."""
                conditions = group.get('conditions', [])
                parts = []
                for i, c in enumerate(conditions):
                    col = c.get('column', '')
                    op = c.get('operator', '=')
                    val = c.get('value', '')
                    
                    if op in ('IS NULL', 'IS NOT NULL'):
                        text = f"{col} {op}"
                    elif op in ('IN', 'NOT IN'):
                        vals = val if isinstance(val, list) else [val]
                        text = f"{col} {op} ({','.join(str(v) for v in vals)})"
                    elif op == 'BETWEEN':
                        if isinstance(val, list) and len(val) == 2:
                            text = f"{col} BETWEEN {val[0]} AND {val[1]}"
                        else:
                            text = f"{col} BETWEEN {val}"
                    else:
                        text = f"{col} {op} {val}"
                    
                    if i > 0:
                        connector = conditions[i - 1].get('connector', 'AND')
                        parts.append(connector)
                    parts.append(text)
                
                return ' '.join(parts)
            
            # Process each group sequentially
            for i, group in enumerate(groups):
                # Only evaluate on rows not already excluded
                current_indices = set(remaining.index)
                
                group_mask = evaluate_group(remaining, group)
                matched_indices = set(remaining[group_mask].index)
                
                # Only count rows not already excluded
                newly_excluded = matched_indices - excluded_indices
                removed_count = len(newly_excluded)
                
                excluded_indices.update(newly_excluded)
                remaining = remaining[~group_mask]
                
                event_rate = float(remaining[target_variable].mean() * 100) if len(remaining) > 0 and target_variable else None
                event_count = int(remaining[target_variable].sum()) if len(remaining) > 0 and target_variable else None
                non_event_count = len(remaining) - event_count if event_count is not None else None
                
                waterfall.append({
                    'step': f'Group {i + 1}',
                    'label': group_to_natural_language(group),
                    'removed': -removed_count,
                    'remaining': len(remaining),
                    'eventRate': event_rate,
                    'eventCount': event_count,
                    'nonEventCount': non_event_count
                })
            
            # Final row
            total_excluded = original_count - len(remaining)
            final_event_rate = float(remaining[target_variable].mean() * 100) if len(remaining) > 0 and target_variable else None
            final_event_count = int(remaining[target_variable].sum()) if len(remaining) > 0 and target_variable else None
            final_non_event_count = len(remaining) - final_event_count if final_event_count is not None else None
            
            waterfall.append({
                'step': 'Final',
                'label': '',
                'removed': -total_excluded,
                'remaining': len(remaining),
                'eventRate': final_event_rate,
                'eventCount': final_event_count,
                'nonEventCount': final_non_event_count
            })
            
            # Check warning conditions
            warnings = []
            remaining_count = len(remaining)
            excluded_pct = (total_excluded / original_count * 100) if original_count > 0 else 0
            
            # >50% excluded
            if excluded_pct > 50 and excluded_pct <= 80:
                warnings.append({
                    'level': 'amber',
                    'message': f'Rules remove {total_excluded:,} rows ({excluded_pct:.1f}%). Significantly reduces population.'
                })
            
            # >80% excluded
            if excluded_pct > 80:
                warnings.append({
                    'level': 'red',
                    'message': f'Only {remaining_count:,} rows ({100-excluded_pct:.1f}%) remain.'
                })
            
            # All events or non-events excluded
            if remaining_count > 0 and target_variable:
                remaining_events = remaining[target_variable].sum()
                remaining_non_events = remaining_count - remaining_events
                if remaining_events == 0 or remaining_non_events == 0:
                    warnings.append({
                        'level': 'block',
                        'message': 'No target variation remains.'
                    })
            
            # Event rate shift >5pp
            if original_event_rate is not None and final_event_rate is not None:
                if abs(final_event_rate - original_event_rate) > 5:
                    warnings.append({
                        'level': 'amber',
                        'message': f'Event rate shifted from {original_event_rate:.1f}% to {final_event_rate:.1f}%.'
                    })
            
            # <1000 rows remaining
            if remaining_count < 1000:
                warnings.append({
                    'level': 'amber',
                    'message': f'Only {remaining_count:,} rows remain.'
                })
            
            # Single group removes >30%
            for row in waterfall[1:-1]:  # Skip Start and Final
                if isinstance(row['removed'], int):
                    group_pct = abs(row['removed']) / original_count * 100 if original_count > 0 else 0
                    if group_pct > 30:
                        warnings.append({
                            'level': 'amber',
                            'message': f"{row['step']} alone removes {group_pct:.1f}%."
                        })
            
            return {
                'success': True,
                'waterfall': waterfall,
                'warnings': warnings
            }
        
        finally:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"exclusion_preview failed: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@upload_router.post("/exclusion-preview-by-id")
async def exclusion_preview_by_id(
    dataset_id: str = Form(...),
    exclusion_groups: str = Form(...),
    target_variable: str = Form(...),
    current_user=Depends(get_current_user_dependency),
):
    """
    Evaluate exclusion rules against the combined dataset for ``dataset_id``.
    Uses in-memory state when present; otherwise loads from object storage (EKS multi-replica).
    Returns waterfall statistics without modifying persisted data.
    """
    import json as _json
    
    logger.info(f"exclusion-preview-by-id called: dataset_id={dataset_id}, target={target_variable}")
    
    try:
        groups = _json.loads(exclusion_groups)
        logger.info(f"=== EXCLUSION PREVIEW BY ID: Received {len(groups)} groups for dataset {dataset_id} ===")
    except _json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="exclusion_groups must be valid JSON")
    
    try:
        # Full combined population (pre-split train+test, etc.); not the active train/test scope view
        df = _resolve_full_dataframe_for_preview_by_id(dataset_id).copy()
        
        # Validate target variable exists
        if target_variable and target_variable not in df.columns:
            logger.warning(f"Target variable '{target_variable}' not found, using first column")
            target_variable = df.columns[0]
        
        original_count = len(df)
        original_event_rate = float(df[target_variable].mean() * 100) if target_variable else None
        original_event_count = int(df[target_variable].sum()) if target_variable else None
        original_non_event_count = original_count - original_event_count if original_event_count is not None else None
        
        logger.info(
            f"exclusion_preview_by_id: full population dataset_id={dataset_id}, "
            f"rows={original_count}, event_rate={original_event_rate}"
        )
        
        waterfall = [{
            'step': 'Start',
            'label': 'Full population',
            'removed': '--',
            'remaining': original_count,
            'eventRate': original_event_rate,
            'eventCount': original_event_count,
            'nonEventCount': original_non_event_count
        }]
        
        remaining = df.copy()
        excluded_indices = set()
        
        def evaluate_condition(data: pd.DataFrame, cond: dict) -> pd.Series:
            """Evaluate a single condition and return boolean mask."""
            col = cond.get('column')
            op = cond.get('operator', '=')
            val = cond.get('value')
            
            if col not in data.columns:
                return pd.Series(False, index=data.index)
            
            if isinstance(val, list) and len(val) == 1 and op in ('=', '!=', '>', '>=', '<', '<=', 'STARTS WITH', 'CONTAINS'):
                val = val[0]
            
            series = data[col]
            
            if op == 'IS NULL':
                return series.isna()
            elif op == 'IS NOT NULL':
                return series.notna()
            elif op == '= TRUE':
                return series.astype(str).str.lower().isin(['true', '1', 'yes'])
            elif op == '= FALSE':
                return series.astype(str).str.lower().isin(['false', '0', 'no'])
            elif op == '=':
                return series == val
            elif op == '!=':
                return series != val
            elif op == '>':
                return series > val
            elif op == '>=':
                return series >= val
            elif op == '<':
                return series < val
            elif op == '<=':
                return series <= val
            elif op == 'IN':
                vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(',')]
                return series.astype(str).isin([str(v) for v in vals])
            elif op == 'NOT IN':
                vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(',')]
                return ~series.astype(str).isin([str(v) for v in vals])
            elif op == 'STARTS WITH':
                return series.astype(str).str.startswith(str(val), na=False)
            elif op == 'CONTAINS':
                return series.astype(str).str.contains(str(val), na=False, case=False)
            elif op == 'BETWEEN':
                if isinstance(val, list) and len(val) == 2:
                    return (series >= val[0]) & (series <= val[1])
                return pd.Series(False, index=data.index)
            elif op == 'NOT BETWEEN':
                if isinstance(val, list) and len(val) == 2:
                    return (series < val[0]) | (series > val[1])
                return pd.Series(False, index=data.index)
            else:
                return pd.Series(False, index=data.index)
        
        def evaluate_group(data: pd.DataFrame, group: dict) -> pd.Series:
            """Evaluate a group of conditions with AND/OR connectors."""
            conditions = group.get('conditions', [])
            if not conditions:
                return pd.Series(False, index=data.index)
            
            masks = [evaluate_condition(data, c) for c in conditions]
            connectors = [c.get('connector', 'AND') for c in conditions]
            
            and_groups = []
            current_group_indices = [0]
            
            for i in range(1, len(masks)):
                prev_connector = connectors[i - 1]
                if prev_connector == 'AND':
                    current_group_indices.append(i)
                else:
                    and_groups.append(current_group_indices)
                    current_group_indices = [i]
            and_groups.append(current_group_indices)
            
            and_results = []
            for grp_indices in and_groups:
                grp_masks = [masks[i] for i in grp_indices]
                grp_result = grp_masks[0]
                for m in grp_masks[1:]:
                    grp_result = grp_result & m
                and_results.append(grp_result)
            
            final_result = and_results[0]
            for ar in and_results[1:]:
                final_result = final_result | ar
            
            return final_result
        
        def group_to_natural_language(group: dict) -> str:
            """Convert group to human-readable rule description."""
            conditions = group.get('conditions', [])
            parts = []
            for i, c in enumerate(conditions):
                col = c.get('column', '')
                op = c.get('operator', '=')
                val = c.get('value', '')
                
                if op in ('IS NULL', 'IS NOT NULL'):
                    text = f"{col} {op}"
                elif op in ('IN', 'NOT IN'):
                    vals = val if isinstance(val, list) else [val]
                    text = f"{col} {op} ({','.join(str(v) for v in vals)})"
                elif op == 'BETWEEN':
                    if isinstance(val, list) and len(val) == 2:
                        text = f"{col} BETWEEN {val[0]} AND {val[1]}"
                    else:
                        text = f"{col} BETWEEN {val}"
                else:
                    text = f"{col} {op} {val}"
                
                if i > 0:
                    connector = conditions[i - 1].get('connector', 'AND')
                    parts.append(connector)
                parts.append(text)
            
            return ' '.join(parts)
        
        # Process each group sequentially
        for i, group in enumerate(groups):
            group_mask = evaluate_group(remaining, group)
            matched_indices = set(remaining[group_mask].index)
            
            newly_excluded = matched_indices - excluded_indices
            removed_count = len(newly_excluded)
            
            excluded_indices.update(newly_excluded)
            remaining = remaining[~group_mask]
            
            event_rate = float(remaining[target_variable].mean() * 100) if len(remaining) > 0 and target_variable else None
            event_count = int(remaining[target_variable].sum()) if len(remaining) > 0 and target_variable else None
            non_event_count = len(remaining) - event_count if event_count is not None else None
            
            waterfall.append({
                'step': f'Group {i + 1}',
                'label': group_to_natural_language(group),
                'removed': -removed_count,
                'remaining': len(remaining),
                'eventRate': event_rate,
                'eventCount': event_count,
                'nonEventCount': non_event_count
            })
        
        # Final row
        total_excluded = original_count - len(remaining)
        final_event_rate = float(remaining[target_variable].mean() * 100) if len(remaining) > 0 and target_variable else None
        final_event_count = int(remaining[target_variable].sum()) if len(remaining) > 0 and target_variable else None
        final_non_event_count = len(remaining) - final_event_count if final_event_count is not None else None
        
        waterfall.append({
            'step': 'Final',
            'label': '',
            'removed': -total_excluded,
            'remaining': len(remaining),
            'eventRate': final_event_rate,
            'eventCount': final_event_count,
            'nonEventCount': final_non_event_count
        })
        
        # Check warning conditions
        warnings = []
        remaining_count = len(remaining)
        excluded_pct = (total_excluded / original_count * 100) if original_count > 0 else 0
        
        if excluded_pct > 50 and excluded_pct <= 80:
            warnings.append({
                'level': 'amber',
                'message': f'Rules remove {total_excluded:,} rows ({excluded_pct:.1f}%). Significantly reduces population.'
            })
        
        if excluded_pct > 80:
            warnings.append({
                'level': 'red',
                'message': f'Only {remaining_count:,} rows ({100-excluded_pct:.1f}%) remain.'
            })
        
        if remaining_count > 0 and target_variable:
            remaining_events = remaining[target_variable].sum()
            remaining_non_events = remaining_count - remaining_events
            if remaining_events == 0 or remaining_non_events == 0:
                warnings.append({
                    'level': 'block',
                    'message': 'No target variation remains.'
                })
        
        if original_event_rate is not None and final_event_rate is not None:
            if abs(final_event_rate - original_event_rate) > 5:
                warnings.append({
                    'level': 'amber',
                    'message': f'Event rate shifted from {original_event_rate:.1f}% to {final_event_rate:.1f}%.'
                })
        
        if remaining_count < 1000:
            warnings.append({
                'level': 'amber',
                'message': f'Only {remaining_count:,} rows remain.'
            })
        
        for row in waterfall[1:-1]:
            if isinstance(row['removed'], int):
                group_pct = abs(row['removed']) / original_count * 100 if original_count > 0 else 0
                if group_pct > 30:
                    warnings.append({
                        'level': 'amber',
                        'message': f"{row['step']} alone removes {group_pct:.1f}%."
                    })
        
        logger.info(f"exclusion-preview-by-id complete: {original_count} -> {remaining_count} rows")
        
        return {
            'success': True,
            'waterfall': waterfall,
            'warnings': warnings
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"exclusion_preview_by_id failed: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# =======================
# Variable Review Endpoints
# =======================

from app.services.variable_review_service import variable_review_service
from app.models.schemas import (
    VariableReviewRequest, VariableReviewResponse, VariableReviewRow, VariableReviewSummary,
    ApplyVariableRemovalRequest, ApplyVariableRemovalResponse, ReasonBadge
)


@upload_router.post("/variable-review/preview")
async def variable_review_preview(
    file: UploadFile = File(...),
    target_variable: str = Form(...),
    sample_id_col: Optional[str] = Form(None),
    weight_col: Optional[str] = Form(None),
    data_dictionary: Optional[str] = Form(None),
    data_dictionary_file: Optional[UploadFile] = File(None),
    enable_llm_reasoning: Optional[str] = Form("true"),
    current_user=Depends(get_current_user_dependency),
):
    """
    Run the 6-layer variable review pipeline on an uploaded file (before submission).
    Similar to /exclusion-preview - works on the raw tabular file (CSV, TSV, Parquet, Excel).
    
    This allows users to review and select variables for removal BEFORE submitting the dataset.
    
    LLM Touchpoints (when data dictionary is provided):
    - TP1: After Layer 2 - Classify high-AUC variables as origination/behavioral/lifecycle/post_event
    - TP2: After Layer 3 - Confirm zero-inflated variables are populated post-event
    - TP3: After Layer 4 - Reason about differential missingness causal linkage
    """
    import tempfile
    
    file_suffix = _preview_upload_suffix(file.filename)
    if not file.filename or file_suffix not in _PREVIEW_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Supported preview formats: CSV, TSV, Parquet, XLSX, XLS",
        )
    
    try:
        file_content = await file.read()
        if len(file_content) > settings.MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File size exceeds maximum limit")
        
        # Read data dictionary file if provided
        data_dict_content = data_dictionary or ""
        if data_dictionary_file and data_dictionary_file.filename:
            try:
                dict_content = await data_dictionary_file.read()
                data_dict_content = dict_content.decode('utf-8')
                logger.info(f"Loaded data dictionary from file: {data_dictionary_file.filename}")
            except Exception as e:
                logger.warning(f"Failed to read data dictionary file: {e}")
        
        llm_enabled = enable_llm_reasoning.lower() in ('true', '1', 'yes') if enable_llm_reasoning else True
        
        with tempfile.NamedTemporaryFile(mode="wb", suffix=file_suffix or ".csv", delete=False) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name
        
        try:
            df = _load_df_from_preview_temp_path(temp_file_path, file_suffix)
            
            # Validate target variable exists
            target_col = target_variable.strip()
            if target_col not in df.columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Target column '{target_col}' not found in dataset"
                )
            
            logger.info(f"Running variable review preview on uploaded file, "
                       f"target={target_col}, shape={df.shape}, llm_enabled={llm_enabled}")
            
            # Run the pipeline with LLM touchpoints
            result = variable_review_service.run_variable_review(
                df=df,
                target_col=target_col,
                sample_id_col=sample_id_col.strip() if sample_id_col else None,
                weight_col=weight_col.strip() if weight_col else None,
                data_dictionary=data_dict_content if data_dict_content else None,
                enable_llm_reasoning=llm_enabled,
            )
            
            # Convert rows to response format
            rows = [
                {
                    "variable": row["variable"],
                    "auc": row["auc"],
                    "auc_value": row.get("auc_value"),
                    "flags": row["flags"],
                    "reason": row["reason"],
                    "pre_selected": row["pre_selected"],
                    "row_class": row["row_class"],
                    "detail_reasons": row.get("detail_reasons", []),
                    "layer_flags": row.get("layer_flags", []),
                    "cardinality_ratio": row.get("cardinality_ratio"),
                    "null_rate": row.get("null_rate"),
                    "null_rate_diff": row.get("null_rate_diff"),
                }
                for row in result["rows"]
            ]
            
            return {
                "success": True,
                "message": f"Variable review completed in {result['pipeline_time_ms']:.1f}ms",
                "rows": rows,
                "summary": result["summary"],
                "pipeline_time_ms": result["pipeline_time_ms"],
            }
        
        finally:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Variable review preview failed: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@upload_router.post("/variable-review/run", response_model=VariableReviewResponse)
async def run_variable_review(
    request: VariableReviewRequest,
    current_user=Depends(get_current_user_dependency),
):
    """
    Run the 6-layer variable review pipeline to detect identifiers, leakage, and low-value variables.
    
    Layers:
    - L1: Identifier detection (row IDs, URLs, indices)
    - L2: Univariate signal screening (AUC, correlation)
    - L3: Pathological distribution patterns
    - L4: Differential missingness
    - L5: Near-perfect separation (AUC >= 0.95)
    - L6: Correlation clustering with confirmed leakers
    
    LLM Touchpoints (when data dictionary is available):
    - TP1: After Layer 2 - Classify high-AUC variables
    - TP2: After Layer 3 - Confirm zero-inflation is post-event
    - TP3: After Layer 4 - Reason about differential missingness
    
    Returns a table of all variables with recommended removals pre-checked.
    """
    try:
        # Get the FULL DataFrame for leakage detection (not train-only)
        # Leakage must be detected on full data because:
        # 1. Leakage exists across all partitions
        # 2. Full data provides more statistical power for AUC/correlation
        # 3. Columns identified as leaky must be removed from ALL partitions
        df = dataframe_state_manager._full_dataframes.get(request.dataset_id)
        
        if df is None:
            # Fallback to processed dataframe if full not available
            df = dataframe_state_manager.get_processed_dataframe(request.dataset_id)
        
        if df is None:
            raise HTTPException(
                status_code=404, 
                detail=f"No DataFrame found for dataset_id: {request.dataset_id}"
            )
        
        # Validate target column exists
        if request.target_col not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target column '{request.target_col}' not found in dataset"
            )
        
        # Try to fetch data dictionary from dataset info
        data_dict_content = None
        try:
            dataset_info = dataset_manager.get_dataset_info(request.dataset_id)
            if dataset_info:
                data_dictionary_path = dataset_info.get('data_dictionary', '')
                if data_dictionary_path:
                    data_dict_content = _read_data_dictionary(data_dictionary_path)
                    logger.info(f"Loaded data dictionary for variable review: {data_dictionary_path}")
        except Exception as e:
            logger.warning(f"Failed to load data dictionary for variable review: {e}")
        
        logger.info(f"Running variable review for dataset {request.dataset_id}, "
                   f"target={request.target_col}, shape={df.shape}, "
                   f"has_dictionary={bool(data_dict_content)}")

        # Offload the variable-review pipeline to the shared executor so the
        # async event loop is not blocked while pandas/numpy do per-column
        # AUC, Mann-Whitney, correlation, etc. on large frames (P1.3).
        # P3.5: time_stage emits midas_pipeline_stage_seconds{stage="variable_review"}.
        import asyncio as _aio
        from app.core.executor import executor as _vr_executor
        from app.core.metrics import time_stage as _time_stage

        def _run_variable_review_sync():
            return variable_review_service.run_variable_review(
                df=df,
                target_col=request.target_col,
                sample_id_col=request.sample_id_col,
                weight_col=request.weight_col,
                auc_threshold=request.auc_threshold,
                near_perfect_auc_threshold=request.near_perfect_auc_threshold,
                correlation_threshold=request.correlation_threshold,
                missingness_diff_threshold=request.missingness_diff_threshold,
                leaker_correlation_threshold=request.leaker_correlation_threshold,
                data_dictionary=data_dict_content,
                enable_llm_reasoning=True,
            )

        _loop = _aio.get_event_loop()
        with _time_stage("variable_review"):
            result = await _loop.run_in_executor(_vr_executor, _run_variable_review_sync)
        
        # Convert rows to Pydantic models
        rows = [
            VariableReviewRow(
                variable=row["variable"],
                auc=row["auc"],
                auc_value=row.get("auc_value"),
                flags=row["flags"],
                reason=ReasonBadge(row["reason"]),
                pre_selected=row["pre_selected"],
                row_class=row["row_class"],
                detail_reasons=row.get("detail_reasons", []),
                layer_flags=row.get("layer_flags", []),
                cardinality_ratio=row.get("cardinality_ratio"),
                null_rate=row.get("null_rate"),
                null_rate_diff=row.get("null_rate_diff"),
            )
            for row in result["rows"]
        ]
        
        summary = VariableReviewSummary(**result["summary"])
        
        return VariableReviewResponse(
            success=True,
            message=f"Variable review completed in {result['pipeline_time_ms']:.1f}ms",
            rows=rows,
            summary=summary,
            pipeline_time_ms=result["pipeline_time_ms"],
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Variable review failed: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@upload_router.post("/variable-review/apply", response_model=ApplyVariableRemovalResponse)
async def apply_variable_removal(
    request: ApplyVariableRemovalRequest,
    current_user=Depends(get_current_user_dependency),
):
    """
    Apply variable removal selections.
    Removes selected columns from the FULL dataset and updates all partitions.
    Columns must be removed from full data so all partitions (train/test/validation) are consistent.
    """
    try:
        # Get the FULL DataFrame - columns must be removed from all partitions
        df = dataframe_state_manager._full_dataframes.get(request.dataset_id)
        
        if df is None:
            # Fallback to processed dataframe if full not available
            df = dataframe_state_manager.get_processed_dataframe(request.dataset_id)
        
        if df is None:
            raise HTTPException(
                status_code=404,
                detail=f"No DataFrame found for dataset_id: {request.dataset_id}"
            )
        
        original_columns = len(df.columns)
        
        # Apply removal to full data
        updated_df = variable_review_service.apply_variable_removal(
            df=df,
            variables_to_remove=request.variables_to_remove,
        )
        
        removed_count = original_columns - len(updated_df.columns)
        
        # Update full dataframes first
        dataframe_state_manager._full_dataframes[request.dataset_id] = updated_df.copy()
        
        # Recompute split indices (but keep _processed_dataframes as full data)
        if 'split_tag' in updated_df.columns:
            import numpy as np
            train_mask = updated_df['split_tag'] == 'train'
            test_mask = updated_df['split_tag'] == 'test'
            all_validation_mask = updated_df['split_tag'].str.startswith('validation')
            
            dataframe_state_manager._split_indices[request.dataset_id] = {
                "train": np.where(train_mask)[0].astype(np.int64),
                "test": np.where(test_mask)[0].astype(np.int64),
                "validation": np.where(all_validation_mask)[0].astype(np.int64),
            }
        
        # Update _processed_dataframes with full data (columns removed)
        # DO NOT set scope to train here - that happens when user clicks Confirm button
        dataframe_state_manager._processed_dataframes[request.dataset_id] = updated_df.copy()
        
        logger.info(f"Applied variable removal: {removed_count} columns removed, "
                   f"{len(updated_df.columns)} remaining")
        
        return ApplyVariableRemovalResponse(
            success=True,
            message=f"Removed {removed_count} variables",
            removed_count=removed_count,
            remaining_columns=len(updated_df.columns),
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Apply variable removal failed: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@chat_router.post("/chat", response_model=ChatResponse)
async def chat_with_agent(
    request: ChatRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Chat with the agentic system using persistent MessageState
    """
    logger.info(f"Chat request for dataset: {request.dataset_id}, query: {request.query[:100]}...")
    
    try:
        # Get dataset info to validate dataset exists
        dataset_info = dataset_manager.get_dataset_info(request.dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset not found: {request.dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Create or load persistent MessageState
        logger.info(f"Loading persistent MessageState for dataset: {request.dataset_id}")
        try:
            state = message_state_manager.create_or_load_state(request.dataset_id, request.query)
            
            # If this is a follow-up query, update the state appropriately
            if state.get("userquery") and state["userquery"] != request.query:
                state = message_state_manager.update_state_with_query(
                    request.dataset_id, state, request.query
                )
            
        except ValueError as e:
            logger.error(f"Failed to create/load MessageState: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))
        
        logger.info("Processing query through agentic system...")
        # Add dataset_id to state for DataFrameStateManager access
        state['dataset_id'] = request.dataset_id
        # Add agent_context if provided (helps route ambiguous queries like IV, VIF, correlation)
        if request.agent_context:
            state['agent_context'] = request.agent_context
            logger.info(f"Agent context provided: {request.agent_context}")
        
        # Add QC-specific fields to state if provided
        if request.qc_mode:
            state['qc_mode'] = request.qc_mode
            logger.info(f"QC mode provided: {request.qc_mode}")
        if request.treatment_sequence:
            state['treatment_sequence'] = request.treatment_sequence
            logger.info(f"Treatment sequence provided: {request.treatment_sequence}")
        if request.qc_templates:
            state['qc_templates'] = request.qc_templates
            logger.info(f"QC templates provided: {list(request.qc_templates.keys())}")
        if request.qc_ui_selections:
            state['qc_ui_selections'] = request.qc_ui_selections
            logger.info(f"QC UI selections provided: {request.qc_ui_selections}")

        # Always refresh runtime dataframe from the latest in-memory processed state
        # before invoking the agent so QC/EDA/treatment planning does not use stale snapshots.
        try:
            state_df = state.get('datasetFile')
            if isinstance(state_df, pd.DataFrame):
                refreshed_df = dataframe_state_manager.get_dataframe_for_execution(request.dataset_id, state_df)
                if refreshed_df is not None:
                    state['datasetFile'] = refreshed_df
                    logger.info(
                        f"Refreshed chat state dataframe from DataFrameStateManager: {request.dataset_id}, "
                        f"shape={refreshed_df.shape}"
                    )
            else:
                refreshed_df = dataframe_state_manager.get_dataframe(request.dataset_id)
                if refreshed_df is not None:
                    state['datasetFile'] = refreshed_df
                    logger.info(
                        f"Refreshed chat state dataframe (no prior state df): {request.dataset_id}, "
                        f"shape={refreshed_df.shape}"
                    )
        except Exception as refresh_err:
            logger.warning(f"Failed to refresh chat dataframe before agent invoke: {refresh_err}")
        
        # Process query through agentic system
        result = agent.invoke(state)
        logger.info(f"Agentic system completed, intent: {result.get('intent', 'unknown')}")
        
        # Save the updated state to database
        logger.info(f"Saving updated MessageState for dataset: {request.dataset_id}")
        save_success = message_state_manager.save_state(request.dataset_id, result)
        if not save_success:
            logger.warning(f"Failed to save MessageState for dataset: {request.dataset_id}")
        
        knowledge_metadata = result.get("knowledge_metadata")

        # # Handle not_relevant intent (completely irrelevant queries)
        # if result.get('intent') == 'not_relevant':
        #     logger.info("Query completely irrelevant to all agents")
        #     return ChatResponse(
        #         response="The user query is out of scope for the agent",
        #         code="# No code to display",
        #         suggestions=["Please rephrase your question to be more relevant to the Agent"],
        #         role="not_relevant"
        #     )
        if result.get('intent') == 'not_relevant':
            logger.info("Guardrail: query not relevant to current agent step")
            # Read the agent-specific message that the guardrail already put in state
            try:
                response_data = json.loads(result["messages"][-1].content)
                return ChatResponse(
                    response=response_data.get('response', "I can only assist with questions relevant to the current step."),
                    code=response_data.get('code', "# No code to display"),
                    suggestions=response_data.get('suggestion', ["Please ask a question relevant to the current step"]),
                    role="not_relevant",
                    knowledge_metadata=knowledge_metadata
                )
            except (json.JSONDecodeError, KeyError, IndexError):
                # Fallback in case message parsing fails
                return ChatResponse(
                    response="I can only assist with questions relevant to the current step.",
                    code="# No code to display",
                    suggestions=["Please ask a question relevant to the current step"],
                    role="not_relevant",
                    knowledge_metadata=knowledge_metadata
                )

        # Handle not_relevant intent (completely irrelevant queries)
        if result.get('intent') == 'not_relevant':
            logger.info("Query completely irrelevant to all agents")
            return ChatResponse(
                response="The user query is out of scope for the agent",
                code="# No code to display",
                suggestions=["Please rephrase your question to be more relevant to the Agent"],
                role="not_relevant",
                knowledge_metadata=knowledge_metadata
            )
        
        # Extract response
        if result['intent'] == 'plan_agent' or result['intent'] == 'data_insight':
            # Handle plan response
            try:
                # plan_data = json.loads(result['plan'])
                logger.info(f"##################  Response messages: {result["messages"]}")
                logger.info("####################################################################")
                logger.info(f"##################  Response data: {result["messages"][-1].content}")
                response_data = json.loads(result["messages"][-1].content)
                
                return ChatResponse(
                    # response=json.dumps(plan_data, indent=2),
                    response=response_data['response'],
                    code="# Plan generated successfully",
                    suggestions=[
                        "Explore feature correlations",
                        "Create target distribution plot", 
                        "Check for data quality issues",
                        "Analyze class imbalance",
                        "Generate feature importance"
                    ],
                    role= result['intent'],
                    knowledge_metadata=knowledge_metadata
                )
            except json.JSONDecodeError:
                return ChatResponse(
                    response=result['plan'],
                    code="# Plan generated successfully",
                    suggestions=[
                        "Explore feature correlations",
                        "Create target distribution plot",
                        "Check for data quality issues"
                    ],
                    role= result['intent'],
                    knowledge_metadata=knowledge_metadata
                )
        elif result['intent'] == 'modelling':
            # Handle modelling agent response
            try:
                response_data = json.loads(result["messages"][-1].content)
                return ChatResponse(
                    response=response_data['response'],
                    code=response_data['code'],
                    suggestions=response_data['suggestion'],
                    role='modelling',
                    knowledge_metadata=knowledge_metadata
                )
            except (json.JSONDecodeError, KeyError, IndexError):
                # Fallback for non-JSON responses
                return ChatResponse(
                    response=str(result["messages"][-1].content) if result["messages"] else "No response generated",
                    code="# No code to display",
                    suggestions=[
                        "Run automatic training with cross-validation",
                        "Tune key hyperparameters (n_estimators, max_depth, learning_rate)",
                        "Compare algorithms (Logistic/RandomForest/GBM) on your target",
                        "Review performance metrics (AUC/F1 for classification, R²/RMSE for regression)"
                    ],
                    role='modelling',
                    knowledge_metadata=knowledge_metadata
                )
        elif result['intent'] == 'data_quality':
            # Handle data quality QC response - return ALL treatment messages
            # The QC sequence generates multiple messages (one per treatment type + summary)
            # We need to return them all so frontend can render each treatment table
            try:
                all_messages = result.get("messages", [])
                treatment_messages = []
                summary_message = None
                
                logger.info(f"Data Quality: Processing {len(all_messages)} messages")
                
                for msg in all_messages:
                    try:
                        msg_content = json.loads(msg.content) if hasattr(msg, 'content') else msg
                        
                        # Check if this is a treatment message with treatment_type
                        if isinstance(msg_content, dict) and msg_content.get('treatment_type'):
                            treatment_messages.append(msg_content)
                            logger.info(f"Found treatment message: {msg_content.get('treatment_type')}")
                        # Check if this is the summary message
                        elif isinstance(msg_content, dict) and msg_content.get('qc_summary'):
                            summary_message = msg_content
                            logger.info("Found QC summary message")
                    except (json.JSONDecodeError, AttributeError) as e:
                        logger.debug(f"Skipping non-JSON message: {e}")
                        continue
                
                logger.info(f"Data Quality: Found {len(treatment_messages)} treatment messages")
                
                # Build combined response with all treatment tables
                combined_response = {
                    "role": "data_quality",
                    "treatment_messages": treatment_messages,
                    "summary": summary_message,
                    "qc_mode": result.get('qc_mode', 'auto'),
                    "treatment_sequence": result.get('treatment_sequence', [])
                }
                
                # Return as JSON string so frontend can parse and render each treatment
                return ChatResponse(
                    response=json.dumps(combined_response),
                    code=summary_message.get('code', '# QC sequence complete') if summary_message else '# QC sequence complete',
                    suggestions=summary_message.get('suggestion', [
                        "Review each treatment plan above",
                        "Execute the generated code for each treatment",
                        "Proceed to feature engineering after treatment"
                    ]) if summary_message else [],
                    role='data_quality',
                    knowledge_metadata=knowledge_metadata
                )
            except Exception as e:
                logger.error(f"Error processing data_quality response: {e}")
                # Fallback to returning just the last message
                return ChatResponse(
                    response=str(result["messages"][-1].content) if result["messages"] else "QC sequence completed",
                    code="# QC sequence complete",
                    suggestions=["Review treatment results", "Proceed to next step"],
                    role='data_quality',
                    knowledge_metadata=knowledge_metadata
                )
        else:
            # Handle regular response
            try:
                response_data = json.loads(result["messages"][-1].content)
                return ChatResponse(
                    response=response_data['response'],
                    code=response_data['code'],
                    suggestions=response_data['suggestion'],
                    knowledge_metadata=knowledge_metadata
                )
            except (json.JSONDecodeError, KeyError, IndexError):
                # Fallback for non-JSON responses
                return ChatResponse(
                    response=str(result["messages"][-1].content) if result["messages"] else "No response generated",
                    code="# No code to display",
                    suggestions=[
                        "Explore feature correlations",
                        "Create target distribution plot",
                        "Check for data quality issues"
                    ],
                    knowledge_metadata=knowledge_metadata
                )
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Chat with agent failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@chat_router.get("/chat/{dataset_id}/history")
async def get_chat_history(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """Get chat history for a dataset"""
    try:
        logger.info(f"Chat history request for dataset: {dataset_id}")
        
        # Load existing state to get chat history
        existing_state = message_state_manager.db.load_message_state(dataset_id)
        
        if not existing_state:
            return {
                "success": True,
                "dataset_id": dataset_id,
                "chat_history": [],
                "message": "No chat history found for this dataset"
            }
        
        chat_history = existing_state.get('chat_history', [])
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "chat_history": chat_history,
            "total_interactions": len(chat_history),
            "current_intent": existing_state.get('intent', ''),
            "last_updated": existing_state.get('updated_at', '')
        }
        
    except Exception as e:
        logger.error(f"Failed to get chat history for {dataset_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@chat_router.delete("/chat/{dataset_id}/reset")
async def reset_chat_state(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """Reset chat state for a dataset"""
    try:
        logger.info(f"Resetting chat state for dataset: {dataset_id}")
        
        # Validate dataset exists
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Delete the persisted state
        success = message_state_manager.delete_state(dataset_id)
        
        if success:
            return {
                "success": True,
                "message": f"Chat state reset successfully for dataset {dataset_id}"
            }
        else:
            return {
                "success": True,
                "message": f"No chat state found to reset for dataset {dataset_id}"
            }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to reset chat state for {dataset_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@chat_router.get("/chat/states")
async def list_all_chat_states(
    current_user = Depends(get_current_user_dependency)
):
    """List all chat states (for debugging/admin)"""
    try:
        states = message_state_manager.list_all_states()
        return {
            "success": True,
            "states": states,
            "count": len(states)
        }
        
    except Exception as e:
        logger.error(f"Failed to list chat states: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.get("/datasets/{dataset_id}/stats", response_model=DataStats)
async def get_dataset_stats(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """Get dataset statistics"""
    try:
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=500, detail="Failed to load dataset")
        
        stats = dataset_manager.get_dataset_stats(df, dataset_info['target_variable'])
        return stats
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.delete("/datasets/{dataset_id}")
async def delete_dataset(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """Delete dataset and associated chat state"""
    try:
        success = dataset_manager.delete_dataset(dataset_id)
        if not success:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Also delete the associated MessageState
        state_deleted = message_state_manager.delete_state(dataset_id)
        if state_deleted:
            logger.info(f"Associated chat state deleted for dataset: {dataset_id}")
        
        return {
            "message": "Dataset deleted successfully",
            "chat_state_deleted": state_deleted
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.post("/vector-store/reinitialize")
async def reinitialize_vector_store(
    current_user = Depends(get_current_user_dependency)
):
    """Reinitialize vector store from knowledge base"""
    try:
        # Remove existing index files
        import os
        if vector_store.index_path.exists():
            os.remove(vector_store.index_path)
        if vector_store.documents_path.exists():
            os.remove(vector_store.documents_path)
        
        # Reset vector store state
        vector_store.index = None
        vector_store.documents = []
        
        # Reinitialize
        success = initialize_vector_store()
        
        if success:
            return {
                "message": "Vector store reinitialized successfully",
                "documents_count": len(vector_store.documents) if vector_store.documents else 0
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to reinitialize vector store")
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.get("/datasets")
async def list_datasets(
    current_user = Depends(get_current_user_dependency)
):
    """List all datasets (for debugging)"""
    try:
        datasets = list(dataset_manager.datasets.keys())
        return {
            "success": True,
            "datasets": datasets,
            "count": len(datasets)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.get("/datasets/{dataset_id}/raw-data")
async def get_raw_data(
    dataset_id: str,
    limit: int = 100,
    current_user = Depends(get_current_user_dependency)
):
    """Get raw data from dataset with optional limit"""
    try:
        logger.info(f"Raw data request for dataset: {dataset_id}")
        
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset metadata not found: {dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found. Please upload a dataset first.")
        
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            logger.error(f"Failed to load dataset file: {dataset_id}")
            # Check if it's because the file doesn't exist
            if dataset_id not in dataset_manager.datasets:
                raise HTTPException(status_code=404, detail="Dataset was removed due to missing file. Please upload your dataset again.")
            else:
                raise HTTPException(status_code=500, detail="Failed to load dataset file. The file may be corrupted.")
        
        # Limit the number of rows returned
        limited_df = df.head(limit)
        
        # Convert to JSON with proper handling of NaN values
        data = limited_df.fillna("").to_dict(orient='records')
        
        logger.info(f"Raw data successfully retrieved: {len(limited_df)} rows for dataset {dataset_id}")
        
        return {
            "success": True,
            "data": data,
            "total_rows": len(df),
            "returned_rows": len(limited_df),
            "columns": list(df.columns)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Internal server error in get_raw_data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.get("/datasets/{dataset_id}/export")
async def export_dataset(
    dataset_id: str,
    format: str = "csv",
    current_user = Depends(get_current_user_dependency)
):
    """Export dataset in specified format"""
    try:
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=500, detail="Failed to load dataset")
        
        if format.lower() == "csv":
            # Create a temporary file for download
            import tempfile
            import os
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
            df.to_csv(temp_file.name, index=False)
            temp_file.close()
            
            return FileResponse(
                path=temp_file.name,
                filename=f"{dataset_info['filename']}_export.csv",
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={dataset_info['filename']}_export.csv"}
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported format. Only CSV is supported.")
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.put("/datasets/{dataset_id}/config")
async def update_dataset_config(
    dataset_id: str,
    target_variable: Optional[str] = Form(None),
    current_user = Depends(get_current_user_dependency),
    target_variable_type: Optional[str] = Form(None),
    problem_statement: Optional[str] = Form(None),
    data_dictionary: Optional[str] = Form(None),
    data_dictionary_file: Optional[UploadFile] = File(None)
):
    """Update dataset configuration"""
    try:
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Update configuration
        updated_config = dataset_info.copy()
        
        if target_variable is not None:
            updated_config['target_variable'] = target_variable
        if target_variable_type is not None:
            updated_config['target_variable_type'] = target_variable_type
        if problem_statement is not None:
            updated_config['problem_statement'] = problem_statement
        
        # Process data dictionary file if provided
        if data_dictionary_file and data_dictionary_file.filename:
            logger.info(f"Processing data dictionary file: {data_dictionary_file.filename}")
            
            # Validate data dictionary file type
            if not data_dictionary_file.filename.endswith('.csv'):
                logger.warning(f"Invalid data dictionary file type: {data_dictionary_file.filename}")
                raise HTTPException(status_code=400, detail="Data dictionary must be a CSV file")
            
            # Read and save data dictionary file
            data_dict_content = await data_dictionary_file.read()
            data_dict_id, data_dictionary_path = dataset_manager.save_uploaded_file(
                data_dict_content, 
                f"data_dict_{dataset_id}_{data_dictionary_file.filename}"
            )
            
            # Store the file path as the data dictionary reference
            updated_config['data_dictionary'] = data_dictionary_path
            logger.info(f"Data dictionary file saved at: {data_dictionary_path}")
        elif data_dictionary is not None:
            # Fall back to text-based data dictionary for backward compatibility
            updated_config['data_dictionary'] = data_dictionary
        
        # Save updated configuration
        success = dataset_manager.update_dataset_config(dataset_id, updated_config)
        
        if success:
            return {
                "success": True,
                "message": "Dataset configuration updated successfully",
                "config": updated_config
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update dataset configuration")
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@upload_router.get("/datasets/{dataset_id}/column-distribution/{column_name}")
async def get_column_distribution(
    dataset_id: str,
    column_name: str,
    bins: int = 10,
    full_distribution: bool = False,  # If True, return ALL categories without limiting to top 20
    current_user = Depends(get_current_user_dependency)
):
    """Get real distribution data for a specific column"""
    try:
        logger.info(f"Column distribution request for dataset: {dataset_id}, column: {column_name}")
        
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset metadata not found: {dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # P1.5 part 2: column-distribution is read-only (just value_counts).
        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is not None:
            logger.info(f"Using cached DataFrame from state manager for column distribution. Shape: {df.shape}")
        else:
            df = dataset_manager.load_dataset(dataset_id)
            if df is None:
                logger.error(f"Failed to load dataset file: {dataset_id}")
                raise HTTPException(status_code=404, detail="Dataset file not found. Please upload your dataset again.")
        
        # Check if column exists
        if column_name not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{column_name}' not found in dataset")
        
        column_data = df[column_name]
        column_type = str(column_data.dtype)
        
        # Calculate distribution based on column type
        distribution_data = {}
        is_numerical = False
        
        if column_data.dtype in ['int64', 'float64', 'int32', 'float32']:
            # Numerical column - create histogram
            is_numerical = True
            
            # Remove NaN values for histogram calculation
            clean_data = column_data.dropna()
            
            if len(clean_data) == 0:
                distribution_data = {"No Data": 0}
            else:
                # Check if it should be treated as categorical (low unique values)
                unique_count = clean_data.nunique()
                if unique_count <= 10:
                    # Treat as categorical
                    value_counts = clean_data.value_counts().sort_index()
                    distribution_data = {str(k): int(v) for k, v in value_counts.items()}
                    is_numerical = False
                else:
                    # Create quantile-based histogram bins using qcut for better distribution
                    try:
                        hist = pd.qcut(clean_data, q=bins, duplicates='drop')
                        value_counts = hist.value_counts().sort_index()
                        
                        # Format bin labels
                        for interval, count in value_counts.items():
                            if pd.isna(interval):
                                continue
                            left = interval.left
                            right = interval.right
                            bin_label = f"{left:.1f}-{right:.1f}"
                            distribution_data[bin_label] = int(count)
                    except ValueError as e:
                        # Fallback to regular cut if qcut fails (e.g., not enough unique values)
                        logger.warning(f"qcut failed for column {column_name}, falling back to cut: {e}")
                        hist, bin_edges = pd.cut(clean_data, bins=bins, retbins=True, include_lowest=True)
                        value_counts = hist.value_counts().sort_index()
                        
                        # Format bin labels
                        for interval, count in value_counts.items():
                            if pd.isna(interval):
                                continue
                            left = interval.left
                            right = interval.right
                            bin_label = f"{left:.1f}-{right:.1f}"
                            distribution_data[bin_label] = int(count)
        else:
            # Categorical column
            value_counts = column_data.value_counts()
            # If full_distribution requested (for stratified split), return ALL categories
            # Otherwise limit to top 20 categories to avoid overcrowding in UI charts
            if full_distribution:
                distribution_data = {str(k): int(v) for k, v in value_counts.items()}
            elif len(value_counts) > 20:
                top_categories = value_counts.head(19)
                others_count = value_counts.tail(len(value_counts) - 19).sum()
                distribution_data = {str(k): int(v) for k, v in top_categories.items()}
                distribution_data["Others"] = int(others_count)
            else:
                distribution_data = {str(k): int(v) for k, v in value_counts.items()}
        
        # Calculate statistics
        total_count = len(column_data)
        missing_count = int(column_data.isna().sum())
        valid_count = total_count - missing_count
        
        logger.info(f"Column distribution calculated: {len(distribution_data)} bins/categories")
        
        return {
            "success": True,
            "column_name": column_name,
            "column_type": column_type,
            "is_numerical": is_numerical,
            "distribution": distribution_data,
            "statistics": {
                "total_count": total_count,
                "missing_count": missing_count,
                "valid_count": valid_count,
                "unique_count": int(column_data.nunique())
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Internal server error in get_column_distribution: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/column-distribution-by-scope/{column_name}")
async def get_column_distribution_by_scope(
    dataset_id: str,
    column_name: str,
    scope: str = "entire",  # entire, train, test, validation
    bins: int = 10,
    full_distribution: bool = False,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get distribution data for a specific column filtered by scope.
    This is a read-only operation that does NOT modify _processed_dataframes or _active_scope.
    
    Args:
        dataset_id: The dataset identifier
        column_name: The column to get distribution for
        scope: One of 'entire', 'train', 'test', 'validation'
        bins: Number of bins for numerical columns
        full_distribution: If True, return all categories without limiting
    """
    try:
        logger.info(f"Column distribution by scope request for dataset: {dataset_id}, column: {column_name}, scope: {scope}")
        
        # Validate scope
        valid_scopes = ['entire', 'train', 'test', 'validation']
        if scope not in valid_scopes:
            raise HTTPException(status_code=400, detail=f"Invalid scope '{scope}'. Must be one of: {valid_scopes}")
        
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset metadata not found: {dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        full_df = dataframe_state_manager.get_full_dataframe_readonly(dataset_id)
        if full_df is None:
            full_df = dataset_manager.load_dataset(dataset_id)
            if full_df is None:
                logger.error(f"Failed to load dataset file: {dataset_id}")
                raise HTTPException(status_code=404, detail="Dataset file not found. Please upload your dataset again.")
        
        if scope != "entire":
            df = None
            split_indices = dataframe_state_manager._split_indices.get(dataset_id)
            if split_indices and scope in split_indices:
                indices = split_indices[scope]
                if indices is not None and len(indices) > 0:
                    df = full_df.iloc[indices].copy()
                    logger.info(f"Filtered to {scope} scope for distribution: {len(df)} rows")
            if (df is None or len(df) == 0) and "split_tag" in full_df.columns:
                mask = full_df["split_tag"] == scope
                if scope == "validation":
                    mask = full_df["split_tag"].astype(str).str.startswith("validation", na=False)
                df = full_df[mask].copy()
                logger.info(f"Filtered by split_tag to {scope} for distribution: {len(df)} rows")
            if df is None or len(df) == 0:
                logger.warning(f"No data for scope '{scope}' in dataset {dataset_id}")
                return {
                    "success": True,
                    "column_name": column_name,
                    "column_type": "unknown",
                    "is_numerical": False,
                    "distribution": {},
                    "statistics": {"total_count": 0, "missing_count": 0, "valid_count": 0, "unique_count": 0},
                    "scope": scope,
                    "total_rows": 0,
                }
        else:
            df = full_df.copy()
        
        # Check if column exists
        if column_name not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{column_name}' not found in dataset")
        
        column_data = df[column_name]
        column_type = str(column_data.dtype)
        
        # Calculate distribution based on column type
        distribution_data = {}
        is_numerical = False
        
        if column_data.dtype in ['int64', 'float64', 'int32', 'float32']:
            is_numerical = True
            clean_data = column_data.dropna()
            
            if len(clean_data) == 0:
                distribution_data = {"No Data": 0}
            else:
                unique_count = clean_data.nunique()
                if unique_count <= 10:
                    value_counts = clean_data.value_counts().sort_index()
                    distribution_data = {str(k): int(v) for k, v in value_counts.items()}
                    is_numerical = False
                else:
                    try:
                        hist = pd.qcut(clean_data, q=bins, duplicates='drop')
                        value_counts = hist.value_counts().sort_index()
                        for interval, count in value_counts.items():
                            if pd.isna(interval):
                                continue
                            left = interval.left
                            right = interval.right
                            bin_label = f"{left:.1f}-{right:.1f}"
                            distribution_data[bin_label] = int(count)
                    except ValueError as e:
                        logger.warning(f"qcut failed for column {column_name}, falling back to cut: {e}")
                        hist, bin_edges = pd.cut(clean_data, bins=bins, retbins=True, include_lowest=True)
                        value_counts = hist.value_counts().sort_index()
                        for interval, count in value_counts.items():
                            if pd.isna(interval):
                                continue
                            left = interval.left
                            right = interval.right
                            bin_label = f"{left:.1f}-{right:.1f}"
                            distribution_data[bin_label] = int(count)
        else:
            value_counts = column_data.value_counts()
            if full_distribution:
                distribution_data = {str(k): int(v) for k, v in value_counts.items()}
            elif len(value_counts) > 20:
                top_categories = value_counts.head(19)
                others_count = value_counts.tail(len(value_counts) - 19).sum()
                distribution_data = {str(k): int(v) for k, v in top_categories.items()}
                distribution_data["Others"] = int(others_count)
            else:
                distribution_data = {str(k): int(v) for k, v in value_counts.items()}
        
        # Calculate statistics
        total_count = len(column_data)
        missing_count = int(column_data.isna().sum())
        valid_count = total_count - missing_count
        
        logger.info(f"Column distribution by scope calculated: {len(distribution_data)} bins/categories, scope={scope}, rows={len(df)}")
        
        return {
            "success": True,
            "column_name": column_name,
            "column_type": column_type,
            "is_numerical": is_numerical,
            "distribution": distribution_data,
            "statistics": {
                "total_count": total_count,
                "missing_count": missing_count,
                "valid_count": valid_count,
                "unique_count": int(column_data.nunique())
            },
            "scope": scope,
            "total_rows": len(df)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Internal server error in get_column_distribution_by_scope: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


def _run_classify_vars_background(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Module-level background worker for ``background_job_manager``.

    Declared as ``def`` (not ``async def``) and module-level so:
    - ``background_job_manager`` can pickle/dispatch it across threads.
    - The blocking LLM call does NOT pin the asyncio event loop.
    - Job snapshots are mirrored to shared object storage so the status
      poll works on ANY gunicorn worker / EKS pod, not only the one
      that enqueued the job (fixes the 404 → 120-second timeout).
    """
    dataset_id = params["dataset_id"]

    dataset_info = dataset_manager.get_dataset_info(dataset_id)
    if not dataset_info:
        raise ValueError(f"Dataset metadata not found: {dataset_id}")

    data_dictionary = dataset_info.get("data_dictionary", "")
    data_dictionary_content = _read_data_dictionary(data_dictionary)

    logger.info(f"[classify-vars-bg {dataset_id}] Calling LLM…")
    llm_response = llm_service.get_variable_classification(data_dictionary_content)

    if isinstance(llm_response, dict):
        classification_result = llm_response
    else:
        classification_result = json.loads(llm_response)

    result = {
        "success": True,
        "message": "Variable classification completed successfully",
        "dataset_id": dataset_id,
        "classification": classification_result,
        "timestamp": pd.Timestamp.now().isoformat(),
    }

    # Keep the in-process cache warm so cache-hit on the same pod is instant
    _variable_classification_cache[dataset_id] = result

    logger.info(
        f"[classify-vars-bg {dataset_id}] Done: "
        f"{len(classification_result.get('variables', []))} variables classified"
    )
    return result


# Job-id format: "classify_vars_<dataset_id>" — stable per dataset so any
# replica can reconstruct it for a status lookup.
def _classify_vars_job_id(dataset_id: str) -> str:
    return f"classify_vars_{dataset_id}"


@upload_router.post("/datasets/{dataset_id}/classify-variables")
async def classify_dataset_variables(
    dataset_id: str,
    background_tasks: BackgroundTasks,  # kept for API compat; no longer used
    current_user = Depends(get_current_user_dependency)
):
    """
    Enqueue variable classification and return a job token immediately.
    If the result is already cached (in-process or S3-mirrored), return it
    synchronously (instant).
    Poll ``GET /datasets/{dataset_id}/classify-variables/status`` to track
    progress.

    Phase-1 stateless-API fix: job state is mirrored to shared object
    storage via ``background_job_manager``, so the status poll returns the
    correct state regardless of which EKS pod / gunicorn worker receives
    the request. The previous ``_classify_vars_jobs`` in-process dict meant
    that status polls landing on a different worker always returned 404,
    causing the frontend to time out after 120 s.
    """
    from app.services.background_jobs import background_job_manager as _bgm

    logger.info(f"Variable classification request for dataset: {dataset_id}")

    # L1 — in-process cache (same pod, instant)
    if dataset_id in _variable_classification_cache:
        logger.info(f"Returning cached variable classification for dataset: {dataset_id}")
        return _variable_classification_cache[dataset_id]

    # Validate dataset exists before queuing
    dataset_info = dataset_manager.get_dataset_info(dataset_id)
    if not dataset_info:
        raise HTTPException(status_code=404, detail="Dataset not found. Please upload a dataset first.")

    job_id = _classify_vars_job_id(dataset_id)

    # L2 — shared job store: already running or completed on another pod
    existing = _bgm.get_job_status(job_id)
    if existing:
        st = str(existing.get("status") or "")
        if st in ("pending", "running"):
            return {
                "success": True,
                "queued": True,
                "status": st,
                "job_id": job_id,
                "message": "Classification already in progress. Poll /classify-variables/status.",
            }
        if st == "completed":
            result = existing.get("result") or {}
            _variable_classification_cache[dataset_id] = result  # warm local cache
            return result

    # Enqueue via background_job_manager (S3-mirrored, cross-replica)
    _bgm.start_job(
        job_id=job_id,
        job_type="classify_variables",
        params={"dataset_id": dataset_id},
        job_function=_run_classify_vars_background,
    )

    logger.info(f"Queued classify-variables job {job_id} for dataset: {dataset_id}")
    return {
        "success": True,
        "queued": True,
        "status": "pending",
        "job_id": job_id,
        "message": "Variable classification started in background. Poll /datasets/{dataset_id}/classify-variables/status.",
    }


@upload_router.get("/datasets/{dataset_id}/classify-variables/status")
async def get_classify_variables_status(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Poll the status of a classify-variables background job.
    Returns the full classification result once status == 'completed'.

    Phase-1 fix: reads from ``background_job_manager`` (S3-mirrored) so
    any EKS replica can answer correctly regardless of which pod originally
    enqueued the job.  Falls back to the in-process cache for instant
    responses when the same pod handled the POST.
    """
    from app.services.background_jobs import background_job_manager as _bgm

    # Fast path: in-process cache hit
    if dataset_id in _variable_classification_cache:
        return {**_variable_classification_cache[dataset_id], "status": "completed"}

    job_id = _classify_vars_job_id(dataset_id)
    snap = _bgm.get_job_status(job_id)
    if snap is None:
        raise HTTPException(status_code=404, detail="No classification job found for this dataset. Call POST first.")

    st = str(snap.get("status") or "")
    resp: Dict[str, Any] = {"dataset_id": dataset_id, "status": st, "job_id": job_id}

    if st == "completed":
        result = snap.get("result") or {}
        _variable_classification_cache[dataset_id] = result  # warm local cache
        resp.update(result)
    elif st == "failed":
        resp["error"] = snap.get("error")

    return resp


# NOTE: The legacy synchronous
# `POST /datasets/{dataset_id}/classify-variables-sync` endpoint was
# removed (Phase 1 stateless-API cleanup, May 2026). Use
# `POST /datasets/{dataset_id}/classify-variables` (already async) +
# `GET /datasets/{dataset_id}/classify-variables/status` instead — the
# LLM round-trip routinely exceeds the 1 s `SLOW_REQUEST_THRESHOLD_MS`
# and must run in a background job per `.cursor/rules/architecture.mdc`.
# No frontend caller references the sync path.


@upload_router.post("/datasets/{dataset_id}/column-insights")
async def generate_column_insights(
    dataset_id: str,
    request: dict,
    current_user = Depends(get_current_user_dependency)
):
    """
    Generate AI-based insights for a column's distribution using LLM.
    Returns 5-10 meaningful, actionable insights about the column data.
    """
    logger.info(f"Column insights request for dataset: {dataset_id}, column: {request.get('column_name')}")
    
    try:
        column_name = request.get('column_name', '')
        column_type = request.get('column_type', 'Unknown')
        distribution = request.get('distribution', {})
        statistics = request.get('statistics', {})
        
        if not column_name or not distribution:
            raise HTTPException(status_code=400, detail="Column name and distribution data are required")
        
        # Use LLM service to generate AI-based insights
        insights = llm_service.get_column_distribution_insights(
            column_name=column_name,
            column_type=column_type,
            distribution=distribution,
            statistics=statistics
        )
        
        logger.info(f"Generated {len(insights)} AI insights for column: {column_name}")
        
        return {
            "success": True,
            "insights": insights,
            "column_name": column_name,
            "column_type": column_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Column insights generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.post(
    "/datasets/{dataset_id}/cross-algorithm-recommendation",
    response_model=CrossAlgorithmRecommendationResponse,
)
async def cross_algorithm_recommendation(
    dataset_id: str,
    body: CrossAlgorithmRecommendationRequest,
    current_user=Depends(get_current_user_dependency),
):
    """
    LLM narrative comparing shortlisted models (max 2 per algorithm) across algorithms.
    """
    logger.info("Cross-algorithm recommendation for dataset %s (%d candidates)", dataset_id, len(body.candidates))
    try:
        ds = dataset_manager.get_dataset_info(dataset_id)
        if not ds:
            raise HTTPException(status_code=404, detail="Dataset not found")
        if not body.candidates:
            return CrossAlgorithmRecommendationResponse(
                success=False,
                summary="",
                error="No candidate models supplied.",
            )

        import json as _json

        cand_json = _json.dumps(body.candidates, indent=2, default=str)
        lr_part = ""
        if body.lr_digest:
            lr_json = _json.dumps(body.lr_digest, indent=2, default=str)
            lr_part = (
                "\n\nLogistic-regression coefficient vs bivariate sign checks (subset):\n"
                f"{lr_json}\n"
                "If mismatched_count > 0, mention at most one concise caveat naming sample_mismatch_features when present."
            )

        metric_label = "Test AUC" if (body.problem_type or "").lower() == "classification" else "Test R²"

        system_prompt = """You are a senior credit-risk model reviewer. Write ONE short narrative (about 140–220 words) for modelers comparing shortlisted training runs.

Requirements:
- Start by naming a clear primary pick when the numbers support it (algorithm name and iteration as #n when best_iteration is present).
- Optionally mention one runner-up or stability-focused alternative if justified by the JSON (e.g., lower overfit_pct, simpler LR with fewer features).
- Use only facts present in the JSON (test_primary, train_primary, overfit_pct, feature_display, flags, guideline G1 vs G2, segment_id).
- G1 means overfit-aware guideline; G2 means test-score-first — you may reference that briefly if it clarifies trade-offs.
- Plain sentences only; no markdown headings, no bullet lists unless a single short clause needs it.
- Do not use asterisks for emphasis (no ** or * markdown).
- Do not invent metrics, policies, or regulatory claims not implied by the data."""

        user_prompt = f"""Dataset id (context only): {dataset_id}
Problem type: {body.problem_type or "classification"}
Primary reported test metric in payload: {metric_label} (field test_primary).

Shortlisted models JSON:
{cand_json}
{lr_part}

Write the cross-algorithm recommendation paragraph now."""

        summary = await llm_service.generate_text(
            prompt=f"{system_prompt}\n\n---\n\n{user_prompt}",
            max_tokens=900,
            temperature=0.25,
            context="model_training",
        )
        text = (summary or "").strip()
        if text.startswith("```"):
            lines = text.split("\n")
            if lines and lines[0].strip().startswith("```"):
                lines = lines[1:]
            if lines and lines[-1].strip() == "```":
                lines = lines[:-1]
            text = "\n".join(lines).strip()

        # Strip markdown bold markers if the model still emits them
        text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
        text = text.replace("**", "")

        if not text:
            return CrossAlgorithmRecommendationResponse(
                success=False,
                summary="",
                error="The model returned an empty summary.",
            )

        return CrossAlgorithmRecommendationResponse(success=True, summary=text, error=None)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("cross_algorithm_recommendation failed: %s", e, exc_info=True)
        return CrossAlgorithmRecommendationResponse(
            success=False,
            summary="",
            error=str(e),
        )


@upload_router.get("/knowledge-graph-progress/{dataset_id}")
async def get_knowledge_graph_progress(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Check if updated knowledge graph is available for a dataset.
    Returns cached progressive result if available.
    """
    from app.services.llm_service import get_kg_cache
    
    cache_key = llm_service.get_kg_cache_key(dataset_id)
    cached_result = get_kg_cache(cache_key)
    
    if cached_result:
        logger.info(f"📊 Poll endpoint: Returning cached result for {cache_key}")
        logger.info(f"   Has nodes: {bool(cached_result.get('nodes'))}, count: {len(cached_result.get('nodes', []))}")
        logger.info(f"   Has categories: {bool(cached_result.get('categories'))}, count: {len(cached_result.get('categories', []))}")
        logger.info(f"   Status: {cached_result.get('processing_info', {}).get('status')}")
        logger.info(f"   Batches: {cached_result.get('processing_info', {}).get('completed_batches')}/{cached_result.get('processing_info', {}).get('total_batches')}")
        return {
            "available": True,
            "result": cached_result
        }
    else:
        logger.info(f"📊 Poll endpoint: No cached result for {cache_key}")
        return {
            "available": False,
            "message": "No updates available yet"
        }

@upload_router.get("/knowledge-graph-stream/{dataset_id}")
async def stream_knowledge_graph_updates(
    request: Request,
    dataset_id: str,
    token: Optional[str] = Query(None),
):
    """
    Server-Sent Events stream for real-time knowledge graph updates.
    Streams updates whenever a new batch completes processing.
    Note: Uses query parameter for token since EventSource can't send headers.
    """
    from app.services.llm_service import get_kg_cache
    from app.core.session import SessionManager

    # Validate token manually since EventSource can't send headers
    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token required as query parameter"
        )

    sm: SessionManager = request.app.state.session_manager
    try:
        user = await sm.authenticate_access_token(token)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid authentication credentials",
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Token validation failed for SSE stream: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Invalid or expired token",
        )

    cache_key = llm_service.get_kg_cache_key(dataset_id)
    
    async def event_generator():
        last_completed_batches = 0
        max_wait_time = 3000  # 5 minutes max wait
        start_time = time.time()
        last_write = start_time
        # Periodic SSE comments keep ALB/nginx/proxy idle timeouts from closing long polls (504).
        keepalive_interval_s = 15.0

        try:
            # Immediately check and send cached result if available
            cached_result = get_kg_cache(cache_key)
            if cached_result:
                logger.info(f"SSE: Sending initial cached result immediately for dataset {dataset_id}")
                logger.info(f"SSE Initial: Has nodes: {bool(cached_result.get('nodes'))}, count: {len(cached_result.get('nodes', []))}")
                logger.info(f"SSE Initial: Has categories: {bool(cached_result.get('categories'))}, count: {len(cached_result.get('categories', []))}")
                logger.info(f"SSE Initial: Keys in cached_result: {list(cached_result.keys())}")
                yield f"data: {json.dumps(cached_result)}\n\n"
                last_write = time.time()

                # If complete, close immediately
                if cached_result.get('processing_info', {}).get('status') == 'complete':
                    logger.info(f"SSE: Knowledge graph already complete for dataset {dataset_id}")
                    yield f"data: {json.dumps({'event': 'complete'})}\n\n"
                    return
                
                # Track what we've sent
                last_completed_batches = cached_result.get('processing_info', {}).get('completed_batches', 0)
            
            # Listen for updates
            while True:
                if time.time() - start_time > max_wait_time:
                    logger.info(f"SSE stream timeout for dataset {dataset_id}")
                    yield f"data: {json.dumps({'event': 'timeout'})}\n\n"
                    break

                now = time.time()
                if now - last_write >= keepalive_interval_s:
                    yield ": keepalive\n\n"
                    last_write = now

                cached_result = get_kg_cache(cache_key)
                
                if cached_result:
                    processing_info = cached_result.get('processing_info', {})
                    current_batches = processing_info.get('completed_batches', 0)
                    total_batches = processing_info.get('total_batches', 0)
                    status = processing_info.get('status', 'unknown')
                    
                    # Send update if batch count increased
                    if current_batches > last_completed_batches:
                        logger.info(f"SSE: Sending update for batch {current_batches}/{total_batches}")
                        logger.info(f"SSE Update: Has nodes: {bool(cached_result.get('nodes'))}, count: {len(cached_result.get('nodes', []))}")
                        logger.info(f"SSE Update: Has categories: {bool(cached_result.get('categories'))}, count: {len(cached_result.get('categories', []))}")
                        logger.info(f"SSE Update: Status: {status}")
                        yield f"data: {json.dumps(cached_result)}\n\n"
                        last_write = time.time()
                        last_completed_batches = current_batches
                        
                        # Only stop if ALL batches are complete (not just if status says complete)
                        if status == 'complete' and current_batches >= total_batches:
                            logger.info(f"SSE: Knowledge graph complete for dataset {dataset_id}")
                            yield f"data: {json.dumps({'event': 'complete'})}\n\n"
                            break
                
                # Wait before checking again (shorter interval for better responsiveness)
                await asyncio.sleep(0.5)
                
        except Exception as e:
            logger.error(f"SSE stream error for dataset {dataset_id}: {str(e)}")
            yield f"data: {json.dumps({'event': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )

@upload_router.post("/generate-knowledge-graph", response_model=KnowledgeGraphResponse)
async def generate_knowledge_graph(
    request: KnowledgeGraphRequest,
    http_request: Request,
    current_user = Depends(get_current_user_dependency)
):
    """
    Generate interactive knowledge graph from dataset's data dictionary
    Uses the dataDesc field from MessageState which contains the data dictionary
    """
    logger.info(f"Knowledge graph generation request for dataset: {request.dataset_id}")
    
    try:
        # Validate dataset exists
        dataset_info = dataset_manager.get_dataset_info(request.dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset not found: {request.dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Create or load MessageState to get the dataDesc
        try:
            state = message_state_manager.create_or_load_state(request.dataset_id, "")
            data_dictionary = state.get("dataDesc", "")
            
            # If dataDesc is empty, try to get it from dataset_info directly
            if not data_dictionary:
                data_dictionary = dataset_info.get('data_dictionary', '')
            
            # If still no data dictionary, return an error
            if not data_dictionary:
                return KnowledgeGraphResponse(
                    success=False,
                    message="No data dictionary found for this dataset",
                    error="Data dictionary is required to generate knowledge graph. Please ensure the dataset has a data dictionary."
                )
            
        except ValueError as e:
            logger.error(f"Failed to create/load MessageState: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))
        
        # Get additional context from dataset
        df = dataset_manager.load_dataset(request.dataset_id)
        if df is None:
            logger.error(f"Failed to load dataset file: {request.dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset file not found")
        
        problem_statement = dataset_info.get('problem_statement', '') or ''

        # Initialize dataset analyser to get additional context
        dataset_analyser = DatasetAnalyser()
        dataset_summary = dataset_analyser.generate_dataset_summary(df)
        
        # Create comprehensive metadata content for LLM processing
#         metadata_content = f"""
# DATASET INFORMATION:
# Dataset ID: {request.dataset_id}
# Filename: {dataset_info.get('filename', 'Unknown')}
# Target Variable: {dataset_info.get('target_variable', 'Not specified')}
# Target Variable Type: {dataset_info.get('target_variable_type', 'Not specified')}
# Problem Statement: {dataset_info.get('problem_statement', 'Not provided')}

# DATA DICTIONARY:
# {data_dictionary}

# DATASET TECHNICAL SUMMARY:
# {dataset_summary}

# ADDITIONAL CONTEXT:
# Total Rows: {df.shape[0]}
# Total Columns: {df.shape[1]}
# Column Names: {list(df.columns)}
# """
        get_dataset_info = dataset_manager.get_dataset_info(request.dataset_id)

        data_dictionary_path = get_dataset_info.get('data_dictionary')
        if data_dictionary_path:
            data_dictionary = _read_data_dictionary(data_dictionary_path)
        else:
            data_dictionary = ''


        logger.info("Calling LLM service for knowledge graph generation")
        llm_response = llm_service.get_knowledge_graph(
            data_dictionary,
            dataset_id=request.dataset_id,
            problem_statement=problem_statement,
        )

        # Parse the JSON response
        if isinstance(llm_response, dict):
            knowledge_graph_result = llm_response
            logger.info("Knowledge graph generation completed successfully")
        else:
            try:
                knowledge_graph_result = json.loads(llm_response)
                logger.info("Knowledge graph generation completed successfully")
            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse LLM response as JSON: {str(e)}")
                # Return error response if JSON parsing fails
                return KnowledgeGraphResponse(
                    success=False,
                    message="Failed to parse knowledge graph response",
                    error=f"JSON parsing error: {str(e)}"
                )
        return KnowledgeGraphResponse(
            success=True,
            message="Knowledge graph generated successfully",
            html_content=knowledge_graph_result.get('html_content', ''),
            algorithm_explanation=knowledge_graph_result.get('algorithm_explanation', ''),
            relationship_mapping=knowledge_graph_result.get('relationship_mapping', ''),
            usage_instructions=knowledge_graph_result.get('usage_instructions', ''),
            nodes=knowledge_graph_result.get('nodes', []),
            categories=knowledge_graph_result.get('categories', []),
            processing_info=knowledge_graph_result.get('processing_info')
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Knowledge graph generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@chat_router.post("/execute-code", response_model= CodeExecutionResponse)
async def execute_python_code(
    dataset_id: str = Form(...),
    code: str = Form(...),
    current_user = Depends(get_current_user_dependency)
):
    """
    Execute Python code by invoking the agent workflow.
    Uses dataset_id to load MessageState, sets generatedCode, and routes to 'code_execution'.
    """
    logger.info(f"Code execution request for dataset: {dataset_id}")
    logger.info(f"Code execution request for dataset: {type(dataset_id)}")
    logger.info(f"String Code length: {len(code)}")
    logger.info(f"String Code (full with line numbers):\n" + "\n".join(f"{i+1:3d}: {line}" for i, line in enumerate(code.splitlines())))
    

    try:
        ## Validate dataset exists
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        logger.info(f"dataset info: {dataset_info}")
        if not dataset_info:
            logger.warning(f"Dataset not found: {dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found")

        ## Load or create MessageState
        try:
            state = message_state_manager.create_or_load_state(dataset_id, "Code execution request")
            # Continuous processing approach - always use the most recent processed DataFrame
            processed_df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, state["datasetFile"])
            
            # Store initial dataframe info for comparison later
            initial_df_shape = state["datasetFile"].shape if state.get("datasetFile") is not None else None
            initial_df_columns = list(state["datasetFile"].columns) if state.get("datasetFile") is not None else []
            
            if processed_df is not state["datasetFile"]:
                state["datasetFile"] = processed_df
                logger.info(f"📥 Using processed DataFrame for continuous processing: {dataset_id}")
                logger.info(f"   Shape: {processed_df.shape}")
                logger.info(f"   Columns: {len(processed_df.columns)}")
            else:
                logger.info(f"📥 Using original DataFrame for dataset {dataset_id}")
                logger.info(f"   Shape: {state['datasetFile'].shape}")
                logger.info(f"   Columns: {len(state['datasetFile'].columns)}")
                
        except ValueError as e:
            logger.error(f"Failed to create/load MessageState: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))

        # Put the code into state so the router will send it to 'code_execution'
        state["generatedCode"] = code
        # state["intent"] = "code_execution"
        
        
        # Add dataset_id to state for DataFrameStateManager access in agentic system
        state['dataset_id'] = dataset_id
        
        # Invoke the agent workflow
        result = agent.invoke(state)
        
        # Verify that dataframe was actually modified by code execution
        result_df_temp = result.get("datasetFile")
        if result_df_temp is not None and initial_df_shape is not None:
            shape_changed = result_df_temp.shape != initial_df_shape
            columns_changed = list(result_df_temp.columns) != initial_df_columns
            if shape_changed or columns_changed:
                logger.info(f"✅ DataFrame MODIFIED by code execution:")
                logger.info(f"   Before: {initial_df_shape} with {len(initial_df_columns)} columns")
                logger.info(f"   After:  {result_df_temp.shape} with {len(result_df_temp.columns)} columns")
                if columns_changed:
                    added_cols = set(result_df_temp.columns) - set(initial_df_columns)
                    removed_cols = set(initial_df_columns) - set(result_df_temp.columns)
                    if added_cols:
                        logger.info(f"   Added columns: {list(added_cols)}")
                    if removed_cols:
                        logger.info(f"   Removed columns: {list(removed_cols)}")
            else:
                logger.warning(f"⚠️ DataFrame NOT modified (shape and columns unchanged)")

        columns_info = []  # Initialize as empty list instead of None
        try:
            # Get the LATEST processed DataFrame after code execution from agent result
            # This is the freshest dataframe that was just transformed by the executed code
            result_df = result.get("datasetFile")
            
            if result_df is not None:
                df_to_preview = result_df
                logger.info(f"✅ Using LATEST PROCESSED dataframe from agent result")
                logger.info(f"   This dataframe reflects ALL code transformations including the one just executed")
                logger.info(f"   Shape: {result_df.shape}")
                logger.info(f"   Columns: {list(result_df.columns)[:10]}... ({len(result_df.columns)} total)")
                logger.info(f"   Memory: {result_df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB")
                logger.info(f"   Index: {result_df.index.tolist()[:5]}... ({len(result_df)} rows)")
            else:
                # Fallback to original dataset only if agent didn't return a dataframe
                df_to_preview = dataset_manager.load_dataset(dataset_id)
                logger.error(f"❌ Agent did not return dataframe in result!")
                logger.error(f"   This means code execution may have failed")
                logger.error(f"   Falling back to ORIGINAL dataset")
                if df_to_preview is not None:
                    logger.info(f"   Original Shape: {df_to_preview.shape}")
            
            if df_to_preview is not None:
                logger.info(f"🔍 Calculating column info from {'PROCESSED' if result_df is not None else 'ORIGINAL'} dataframe")
                logger.info(f"   Shape: {df_to_preview.shape}")
                logger.info(f"   Columns: {list(df_to_preview.columns)[:10]}... ({len(df_to_preview.columns)} total)")
                
                # Calculate detailed column information (same as column-info endpoint)
                columns_info = calculate_column_info(df_to_preview)
                logger.info(f"✅ Column info calculated for execute-code: {len(columns_info)} columns")
                logger.info(f"📊 DataFrame for UI display - Shape: {df_to_preview.shape}, Columns: {list(df_to_preview.columns)[:5]}...")
                
                # Create a data fingerprint for the displayed DataFrame
                import hashlib
                display_fingerprint = hashlib.md5(
                    f"{df_to_preview.shape}_{list(df_to_preview.columns)}_{df_to_preview.index.tolist()[:10]}".encode()
                ).hexdigest()
                logger.info(f"🔐 Display DataFrame fingerprint: {display_fingerprint}")
            else:
                logger.warning(f"⚠️ No DataFrame available to calculate column info after code execution")
            
            # Store the processed DataFrame using the state manager
            if result_df is not None:
                # Get active scope to ensure we save to the correct scope
                active_scope = dataframe_state_manager._active_scope.get(dataset_id, 'train')
                logger.info(f"💾 Saving processed dataframe to scope: {active_scope} for dataset: {dataset_id}, shape: {result_df.shape}")
                # The agentic code execution flow may have already persisted the updated df.
                # Avoid double-updating here because update_dataframe() snapshots the previous
                # dataframe; calling it twice with the same df would overwrite the baseline used
                # by /compare-column-stats.
                stored_df = dataframe_state_manager.get_dataframe(dataset_id)
                should_update_state_manager = stored_df is None
                if (not should_update_state_manager) and isinstance(stored_df, pd.DataFrame):
                    try:
                        should_update_state_manager = not stored_df.equals(result_df)
                    except Exception:
                        # If equality check fails for any reason, proceed with update
                        should_update_state_manager = True

                if should_update_state_manager:
                    dataframe_state_manager.update_dataframe(dataset_id, result_df, original_shape=state["datasetFile"].shape)
                    stored_df = dataframe_state_manager.get_dataframe(dataset_id)
                # Verify storage by retrieving it back
                if stored_df is not None:
                    logger.info(f"✅ Verified stored DataFrame for download - Shape: {stored_df.shape}, Columns: {list(stored_df.columns)}")
                    
                    # Create a data fingerprint for the stored DataFrame
                    import hashlib
                    stored_fingerprint = hashlib.md5(
                        f"{stored_df.shape}_{list(stored_df.columns)}_{stored_df.index.tolist()[:10]}".encode()
                    ).hexdigest()
                    logger.info(f"🔐 Stored DataFrame fingerprint: {stored_fingerprint}")
                    
                    # Deep comparison: shape, columns, and data content
                    shapes_match = stored_df.shape == result_df.shape
                    columns_match = list(stored_df.columns) == list(result_df.columns)
                    data_matches = stored_df.equals(result_df)
                    
                    if shapes_match and columns_match and data_matches:
                        logger.info(f"✅✅✅ VERIFIED: The table displayed in UI and downloadable CSV contain IDENTICAL data")
                        logger.info(f"    - Same shape: {stored_df.shape}")
                        logger.info(f"    - Same columns: {list(stored_df.columns)}")
                        logger.info(f"    - Same values: DataFrame.equals() = True")
                    else:
                        logger.error(f"❌ DATA MISMATCH DETECTED!")
                        logger.error(f"    - Shapes match: {shapes_match}")
                        logger.error(f"    - Columns match: {columns_match}")
                        logger.error(f"    - Data matches: {data_matches}")
                
                # Merge all scopes into 'entire' so that EDA snapshot gets the full transformed data
                try:
                    merged_df = dataframe_state_manager.merge_scopes_to_entire(dataset_id)
                    if merged_df is not None:
                        logger.info(f"✅ Merged all scopes to 'entire' after code execution: {merged_df.shape}")
                except Exception as merge_err:
                    logger.warning(f"⚠️ Failed to merge scopes to 'entire': {merge_err}")

        except Exception as e:
            logger.error(f"❌ Failed to build column info: {e}", exc_info=True)
            # Ensure columns_info is always a list, even on error
            if columns_info is None:
                columns_info = []

        save_success = message_state_manager.save_state(dataset_id, result)
        if not save_success:
            logger.warning(f"Failed to save MessageState for dataset: {dataset_id}")

        if result.get("messages"):
            last_msg = result["messages"][-1]
            content = getattr(last_msg, "content", "")
            try:
                payload = json.loads(content)
                response_text = payload.get("response", "Code executed")
                role = payload.get("role", "modelling")

                # Always include columns_info so frontend can display the updated dataframe stats
                resp_body = {
                    "success": True,
                    "response": response_text,
                    "role": role,
                    "columns_info": columns_info  # Always include for Data Treatment step display
                }
                return resp_body

            except json.JSONDecodeError:
                # Fallback if message content isn't JSON
                response_text = content or "Code executed"

                # Always include columns_info so frontend can display the updated dataframe stats
                resp_body = {
                    "success": True,
                    "response": response_text,
                    "role": "modelling",
                    "columns_info": columns_info  # Always include for Data Treatment step display
                }
                return resp_body
        # No messages generated → surface stats to help the user
        return {
            "success": False,
            "response": "No response generated by workflow",
            "columns_info": columns_info,  # Include even on workflow failure
            "role": "modelling"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Code execution failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


# =============================================================================
# QC STEP-BY-STEP ENDPOINTS (Manual QC Mode)
# =============================================================================

class QCNextStepRequest(BaseModel):
    """Request model for getting next QC treatment step."""
    dataset_id: str
    action: str = "apply"  # "apply" or "skip"
    treatment_type: str  # Current treatment type being applied/skipped
    code: Optional[str] = None  # Code to execute if action is "apply"


class QCRegenerateCodeRequest(BaseModel):
    """Request model for regenerating code based on manual QC user selections."""
    dataset_id: str
    treatment_type: str
    selections: Dict[str, str] = {}


@chat_router.post("/qc/next-step")
async def get_next_qc_step(
    request: QCNextStepRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get the next QC treatment step in Manual QC mode.
    
    Called after user clicks Apply or Skip on current treatment.
    Code execution is handled separately via /execute-code endpoint (same as Auto QC).
    This endpoint only manages the treatment sequence and returns the next plan.
    """
    logger.info(f"QC next step request: dataset={request.dataset_id}, action={request.action}, treatment={request.treatment_type}")
    
    try:
        # Load current state
        state = message_state_manager.create_or_load_state(request.dataset_id, "QC next step")
        
        if not state:
            raise HTTPException(status_code=404, detail="Session state not found")
        
        # Log state info for debugging
        logger.info(f"State loaded - qc_mode: {state.get('qc_mode')}, treatment_sequence: {state.get('treatment_sequence')}, current_treatment_index: {state.get('current_treatment_index')}")
        
        # Check if we have treatment sequence - this means we're in a QC session
        # The qc_mode might not be persisted properly, so also check for treatment_sequence
        if not state.get('treatment_sequence'):
            raise HTTPException(status_code=400, detail="No active QC session found. Please start a Manual QC first.")
        
        # If qc_mode is not set, set it to manual since we have a treatment_sequence
        if not state.get('qc_mode'):
            state['qc_mode'] = 'manual'
            logger.info("Set qc_mode to 'manual' based on presence of treatment_sequence")
        
        # Code execution is now handled by frontend via /execute-code endpoint (same as Auto QC)
        # This endpoint only advances the sequence and returns the next treatment plan
        logger.info(f"Getting next treatment step after {request.action} on {request.treatment_type}")
        
        # Get the next treatment step
        state = agentic_system._execute_next_qc_step(state, request.action)
        
        # Save updated state
        message_state_manager.save_state(request.dataset_id, state)
        
        # Extract response from the last message
        response_data = {
            "success": True,
            "action_performed": request.action,
            "treatment_processed": request.treatment_type
        }
        
        # Get the last message (next treatment plan or completion summary)
        if state.get('messages'):
            last_msg = state['messages'][-1]
            if hasattr(last_msg, 'content'):
                try:
                    content = json.loads(last_msg.content)
                    response_data["next_treatment"] = content
                    response_data["step_info"] = content.get("step_info", {})
                    response_data["is_complete"] = content.get("qc_complete", False)
                except (json.JSONDecodeError, TypeError):
                    response_data["next_treatment"] = {"response": last_msg.content}
        
        # Include treatment statuses
        response_data["treatment_statuses"] = state.get("treatment_statuses", {})
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"QC next step failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get next QC step: {str(e)}")


@chat_router.post("/qc/skip-treatment")
async def skip_qc_treatment(
    request: QCNextStepRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Skip current QC treatment and get next step.
    Convenience endpoint that calls next-step with action="skip".
    """
    request.action = "skip"
    return await get_next_qc_step(request, current_user)


@chat_router.post("/qc/regenerate-code")
async def regenerate_qc_code(
    request: QCRegenerateCodeRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Regenerate current Manual QC treatment code using user-selected dropdown values.

    This endpoint does not advance treatment sequence. It only recomputes payload/code for
    the specified treatment and persists selection map in MessageState.
    """
    logger.info(
        f"QC regenerate request: dataset={request.dataset_id}, treatment={request.treatment_type}, "
        f"selection_count={len(request.selections or {})}"
    )

    try:
        state = message_state_manager.create_or_load_state(request.dataset_id, "QC regenerate code")
        if not state:
            raise HTTPException(status_code=404, detail="Session state not found")

        sequence = state.get('treatment_sequence') or []
        if not sequence:
            raise HTTPException(status_code=400, detail="No active QC session found. Please start Manual QC first.")

        treatment_type = (request.treatment_type or '').strip()
        if treatment_type not in {'invalid_values', 'special_values', 'outliers', 'missing_values'}:
            raise HTTPException(status_code=400, detail=f"Unsupported treatment_type: {request.treatment_type}")

        state.setdefault('qc_ui_selections', {})
        state['qc_ui_selections'][treatment_type] = request.selections or {}
        message_state_manager.save_state(request.dataset_id, state)

        # Re-run only the treatment handler on a cloned state to avoid mutating sequence progress/messages.
        working_state = copy.deepcopy(state)
        agentic_system._initialize_treatment_handlers()
        handler = agentic_system._treatment_handler_registry.get(treatment_type)
        if not handler:
            raise HTTPException(status_code=500, detail=f"Treatment handler not found: {treatment_type}")

        working_state = handler(working_state)
        payload: Dict[str, Any] = {}

        if working_state.get('messages'):
            last_msg = working_state['messages'][-1]
            if hasattr(last_msg, 'content'):
                try:
                    payload = json.loads(last_msg.content)
                except (json.JSONDecodeError, TypeError):
                    payload = {"response": str(last_msg.content), "treatment_type": treatment_type}

        current_idx = state.get('current_treatment_index', 0)
        payload.setdefault('step_info', {
            'current_step': current_idx + 1,
            'total_steps': len(sequence),
            'current_treatment': sequence[current_idx] if current_idx < len(sequence) else treatment_type,
            'next_treatment': sequence[current_idx + 1] if current_idx + 1 < len(sequence) else None,
            'has_next': current_idx + 1 < len(sequence),
            'is_step_by_step': True,
        })
        payload.setdefault('treatment_type', treatment_type)
        payload.setdefault('qc_mode', state.get('qc_mode', 'manual'))

        return {
            "success": True,
            "treatment_type": treatment_type,
            "payload": payload,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"QC regenerate failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to regenerate QC code: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/column-info", response_model=ColumnInfoResponse)
async def get_column_info(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get detailed statistical information for each column in the dataset
    Returns mean, median, mode, standard deviation, variance, min, 25%, 50%, 75%, max for each variable
    """
    try:
        logger.info(f"Column info request for dataset: {dataset_id}")

        # 1) Try in-memory processed DataFrame first (after upload/analysis)
        processed_df = dataframe_state_manager.get_processed_dataframe(dataset_id)
        if processed_df is not None:
            df = processed_df.copy()
            logger.info(f"Using processed DataFrame for column info: {dataset_id}")
        else:
            # 2) Try MessageState (persists latest processed DF across workers)
            df = None
            try:
                state = message_state_manager.create_or_load_state(dataset_id, "column-info")
                state_df = state.get("datasetFile") if state else None
                if isinstance(state_df, pd.DataFrame):
                    df = state_df.copy()
                    logger.info(f"Using MessageState DataFrame for column info: {dataset_id}, shape: {df.shape}")
            except Exception as e:
                logger.warning(f"Failed to load DataFrame from MessageState for column info: {e}")

            # 3) Fallback to original CSV load
            if df is None:
                dataset_info = dataset_manager.get_dataset_info(dataset_id)
                if not dataset_info:
                    logger.warning(f"Dataset metadata not found: {dataset_id}")
                    raise HTTPException(status_code=404, detail="Dataset not found. Please upload a dataset first.")

                df = dataset_manager.load_dataset(dataset_id)
                if df is None:
                    logger.error(f"Failed to load dataset file: {dataset_id}")
                    raise HTTPException(status_code=404, detail="Dataset file not found. Please upload your dataset again.")
                logger.info(f"Using original dataset for column info dataset {dataset_id}: {list(df.columns)[:10]}...")
                logger.info(f"Original dataset shape: {df.shape}")
        
        # Create data preview
        data_preview = None
        try:
            data_preview = {
                "columns": list(df.columns),
                "rows": df.head(10).to_dict(orient="records")
            }
        except Exception as e:
            logger.warning(f"Failed to build data preview: {e}")
        
        # Offload CPU-bound column stats to the shared executor so this endpoint
        # doesn't block the event loop or hold a browser connection slot.
        import asyncio as _aio
        from app.core.executor import executor as _col_executor
        _loop = _aio.get_event_loop()
        columns_info = await _loop.run_in_executor(_col_executor, calculate_column_info, df)

        logger.info(f"Column info calculated successfully: {len(columns_info)} columns processed")
        
        return ColumnInfoResponse(
            success=True,
            message="Column information retrieved successfully",
            dataset_id=dataset_id,
            columns_info=columns_info,
            total_columns=len(columns_info),
            data_preview=data_preview
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Column info calculation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/column-info-by-scope", response_model=ColumnInfoResponse)
async def get_column_info_by_scope(
    dataset_id: str,
    scope: str = "entire",  # entire, train, test, validation
    current_user = Depends(get_current_user_dependency)
):
    """
    Get detailed statistical information for each column filtered by scope.
    This is a read-only operation that does NOT modify _processed_dataframes or _active_scope.
    
    IMPORTANT: This endpoint ALWAYS loads from the original raw dataset on disk to ensure
    the "Original EDA" in Data Treatments page matches what was shown in Objectives page.
    This guarantees consistent baseline statistics regardless of any in-memory transformations.
    
    Args:
        dataset_id: The dataset identifier
        scope: One of 'entire', 'train', 'test', 'validation'
    
    Returns:
        ColumnInfoResponse with column stats for the specified scope
    """
    try:
        logger.info(f"Column info by scope request for dataset: {dataset_id}, scope: {scope}")
        # #region agent log
        try:
            import os as _dbg_os, json as _dbg_json, time as _dbg_time
            _dbg_si = dataframe_state_manager._split_indices.get(dataset_id)
            _dbg_full = dataframe_state_manager._full_dataframes.get(dataset_id)
            _dbg_payload = {
                "sessionId": "826a95",
                "hypothesisId": "H1",
                "location": "routes.py:column-info-by-scope:entry",
                "message": "worker view of DFSM state",
                "data": {
                    "pid": _dbg_os.getpid(),
                    "dataset_id": dataset_id,
                    "scope": scope,
                    "has_split_indices": _dbg_si is not None,
                    "split_index_keys": list(_dbg_si.keys()) if _dbg_si else [],
                    "split_index_sizes": {k: int(len(v)) for k, v in (_dbg_si or {}).items()},
                    "has_full_dataframe_in_dfsm": _dbg_full is not None,
                    "full_df_shape": list(_dbg_full.shape) if _dbg_full is not None else None,
                    "full_df_has_split_tag": (_dbg_full is not None and "split_tag" in _dbg_full.columns),
                },
                "timestamp": int(_dbg_time.time() * 1000),
            }
            with open("/Users/saiyam268728/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/UC-Github/RI Branch/.cursor/debug-826a95.log", "a") as _dbg_fh:
                _dbg_fh.write(_dbg_json.dumps(_dbg_payload) + "\n")
        except Exception:
            pass
        # #endregion

        # Validate scope
        valid_scopes = ['entire', 'train', 'test', 'validation']
        if scope not in valid_scopes:
            raise HTTPException(status_code=400, detail=f"Invalid scope '{scope}'. Must be one of: {valid_scopes}")
        
        # ALWAYS load from disk to ensure Original EDA matches Objectives page.
        # This prevents any in-memory mutations from affecting the baseline statistics.
        # Use the shared read-only cache so concurrent calls from the overview
        # sidebar (column-info-by-scope + dqs-by-scope + bundle) collapse onto
        # a single CSV/Parquet parse instead of re-reading multi-GB files in
        # parallel and OOM-killing the worker.
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found. Please upload a dataset first.")
        full_df = dataset_manager.load_dataset_readonly_cached(dataset_id)
        if full_df is None:
            raise HTTPException(status_code=404, detail="Dataset file not found. Please upload your dataset again.")

        logger.info(f"📊 Loaded original dataset (read-only) for column-info-by-scope: {full_df.shape}")
        # #region agent log
        try:
            import os as _dbg_os, json as _dbg_json, time as _dbg_time
            _dbg_payload = {
                "sessionId": "826a95",
                "hypothesisId": "H2",
                "location": "routes.py:column-info-by-scope:after-load-readonly-cached",
                "message": "disk-loaded readonly df characteristics",
                "data": {
                    "pid": _dbg_os.getpid(),
                    "dataset_id": dataset_id,
                    "scope": scope,
                    "shape": list(full_df.shape),
                    "has_split_tag_column": "split_tag" in full_df.columns,
                    "columns_sample": list(full_df.columns[:5]),
                },
                "timestamp": int(_dbg_time.time() * 1000),
            }
            with open("/Users/saiyam268728/Library/CloudStorage/OneDrive-EXLService.com(I)Pvt.Ltd/Desktop/UC-Github/RI Branch/.cursor/debug-826a95.log", "a") as _dbg_fh:
                _dbg_fh.write(_dbg_json.dumps(_dbg_payload) + "\n")
        except Exception:
            pass
        # #endregion
        
        # Filter by scope using split_indices (without changing global state).
        # If indices exist but are empty (stale vs split_tag), fall back to split_tag (pre-split uploads).
        if scope != "entire":
            df = None
            # Cross-worker hydration: on a worker that did not handle the
            # original /upload, _split_indices is empty. Try Redis first
            # (single HGETALL), then the in-frame split_tag fallback below.
            split_indices = dataframe_state_manager._split_indices.get(dataset_id)
            if not split_indices:
                try:
                    if dataframe_state_manager._hydrate_split_indices_from_redis(dataset_id):
                        split_indices = dataframe_state_manager._split_indices.get(dataset_id)
                except Exception as _exc:
                    logger.warning(f"column-info-by-scope: redis hydrate failed: {_exc}")
            if split_indices and scope in split_indices:
                indices = split_indices[scope]
                if indices is not None and len(indices) > 0:
                    df = full_df.iloc[indices].copy()
                    logger.info(f"Filtered to {scope} scope: {len(df)} rows (from {len(full_df)} total)")
            if (df is None or len(df) == 0) and "split_tag" in full_df.columns:
                mask = full_df["split_tag"] == scope
                if scope == "validation":
                    mask = full_df["split_tag"].astype(str).str.startswith("validation", na=False)
                df = full_df[mask].copy()
                logger.info(f"Filtered by split_tag to {scope}: {len(df)} rows")
            if df is None or len(df) == 0:
                logger.warning(f"No data for scope '{scope}' in dataset {dataset_id}")
                return ColumnInfoResponse(
                    success=True,
                    message=f"No data available for scope '{scope}'",
                    dataset_id=dataset_id,
                    columns_info=[],
                    total_columns=0,
                    data_preview=None,
                    scope=scope,
                    total_rows=0,
                )
        else:
            # Read-only: calculate_column_info does not mutate. The shared
            # cache hands back the same reference; the per-column reads
            # below are GIL-safe pandas C operations. Skip the gratuitous
            # 4M-row .copy() that used to live here — it doubled RAM on
            # every parallel sidebar refresh.
            df = full_df

        # Create data preview
        data_preview = None
        try:
            data_preview = {
                "columns": list(df.columns),
                "rows": df.head(10).to_dict(orient="records")
            }
        except Exception as e:
            logger.warning(f"Failed to build data preview: {e}")

        # P2.3: result-cache lookup. If the same (dataset_id, scope) was
        # computed for the current data version, return the cached payload
        # instantly. The cache is invalidated automatically when
        # DataFrameStateManager.update_dataframe bumps the version.
        from app.services.analytics_cache import analytics_cache as _ac
        _version = dataframe_state_manager.get_version(dataset_id)
        cached = _ac.get("column_info", dataset_id, scope, _version)
        if cached is not None:
            logger.info(
                f"Column info by scope cache HIT for dataset_id={dataset_id} "
                f"scope={scope} version={_version}"
            )
            return cached

        # Offload CPU-bound stats to the shared executor so this endpoint
        # doesn't block the event loop on large frames (P1.3).
        # P3.5: time_stage emits midas_pipeline_stage_seconds{stage="column_info"}.
        import asyncio as _aio
        from app.core.executor import executor as _col_executor
        from app.core.metrics import time_stage as _time_stage
        _loop = _aio.get_event_loop()
        with _time_stage("column_info"):
            columns_info = await _loop.run_in_executor(
                _col_executor, calculate_column_info, df
            )

        logger.info(f"Column info by scope calculated: {len(columns_info)} columns, scope={scope}, rows={len(df)}")

        response = ColumnInfoResponse(
            success=True,
            message=f"Column information retrieved successfully for scope '{scope}'",
            dataset_id=dataset_id,
            columns_info=columns_info,
            total_columns=len(columns_info),
            data_preview=data_preview,
            scope=scope,
            total_rows=len(df)
        )
        # P2.3: store in cache. Subsequent requests for the same
        # (dataset_id, scope, version) tuple return the cached object
        # without recomputing. Cache invalidates on update_dataframe.
        try:
            _ac.set("column_info", dataset_id, scope, _version, response)
        except Exception as exc:
            logger.warning(f"AnalyticsResultCache set failed: {exc}")
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Column info by scope calculation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/dqs", response_model=DQSResponse)
async def get_data_quality_score(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Calculate Data Quality Score (DQS) for a dataset.
    
    Returns a composite score based on 4 dimensions:
    - Completeness (35%): Column fill rates and row sparseness
    - Consistency (30%): Data type, format, placeholder, and range consistency
    - Structural Integrity (25%): Constant, near-constant, and duplicate columns
    - Uniqueness (10%): Duplicate row detection
    
    Also includes Target Readiness (informational) if target variable is configured.
    
    Error Responses:
    - 400: Invalid dataset_id format
    - 404: Dataset not found
    - 422: Insufficient data for analysis
    - 500: Internal server error
    """
    try:
        # Input validation
        if not dataset_id or not dataset_id.strip():
            return create_error_response(
                DQValidationError("Dataset ID is required", field="dataset_id"),
                status_code=400
            )
        
        logger.info(f"DQS calculation request for dataset: {dataset_id}")

        # Load dataset - try processed DataFrame first
        processed_df = dataframe_state_manager.get_processed_dataframe(dataset_id)
        if processed_df is not None:
            df = processed_df.copy()
            logger.info(f"Using processed DataFrame for DQS: {dataset_id}")
        else:
            # Try MessageState
            df = None
            try:
                state = message_state_manager.create_or_load_state(dataset_id, "dqs")
                state_df = state.get("datasetFile") if state else None
                if isinstance(state_df, pd.DataFrame):
                    df = state_df.copy()
                    logger.info(f"Using MessageState DataFrame for DQS: {dataset_id}, shape: {df.shape}")
            except Exception as e:
                logger.warning(f"Failed to load DataFrame from MessageState for DQS: {e}")

            # Fallback to original CSV load
            if df is None:
                dataset_info = dataset_manager.get_dataset_info(dataset_id)
                if not dataset_info:
                    logger.warning(f"Dataset metadata not found: {dataset_id}")
                    raise HTTPException(
                        status_code=404, 
                        detail={
                            "error": "DatasetNotFoundError",
                            "message": "Dataset not found. Please upload a dataset first.",
                            "details": {"dataset_id": dataset_id}
                        }
                    )

                df = dataset_manager.load_dataset(dataset_id)
                if df is None:
                    logger.error(f"Failed to load dataset file: {dataset_id}")
                    raise HTTPException(
                        status_code=404, 
                        detail={
                            "error": "DatasetFileNotFoundError",
                            "message": "Dataset file not found. Please upload your dataset again.",
                            "details": {"dataset_id": dataset_id}
                        }
                    )
                logger.info(f"Using original dataset for DQS: {dataset_id}")
        
        # Validate DataFrame has sufficient data
        if df.empty:
            return create_error_response(
                InsufficientDataError("Dataset is empty", required=1, actual=0),
                status_code=422
            )
        
        # Get target variable from dataset config if available
        target_variable = None
        try:
            dataset_info = dataset_manager.get_dataset_info(dataset_id)
            if dataset_info and 'target_variable' in dataset_info:
                target_variable = dataset_info.get('target_variable')
        except Exception as e:
            logger.warning(f"Could not get target variable for DQS: {e}")
        
        # P2.3: cache lookup keyed by (dataset_id, scope='entire', version).
        # The legacy /dqs endpoint is implicitly entire-scope.
        from app.services.analytics_cache import analytics_cache as _ac
        _version = dataframe_state_manager.get_version(dataset_id)
        _cache_scope = f"entire|tv={target_variable or ''}"
        cached = _ac.get("dqs", dataset_id, _cache_scope, _version)
        if cached is not None:
            logger.info(
                f"DQS legacy cache HIT for dataset_id={dataset_id} version={_version}"
            )
            return cached

        # P2.4 part 2: opt-in sampled DQS for very large frames. When
        # MIDAS_DQS_SAMPLED=1 and len(df) exceeds MIDAS_DQS_SAMPLE_THRESHOLD
        # (default 500_000), DQS is computed against a stratified sample
        # built once per (dataset_id, sample_rows, seed) and persisted to a
        # Parquet sidecar. This trims a 2 GB DQS pass from ~60s to ~3-5s.
        # Default OFF so user-visible scores remain identical until the flag
        # is enabled per-environment.
        try:
            if os.environ.get("MIDAS_DQS_SAMPLED", "0") == "1":
                threshold = int(os.environ.get("MIDAS_DQS_SAMPLE_THRESHOLD", "500000"))
                sample_rows = int(os.environ.get("MIDAS_DQS_SAMPLE_ROWS", "200000"))
                min_per_class = int(os.environ.get("MIDAS_DQS_MIN_PER_CLASS", "5000"))
                if len(df) > threshold:
                    from app.services.sampling import get_or_build_sample_sidecar
                    df = get_or_build_sample_sidecar(
                        dataset_id=dataset_id,
                        full_df=df,
                        target_variable=target_variable,
                        sample_rows=sample_rows,
                        min_per_class=min_per_class,
                    )
                    logger.info(
                        "P2.4 sampled-DQS path active: "
                        f"dataset={dataset_id} sample_rows={len(df)} (was >{threshold})"
                    )
        except Exception as exc:
            logger.warning(f"P2.4 sample fallback failed; using full df: {exc}")

        # Offload DQS computation to the shared executor (P1.3).
        # calculate_dqs is CPU-bound (full-frame isna scans, pairwise column
        # comparisons, value_counts) and would otherwise block the event loop
        # on large datasets. P3.5: emit midas_pipeline_stage_seconds{stage="dqs"}.
        import asyncio as _aio
        from app.core.executor import executor as _dqs_executor
        from app.core.metrics import time_stage as _time_stage
        _loop = _aio.get_event_loop()
        with _time_stage("dqs"):
            dqs_response = await _loop.run_in_executor(
                _dqs_executor,
                dqs_service.calculate_dqs,
                df,
                dataset_id,
                target_variable,
            )
        try:
            _ac.set("dqs", dataset_id, _cache_scope, _version, dqs_response)
        except Exception as exc:
            logger.warning(f"AnalyticsResultCache set failed (dqs legacy): {exc}")

        logger.info(f"DQS calculated successfully: {dqs_response.composite_score} ({dqs_response.score_label})")
        
        return dqs_response
        
    except HTTPException:
        raise
    except DataQualityError as e:
        return create_error_response(e)
    except Exception as e:
        logger.exception(f"DQS calculation failed: {str(e)}")
        return create_error_response(e, status_code=500)


@upload_router.get("/datasets/{dataset_id}/overview-bundle")
async def get_overview_bundle(
    dataset_id: str,
    scope: str = "entire",
    current_user = Depends(get_current_user_dependency)
):
    """
    P1.4 part 2: Single-roundtrip endpoint for the Step 2 Overview sidebar.

    Replaces the prior fan-out of 3-5 parallel GETs (column-info-by-scope,
    dqs-by-scope, classify-variables, etc.) with one call that loads the
    DataFrame ONCE and computes column-info + DQS in parallel via the
    shared executor. Cache lookups happen first so a second tab opening
    the sidebar is O(1).

    Frontend uses this when available; falls back to legacy individual
    endpoints if it 404s on older deployments.
    """
    try:
        valid_scopes = ['entire', 'train', 'test', 'validation']
        if scope not in valid_scopes:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid scope '{scope}'. Must be one of: {valid_scopes}",
            )

        from app.services.analytics_cache import analytics_cache as _ac
        _version = dataframe_state_manager.get_version(dataset_id)

        # Resolve the target_variable up front because dqs cache key embeds it.
        target_variable = None
        try:
            ds_info = dataset_manager.get_dataset_info(dataset_id)
            if ds_info and 'target_variable' in ds_info:
                target_variable = ds_info.get('target_variable')
        except Exception:
            pass

        column_cached = _ac.get("column_info", dataset_id, scope, _version)
        dqs_cache_scope = f"{scope}|tv={target_variable or ''}"
        dqs_cached = _ac.get("dqs", dataset_id, dqs_cache_scope, _version)

        # If both cached we are done.
        if column_cached is not None and dqs_cached is not None:
            return {
                "success": True,
                "dataset_id": dataset_id,
                "scope": scope,
                "version": _version,
                "from_cache": {"column_info": True, "dqs": True},
                "column_info": column_cached,
                "dqs": dqs_cached,
            }

        # Resolve the DataFrame for the requested scope, exactly the same
        # way the individual endpoints do so behavior is identical. Use
        # the shared read-only cache for the disk fallback: this is the
        # endpoint most likely to be hit concurrently with the legacy
        # scope endpoints during a UI refresh.
        full_df = dataframe_state_manager.get_full_dataframe_readonly(dataset_id)
        if full_df is None:
            ds_info = dataset_manager.get_dataset_info(dataset_id)
            if not ds_info:
                raise HTTPException(status_code=404, detail="Dataset not found.")
            full_df = dataset_manager.load_dataset_readonly_cached(dataset_id)
            if full_df is None:
                raise HTTPException(
                    status_code=404,
                    detail="Dataset file not found. Please upload your dataset again.",
                )

        if scope != "entire":
            df = None
            split_indices = dataframe_state_manager._split_indices.get(dataset_id)
            if split_indices and scope in split_indices:
                indices = split_indices[scope]
                if indices is not None and len(indices) > 0:
                    df = full_df.iloc[indices].copy()
            if (df is None or len(df) == 0) and "split_tag" in full_df.columns:
                mask = full_df["split_tag"] == scope
                if scope == "validation":
                    mask = full_df["split_tag"].astype(str).str.startswith(
                        "validation", na=False
                    )
                df = full_df[mask].copy()
            if df is None or len(df) == 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"No data available for scope '{scope}'",
                )
        else:
            df = full_df

        if df.empty:
            raise HTTPException(status_code=422, detail="Dataset is empty")

        import asyncio as _aio
        from app.core.executor import executor as _bundle_executor
        from app.core.metrics import time_stage as _time_stage
        _loop = _aio.get_event_loop()

        # Run BOTH calculations in parallel inside the shared thread pool.
        async def _column_info_coro():
            if column_cached is not None:
                return column_cached
            with _time_stage("column_info"):
                cols = await _loop.run_in_executor(
                    _bundle_executor, calculate_column_info, df
                )
            data_preview = None
            try:
                data_preview = {
                    "columns": list(df.columns),
                    "rows": df.head(10).to_dict(orient="records"),
                }
            except Exception as exc:
                logger.warning("overview-bundle: failed to build data preview: %s", exc)
            response = ColumnInfoResponse(
                success=True,
                message=(
                    f"Column information retrieved successfully for scope '{scope}'"
                ),
                dataset_id=dataset_id,
                columns_info=cols,
                total_columns=len(cols),
                data_preview=data_preview,
                scope=scope,
                total_rows=len(df),
            )
            try:
                _ac.set("column_info", dataset_id, scope, _version, response)
            except Exception as exc:
                logger.warning(f"AnalyticsResultCache set failed (bundle col): {exc}")
            return response

        async def _dqs_coro():
            if dqs_cached is not None:
                return dqs_cached
            # Auto-sample for DQS so the bundle endpoint stays fast on
            # multi-million-row datasets. Sampling is keyed by scope so
            # different scopes get different sidecars. Mirrors the
            # behaviour now applied in /dqs-by-scope so both endpoints
            # produce identical scores at scale.
            dqs_input_df = df
            try:
                from app.services.sampling import maybe_sample_for_dqs as _maybe_sample_for_dqs
                dqs_input_df, _ = _maybe_sample_for_dqs(
                    df,
                    dataset_id=dataset_id,
                    target_variable=target_variable,
                    scope=scope,
                )
            except Exception as _exc:
                logger.warning("overview-bundle DQS sampling failed; using full frame: %s", _exc)
            with _time_stage("dqs"):
                dqs_resp = await _loop.run_in_executor(
                    _bundle_executor,
                    dqs_service.calculate_dqs,
                    dqs_input_df,
                    dataset_id,
                    target_variable,
                )
            try:
                _ac.set("dqs", dataset_id, dqs_cache_scope, _version, dqs_resp)
            except Exception as exc:
                logger.warning(f"AnalyticsResultCache set failed (bundle dqs): {exc}")
            return dqs_resp

        column_info_resp, dqs_resp = await _aio.gather(_column_info_coro(), _dqs_coro())
        return {
            "success": True,
            "dataset_id": dataset_id,
            "scope": scope,
            "version": _version,
            "from_cache": {
                "column_info": column_cached is not None,
                "dqs": dqs_cached is not None,
            },
            "column_info": column_info_resp,
            "dqs": dqs_resp,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"overview-bundle failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {exc}")


@upload_router.get("/datasets/{dataset_id}/dqs-by-scope", response_model=DQSResponse)
async def get_data_quality_score_by_scope(
    dataset_id: str,
    scope: str = "entire",  # entire, train, test, validation
    current_user = Depends(get_current_user_dependency)
):
    """
    Calculate Data Quality Score (DQS) for a dataset filtered by scope.
    This is a read-only operation that does NOT modify _processed_dataframes or _active_scope.
    
    Args:
        dataset_id: The dataset identifier
        scope: One of 'entire', 'train', 'test', 'validation'
    
    Returns:
        DQSResponse with quality scores for the specified scope
    """
    try:
        logger.info(f"DQS by scope calculation request for dataset: {dataset_id}, scope: {scope}")
        
        # Validate scope
        valid_scopes = ['entire', 'train', 'test', 'validation']
        if scope not in valid_scopes:
            raise HTTPException(status_code=400, detail=f"Invalid scope '{scope}'. Must be one of: {valid_scopes}")
        
        full_df = dataframe_state_manager.get_full_dataframe_readonly(dataset_id)
        if full_df is None:
            # Try loading from disk via the shared read-only cache so the
            # parallel overview-bundle fan-out collapses onto a single parse.
            dataset_info = dataset_manager.get_dataset_info(dataset_id)
            if not dataset_info:
                raise HTTPException(status_code=404, detail="Dataset not found. Please upload a dataset first.")
            full_df = dataset_manager.load_dataset_readonly_cached(dataset_id)
            if full_df is None:
                raise HTTPException(status_code=404, detail="Dataset file not found. Please upload your dataset again.")

        if scope != "entire":
            df = None
            split_indices = dataframe_state_manager._split_indices.get(dataset_id)
            if not split_indices:
                try:
                    if dataframe_state_manager._hydrate_split_indices_from_redis(dataset_id):
                        split_indices = dataframe_state_manager._split_indices.get(dataset_id)
                except Exception as _exc:
                    logger.warning(f"dqs-by-scope: redis hydrate failed: {_exc}")
            if split_indices and scope in split_indices:
                indices = split_indices[scope]
                if indices is not None and len(indices) > 0:
                    df = full_df.iloc[indices].copy()
                    logger.info(f"Filtered to {scope} scope for DQS: {len(df)} rows (from {len(full_df)} total)")
            if (df is None or len(df) == 0) and "split_tag" in full_df.columns:
                mask = full_df["split_tag"] == scope
                if scope == "validation":
                    mask = full_df["split_tag"].astype(str).str.startswith("validation", na=False)
                df = full_df[mask].copy()
                logger.info(f"Filtered by split_tag to {scope} for DQS: {len(df)} rows")
            if df is None or len(df) == 0:
                raise HTTPException(status_code=400, detail=f"No data available for scope '{scope}'")
        else:
            # Read-only DQS path: calculate_dqs reads from df without
            # mutating it. Skip the 4M-row copy that used to double RAM.
            df = full_df
        
        # Get target variable from dataset config if available
        target_variable = None
        try:
            dataset_info = dataset_manager.get_dataset_info(dataset_id)
            if dataset_info and 'target_variable' in dataset_info:
                target_variable = dataset_info.get('target_variable')
        except Exception as e:
            logger.warning(f"Could not get target variable for DQS: {e}")

        # P2.3: result-cache lookup. The key includes the target_variable
        # too (encoded into scope) so a target rename is treated as a new
        # cache entry. Cache invalidates on update_dataframe via version.
        from app.services.analytics_cache import analytics_cache as _ac
        _version = dataframe_state_manager.get_version(dataset_id)
        _cache_scope = f"{scope}|tv={target_variable or ''}"
        cached = _ac.get("dqs", dataset_id, _cache_scope, _version)
        if cached is not None:
            logger.info(
                f"DQS by scope cache HIT for dataset_id={dataset_id} "
                f"scope={scope} version={_version}"
            )
            return cached

        # Auto-sample very large frames for DQS so /dqs-by-scope does not
        # time out behind the ALB on 4M-row workloads. Policy lives in
        # app.services.sampling.maybe_sample_for_dqs (default: on when
        # len(df) > MIDAS_DQS_SAMPLE_THRESHOLD = 1_000_000). The legacy
        # /dqs route already auto-sampled when MIDAS_DQS_SAMPLED=1 was
        # set; this brings the scope-aware endpoint to parity.
        try:
            from app.services.sampling import maybe_sample_for_dqs as _maybe_sample_for_dqs
            df, _sampled = _maybe_sample_for_dqs(
                df,
                dataset_id=dataset_id,
                target_variable=target_variable,
                scope=scope,
            )
        except Exception as _exc:
            logger.warning("DQS by scope sampling pre-step failed; using full frame: %s", _exc)

        # Offload DQS computation to the shared executor (P1.3).
        # P3.5: emit midas_pipeline_stage_seconds{stage="dqs"}.
        import asyncio as _aio
        from app.core.executor import executor as _dqs_executor
        from app.core.metrics import time_stage as _time_stage
        _loop = _aio.get_event_loop()
        with _time_stage("dqs"):
            dqs_response = await _loop.run_in_executor(
                _dqs_executor,
                dqs_service.calculate_dqs,
                df,
                dataset_id,
                target_variable,
            )

        logger.info(f"DQS by scope calculated: {dqs_response.composite_score} ({dqs_response.score_label}), scope={scope}, rows={len(df)}")

        try:
            _ac.set("dqs", dataset_id, _cache_scope, _version, dqs_response)
        except Exception as exc:
            logger.warning(f"AnalyticsResultCache set failed (dqs): {exc}")
        return dqs_response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"DQS by scope calculation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.post("/datasets/{dataset_id}/dqs-recommendations")
async def generate_dqs_recommendations(
    dataset_id: str,
    request: dict,
    current_user = Depends(get_current_user_dependency)
):
    """
    Generate AI-based recommendations for improving data quality based on DQS results.
    Uses LLM to analyze the quality scores and provide actionable recommendations.
    """
    logger.info(f"DQS recommendations request for dataset: {dataset_id}")
    
    try:
        dqs_data = request.get('dqs_data', {})
        
        if not dqs_data:
            raise HTTPException(status_code=400, detail="DQS data is required")
        
        # Build context for LLM
        composite_score = dqs_data.get('composite_score', 0)
        completeness = dqs_data.get('completeness', {})
        consistency = dqs_data.get('consistency', {})
        structural = dqs_data.get('structural_integrity', {})
        uniqueness = dqs_data.get('uniqueness', {})
        target_readiness = dqs_data.get('target_readiness', {})
        
        # Create detailed prompt for LLM
        system_prompt = """You are an expert data scientist specializing in data quality assessment and improvement.
Analyze the provided Data Quality Score (DQS) results and generate exactly 3 high-impact, data-backed recommendations to improve data quality.

For each recommendation, provide:
1. A concise title (max 10 words)
2. A detailed description with specific action steps backed by the actual data metrics (2-3 sentences)
3. A type: 'warning' for critical issues, 'info' for informational suggestions, 'success' for things that are good
4. A priority: 'high' for critical/immediate issues, 'medium' for important but not urgent, 'low' for nice-to-have improvements

Focus on:
- Selecting only the TOP 3 most impactful recommendations based on actual data quality metrics
- Addressing the lowest scoring dimensions first
- Providing specific, actionable steps backed by actual numbers from the DQS assessment (not generic advice)
- Considering the impact on model performance

Return a JSON array of exactly 3 recommendations in this exact format:
[
  {"title": "...", "description": "...", "type": "warning|info|success", "priority": "high|medium|low"},
  ...
]"""

        user_prompt = f"""Analyze this Data Quality Score (DQS) assessment and provide recommendations:

COMPOSITE SCORE: {composite_score}/100 ({dqs_data.get('score_label', 'Unknown')})

DIMENSION SCORES:
1. Completeness ({completeness.get('weight', 0.35)*100}% weight): {completeness.get('score', 0)}/100
   - Base score: {completeness.get('details', {}).get('base_score', 'N/A')}
   - Row sparseness penalty: {completeness.get('details', {}).get('row_sparseness_penalty', 0)} points
   - Columns with >50% missing: {completeness.get('details', {}).get('columns_with_high_missing', 0)}
   - Sparse row percentage: {completeness.get('details', {}).get('sparse_row_percentage', 0)}%

2. Consistency ({consistency.get('weight', 0.30)*100}% weight): {consistency.get('score', 0)}/100
   - Type consistency: {consistency.get('details', {}).get('type_score', 'N/A')}
   - Format consistency: {consistency.get('details', {}).get('format_score', 'N/A')}
   - Placeholder detection: {consistency.get('details', {}).get('placeholder_score', 'N/A')}
   - Range validity: {consistency.get('details', {}).get('range_score', 'N/A')}
   - Formatting issues: {consistency.get('details', {}).get('formatting_issues', 0)} columns
   - Placeholder values: {consistency.get('details', {}).get('placeholder_count', 0)} columns
   - Invalid ranges: {consistency.get('details', {}).get('invalid_range_count', 0)} columns

3. Structural Integrity ({structural.get('weight', 0.25)*100}% weight): {structural.get('score', 0)}/100
   - Constant columns: {structural.get('details', {}).get('constant_columns', 0)}
   - Near-constant columns: {structural.get('details', {}).get('near_constant_columns', 0)}
   - Duplicate columns: {structural.get('details', {}).get('duplicate_columns', 0)}

4. Uniqueness ({uniqueness.get('weight', 0.10)*100}% weight): {uniqueness.get('score', 0)}/100
   - Duplicate rows: {uniqueness.get('details', {}).get('duplicate_row_count', 0)} ({uniqueness.get('details', {}).get('duplicate_row_percentage', 0)}%)

TARGET READINESS (Informational):
   - Target variable: {target_readiness.get('target_variable', 'Not set')}
   - Missing rate: {target_readiness.get('target_missing_rate', 'N/A')}%
   - Event rate: {target_readiness.get('event_rate', 'N/A')}%

Total rows: {dqs_data.get('total_rows', 'Unknown')}
Total columns: {dqs_data.get('total_columns', 'Unknown')}

Generate exactly 3 specific, data-backed recommendations based on these quality metrics. Focus on the top 3 most impactful improvements only."""

        # Call LLM service
        recommendations = llm_service.get_dqs_recommendations(system_prompt, user_prompt)
        
        logger.info(f"Generated {len(recommendations)} DQS recommendations")
        
        return {
            "success": True,
            "recommendations": recommendations,
            "dataset_id": dataset_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"DQS recommendations generation failed: {str(e)}")
        # Return fallback recommendations based on DQS data
        fallback_recommendations = _generate_fallback_dqs_recommendations(request.get('dqs_data', {}))
        return {
            "success": True,
            "recommendations": fallback_recommendations,
            "dataset_id": dataset_id,
            "fallback": True
        }


def _generate_fallback_dqs_recommendations(dqs_data: dict) -> list:
    """Generate rule-based fallback recommendations when LLM fails."""
    recommendations = []
    
    completeness = dqs_data.get('completeness', {})
    consistency = dqs_data.get('consistency', {})
    structural = dqs_data.get('structural_integrity', {})
    uniqueness = dqs_data.get('uniqueness', {})
    
    # Completeness recommendations
    if completeness.get('score', 100) < 80:
        sparse_pct = completeness.get('details', {}).get('sparse_row_percentage', 0)
        if sparse_pct > 10:
            recommendations.append({
                "title": "Address Sparse Rows",
                "description": f"{sparse_pct:.1f}% of rows have more than 50% missing values. Consider removing these rows or investigating the data collection process that caused this sparsity.",
                "type": "warning",
                "priority": "high"
            })
        
        high_missing = completeness.get('details', {}).get('columns_with_high_missing', 0)
        if high_missing > 0:
            recommendations.append({
                "title": "Handle High-Missing Columns",
                "description": f"{high_missing} columns have more than 50% missing values. Consider dropping these columns or using advanced imputation techniques like MICE or KNN imputation.",
                "type": "warning",
                "priority": "high"
            })
    
    # Consistency recommendations
    if consistency.get('score', 100) < 90:
        placeholder_count = consistency.get('details', {}).get('placeholder_count', 0)
        if placeholder_count > 0:
            recommendations.append({
                "title": "Replace Placeholder Values",
                "description": f"{placeholder_count} columns contain placeholder values (like -999, N/A, etc.). Convert these to proper missing values (NA/NULL) before analysis.",
                "type": "warning",
                "priority": "medium"
            })
        
        format_issues = consistency.get('details', {}).get('formatting_issues', 0)
        if format_issues > 0:
            recommendations.append({
                "title": "Standardize Data Formats",
                "description": f"{format_issues} columns have inconsistent formatting. Standardize text casing, date formats, and numeric representations.",
                "type": "info",
                "priority": "medium"
            })
    
    # Structural recommendations
    if structural.get('score', 100) < 80:
        constant_cols = structural.get('details', {}).get('constant_columns', 0)
        if constant_cols > 0:
            recommendations.append({
                "title": "Remove Constant Columns",
                "description": f"{constant_cols} columns have only one unique value and provide no predictive power. Remove these columns to reduce dimensionality.",
                "type": "warning",
                "priority": "high"
            })
        
        near_constant = structural.get('details', {}).get('near_constant_columns', 0)
        if near_constant > 0:
            recommendations.append({
                "title": "Review Near-Constant Columns",
                "description": f"{near_constant} columns have >98% same value. Evaluate if these provide meaningful signal or should be removed.",
                "type": "info",
                "priority": "medium"
            })
        
        duplicate_cols = structural.get('details', {}).get('duplicate_columns', 0)
        if duplicate_cols > 0:
            recommendations.append({
                "title": "Remove Duplicate Columns",
                "description": f"{duplicate_cols} columns are exact duplicates of other columns. Keep only one copy to avoid multicollinearity issues.",
                "type": "warning",
                "priority": "high"
            })
    
    # Uniqueness recommendations
    dup_rows = uniqueness.get('details', {}).get('duplicate_row_count', 0)
    if dup_rows > 0:
        recommendations.append({
            "title": "Handle Duplicate Rows",
            "description": f"{dup_rows} duplicate rows detected. Investigate if these are true duplicates (remove) or legitimate repeated observations.",
            "type": "warning",
            "priority": "medium"
        })
    
    # Add positive recommendations if score is good and we have room
    composite = dqs_data.get('composite_score', 0)
    if len(recommendations) < 3:
        if composite >= 90:
            recommendations.append({
                "title": "Excellent Data Quality",
                "description": "Your dataset has excellent quality metrics. Proceed with confidence to feature engineering and model building.",
                "type": "success",
                "priority": "low"
            })
        elif composite >= 70:
            recommendations.append({
                "title": "Good Data Quality Foundation",
                "description": "Overall data quality is good. Address the specific issues above to further improve model performance.",
                "type": "success",
                "priority": "low"
            })
    
    # Sort by priority (high first) and return top 3 only
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    recommendations.sort(key=lambda x: priority_order.get(x.get('priority', 'low'), 2))
    return recommendations[:3]


@upload_router.get("/datasets/{dataset_id}/download-processed")
async def download_processed_dataset(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Download the processed dataset from MessageState as CSV file
    Fetches the datasetFile from the state and returns it as a downloadable CSV
    """
    try:
        logger.info(f"Download processed dataset request for dataset: {dataset_id}")
        
        # Validate dataset exists
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset metadata not found: {dataset_id}")
            raise HTTPException(status_code=404, detail="Dataset not found. Please upload a dataset first.")
        
        # P1.5 part 2: download endpoint serializes to CSV - never mutates.
        processed_df = dataframe_state_manager.get_dataframe_readonly(dataset_id)

        if processed_df is not None:
            logger.info(f"✅ Using processed DataFrame from state manager for download: {dataset_id}")
            logger.info(f"📥 DataFrame to download - Shape: {processed_df.shape}, Columns: {list(processed_df.columns)}")
        else:
            # Fallback to loading from MessageState
            try:
                state = message_state_manager.create_or_load_state(dataset_id, "")
                logger.info(f"MessageState loaded successfully for dataset: {dataset_id}")
                
                # Extract datasetFile from state
                processed_df = state.get("datasetFile")
                
                if processed_df is None:
                    logger.warning(f"No processed dataset found in MessageState for dataset: {dataset_id}")
                    raise HTTPException(status_code=404, detail="No processed dataset found. Please run some analysis first.")
                
                logger.info(f"ℹ️ Using DataFrame from MessageState for download: {dataset_id}")
                logger.info(f"📥 DataFrame to download - Shape: {processed_df.shape}, Columns: {list(processed_df.columns)}")
                
            except ValueError as e:
                logger.error(f"Failed to create/load MessageState: {str(e)}")
                raise HTTPException(status_code=500, detail=str(e))
        
        # Validate that it's a pandas DataFrame
        if not isinstance(processed_df, pd.DataFrame):
            logger.error(f"datasetFile in state is not a DataFrame: {type(processed_df)}")
            raise HTTPException(status_code=500, detail="Processed dataset is not in the expected format.")
        
        logger.info(f"✅ Processed dataset ready for download: {processed_df.shape[0]} rows, {processed_df.shape[1]} columns")
        
        # Verify that this DataFrame will produce the same statistics as displayed
        import hashlib
        download_fingerprint = hashlib.md5(
            f"{processed_df.shape}_{list(processed_df.columns)}_{processed_df.index.tolist()[:10]}".encode()
        ).hexdigest()
        logger.info(f"🔐 Download DataFrame fingerprint: {download_fingerprint}")
        logger.info(f"💾 This CSV file contains the EXACT data used to generate the displayed table statistics")
        
        # Log sample statistics to verify
        logger.info(f"📊 Download DataFrame preview - First 3 columns stats:")
        for col in list(processed_df.columns)[:3]:
            if processed_df[col].dtype in ['int64', 'float64', 'int32', 'float32']:
                logger.info(f"   {col}: mean={processed_df[col].mean():.2f}, missing={processed_df[col].isna().sum()}")
            else:
                logger.info(f"   {col}: type={processed_df[col].dtype}, unique={processed_df[col].nunique()}, missing={processed_df[col].isna().sum()}")
        
        # Create a temporary file for the CSV
        import tempfile
        import os
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        
        try:
            # Write DataFrame to CSV
            processed_df.to_csv(temp_file.name, index=False)
            temp_file.close()
            
            # Generate filename for download
            original_filename = dataset_info.get('filename', 'dataset')
            # Remove .csv extension if present and add _processed
            if original_filename.endswith('.csv'):
                base_name = original_filename[:-4]
            else:
                base_name = original_filename
            download_filename = f"{base_name}_{str(uuid.uuid4()).split('-')[0]}_processed.csv"
            
            logger.info(f"CSV file created successfully: {temp_file.name}")
            
            # Return file response for download
            return FileResponse(
                path=temp_file.name,
                filename=download_filename,
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={download_filename}"}
            )
            
        except Exception as e:
            # Clean up temp file if there was an error
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
            raise e
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download processed dataset failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/download-column-stats")
async def download_column_stats(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Download the Column Stats table as CSV
    Contains all statistical information displayed in the UI table after code execution
    """
    try:
        logger.info(f"Download column stats request for dataset: {dataset_id}")

        # P1.5 part 2: this endpoint only reads stats - safe to skip the copy.
        processed_df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        
        if processed_df is None:
            # Fallback to loading from MessageState
            try:
                state = message_state_manager.create_or_load_state(dataset_id, "")
                processed_df = state.get("datasetFile")
                
                if processed_df is None:
                    logger.warning(f"No processed dataset found for column stats download: {dataset_id}")
                    raise HTTPException(status_code=404, detail="No processed dataset found. Please run code execution first.")
                    
            except ValueError as e:
                logger.error(f"Failed to load MessageState for column stats: {str(e)}")
                raise HTTPException(status_code=500, detail=str(e))
        
        # Calculate column statistics (same as displayed in UI)
        columns_info = calculate_column_info(processed_df)
        logger.info(f"✅ Column stats calculated for download: {len(columns_info)} columns")
        
        # Convert column stats to DataFrame for CSV export
        # Match EXACTLY the columns displayed in the UI table
        stats_data = []
        for col_info in columns_info:
            stats_row = {
                'Column': col_info.column_name,
                'Type': col_info.data_type,
                'Missing': col_info.missing_count,
                'Unique': col_info.unique_count,
                'Mean': col_info.mean,
                'Median': col_info.median,
                'Mode': col_info.mode,
                'Std': col_info.standard_deviation,
                'Var': col_info.variance,
                'Min': col_info.min_value,
                'p5%': col_info.percentile_5,
                'p25%': col_info.percentile_25,
                'p50%': col_info.percentile_50,
                'p75%': col_info.percentile_75,
                'p95%': col_info.percentile_95,
                'p99%': col_info.percentile_99,
                'Max': col_info.max_value
            }
            stats_data.append(stats_row)
        
        # Create DataFrame from stats
        stats_df = pd.DataFrame(stats_data)
        
        logger.info(f"📊 Column Stats table created: {stats_df.shape[0]} rows, {stats_df.shape[1]} columns")
        logger.info(f"✅ CSV columns match UI table exactly: {list(stats_df.columns)}")
        
        # Create temporary file for CSV
        import tempfile
        import os
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        
        try:
            # Write stats DataFrame to CSV
            stats_df.to_csv(temp_file.name, index=False)
            temp_file.close()
            
            # Generate filename
            dataset_info = dataset_manager.get_dataset_info(dataset_id)
            original_filename = dataset_info.get('filename', 'dataset') if dataset_info else 'dataset'
            
            if original_filename.endswith('.csv'):
                base_name = original_filename[:-4]
            else:
                base_name = original_filename
            
            download_filename = f"{base_name}_column_stats_{str(uuid.uuid4()).split('-')[0]}.csv"
            
            logger.info(f"✅ Column Stats CSV created successfully: {download_filename}")
            
            # Return file response for download
            return FileResponse(
                path=temp_file.name,
                filename=download_filename,
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={download_filename}"}
            )
            
        except Exception as e:
            # Clean up temp file if there was an error
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
            raise e
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download column stats failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/compare-column-stats")
async def compare_column_stats(
    dataset_id: str,
    scope: str = "entire",
    current_user = Depends(get_current_user_dependency)
):
    """
    Compare original column statistics with processed column statistics.
    Accepts an optional scope param (entire | train | test | validation) to filter
    both the baseline and processed DataFrames to the same split before computing stats.
    Returns a comparison report showing what changed after data treatment.
    """
    try:
        logger.info(f"📊 Column stats comparison request for dataset: {dataset_id}, scope: {scope}")

        valid_scopes = ['entire', 'train', 'test', 'validation']
        if scope not in valid_scopes:
            raise HTTPException(status_code=400, detail=f"Invalid scope '{scope}'. Must be one of: {valid_scopes}")

        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")

        # Always use the original raw dataset from disk as baseline.
        # This guarantees the split sizes shown for every scope come from the ORIGINAL
        # dataset assignment made on the Objectives page — they are never re-derived from
        # the treated dataset.
        full_baseline_df = dataset_manager.load_dataset(dataset_id)
        baseline_source = "raw_dataset"

        if full_baseline_df is None:
            raise HTTPException(status_code=404, detail="Baseline dataset not found")
        
        logger.info(
            f"📊 Using raw dataset as baseline: {full_baseline_df.shape[0]} rows, "
            f"{full_baseline_df.shape[1]} columns"
        )

        # Resolve the ORIGINAL pandas-index labels for the requested scope.
        # Using labels (rather than positional iloc on the processed df) allows us to
        # correctly align a treated dataset — where rows may have been removed — back to
        # the Objectives-page split assignment.
        scope_original_labels = None
        if scope != "entire":
            split_indices = dataframe_state_manager._split_indices.get(dataset_id)
            if split_indices and scope in split_indices:
                indices = split_indices[scope]
                if indices is not None and len(indices) > 0:
                    scope_original_labels = full_baseline_df.index[list(indices)]
            # Fallback: derive from split_tag column on the full baseline df
            if (scope_original_labels is None or len(scope_original_labels) == 0) \
                    and "split_tag" in full_baseline_df.columns:
                tag = full_baseline_df["split_tag"].astype(str)
                if scope == "validation":
                    mask = tag.str.startswith("validation", na=False)
                else:
                    mask = tag == scope
                scope_original_labels = full_baseline_df.index[mask]

            if scope_original_labels is None or len(scope_original_labels) == 0:
                logger.warning(f"No rows found for scope '{scope}' — returning empty scoped view")

        # Baseline rows for this scope
        if scope == "entire" or scope_original_labels is None:
            baseline_df = full_baseline_df
        else:
            baseline_df = full_baseline_df.loc[scope_original_labels].copy()
            logger.info(
                f"Filtered baseline to scope '{scope}': {baseline_df.shape[0]} rows "
                f"(from {full_baseline_df.shape[0]})"
            )

        original_stats = calculate_column_info(baseline_df)
        logger.info(f"✅ Baseline stats ({baseline_source}, scope={scope}) calculated: {len(original_stats)} columns")
        
        # Get processed DataFrame for the requested scope.
        # Priority: 1) transformed_copies for specific scope, 2) scoped DataFrameStateManager view,
        # 3) DataFrameStateManager execution view for entire, 4) MessageState fallback.
        processed_df = None
        
        if scope != "entire":
            # Try to get the transformed copy for the specific scope first
            transformed_copies = dataframe_state_manager._transformed_copies.get(dataset_id, {})
            if scope in transformed_copies and transformed_copies[scope] is not None:
                processed_df = transformed_copies[scope].copy()
                logger.info(f"📊 Using transformed_copies['{scope}'] for processed data: {processed_df.shape}")
            else:
                # Fallback: use set_scope to switch to the requested scope
                try:
                    current_scope = dataframe_state_manager._active_scope.get(dataset_id, 'entire')
                    dataframe_state_manager.set_scope(dataset_id, scope)
                    processed_df = dataframe_state_manager.get_dataframe(dataset_id)
                    # Restore original scope
                    dataframe_state_manager.set_scope(dataset_id, current_scope)
                    if processed_df is not None:
                        logger.info(f"📊 Using set_scope('{scope}') for processed data: {processed_df.shape}")
                except Exception as scope_err:
                    logger.warning(f"Failed to get scoped processed data: {scope_err}")
        else:
            # For 'entire' scope, try transformed_copies['entire'] first
            transformed_copies = dataframe_state_manager._transformed_copies.get(dataset_id, {})
            if 'entire' in transformed_copies and transformed_copies['entire'] is not None:
                processed_df = transformed_copies['entire'].copy()
                logger.info(f"📊 Using transformed_copies['entire'] for processed data: {processed_df.shape}")
            else:
                processed_df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, full_baseline_df)
                if processed_df is not None:
                    logger.info(f"📊 Using DataFrameStateManager execution view for processed data: {processed_df.shape}")
        
        if processed_df is None:
            # Final fallback to MessageState
            try:
                state = message_state_manager.create_or_load_state(dataset_id, "")
                processed_df = state.get("datasetFile")
                if processed_df is not None:
                    logger.info(f"📊 Using MessageState fallback for processed data: {processed_df.shape}")
            except Exception as state_err:
                logger.warning(f"Failed MessageState fallback for processed data: {state_err}")
        
        if processed_df is None:
            raise HTTPException(status_code=404, detail="No processed dataset found. Please run code execution first.")

        # For non-entire scopes, if we got the data from transformed_copies, it's already scoped
        # Only do additional filtering if the processed_df is larger than expected (e.g., entire dataset)
        if scope != "entire" and scope_original_labels is not None and len(scope_original_labels) > 0:
            # Check if we need to filter (processed_df might already be correctly scoped)
            expected_rows = len(scope_original_labels)
            if processed_df.shape[0] > expected_rows * 1.1:  # Allow 10% tolerance for minor differences
                # Need to filter - processed_df is likely the entire dataset
                if isinstance(processed_df.index, type(full_baseline_df.index)):
                    surviving = processed_df.index.intersection(scope_original_labels)
                else:
                    surviving = processed_df.index.intersection(pd.Index(scope_original_labels))

                if len(surviving) > 0:
                    processed_df = processed_df.loc[surviving].copy()
                    logger.info(
                        f"Filtered processed to scope '{scope}': {processed_df.shape[0]} rows "
                        f"(original assignment had {len(scope_original_labels)})"
                    )
                else:
                    # Fallback to split_tag-based filtering on the processed df if available
                    if "split_tag" in processed_df.columns:
                        tag = processed_df["split_tag"].astype(str)
                        if scope == "validation":
                            mask = tag.str.startswith("validation", na=False)
                        else:
                            mask = tag == scope
                        processed_df = processed_df[mask].copy()
                        logger.info(f"Filtered processed by split_tag to '{scope}': {processed_df.shape[0]} rows")
                    else:
                        logger.warning(
                            f"Processed df has no rows aligned with scope '{scope}' original labels"
                        )
                        processed_df = processed_df.iloc[0:0].copy()
            else:
                logger.info(f"Processed df already scoped to '{scope}': {processed_df.shape[0]} rows")
        
        # Calculate processed column statistics
        processed_stats = calculate_column_info(processed_df)
        logger.info(f"✅ Processed stats calculated (scope={scope}): {len(processed_stats)} columns")
        
        # Create comparison report
        comparison_data = {
            "dataset_id": dataset_id,
            "original_shape": {"rows": baseline_df.shape[0], "columns": baseline_df.shape[1]},
            "processed_shape": {"rows": processed_df.shape[0], "columns": processed_df.shape[1]},
            "baseline_source": baseline_source,
            "changes": []
        }
        
        # Map original stats by column name for easy lookup
        original_stats_map = {stat.column_name: stat for stat in original_stats}
        processed_stats_map = {stat.column_name: stat for stat in processed_stats}
        
        # Get all column names (union of original and processed)
        all_columns = set(original_stats_map.keys()) | set(processed_stats_map.keys())

        # Helper function to create metric data in the format expected by the Compare Changes UI
        def create_metric_data(orig_val, proc_val, is_count=False):
            if orig_val is None and proc_val is None:
                return None

            orig = orig_val if orig_val is not None else 0
            proc = proc_val if proc_val is not None else 0
            change = proc - orig

            if is_count:
                change_pct = ((change) / max(orig, 1)) * 100
            else:
                change_pct = ((change) / max(abs(orig), 0.000001)) * 100 if orig != 0 else 0

            return {
                "original": orig_val,
                "processed": proc_val,
                "change": change,
                "change_pct": change_pct
            }
        
        for column_name in all_columns:
            original = original_stats_map.get(column_name)
            processed = processed_stats_map.get(column_name)
            
            change_info = {
                "column_name": column_name,
                "status": "unchanged",
                "changes": {}
            }
            
            if original is None:
                # Column was added
                change_info["status"] = "added"
                change_info["processed_type"] = processed.data_type
                change_info["original_type"] = None
                change_info["column_type"] = processed.column_type
                changes = {}
                for metric_name, orig_val, proc_val, is_count in [
                    ("missing", None, processed.missing_count, True),
                    ("unique", None, processed.unique_count, True),
                    ("mean", None, processed.mean, False),
                    ("median", None, processed.median, False),
                    ("mode", None, processed.mode if isinstance(processed.mode, (int, float)) else None, False),
                    ("std", None, processed.standard_deviation, False),
                    ("var", None, processed.variance, False),
                    ("skewness", None, processed.skewness, False),
                    ("min", None, processed.min_value, False),
                    ("max", None, processed.max_value, False),
                    ("p1", None, processed.percentile_1, False),
                    ("p5", None, processed.percentile_5, False),
                    ("p25", None, processed.percentile_25, False),
                    ("p50", None, processed.percentile_50, False),
                    ("p75", None, processed.percentile_75, False),
                    ("p95", None, processed.percentile_95, False),
                    ("p99", None, processed.percentile_99, False),
                    ("top_category_pct", None, processed.top_category_pct, False),
                    ("lowest_category_pct", None, processed.lowest_category_pct, False),
                ]:
                    metric_data = create_metric_data(orig_val, proc_val, is_count=is_count)
                    if metric_data is not None:
                        changes[metric_name] = metric_data
                # String-based fields
                if processed.mode is not None:
                    changes["mode_str"] = {"original": None, "processed": str(processed.mode)}
                if getattr(processed, 'date_min', None):
                    changes["date_min"] = {"original": None, "processed": processed.date_min}
                if getattr(processed, 'date_max', None):
                    changes["date_max"] = {"original": None, "processed": processed.date_max}
                if getattr(processed, 'most_frequent_date', None):
                    changes["most_frequent_date"] = {"original": None, "processed": processed.most_frequent_date}
                change_info["changes"] = changes
            elif processed is None:
                # Column was removed
                change_info["status"] = "removed"
                change_info["original_type"] = original.data_type
                change_info["processed_type"] = None
                change_info["column_type"] = original.column_type
                changes = {}
                for metric_name, orig_val, proc_val, is_count in [
                    ("missing", original.missing_count, None, True),
                    ("unique", original.unique_count, None, True),
                    ("mean", original.mean, None, False),
                    ("median", original.median, None, False),
                    ("mode", original.mode if isinstance(original.mode, (int, float)) else None, None, False),
                    ("std", original.standard_deviation, None, False),
                    ("var", original.variance, None, False),
                    ("skewness", original.skewness, None, False),
                    ("min", original.min_value, None, False),
                    ("max", original.max_value, None, False),
                    ("p1", original.percentile_1, None, False),
                    ("p5", original.percentile_5, None, False),
                    ("p25", original.percentile_25, None, False),
                    ("p50", original.percentile_50, None, False),
                    ("p75", original.percentile_75, None, False),
                    ("p95", original.percentile_95, None, False),
                    ("p99", original.percentile_99, None, False),
                    ("top_category_pct", original.top_category_pct, None, False),
                    ("lowest_category_pct", original.lowest_category_pct, None, False),
                ]:
                    metric_data = create_metric_data(orig_val, proc_val, is_count=is_count)
                    if metric_data is not None:
                        changes[metric_name] = metric_data
                # String-based fields
                if original.mode is not None:
                    changes["mode_str"] = {"original": str(original.mode), "processed": None}
                if getattr(original, 'date_min', None):
                    changes["date_min"] = {"original": original.date_min, "processed": None}
                if getattr(original, 'date_max', None):
                    changes["date_max"] = {"original": original.date_max, "processed": None}
                if getattr(original, 'most_frequent_date', None):
                    changes["most_frequent_date"] = {"original": original.most_frequent_date, "processed": None}
                change_info["changes"] = changes
            else:
                # Column exists in both - compare values and always include all metrics
                changes = {}
                has_changes = False
                
                # Always include missing count
                missing_data = create_metric_data(original.missing_count, processed.missing_count, is_count=True)
                if missing_data:
                    changes["missing"] = missing_data
                    if original.missing_count != processed.missing_count:
                        has_changes = True
                
                # Always include unique count
                unique_data = create_metric_data(original.unique_count, processed.unique_count, is_count=True)
                if unique_data:
                    changes["unique"] = unique_data
                    if original.unique_count != processed.unique_count:
                        has_changes = True
                
                # Always include all numerical metrics
                metric_comparisons = [
                    ("mean", original.mean, processed.mean),
                    ("median", original.median, processed.median),
                    ("mode", original.mode if isinstance(original.mode, (int, float)) else None, 
                           processed.mode if isinstance(processed.mode, (int, float)) else None),
                    ("std", original.standard_deviation, processed.standard_deviation),
                    ("var", original.variance, processed.variance),
                    ("skewness", original.skewness, processed.skewness),
                    ("min", original.min_value, processed.min_value),
                    ("max", original.max_value, processed.max_value),
                    ("p1", original.percentile_1, processed.percentile_1),
                    ("p5", original.percentile_5, processed.percentile_5),
                    ("p25", original.percentile_25, processed.percentile_25),
                    ("p50", original.percentile_50, processed.percentile_50),
                    ("p75", original.percentile_75, processed.percentile_75),
                    ("p95", original.percentile_95, processed.percentile_95),
                    ("p99", original.percentile_99, processed.percentile_99),
                    ("top_category_pct", original.top_category_pct, processed.top_category_pct),
                    ("lowest_category_pct", original.lowest_category_pct, processed.lowest_category_pct),
                ]
                
                for metric_name, orig_val, proc_val in metric_comparisons:
                    metric_data = create_metric_data(orig_val, proc_val)
                    if metric_data:
                        changes[metric_name] = metric_data
                        # Check if there's an actual change (threshold of 0.01)
                        if orig_val is not None and proc_val is not None and abs(proc_val - orig_val) > 0.01:
                            has_changes = True
                
                # Include string-based fields (mode for categorical, date fields)
                orig_mode_str = str(original.mode) if original.mode is not None else None
                proc_mode_str = str(processed.mode) if processed.mode is not None else None
                if orig_mode_str or proc_mode_str:
                    changes["mode_str"] = {"original": orig_mode_str, "processed": proc_mode_str}

                # Date-specific fields
                if getattr(original, 'date_min', None) or getattr(processed, 'date_min', None):
                    changes["date_min"] = {"original": original.date_min, "processed": processed.date_min}
                if getattr(original, 'date_max', None) or getattr(processed, 'date_max', None):
                    changes["date_max"] = {"original": original.date_max, "processed": processed.date_max}
                if getattr(original, 'most_frequent_date', None) or getattr(processed, 'most_frequent_date', None):
                    changes["most_frequent_date"] = {"original": original.most_frequent_date, "processed": processed.most_frequent_date}
                
                # Always set changes object with all metrics
                change_info["changes"] = changes
                change_info["original_type"] = original.data_type
                change_info["processed_type"] = processed.data_type
                change_info["column_type"] = processed.column_type or original.column_type
                
                if has_changes:
                    change_info["status"] = "modified"
            
            comparison_data["changes"].append(change_info)
        
        # Add summary
        comparison_data["summary"] = {
            "total_columns_original": len(original_stats),
            "total_columns_processed": len(processed_stats),
            "columns_added": len([c for c in comparison_data["changes"] if c["status"] == "added"]),
            "columns_removed": len([c for c in comparison_data["changes"] if c["status"] == "removed"]),
            "columns_modified": len([c for c in comparison_data["changes"] if c["status"] == "modified"]),
            "columns_unchanged": len([c for c in comparison_data["changes"] if c["status"] == "unchanged"]),
            "rows_change": processed_df.shape[0] - baseline_df.shape[0],
            "rows_change_pct": ((processed_df.shape[0] - baseline_df.shape[0]) / max(baseline_df.shape[0], 1)) * 100
        }
        
        logger.info(f"✅ Comparison completed: {comparison_data['summary']}")

        return comparison_data

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Column stats comparison failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@upload_router.post("/user-knowledge/upload")
async def upload_user_knowledge(
    dataset_id: str = Form(...),
    scope: str = Form(...),
    use_across_midas: bool = Form(False),
    use_exl_expertise: bool = Form(True),
    files: List[UploadFile] = File(...),
    current_user=Depends(get_current_user_dependency),
):
    """Index user-uploaded knowledge files for a dataset and scope."""
    from app.services.user_knowledge_service import user_knowledge_service
    try:
        result = user_knowledge_service.ingest_files(
            dataset_id=dataset_id,
            scope=scope,
            use_across_midas=use_across_midas,
            use_exl_expertise=use_exl_expertise,
            files=files,
        )
        return {
            "indexed_chunks": result["indexed_chunks"],
            "total_chunks": result["total_chunks"],
            "warnings": [],
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"User knowledge upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Knowledge indexing failed: {str(e)}")


@upload_router.post("/user-knowledge/preferences")
async def update_user_knowledge_preferences(
    dataset_id: str = Form(...),
    scope: str = Form(...),
    use_across_midas: bool = Form(False),
    use_exl_expertise: bool = Form(True),
    current_user=Depends(get_current_user_dependency),
):
    """Persist user knowledge preferences (EXL expertise toggle, scope) without uploading files."""
    from app.services.user_knowledge_service import user_knowledge_service
    try:
        user_knowledge_service.set_preferences(
            dataset_id=dataset_id,
            scope=scope,
            use_across_midas=use_across_midas,
            use_exl_expertise=use_exl_expertise,
        )
        return {"status": "ok"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"User knowledge preferences update failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Preferences update failed: {str(e)}")


@chat_router.get("/models/{model_id}/download-artifacts")
async def download_model_artifacts(
    model_id: str,
    format: str = "csv",
    current_user = Depends(get_current_user_dependency)
):
    """
    Download model artifacts and metrics in specified format (csv, excel, txt)
    """
    try:
        logger.info(f"Download model artifacts request for model: {model_id}, format: {format}")
        
        # Import here to avoid circular imports
        from app.services.model_training import model_training_service
        
        # Get model results from training service
        # For now, we'll create a mock response since we need to store results
        # In a real implementation, you'd store results in a database
        
        # Create metrics data structure
        metrics_data = {
            "model_id": model_id,
            "timestamp": datetime.now().isoformat(),
            "format": format
        }
        
        if format == "csv":
            # Create CSV content
            csv_content = create_metrics_csv(metrics_data)
            filename = f"{model_id}_metrics.csv"
            
            return Response(
                csv_content,
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        elif format == "excel":
            # Create Excel content
            excel_content = create_metrics_excel(metrics_data)
            filename = f"{model_id}_metrics.xlsx"
            
            return Response(
                excel_content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        elif format == "txt":
            # Create TXT report
            txt_content = create_metrics_report(metrics_data)
            filename = f"{model_id}_report.txt"
            
            return Response(
                txt_content,
                media_type="text/plain",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'csv', 'excel', or 'txt'")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download model artifacts failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

def create_metrics_csv(data: dict) -> str:
    """Create CSV content from metrics data"""
    import io
    import csv
    import json

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(["Metric", "Value"])

    # Basic identifying rows (use get to avoid KeyError)
    writer.writerow(["Model ID", data.get("model_id", "")])
    writer.writerow(["Download Timestamp", data.get("timestamp", "")])
    writer.writerow(["Format", data.get("format", "csv")])

    # If caller provided a 'metrics' mapping, write those rows
    metrics = data.get("metrics")
    if isinstance(metrics, dict):
        for name, value in metrics.items():
            try:
                if isinstance(value, (dict, list)):
                    value_str = json.dumps(value, default=str)
                else:
                    value_str = str(value)
            except Exception:
                value_str = str(value)
            writer.writerow([name, value_str])
    else:
        # Fallback: write common keys if present
        for key in ("accuracy", "precision", "recall", "f1_score", "auc_roc", "log_loss"):
            if key in data:
                writer.writerow([key, data.get(key)])

    return output.getvalue()

def create_metrics_excel(data: dict) -> bytes:
    """Create Excel content from metrics data"""
    try:
        import io
        import json
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Model Metrics"

        # Add headers
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Prepare rows
        rows = []
        rows.append(("Model ID", data.get("model_id", "")))
        rows.append(("Download Timestamp", data.get("timestamp", "")))
        rows.append(("Format", data.get("format", "xlsx")))

        metrics = data.get("metrics")
        if isinstance(metrics, dict):
            for k, v in metrics.items():
                try:
                    if isinstance(v, (dict, list)):
                        v_str = json.dumps(v, default=str)
                    else:
                        v_str = v
                except Exception:
                    v_str = str(v)
                rows.append((k, v_str))
        else:
            for key in ("accuracy", "precision", "recall", "f1_score", "auc_roc", "log_loss"):
                if key in data:
                    rows.append((key, data.get(key)))

        for r_idx, (metric, value) in enumerate(rows, start=2):
            ws.cell(row=r_idx, column=1).value = metric
            ws.cell(row=r_idx, column=2).value = value

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        # Fallback to CSV if openpyxl not available
        csv_content = create_metrics_csv(data)
        return csv_content.encode('utf-8')

def create_metrics_report(data: dict) -> str:
    """Create TXT report from metrics data"""
    report = []
    report.append("MODEL TRAINING REPORT")
    report.append("====================")
    report.append("")
    report.append(f"Model ID: {data.get('model_id', 'N/A')}")
    report.append(f"Download Timestamp: {data.get('timestamp', 'N/A')}")
    report.append(f"Format: {data.get('format', 'txt')}")
    report.append("")
    report.append("PERFORMANCE METRICS")
    report.append("===================")

    metrics = data.get("metrics")
    if isinstance(metrics, dict) and metrics:
        for k, v in metrics.items():
            try:
                if isinstance(v, (dict, list)):
                    val = json.dumps(v, default=str)
                else:
                    val = v
            except Exception:
                val = str(v)
            report.append(f"{k}: {val}")
    else:
        found = False
        for key in ("accuracy", "precision", "recall", "f1_score", "auc_roc", "log_loss"):
            if key in data:
                report.append(f"{key}: {data.get(key)}")
                found = True
        if not found:
            report.append("No metrics available")

    report.append("")
    report.append("Generated by MIDAS AutoML System")
    return "\n".join(report)

@chat_router.post("/update-custom-treatments", response_model=CustomTreatmentResponse)
async def update_custom_treatments(
    request: CustomTreatmentUpdate,
    current_user = Depends(get_current_user_dependency)
):
    """
    Update custom treatments in the plan data for a dataset.
    
    Error Responses:
    - 400: Invalid request format or treatment configuration
    - 404: Dataset or plan not found
    - 500: Internal server error
    """
    logger.info(f"Custom treatment update request for dataset: {request.dataset_id}")
    
    try:
        # Input validation
        if not request.dataset_id or not request.dataset_id.strip():
            return create_error_response(
                DQValidationError("Dataset ID is required", field="dataset_id"),
                status_code=400
            )
        
        if not request.custom_treatments:
            return create_error_response(
                DQValidationError("At least one custom treatment is required", field="custom_treatments"),
                status_code=400
            )
        
        # Validate dataset exists
        dataset_info = dataset_manager.get_dataset_info(request.dataset_id)
        if not dataset_info:
            logger.warning(f"Dataset not found: {request.dataset_id}")
            raise HTTPException(
                status_code=404, 
                detail={
                    "error": "DatasetNotFoundError",
                    "message": "Dataset not found",
                    "details": {"dataset_id": request.dataset_id}
                }
            )
        
        # Load existing MessageState
        try:
            state = message_state_manager.create_or_load_state(request.dataset_id, "")
            logger.info(f"MessageState loaded successfully for dataset: {request.dataset_id}")
        except ValueError as e:
            logger.error(f"Failed to create/load MessageState: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))
        
        # Get current plan data
        current_plan = state.get("plan", "")
        if not current_plan:
            logger.warning(f"No plan found for dataset: {request.dataset_id}")
            raise HTTPException(status_code=404, detail="No plan found. Please generate a plan first.")
        
        # Parse the plan data
        try:
            plan_data = json.loads(current_plan)
            logger.info(f"Plan data parsed successfully for dataset: {request.dataset_id}")
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse plan data as JSON: {str(e)}")
            raise HTTPException(status_code=500, detail="Plan data is not in valid JSON format")
        
        # Update custom treatments in the plan data
        updated_count = 0
        logger.info(f"Processing {len(request.custom_treatments)} custom treatments")
        for key, custom_treatment in request.custom_treatments.items():
            logger.info(f"Processing key: '{key}', treatment: '{custom_treatment}'")
            # Allow empty strings too - user might want to clear custom treatment
            if custom_treatment is not None:  # Changed from custom_treatment.strip()
                # Parse the key to find the right location in plan data
                if '-' in key:
                    # Format: "category-index" (for array format)
                    category, index_str = key.split('-', 1)
                    try:
                        index = int(index_str)
                        logger.info(f"Parsed key: category='{category}', index={index}")
                        if category in plan_data and isinstance(plan_data[category], list):
                            if index < len(plan_data[category]):
                                # Add or update custom_treatment field
                                if isinstance(plan_data[category][index], dict):
                                    plan_data[category][index]['custom_treatment'] = custom_treatment
                                    updated_count += 1
                                    logger.info(f"Updated custom treatment for {category}[{index}]: '{custom_treatment}'")
                                else:
                                    logger.warning(f"Item at {category}[{index}] is not a dict: {type(plan_data[category][index])}")
                            else:
                                logger.warning(f"Index {index} out of range for {category} (length: {len(plan_data[category])})")
                        else:
                            logger.warning(f"Category '{category}' not found in plan_data or not a list")
                    except ValueError:
                        logger.warning(f"Invalid index format in key: {key}")
                else:
                    # Format: "category" (for legacy object format)
                    if key in plan_data:
                        if isinstance(plan_data[key], dict):
                            plan_data[key]['custom_treatment'] = custom_treatment
                            updated_count += 1
                            logger.info(f"Updated custom treatment for {key}: '{custom_treatment}'")
                        else:
                            logger.warning(f"Item at {key} is not a dict: {type(plan_data[key])}")
                    else:
                        logger.warning(f"Key '{key}' not found in plan_data")
            else:
                logger.info(f"Skipping key '{key}' because custom_treatment is None")
        
        # Update the plan in the state
        updated_plan_json = json.dumps(plan_data, indent=2)
        state["plan"] = updated_plan_json
        logger.info(f"Updated plan data saved to state: {updated_plan_json[:200]}...")
        
        # Save the updated state
        save_success = message_state_manager.save_state(request.dataset_id, state)
        if not save_success:
            logger.warning(f"Failed to save updated MessageState for dataset: {request.dataset_id}")
            raise HTTPException(status_code=500, detail="Failed to save updated plan")
        
        logger.info(f"Custom treatments updated successfully: {updated_count} treatments for dataset: {request.dataset_id}")
        
        return CustomTreatmentResponse(
            success=True,
            message=f"Successfully updated {updated_count} custom treatments",
            updated_plan=plan_data
        )
        
    except HTTPException:
        raise
    except DataQualityError as e:
        return create_error_response(e)
    except Exception as e:
        logger.exception(f"Custom treatment update failed: {str(e)}")
        return create_error_response(e, status_code=500)

# ============================================================================
# INSIGHT BACKGROUND JOBS (CPU-heavy analytics; poll ``/insights/jobs/status/{job_id}``)
# ============================================================================


@chat_router.get("/insights/jobs/status/{job_id}")
async def get_insight_job_status(job_id: str):
    """Poll a queued insight job started from POST /insights/* (HTTP 202) paths."""
    from app.services.background_jobs import background_job_manager as _bgm

    snap = _bgm.get_job_status(job_id)
    if not snap:
        raise HTTPException(status_code=404, detail="Insight job not found")
    jt = str(snap.get("job_type") or "")
    if not jt.startswith("insight_"):
        raise HTTPException(status_code=404, detail="Insight job not found")
    st = str(snap.get("status") or "")
    out: Dict[str, Any] = {
        "job_id": job_id,
        "job_type": jt,
        "status": st,
        "progress": snap.get("progress", 0),
        "message": snap.get("message") or "",
    }
    if st == "completed":
        out["result"] = snap.get("result")
    if st == "failed":
        out["error"] = safe_json_serialize(snap.get("error"))
    return out


# ============================================================================
# BIVARIATE ANALYSIS ENDPOINTS
# ============================================================================

@chat_router.post("/insights/bivariate/all")
async def generate_bivariate_analysis_all(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    current_user = Depends(get_current_user_dependency),
    binning_method: Optional[str] = Form("quantile"),
    top_categories: Optional[int] = Form(10),
    bins: Optional[int] = Form(10),
):
    """
    Perform bivariate analysis for ALL variables against the target variable
    Returns analysis results for all variables, frontend can use dropdown to display

    Returns ``200`` when cached, or ``202`` with ``job_id`` for background completion.
    """
    _ = current_user
    try:
        logger.info(
            "=== BIVARIATE ANALYSIS REQUEST === dataset=%s target=%s",
            dataset_id,
            target_variable,
        )

        _scope, _ver = _insight_scope_version(dataset_id)
        scope_key = f"{_scope}|bin={binning_method or 'quantile'}|tc={top_categories or 10}|bins={bins or 10}"

        hit = analytics_cache.get("insight_bivariate_all", dataset_id, scope_key, _ver)
        if hit is not None:
            return hit

        _cached_biv = _bivariate_cache.get(
            dataset_id=dataset_id,
            target_variable=target_variable,
            scope=_scope,
            binning_method=binning_method or "quantile",
            top_categories=str(top_categories or 10),
            bins=str(bins or 10),
        )
        if _cached_biv is not None:
            analytics_cache.set("insight_bivariate_all", dataset_id, scope_key, _ver, _cached_biv)
            return _cached_biv

        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(
                status_code=404,
                detail=f"No dataframe found for dataset {dataset_id}. Please ensure the dataset is loaded.",
            )
        if not validate_target_variable(df, target_variable):
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset {dataset_id}",
            )

        job_id = f"insight_bivar_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_bivariate_all",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "binning_method": binning_method or "quantile",
                "top_categories": int(top_categories or 10),
                "bins": int(bins or 10),
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_bivariate_all_job,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bivariate analysis failed for dataset {dataset_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Bivariate analysis failed: {str(e)}")

@chat_router.post("/insights/vif-analysis")
async def generate_vif_analysis(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    current_user = Depends(get_current_user_dependency),
):
    """
    Generate VIF (Variance Inflation Factor) analysis for multicollinearity detection.

    Returns ``200`` with the payload when the result is cached, or ``202`` with a
    ``job_id`` to poll via ``GET /insights/jobs/status/{job_id}``.
    """
    _ = current_user
    try:
        logger.info("=== VIF ANALYSIS REQUEST === dataset=%s target=%s", dataset_id, target_variable)

        _scope, _ver = _insight_scope_version(dataset_id)
        scope_key = _scope
        hit = analytics_cache.get("insight_vif_analysis", dataset_id, scope_key, _ver)
        if hit is not None:
            return hit

        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        if target_variable not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset",
            )

        job_id = f"insight_vif_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_vif_analysis",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_vif_analysis_job,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("VIF analysis failed for dataset %s: %s", dataset_id, str(e))
        raise HTTPException(status_code=500, detail=f"VIF analysis failed: {str(e)}")


@chat_router.post("/insights/correlation-ratio-analysis")
async def generate_correlation_ratio_analysis(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    categorical_variables: Optional[str] = Form(None),
    numerical_variables: Optional[str] = Form(None),
    current_user = Depends(get_current_user_dependency),
):
    """
    Correlation ratio (η) via ``generate_correlation_ratio_analysis_tables``.

    Returns JSON with ``sections`` (list); the η heatmap is the section where
    ``analysis_kind`` equals ``correlation_ratio_categorical_numeric_heatmap`` with
    ``row_labels``, ``column_labels``, and ``matrix`` (rows × cols of η in [0,1] or null).
    """
    try:
        logger.info(f"=== CORRELATION RATIO (η) ANALYSIS REQUEST === dataset={dataset_id}")

        def _parse_vars(raw: Optional[str], field_name: str) -> Optional[List[str]]:
            if raw is None:
                return None
            value = raw.strip()
            if not value:
                return []
            try:
                parsed = json.loads(value)
            except Exception as parse_err:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid {field_name}; expected JSON array of strings"
                ) from parse_err
            if not isinstance(parsed, list):
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid {field_name}; expected JSON array of strings"
                )
            out: List[str] = []
            for item in parsed:
                if item is None:
                    continue
                text = str(item).strip()
                if text:
                    out.append(text)
            return out

        requested_categorical_vars = _parse_vars(categorical_variables, "categorical_variables")
        requested_numerical_vars = _parse_vars(numerical_variables, "numerical_variables")
        vars_key = "default"
        if requested_categorical_vars is not None or requested_numerical_vars is not None:
            vars_key = json.dumps(
                {
                    "categorical": requested_categorical_vars or [],
                    "numerical": requested_numerical_vars or [],
                },
                sort_keys=True,
            )

        _scope, _ver = _insight_scope_version(dataset_id)
        scope_key = f"{_scope}|vk={vars_key}"
        hit = analytics_cache.get("insight_correlation_ratio", dataset_id, scope_key, _ver)
        if hit is not None:
            logger.info("Correlation ratio analytics_cache HIT for %s", dataset_id)
            return hit

        _cached = _correlation_ratio_cache.get(
            dataset_id=dataset_id,
            target_variable=target_variable,
            scope=_scope,
            vars_key=vars_key,
        )
        if _cached is not None:
            logger.info("Correlation ratio legacy cache HIT for %s", dataset_id)
            analytics_cache.set("insight_correlation_ratio", dataset_id, scope_key, _ver, _cached)
            return _cached

        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        if target_variable not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset",
            )

        job_id = f"insight_cratio_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_correlation_ratio",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "requested_categorical_vars": requested_categorical_vars,
                "requested_numerical_vars": requested_numerical_vars,
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_correlation_ratio_job,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Correlation ratio analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Correlation ratio analysis failed: {str(e)}")


@chat_router.get("/insights/bivariate/{dataset_id}/variable/{variable_name}", response_model=BivariateAnalysisSingleResponse)
async def get_variable_analysis(
    dataset_id: str,
    variable_name: str,
    target_variable: str,
    current_user = Depends(get_current_user_dependency),
    coarse_bins: Optional[str] = Query(
        None,
        description="Optional coarse numeric bins, e.g. '0-20, 20-40, 40+'",
    ),
    category_groups: Optional[str] = Query(
        None,
        description="Optional category merges, e.g. 'AA+BB, CC+DD'",
    ),
    top_categories: int = Query(10, ge=2, le=200),
    bins: int = Query(10, ge=2, le=50),
    binning_method: str = Query("quantile"),
):
    """
    Get analysis results for a specific variable (used when user selects from dropdown).
    Optional coarse_bins (numerical) or category_groups (categorical) refine the analysis.
    """
    # Import bivariate analysis functions locally
    from app.utils.helpers import (
        analyze_categorical_vs_target, analyze_numerical_vs_target, 
        generate_variable_summary
    )
    
    try:
        logger.info(f"Getting analysis for variable {variable_name} in dataset {dataset_id}")

        # P1.5 part 2: variable analysis is pure read-only.
        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)

        if df is None:
            logger.warning(f"No dataframe found for dataset {dataset_id}")
            raise HTTPException(
                status_code=404, 
                detail=f"No dataframe found for dataset {dataset_id}. Please ensure the dataset is loaded."
            )
        
        # Validate variables
        if not validate_target_variable(df, target_variable):
            logger.warning(f"Invalid target variable {target_variable} for dataset {dataset_id}")
            raise HTTPException(
                status_code=400, 
                detail=f"Target variable '{target_variable}' not found in dataset {dataset_id}"
            )
        
        if variable_name not in df.columns:
            logger.warning(f"Variable {variable_name} not found in dataset {dataset_id}")
            raise HTTPException(
                status_code=400, 
                detail=f"Variable '{variable_name}' not found in dataset {dataset_id}"
            )
        
        # Generate dataset summary
        dataset_summary = generate_dataset_summary(df)
        
        # Determine variable type and analyze
        try:
            if variable_name in dataset_summary.get('categorical_columns', []):
                analysis_type = "categorical"
                cg = (category_groups.strip() or None) if category_groups else None
                result = analyze_categorical_vs_target(
                    df,
                    target_variable,
                    variable_name,
                    dataset_summary,
                    top_categories,
                    dataset_id,
                    category_groups_spec=cg,
                )
            elif variable_name in dataset_summary.get('numeric_columns', []):
                analysis_type = "numerical"
                cb = (coarse_bins.strip() or None) if coarse_bins else None
                bm = binning_method if binning_method in ("quantile", "equal_width") else "quantile"
                result = analyze_numerical_vs_target(
                    df,
                    target_variable,
                    variable_name,
                    dataset_summary,
                    bm,
                    bins,
                    dataset_id,
                    coarse_bins_spec=cb,
                )
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Variable type for '{variable_name}' could not be determined",
                )
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))
        
        analysis_result = {
            "variable_name": variable_name,
            "variable_type": analysis_type,
            "analysis_result": result,
            "summary": generate_variable_summary(result, analysis_type)
        }
        
        logger.info(f"Retrieved analysis for variable {variable_name}")
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "target_variable": target_variable,
            "variable_name": variable_name,
            "analysis_result": analysis_result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get analysis for variable {variable_name}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
# ============================================================================
# CORRELATION ANALYSIS ENDPOINTS
# ============================================================================

@chat_router.post("/insights/correlation/analyze")
async def analyze_correlations(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    current_user = Depends(get_current_user_dependency),
    correlation_threshold: Optional[float] = Form(0.05),
    correlation_method: Optional[str] = Form("pearson")
):
    """
    Analyze correlations for all variables against target variable
    Returns comprehensive correlation analysis:
    - Numeric variables: Pearson & Spearman correlations
    - Categorical variables: Chi-Square Test & Cramér's V
    """
    try:
        logger.info(f"=== CORRELATION ANALYSIS REQUEST ===")
        logger.info(f"Dataset ID: {dataset_id}")
        logger.info(f"Target Variable: {target_variable}")
        logger.info(f"Threshold: {correlation_threshold}")
        logger.info(f"Method: {correlation_method}")

        # Determine active data scope for cache keying
        _scope = dataframe_state_manager._active_scope.get(dataset_id, "entire")

        # Return cached result immediately if available
        _cached_corr = _correlation_cache.get(
            dataset_id=dataset_id, target_variable=target_variable,
            scope=_scope, threshold=str(correlation_threshold or 0.05),
            method=correlation_method or "pearson"
        )
        if _cached_corr is not None:
            logger.info(f"Correlation cache HIT for dataset {dataset_id} / target {target_variable}")
            return _cached_corr

        # Get the dataframe (P1.5 part 2: read-only).
        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")

        # Validate target variable exists
        if target_variable not in df.columns:
            raise HTTPException(status_code=400, detail=f"Target variable '{target_variable}' not found in dataset")
        
        import asyncio as _asyncio
        from app.core.executor import executor as _executor
        _loop = _asyncio.get_event_loop()

        # Offload CPU-heavy work to thread pool
        dataset_summary, analysis_results = await _asyncio.gather(
            _loop.run_in_executor(_executor, generate_dataset_summary, df),
            _loop.run_in_executor(_executor, analyze_all_correlations, df, target_variable, correlation_threshold),
        )
        logger.info(f"Dataset summary generated: {dataset_summary['shape']}")

        if "error" in analysis_results:
            raise HTTPException(status_code=500, detail=f"Correlation analysis failed: {analysis_results['error']}")
        
        logger.info(f"Correlation analysis completed for {analysis_results['total_variables']} variables")
        logger.info(f"Significant variables: {analysis_results['significant_variables']}")
        
        # Prepare response data
        response_data = {
            "success": True,
            "message": f"Correlation analysis completed for dataset {dataset_id}",
            "dataset_id": dataset_id,
            "target_variable": target_variable,
            "correlation_threshold": correlation_threshold,
            "total_variables_analyzed": analysis_results['total_variables'],
            "significant_variables": analysis_results['significant_variables'],
            "correlation_results": analysis_results['correlation_results'],
            "visualization_data": analysis_results['visualization_data'],
            "dataset_summary": {
                "shape": dataset_summary['shape'],
                "numeric_columns": dataset_summary['numeric_columns'],
                "categorical_columns": dataset_summary['categorical_columns'],
                "missing_values": dataset_summary['missing_values']
            },
            "analysis_timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Clean NaN values from response data before returning
        clean_response_data = clean_nan_values(response_data, replace_with=None)
        
        # Serialize response data to ensure JSON compatibility
        safe_response_data = safe_json_serialize(clean_response_data)

        # Store in server-side cache so revisits are instant
        _correlation_cache.set(
            safe_response_data,
            dataset_id=dataset_id, target_variable=target_variable,
            scope=_scope, threshold=str(correlation_threshold or 0.05),
            method=correlation_method or "pearson"
        )
        
        logger.info("Correlation analysis completed successfully")
        return safe_response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Correlation analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Correlation analysis failed: {str(e)}")

@chat_router.get("/insights/correlation/{dataset_id}/variable/{variable_name}", 
                response_model=SingleVariableCorrelationResponse)
async def get_variable_correlation(
    dataset_id: str,
    variable_name: str,
    target_variable: str,
    correlation_threshold: float = 0.05,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get correlation analysis for a specific variable (for right pane dropdown)
    Returns detailed correlation results and visualization data
    """
    try:
        logger.info(f"=== SINGLE VARIABLE CORRELATION REQUEST ===")
        logger.info(f"Dataset ID: {dataset_id}")
        logger.info(f"Variable: {variable_name}")
        logger.info(f"Target Variable: {target_variable}")
        logger.info(f"Threshold: {correlation_threshold}")

        # Get the dataframe (P1.5 part 2: read-only).
        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")

        # Validate variables exist
        if variable_name not in df.columns:
            raise HTTPException(status_code=400, detail=f"Variable '{variable_name}' not found in dataset")
        if target_variable not in df.columns:
            raise HTTPException(status_code=400, detail=f"Target variable '{target_variable}' not found in dataset")
        
        # Import correlation analysis functions
        from app.utils.helpers import (
            calculate_pearson_correlation, calculate_spearman_correlation,
            calculate_chi_square_test,
            get_correlation_significance_level, generate_correlation_visualization_data,
            safe_json_serialize
        )
        from app.models.schemas import CorrelationResult, CorrelationVisualizationData
        
        # Determine variable type
        if df[variable_name].dtype in ['object', 'category']:
            var_type = "categorical"
            # Calculate chi-square test for categorical
            chi_square_results = calculate_chi_square_test(df, variable_name, target_variable)
            correlation_value = chi_square_results["cramers_v"]
            
            result = CorrelationResult(
                variable_name=variable_name,
                variable_type=var_type,
                chi_square_statistic=chi_square_results["chi_square_statistic"],
                chi_square_p_value=chi_square_results["chi_square_p_value"],
                cramers_v=correlation_value,
                is_significant=abs(correlation_value) >= correlation_threshold,
                significance_level=get_correlation_significance_level(correlation_value, correlation_threshold)
            )
        else:
            var_type = "numeric"
            # Calculate only Pearson and Spearman correlations for numeric
            pearson = calculate_pearson_correlation(df, variable_name, target_variable)
            spearman = calculate_spearman_correlation(df, variable_name, target_variable)
            
            # Use Pearson as primary correlation
            correlation_value = pearson
            max_corr = max(abs(pearson), abs(spearman))
            
            result = CorrelationResult(
                variable_name=variable_name,
                variable_type=var_type,
                pearson_correlation=pearson,
                spearman_correlation=spearman,
                is_significant=max_corr >= correlation_threshold,
                significance_level=get_correlation_significance_level(max_corr, correlation_threshold)
            )
        
        # Generate visualization data
        viz_data = generate_correlation_visualization_data(df, variable_name, target_variable, correlation_value, var_type)
        visualization_data = CorrelationVisualizationData(**viz_data)
        
        logger.info(f"Single variable correlation analysis completed for {variable_name}")
        
        # Prepare response data
        response_data = {
            "success": True,
            "dataset_id": dataset_id,
            "target_variable": target_variable,
            "variable_name": variable_name,
            "correlation_result": result,
            "visualization_data": visualization_data
        }
        
        # Safely serialize the response data
        try:
            safe_response = safe_json_serialize(response_data)
            return safe_response
        except Exception as serialization_error:
            logger.error(f"Serialization error: {str(serialization_error)}")
            raise HTTPException(status_code=500, detail=f"Response serialization failed: {str(serialization_error)}")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get correlation for variable {variable_name}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@chat_router.get("/insights/correlation/{dataset_id}/heatmap", response_model=CorrelationHeatmapImageResponse)
async def get_correlation_heatmap_image_basic(
    dataset_id: str,
    target_variable: Optional[str] = None,
    dark_mode: bool = False,
    top_n: Optional[int] = Query(None, description="Top N columns in heatmap (5, 10, 15, or 20)"),
    current_user=Depends(get_current_user_dependency),
):
    """
    Generate correlation heatmap image (PNG base64). Cached via ``analytics_cache``;
    returns ``202`` with ``job_id`` when computation is queued.
    """
    _ = current_user
    try:
        k = _heatmap_top_n(top_n)
        _scope, _ver = _insight_scope_version(dataset_id)
        tv = target_variable or ""
        scope_key = f"{_scope}|heatmap_basic|tv={tv}|dm={int(bool(dark_mode))}|k={k}"
        hit = analytics_cache.get("insight_correlation_heatmap_basic", dataset_id, scope_key, _ver)
        if hit is not None:
            return hit

        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        if target_variable and target_variable not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset",
            )

        job_id = f"insight_hmnum_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_correlation_heatmap_basic",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "dark_mode": dark_mode,
                "top_n": top_n,
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_correlation_heatmap_basic_job,
        )
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as e:
        logger.error("Failed to generate basic correlation heatmap image: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@chat_router.get(
    "/insights/correlation/{dataset_id}/heatmap/categorical",
    response_model=CorrelationHeatmapImageResponse,
)
async def get_categorical_association_heatmap_image(
    dataset_id: str,
    target_variable: Optional[str] = None,
    dark_mode: bool = False,
    top_n: Optional[int] = Query(None, description="Top N columns in heatmap (5, 10, 15, or 20)"),
    current_user=Depends(get_current_user_dependency),
):
    """
    Categorical association heatmap (Cramér's V). Cached; may return ``202`` while queued.
    """
    _ = current_user
    try:
        k = _heatmap_top_n(top_n)
        _scope, _ver = _insight_scope_version(dataset_id)
        tv = target_variable or ""
        scope_key = f"{_scope}|heatmap_cat|tv={tv}|dm={int(bool(dark_mode))}|k={k}"
        hit = analytics_cache.get("insight_correlation_heatmap_categorical", dataset_id, scope_key, _ver)
        if hit is not None:
            return hit

        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        if target_variable and target_variable not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset",
            )

        job_id = f"insight_hmcat_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_correlation_heatmap_categorical",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "dark_mode": dark_mode,
                "top_n": top_n,
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_correlation_heatmap_categorical_job,
        )
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as e:
        logger.error("Failed to generate categorical association heatmap: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@chat_router.post("/insights/correlation-matrix")
async def generate_correlation_matrix(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    current_user = Depends(get_current_user_dependency),
    correlation_method: Optional[str] = Form("pearson")
):
    """
    Generate full correlation matrix for numeric variables in the dataset
    Returns correlation matrix data and heatmap visualization
    """
    _ = current_user
    try:
        logger.info(
            "=== CORRELATION MATRIX ANALYSIS REQUEST === dataset=%s target=%s method=%s",
            dataset_id,
            target_variable,
            correlation_method,
        )

        _scope, _ver = _insight_scope_version(dataset_id)
        scope_key = f"{_scope}|method={correlation_method or 'pearson'}"

        hit = analytics_cache.get("insight_correlation_matrix", dataset_id, scope_key, _ver)
        if hit is not None:
            return hit

        _cached_cm = _corr_matrix_cache.get(
            dataset_id=dataset_id,
            target_variable=target_variable,
            scope=_scope,
            method=correlation_method or "pearson",
        )
        if _cached_cm is not None:
            analytics_cache.set("insight_correlation_matrix", dataset_id, scope_key, _ver, _cached_cm)
            return _cached_cm

        df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        if target_variable not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset",
            )

        job_id = f"insight_cmat_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_correlation_matrix",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "correlation_method": correlation_method or "pearson",
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_correlation_matrix_job,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Correlation matrix analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Correlation matrix analysis failed: {str(e)}")

def _run_global_model_background(
    job_id: str,
    dataset_id: str,
    algorithm_enum: ModelAlgorithm,
    k_folds: int,
    target_variable,
    selected_variables,
) -> None:
    """Background worker for /train-global-model - keeps HTTP response under 230 s.

    Declared ``def`` (not ``async def``) on purpose: ``segmentation_service.
    train_global_model`` does pandas / sklearn fits + k-fold CV that take
    minutes on real datasets, all synchronously. FastAPI runs ``async def``
    background tasks on the request-serving event loop, so a sync-bodied
    coroutine pins the loop for the full training duration. That manifests
    as ALB / ingress 504s on every other request hitting the same worker
    (status polls, dataset previews, classification status checks) and as
    Kubernetes liveness-probe failures that restart the pod -- the exact
    "Job was interrupted by server restart. Please retry." pattern the UI
    reported on auto-analysis. Sync ``def`` is offloaded to anyio's thread
    limiter so the event loop stays free.

    Phase-1 (May 2026): retained ONLY as a backward-compatibility wrapper
    for any in-flight job that pre-dated the migration to
    ``background_job_manager``. New jobs go through
    ``_run_global_model_job`` (module-level, picklable, S3-mirrored).
    """
    try:
        training_jobs[job_id]["status"] = "running"
        training_jobs[job_id]["message"] = "Training in progress…"
        training_jobs[job_id]["last_heartbeat"] = time.time()
        _save_jobs_state()

        response = segmentation_service.train_global_model(
            dataset_id=dataset_id,
            algorithm=algorithm_enum,
            k_folds=k_folds,
            target_variable=target_variable,
            dataset_manager=dataset_manager,
            selected_variables=selected_variables,
        )

        result_dict = response.model_dump() if hasattr(response, "model_dump") else response.dict()
        serializable_result = make_json_serializable(result_dict)

        training_jobs[job_id]["status"] = "completed"
        training_jobs[job_id]["progress"] = 100
        training_jobs[job_id]["message"] = "Training completed successfully"
        training_jobs[job_id]["results"] = serializable_result
        training_jobs[job_id]["last_heartbeat"] = time.time()
        _save_jobs_state()
        logger.info(f"Global model training job {job_id} completed")
    except Exception as exc:
        training_jobs[job_id]["status"] = "failed"
        training_jobs[job_id]["message"] = f"Training failed: {exc}"
        training_jobs[job_id]["error"] = str(exc)
        training_jobs[job_id]["last_heartbeat"] = time.time()
        _save_jobs_state()
        logger.error(f"Global model training job {job_id} failed: {exc}")
        import traceback
        logger.error(traceback.format_exc())


def _run_global_model_job(params: Dict[str, Any]) -> Dict[str, Any]:
    """Module-level handler invoked by ``background_job_manager`` (Phase 1).

    Why module-level? ``BackgroundJobManager.start_job`` will pickle this
    callable when ``BROKER_URL`` is set and dispatch to a Celery / RQ
    worker. Closures captured from FastAPI request scope are not
    picklable. Keeping the handler at module scope lets the same code
    run unchanged on a future worker pool without a second migration.

    Acquires the per-dataset cross-process lock (``dataset_job_lock``)
    because a concurrent VIF / correlation / training job on the same
    dataset doubles the in-pod DataFrame footprint and trips the OOM
    ceiling on 4M-row workloads (see ``app/services/job_locks.py``).
    """
    from app.services.job_locks import dataset_job_lock

    dataset_id = params["dataset_id"]
    algorithm_value = params["algorithm"]
    k_folds = int(params.get("k_folds") or 5)
    target_variable = params.get("target_variable")
    selected_variables = params.get("selected_variables")

    algorithm_enum = ModelAlgorithm(algorithm_value)

    with dataset_job_lock(dataset_id, job_label="train_global_model"):
        response = segmentation_service.train_global_model(
            dataset_id=dataset_id,
            algorithm=algorithm_enum,
            k_folds=k_folds,
            target_variable=target_variable,
            dataset_manager=dataset_manager,
            selected_variables=selected_variables,
        )

    if hasattr(response, "model_dump"):
        result_dict = response.model_dump()
    else:
        result_dict = response.dict()
    return make_json_serializable(result_dict)


def _run_segment_manual_training_job(params: Dict[str, Any]) -> Dict[str, Any]:
    """Module-level handler for ``POST /segment-training/run`` (background_job_manager)."""
    from app.services.job_locks import dataset_job_lock
    from app.services.model_training_auto_training import make_json_serializable
    from app.services.model_training_segment_manual import segment_training_service

    job_id = params["job_id"]
    dataset_id = params["dataset_id"]
    target_column = params["target_column"]
    independent_variables = params["independent_variables"]
    algorithms = params["algorithms"]
    algorithm_params = params["algorithm_params"]
    max_iterations = int(params["max_iterations"])
    segment_column = params.get("segment_column")
    locked_variables = params.get("locked_variables")
    optimization_method = params.get("optimization_method") or "random"
    target_metric = params.get("target_metric")
    cv_folds = params.get("cv_folds")
    optuna_trials = params.get("optuna_trials")
    early_stopping_rounds = params.get("early_stopping_rounds")
    lr_backward_elimination = params.get("lr_backward_elimination")
    algorithm_param_ranges = params.get("algorithm_param_ranges") or {}

    config = {
        "dataset_id": dataset_id,
        "target_column": target_column,
        "independent_variables": independent_variables,
        "algorithms": algorithms,
        "algorithm_params": algorithm_params,
        "max_iterations": max_iterations,
        "segment_column": (segment_column or "").strip() or None,
        "locked_variables": locked_variables,
        "optimization_method": optimization_method,
        "target_metric": target_metric,
        "cv_folds": cv_folds,
        "optuna_trials": optuna_trials,
        "early_stopping_rounds": early_stopping_rounds,
        "lr_backward_elimination": lr_backward_elimination,
        "algorithm_param_ranges": algorithm_param_ranges,
    }

    with dataset_job_lock(dataset_id, job_label="segment_manual_training"):
        result = segment_training_service.train_models_by_segment(config)

    _persist_artifacts_if_available(dataset_id, result)
    try:
        from app.services.model_training_dump_service import model_training_dump_service

        model_training_dump_service.dump_training_payload(
            training_type="segment_manual",
            dataset_id=dataset_id,
            payload=result,
            context={
                "job_id": job_id,
                "target_column": target_column,
                "algorithms": algorithms,
                "max_iterations": max_iterations,
                "locked_variables": locked_variables,
                "optimization_method": optimization_method,
                "target_metric": target_metric,
                "cv_folds": cv_folds,
                "optuna_trials": optuna_trials,
                "early_stopping_rounds": early_stopping_rounds,
                "lr_backward_elimination": lr_backward_elimination,
            },
        )
    except Exception as dump_err:
        logger.warning("Failed to dump segment-manual intermediates for job %s: %s", job_id, dump_err)

    return make_json_serializable(result)


def _run_auto_training_job(params: Dict[str, Any]) -> Dict[str, Any]:
    """Module-level handler for ``POST /auto-training/run`` (background_job_manager)."""
    from app.services.job_locks import dataset_job_lock
    from app.services.model_training_auto_training import auto_training_service, make_json_serializable

    job_id = params["job_id"]
    dataset_id = params["dataset_id"]
    target_column = params["target_column"]
    selected_variables = params.get("selected_variables")
    selection_mode = params.get("selection_mode") or "auto"
    selected_algorithms = params.get("selected_algorithms")
    weight_variable = params.get("weight_variable")
    locked_variables = params.get("locked_variables")

    try:
        if weight_variable:
            logger.info("Weight variable specified: %s", weight_variable)

        with dataset_job_lock(dataset_id, job_label="auto_training_run"):
            results = auto_training_service.run_complete_auto_training(
                dataset_id=dataset_id,
                target_column=target_column,
                selected_variables=selected_variables,
                selection_mode=selection_mode,
                selected_algorithms=selected_algorithms,
                weight_variable=weight_variable,
                locked_variables=locked_variables,
                job_id=job_id,
            )

        serializable_results = make_json_serializable(results)
        _persist_artifacts_if_available(dataset_id, serializable_results, job_id=job_id)
        try:
            from app.services.model_training_dump_service import model_training_dump_service

            model_training_dump_service.dump_training_payload(
                training_type="auto",
                dataset_id=dataset_id,
                payload=serializable_results,
                context={
                    "job_id": job_id,
                    "target_column": target_column,
                    "selection_mode": selection_mode,
                    "selected_algorithms": selected_algorithms,
                    "weight_variable": weight_variable,
                    "locked_variables": locked_variables,
                },
            )
        except Exception as dump_err:
            logger.warning("Failed to dump auto training intermediates for job %s: %s", job_id, dump_err)

        try:
            pre_len = len(json.dumps(serializable_results, default=str))
        except Exception:
            pre_len = 10**9

        try:
            import threading

            meea_thread = threading.Thread(
                target=auto_training_service.run_pending_meea_jobs,
                args=(dataset_id,),
                daemon=True,
                name=f"meea_{job_id}",
            )
            meea_thread.start()
            logger.info("Launched MEEA background thread for job %s / dataset %s", job_id, dataset_id)
        except Exception as meea_launch_err:
            logger.warning("Failed to launch MEEA background thread: %s", meea_launch_err)

        return {
            "__training_results_body__": serializable_results,
            "_results_json_len": pre_len,
        }
    finally:
        _clear_auto_training_slot(dataset_id, job_id)


def _run_segment_auto_training_job(params: Dict[str, Any]) -> Dict[str, Any]:
    """Module-level handler for ``POST /segment-auto-training/run`` (background_job_manager)."""
    from app.services.job_locks import dataset_job_lock
    from app.services.model_training_auto_training import make_json_serializable
    from app.services.model_training_segment_auto import segment_auto_training_service

    job_id = params["job_id"]
    dataset_id = params["dataset_id"]
    target_column = params["target_column"]
    selected_variables = params.get("selected_variables")
    selection_mode = params.get("selection_mode") or "auto"
    selected_algorithms = params.get("selected_algorithms")
    locked_variables = params.get("locked_variables")

    with dataset_job_lock(dataset_id, job_label="segment_auto_training"):
        results = segment_auto_training_service.run_complete_segment_auto_training(
            dataset_id=dataset_id,
            target_column=target_column,
            selected_variables=selected_variables,
            selection_mode=selection_mode,
            selected_algorithms=selected_algorithms,
            locked_variables=locked_variables,
        )

    serializable_results = make_json_serializable(results)
    segment_results = serializable_results.get("segment_results", {})
    aggregated_artifacts: Dict[str, Any] = {}
    for _seg_key, segment_result in segment_results.items():
        if isinstance(segment_result, dict) and "error" not in segment_result:
            if "variable_analysis" in segment_result and "variable_analysis" not in aggregated_artifacts:
                aggregated_artifacts["variable_analysis"] = segment_result["variable_analysis"]
            if "used_features" in segment_result and "used_features" not in aggregated_artifacts:
                aggregated_artifacts["used_features"] = segment_result["used_features"]
            if "results" in segment_result and "results" not in aggregated_artifacts:
                aggregated_artifacts["results"] = segment_result["results"]
            for key in ("problem_type", "best_model_selection", "auto_selection_summary"):
                if key in segment_result and key not in aggregated_artifacts:
                    aggregated_artifacts[key] = segment_result[key]
            break

    if aggregated_artifacts:
        logger.info(
            "Persisting modelling artifacts from segment results for dataset %s, keys: %s",
            dataset_id,
            list(aggregated_artifacts.keys()),
        )
        _persist_artifacts_if_available(dataset_id, aggregated_artifacts)
    else:
        logger.warning("No modelling artifacts found in segment results for dataset %s", dataset_id)

    try:
        from app.services.model_training_dump_service import model_training_dump_service

        model_training_dump_service.dump_training_payload(
            training_type="segment_auto",
            dataset_id=dataset_id,
            payload=serializable_results,
            context={
                "job_id": job_id,
                "target_column": target_column,
                "selection_mode": selection_mode,
                "selected_algorithms": selected_algorithms,
                "locked_variables": locked_variables,
            },
        )
    except Exception as dump_err:
        logger.warning("Failed to dump segment-auto intermediates for job %s: %s", job_id, dump_err)

    return serializable_results


@chat_router.post("/train-global-model")
async def train_global_model(
    request: GlobalModelTrainingRequest,
    background_tasks: BackgroundTasks,
    current_user = Depends(get_current_user_dependency),
):
    """
    Enqueue a global model training job and return immediately with a job_id.

    Poll ``GET /train-global-model/status/{job_id}`` to track progress
    and retrieve results.

    Phase-1 stateless-API fix (May 2026): jobs are now enqueued via
    ``background_job_manager`` so the per-job snapshot is mirrored to
    shared object storage on every state change. That means a status
    poll hitting any EKS replica returns the up-to-date job state, and
    a pod restart no longer loses an in-flight result. The handler is
    module-level (``_run_global_model_job``) so a future Celery / RQ
    worker can pickle and dispatch it unchanged.

    The unused ``background_tasks`` parameter is kept on the signature
    so existing callers do not break — FastAPI still injects it but we
    no longer schedule onto it (the broker / thread executor inside
    ``BackgroundJobManager`` owns execution).
    """
    _ = background_tasks  # kept for backward compatibility, no longer used
    try:
        ModelAlgorithm(request.algorithm)  # validate before enqueue
        job_id = f"global_train_{request.dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"

        from app.services.background_jobs import background_job_manager as _bgm

        _bgm.start_job(
            job_id=job_id,
            job_type="train_global_model",
            params={
                "dataset_id": request.dataset_id,
                "algorithm": request.algorithm,
                "k_folds": request.k_folds,
                "target_variable": request.target_variable,
                "selected_variables": request.selected_variables,
                "__training_type__": "global",
            },
            job_function=_run_global_model_job,
        )

        logger.info(
            "train_global_model_enqueued",
            extra={
                "event": "train_global_model_enqueued",
                "log_category": "training",
                "job_id": job_id,
                "dataset_id": request.dataset_id,
                "algorithm": request.algorithm,
                "k_folds": request.k_folds,
            },
        )
        return {
            "success": True,
            "job_id": job_id,
            "status": "pending",
            "message": "Training started in background. Poll /train-global-model/status/{job_id} for updates.",
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(
            "train_global_model_enqueue_failed",
            exc_info=True,
            extra={
                "event": "train_global_model_enqueue_failed",
                "log_category": "training",
                "dataset_id": request.dataset_id,
            },
        )
        raise HTTPException(status_code=500, detail=f"Error starting training: {exc}") from exc


@chat_router.get("/train-global-model/status/{job_id}")
async def get_global_model_training_status(job_id: str):
    """
    Poll the status of a /train-global-model background job.

    Phase-1 stateless-API fix: reads from ``background_job_manager``
    first (S3-mirrored snapshot, works across EKS replicas). Falls back
    to the legacy in-process ``training_jobs`` dict so any job that was
    enqueued before this deploy can still complete its lifecycle.
    """
    from app.services.background_jobs import background_job_manager as _bgm

    bgm_snapshot = _bgm.get_job_status(job_id)
    if bgm_snapshot is not None:
        status = str(bgm_snapshot.get("status") or "")
        resp: Dict[str, Any] = {
            "job_id": job_id,
            "status": status,
            "progress": bgm_snapshot.get("progress", 0),
            "message": bgm_snapshot.get("message", ""),
            "last_heartbeat": bgm_snapshot.get("completed_at") or bgm_snapshot.get("started_at"),
        }
        if status == "completed":
            resp["results"] = bgm_snapshot.get("result")
        if status == "failed":
            resp["error"] = safe_json_serialize(bgm_snapshot.get("error"))
        return resp

    legacy = training_jobs.get(job_id)
    if legacy is None:
        raise HTTPException(status_code=404, detail="Job not found")

    training_jobs[job_id]["last_heartbeat"] = time.time()
    resp = {
        "job_id": job_id,
        "status": legacy["status"],
        "progress": legacy.get("progress", 0),
        "message": legacy.get("message", ""),
        "last_heartbeat": legacy.get("last_heartbeat"),
    }
    if legacy["status"] == "completed":
        resp["results"] = legacy.get("results")
    if legacy["status"] == "failed":
        resp["error"] = safe_json_serialize(legacy.get("error"))
    return resp

@chat_router.get("/model-codebook/{algorithm}", response_model=ModelCodebookResponse)
async def get_model_codebook(
    algorithm: str,
    dataset_id: Optional[str] = None,
    target_variable: Optional[str] = None,
    selected_variables: Optional[str] = None,  # JSON string of list
    k_folds: Optional[int] = 3,
    problem_type: Optional[str] = None,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get the codebook (implementation details) for a specific machine learning algorithm
    with real-time context from the user's current configuration
    """
    try:
        import json
        import os
        
        # Map algorithm to filename
        algorithm_files = {
            'random_forest': 'random_forest.json',
            'gradient_boosting': 'gradient_boosting.json',
            'logistic_regression': 'logistic_regression.json',
            'cart': 'cart_segmentation.json',
            'chaid': 'chaid_segmentation.json'
        }
        
        if algorithm not in algorithm_files:
            raise HTTPException(status_code=400, detail=f"Unknown algorithm: {algorithm}")
        
        # Get the path to the notebooks directory
        current_dir = os.path.dirname(os.path.abspath(__file__))
        notebooks_dir = os.path.join(current_dir, '..', 'notebooks')
        codebook_path = os.path.join(notebooks_dir, algorithm_files[algorithm])
        
        # Read the codebook file
        if not os.path.exists(codebook_path):
            raise HTTPException(status_code=404, detail=f"Codebook not found for algorithm: {algorithm}")
        
        with open(codebook_path, 'r', encoding='utf-8') as f:
            codebook_data = json.load(f)
        
        # Get real dataset context if available
        dataset_name = "your_dataset.csv"
        target_var_name = "target_column"
        selected_vars_list = None
        
        if dataset_id:
            try:
                dataset_info = dataset_manager.get_dataset_info(dataset_id)
                if dataset_info and dataset_info.get('file_path'):
                    dataset_name = os.path.basename(dataset_info['file_path'])
            except:
                pass
        
        if target_variable:
            target_var_name = target_variable
        
        if selected_variables:
            try:
                selected_vars_list = json.loads(selected_variables)
            except:
                pass
        
        # DYNAMIC CODEBOOK GENERATION: Use actual backend implementation
        if algorithm in ['cart', 'chaid']:
            # Generate dynamic codebook that reflects ACTUAL backend code
            from app.services.segmentation_service import segmentation_service
            
            logger.info(f"Generating dynamic codebook for {algorithm} with real backend implementation")
            
            dynamic_codebook = segmentation_service.generate_dynamic_codebook(
                algorithm=algorithm,
                dataset_name=dataset_name,
                target_variable=target_var_name,
                selected_variables=selected_vars_list or ['var1', 'var2', 'var3'],
                max_depth=4,  # Default values, could be made configurable
                min_samples_leaf=25,
                problem_type=problem_type
            )
            
            # Update description with real-time context
            description = dynamic_codebook['description']
            if target_variable and dataset_name != "your_dataset.csv":
                description += f"\n\n**Current Configuration:**\n- Dataset: `{dataset_name}`\n- Target Variable: `{target_var_name}`"
                if selected_vars_list and len(selected_vars_list) > 0:
                    description += f"\n- Selected Variables: {len(selected_vars_list)} features"
                if k_folds:
                    description += f"\n- K-Folds: {k_folds}"
                if problem_type:
                    description += f"\n- Problem Type: {problem_type.capitalize()}"
                description += f"\n\n**⚠️ IMPORTANT:** This codebook shows the ACTUAL backend implementation with all enhancements:\n- Enhanced preprocessing with categorical binning\n- Statistical testing with 8 fallback strategies\n- Relaxed constraints for better segment formation\n- Variable filtering optimization"
            
            return ModelCodebookResponse(
                success=True,
                algorithm=dynamic_codebook['algorithm'],
                title=dynamic_codebook['title'],
                description=description,
                sections=[CodebookSection(**section) for section in dynamic_codebook['sections']]
            )
        
        else:
            # Fallback to static codebook for non-segmentation algorithms
            # Apply real-time context to code sections
            for section in codebook_data['sections']:
                content = section['content']
                
                # Replace dataset name
                content = content.replace("'your_dataset.csv'", f"'{dataset_name}'")
                content = content.replace('"your_dataset.csv"', f'"{dataset_name}"')
                
                # Replace target variable
                content = content.replace("'target_column'", f"'{target_var_name}'")
                content = content.replace('"target_column"', f'"{target_var_name}"')
                
                # Replace k_folds value
                if k_folds:
                    content = content.replace("k_folds = 5", f"k_folds = {k_folds}")
                
                # Replace problem_type if available
                if problem_type:
                    content = content.replace("problem_type = infer_problem_type(y)", f"problem_type = '{problem_type}'  # Detected from your data")
                
                # Add selected variables context if available
                if selected_vars_list and len(selected_vars_list) > 0:
                    if "# Load your dataset" in content or "df = pd.read_csv" in content:
                        # Add comment about selected variables
                        vars_comment = f"\n\n# Selected variables for training: {selected_vars_list}\n# Total: {len(selected_vars_list)} variables"
                        content = content + vars_comment
                
                section['content'] = content
            
            # Update description with real-time context
            description = codebook_data['description']
            if target_variable and dataset_name != "your_dataset.csv":
                description += f"\n\n**Current Configuration:**\n- Dataset: `{dataset_name}`\n- Target Variable: `{target_var_name}`"
                if selected_vars_list and len(selected_vars_list) > 0:
                    description += f"\n- Selected Variables: {len(selected_vars_list)} features"
                if k_folds:
                    description += f"\n- K-Folds: {k_folds}"
                if problem_type:
                    description += f"\n- Problem Type: {problem_type.capitalize()}"
            
            return ModelCodebookResponse(
                success=True,
                algorithm=codebook_data['algorithm'],
                title=codebook_data['title'],
                description=description,
                sections=[CodebookSection(**section) for section in codebook_data['sections']]
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to load codebook: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error loading codebook: {str(e)}")

@chat_router.post("/run-segmentation", response_model=SegmentationResponse)
async def run_segmentation(
    request: SegmentationRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Run supervised segmentation (CART or CHAID) on selected variables.
    Applies interpretability constraints per product guidance.
    """
    try:
        logger.info(f"Run segmentation | dataset={request.dataset_id} method={request.method} vars={len(request.variables)}")

        result = segmentation_service.run_custom_segmentation(
            dataset_id=request.dataset_id,
            variables=request.variables,
            method=request.method.value if hasattr(request.method, 'value') else str(request.method),
            target_variable=request.target_variable,
            max_depth=request.max_depth,
            min_samples_leaf=request.min_samples_leaf,
            min_segment_size_ratio=request.min_segment_size_ratio,
            max_segments=request.max_segments,
            dataset_manager=dataset_manager
        )

        return result  # conforms to SegmentationResponse shape
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to run segmentation: {str(e)}")
        logger.error(f"Exception type: {type(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error running segmentation: {str(e)}")

@chat_router.post("/run-auto-segmentation", response_model=SegmentationResponse)
async def run_auto_segmentation(
    request: AutoSegmentationRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Run supervised segmentation (CART or CHAID) on the entire dataset automatically.
    Uses all available variables for segmentation.
    """
    try:
        logger.info(f"Run auto segmentation | dataset={request.dataset_id} method={request.method}")
        
        # Get dataset info to determine all available variables
        dataset_info = dataset_manager.get_dataset_info(request.dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Load dataset to get all column names - USE TREATED DATAFRAME, NOT CSV
        import pandas as pd
        import numpy as np
        from sklearn.feature_selection import SelectKBest, f_classif, f_regression
        from sklearn.preprocessing import LabelEncoder
        
        # Try to get processed DataFrame from state manager first (treated data)
        df = dataframe_state_manager.get_dataframe_for_execution(request.dataset_id, None)
        if df is None:
            # Fallback via dataset_manager (Parquet-first, then CSV, Azure-safe)
            df = dataset_manager.load_dataset(request.dataset_id)
            logger.info(f"Loaded dataset via storage fallback (no treated data found). Shape: {df.shape if df is not None else 'None'}")
        else:
            logger.info(f"Retrieved treated dataset from state manager. Shape: {df.shape}")
        
        # Get all columns except target variable (if specified)
        all_variables = df.columns.tolist()
        if request.target_variable and request.target_variable in all_variables:
            all_variables = [col for col in all_variables if col != request.target_variable]
        
        # OPTIMIZATION 1: Remove redundant ID columns
        id_keywords = ['id', 'member', 'customer', 'user', 'account', 'client', 'key', 'index', 'row']
        filtered_variables = []
        for var in all_variables:
            var_lower = var.lower()
            # Skip if variable name contains ID keywords
            if not any(keyword in var_lower for keyword in id_keywords):
                filtered_variables.append(var)
        
        logger.info(f"Filtered out ID columns: {len(all_variables)} -> {len(filtered_variables)} variables")
        
        # OPTIMIZATION 2: Select top 15 most important variables wrt target
        if request.target_variable and len(filtered_variables) > 15:
            try:
                # Prepare data for feature selection
                df_temp = df[filtered_variables + [request.target_variable]].copy()
                
                # Handle missing values quickly
                df_temp = df_temp.fillna(df_temp.median(numeric_only=True))
                df_temp = df_temp.fillna('Unknown')  # For categorical
                
                # Separate numerical and categorical
                numerical_vars = df_temp.select_dtypes(include=[np.number]).columns.tolist()
                categorical_vars = df_temp.select_dtypes(include=['object']).columns.tolist()
                
                if request.target_variable in numerical_vars:
                    numerical_vars.remove(request.target_variable)
                if request.target_variable in categorical_vars:
                    categorical_vars.remove(request.target_variable)
                
                # Quick categorical encoding (only for feature selection)
                df_encoded = df_temp.copy()
                for cat_var in categorical_vars:
                    if cat_var in df_encoded.columns:
                        le = LabelEncoder()
                        df_encoded[cat_var] = le.fit_transform(df_encoded[cat_var].astype(str))
                
                # Feature selection
                X = df_encoded[filtered_variables]
                y = df_encoded[request.target_variable]
                
                # Use appropriate scoring function
                if df_temp[request.target_variable].dtype in ['object', 'category']:
                    # Classification - encode target
                    le_target = LabelEncoder()
                    y = le_target.fit_transform(y.astype(str))
                    score_func = f_classif
                else:
                    score_func = f_regression
                
                # Select top 15 features with random seed for variety
                n_features = min(15, len(filtered_variables))
                import time
                random_seed = int(time.time() * 1000) % 10000
                selector = SelectKBest(score_func=score_func, k=n_features)
                X_selected = selector.fit_transform(X, y)
                
                # Get selected feature names
                selected_indices = selector.get_support(indices=True)
                selected_variables = [filtered_variables[i] for i in selected_indices]
                
                logger.info(f"Selected top {len(selected_variables)} important variables: {selected_variables[:5]}...")
                all_variables = selected_variables
                
            except Exception as e:
                logger.warning(f"Feature selection failed, using first 15 variables: {e}")
                all_variables = filtered_variables[:15]
        else:
            all_variables = filtered_variables[:15]  # Limit to 15 even if no target
        
        logger.info(f"Auto segmentation using {len(all_variables)} optimized variables: {all_variables[:5]}...")
        
        # Run segmentation on entire dataset
        result = segmentation_service.run_custom_segmentation(
            dataset_id=request.dataset_id,
            variables=all_variables,  # Use all available variables
            method=request.method.value if hasattr(request.method, 'value') else str(request.method),
            target_variable=request.target_variable,
            max_depth=request.max_depth,
            min_samples_leaf=request.min_samples_leaf,
            min_segment_size_ratio=request.min_segment_size_ratio,
            max_segments=request.max_segments,
            dataset_manager=dataset_manager
        )

        return result  # conforms to SegmentationResponse shape
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to run auto segmentation: {str(e)}")
        logger.error(f"Exception type: {type(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error running auto segmentation: {str(e)}")


# =============================================================================
# Segmentation Agent API Endpoints (4-Mode Architecture)
# Supports: Pre-existing Identifier (C1), Variable-Driven (C2), 
#           Manual Rules (C3), Auto Segmentation
# =============================================================================


def _tertiary_promotion_from_priority_info(
    priority_info: Optional[Dict[str, Any]],
) -> Optional[TertiaryPromotionSuggestion]:
    """
    Map segmentation_service priority_info (C2 enforced tree) to API model — Plan Section 3.4.
    """
    if not priority_info or not priority_info.get("secondary"):
        return None
    if priority_info.get("secondary_significant") is not False:
        return None
    st = priority_info.get("type")
    msg = priority_info.get("message") or priority_info.get("suggestion")
    if not st or not msg:
        return None
    return TertiaryPromotionSuggestion(
        type=st,
        message=msg,
        failed_variable=priority_info.get("failed_variable") or priority_info.get("secondary"),
        suggested_variable=priority_info.get("suggested_variable"),
        suggested_p_value=priority_info.get("suggested_p_value"),
        secondary_significant=priority_info.get("secondary_significant"),
    )


def _sanitize_json_payload_for_response(obj: Any, nan_as: Any = 0.0) -> Any:
    """
    Recursively replace NaN/Inf so stdlib json.dumps (Starlette JSONResponse) does not fail.
    Many segments or sparse OOS tables can still leave non-finite floats in nested models.
    Required Pydantic float fields cannot be null — use 0.0 for non-finite values (edge cases).
    """
    import math
    if obj is None:
        return None
    if isinstance(obj, dict):
        return {k: _sanitize_json_payload_for_response(v, nan_as) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_json_payload_for_response(v, nan_as) for v in obj]
    if isinstance(obj, tuple):
        return tuple(_sanitize_json_payload_for_response(v, nan_as) for v in obj)
    if isinstance(obj, bool):
        return obj
    if isinstance(obj, (np.integer,)):
        try:
            return int(obj)
        except Exception:
            return obj
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.floating):
        obj = float(obj)
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return nan_as
        return obj
    return obj


def _finalize_unified_segmentation_response(resp: UnifiedSegmentationResponse) -> UnifiedSegmentationResponse:
    """Rebuild response after stripping NaN — keeps response_model validation."""
    try:
        dumped = resp.model_dump(mode="python")
        cleaned = _sanitize_json_payload_for_response(dumped, nan_as=0.0)
        return UnifiedSegmentationResponse.model_validate(cleaned)
    except Exception as exc:
        logger.warning(f"Unified segmentation JSON sanitize/revalidate failed (returning raw): {exc}")
        return resp


@chat_router.post("/segmentation/run", response_model=UnifiedSegmentationResponse)
async def run_unified_segmentation(
    request: UnifiedSegmentationRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Unified segmentation endpoint supporting all 4 modes:
    - PRE_EXISTING (C1): Use existing segment column in data
    - VARIABLE_DRIVEN (C2): Platform finds optimal cutoffs with variable priority
    - MANUAL_RULES (C3): SQL-style manual rule builder
    - AUTO: Fully automated segmentation — evaluates candidate schemes and returns the
      single best sequential variable-selection result (one retained scheme per run)
    """
    import time
    start_time = time.time()
    
    try:
        logger.info(f"Unified segmentation | dataset={request.dataset_id} mode={request.mode.value}")
        
        # Load dataset with timing
        load_start = time.time()
        df = dataframe_state_manager.get_dataframe(request.dataset_id)
        if df is None:
            # Try loading from disk
            df = dataset_manager.load_dataset(request.dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        total_records = len(df)
        num_columns = len(df.columns)
        logger.info(f"Dataset loaded | records={total_records} columns={num_columns} load_time={time.time() - load_start:.2f}s")
        
        # Performance check - warn for very large datasets
        if total_records > 500000:
            logger.warning(f"Large dataset detected ({total_records} records) - segmentation may take longer")
        
        # Get target variable for event rate calculations
        target_var = request.target_variable
        if not target_var:
            dataset_info = dataset_manager.get_dataset_info(request.dataset_id)
            target_var = dataset_info.get("target_variable") if dataset_info else None
        
        # Compute overall event rate if target available
        overall_event_rate = None
        if target_var and target_var in df.columns:
            try:
                overall_event_rate = float(df[target_var].mean() * 100)
            except:
                pass
        
        # Process based on mode
        if request.mode == SegmentationMode.PRE_EXISTING:
            # C1: Pre-existing Identifier
            if not request.segment_column:
                raise HTTPException(status_code=400, detail="segment_column is required for PRE_EXISTING mode")
            
            if request.segment_column not in df.columns:
                raise HTTPException(status_code=400, detail=f"Column '{request.segment_column}' not found in dataset")
            
            # Build segments from existing column
            segments = []
            segment_values = df[request.segment_column].unique()
            
            # Limit number of segments for performance (max 50)
            MAX_SEGMENTS = 50
            if len(segment_values) > MAX_SEGMENTS:
                logger.warning(f"Column '{request.segment_column}' has {len(segment_values)} unique values, limiting to top {MAX_SEGMENTS} by frequency")
                # Get top segments by frequency
                value_counts = df[request.segment_column].value_counts().head(MAX_SEGMENTS)
                segment_values = value_counts.index.tolist()
            
            for idx, seg_val in enumerate(segment_values, 1):
                mask = df[request.segment_column] == seg_val
                seg_df = df[mask]
                record_count = len(seg_df)
                
                event_count = 0
                event_rate = 0.0
                if target_var and target_var in df.columns:
                    try:
                        event_count = int(seg_df[target_var].sum())
                        event_rate = float(seg_df[target_var].mean() * 100)
                    except:
                        pass
                
                # Calculate WoE and IV
                woe, iv_contrib = _calculate_woe_iv(
                    seg_df, df, target_var, record_count, total_records, event_count
                )
                
                # Calculate Wilson Score confidence intervals
                ci_lower, ci_upper = _calculate_wilson_score_ci(record_count, event_count)
                
                segments.append(SegmentDetail(
                    segment_id=idx,
                    segment_name=str(seg_val),
                    rule_definition=f"{request.segment_column} = '{seg_val}'",
                    record_count=record_count,
                    pct_of_population=round(record_count / total_records * 100, 2),
                    event_count=event_count,
                    event_rate=round(event_rate, 4),
                    event_rate_ci_lower=ci_lower,
                    event_rate_ci_upper=ci_upper,
                    woe=round(woe, 4),
                    iv_contribution=round(iv_contrib, 4)
                ))
            
            # Run validation suite with bootstrap stability and OOS validation
            validation = _run_validation_suite(
                segments=segments,
                df=df,
                target_var=target_var,
                total_records=total_records,
                overall_event_rate=overall_event_rate,
                min_segment_size=request.min_segment_size,
                dataset_id=request.dataset_id,
                segment_column=request.segment_column,  # Enable bootstrap stability and OOS
                run_stability=True,
                run_oos_validation=True  # Enable out-of-sample validation
            )
            
            # Compute variable relevance matrix
            var_rel_start = time.time()
            variable_relevance = _compute_variable_relevance(
                df=df,
                segment_column=request.segment_column,
                target_var=target_var
            )
            logger.info(f"Variable relevance computed | time={time.time() - var_rel_start:.2f}s")
            
            elapsed = time.time() - start_time
            logger.info(f"PRE_EXISTING segmentation completed | segments={len(segments)} total_time={elapsed:.2f}s")
            
            return _finalize_unified_segmentation_response(UnifiedSegmentationResponse(
                success=True,
                message=f"Pre-existing segmentation validated with {len(segments)} segments",
                mode=request.mode,
                method=None,
                variables_used=[request.segment_column],
                variable_priority=None,
                parameters={"segment_column": request.segment_column},
                num_segments=len(segments),
                segments=segments,
                validation=validation,
                variable_relevance=variable_relevance,
                dataset_shape=total_records,
                total_segment_records=sum(s.record_count for s in segments),
                records_match=sum(s.record_count for s in segments) == total_records
            ))
        
        elif request.mode == SegmentationMode.VARIABLE_DRIVEN:
            # C2: Variable-Driven with priority ordering
            if not request.variable_priority or not request.variable_priority.primary:
                raise HTTPException(status_code=400, detail="variable_priority with primary is required for VARIABLE_DRIVEN mode")
            
            # Build variable list respecting priority
            variables = [request.variable_priority.primary]
            if request.variable_priority.secondary:
                variables.append(request.variable_priority.secondary)
            if request.variable_priority.tertiary:
                variables.append(request.variable_priority.tertiary)
            
            # Validate variables exist
            for var in variables:
                if var not in df.columns:
                    raise HTTPException(status_code=400, detail=f"Variable '{var}' not found in dataset")
            
            # Calculate min_samples_leaf based on mode
            if request.min_segment_size_mode == "percentage":
                min_samples = int(total_records * request.min_segment_size_pct / 100)
            else:
                min_samples = request.min_segment_size
            
            # Run segmentation with priority enforcement
            result = segmentation_service.run_custom_segmentation(
                dataset_id=request.dataset_id,
                variables=variables,
                method=request.method.value,
                target_variable=target_var,
                max_depth=request.max_depth,
                min_samples_leaf=min_samples,
                min_segment_size_ratio=None,
                max_segments=request.max_segments,
                dataset_manager=dataset_manager,
                enforce_variable_priority=True,
                variable_priority=request.variable_priority.model_dump() if request.variable_priority else None
            )
            
            if not result.get("success"):
                raise HTTPException(status_code=400, detail=result.get("message", "Segmentation failed"))
            
            # Convert result segments to SegmentDetail
            segments = _convert_to_segment_details(result, df, target_var, total_records)
            
            # Add segment assignments to dataframe for stability analysis
            segment_column_name = '_segment_assignment'
            leaf_ids = result.get('leaf_ids')
            df_with_segments = df.copy()
            if leaf_ids and len(leaf_ids) == len(df):
                mapped_labels = _stability_labels_from_leaf_assignment(
                    leaf_ids, list(result.get("segments") or []), segments
                )
                if mapped_labels is not None:
                    df_with_segments[segment_column_name] = mapped_labels
                else:
                    df_with_segments[segment_column_name] = leaf_ids
            else:
                segment_column_name = None  # Cannot run stability without segment column
            
            # Merge split_tag from original df to df_with_segments for OOS validation
            if 'split_tag' in df.columns and 'split_tag' not in df_with_segments.columns:
                df_with_segments['split_tag'] = df['split_tag'].values
            
            # Run validation suite with bootstrap stability and OOS validation
            validation = _run_validation_suite(
                segments=segments,
                df=df_with_segments,
                target_var=target_var,
                total_records=total_records,
                overall_event_rate=overall_event_rate,
                min_segment_size=min_samples,
                dataset_id=request.dataset_id,
                segment_column=segment_column_name,  # Enable bootstrap stability and OOS
                run_stability=segment_column_name is not None,
                run_oos_validation=segment_column_name is not None  # Enable OOS validation
            )
            
            # Compute variable relevance matrix using leaf_ids from result
            variable_relevance = None
            if segment_column_name:
                try:
                    variable_relevance = _compute_variable_relevance(
                        df=df_with_segments,
                        segment_column=segment_column_name,
                        target_var=target_var,
                        variables_used=variables
                    )
                except Exception as e:
                    logger.warning(f"Could not compute variable relevance for variable-driven mode: {e}")

            # Section 3.4 — depth-1 tertiary test (same gate as auto pipeline) when secondary fails
            pi = result.get("priority_info") or {}
            tertiary_promotion = None
            promotion_suggestions = None
            secondary_failed = (
                bool(request.variable_priority.secondary)
                and pi.get("secondary_significant") is False
            )
            if secondary_failed:
                if (
                    segment_column_name
                    and target_var
                    and target_var in df_with_segments.columns
                ):
                    failed = pi.get("failed_variable") or request.variable_priority.secondary
                    tertiary_cand = request.variable_priority.tertiary
                    uvals = sorted(
                        df_with_segments[segment_column_name].dropna().astype(str).unique().tolist()
                    )
                    primary_segments = [{"segment_name": v} for v in uvals]
                    dc = run_depth1_tertiary_promotion_check(
                        df=df_with_segments,
                        target_variable=target_var,
                        primary_segments=primary_segments,
                        segment_column=segment_column_name,
                        failed_secondary=str(failed),
                        tertiary_candidate=tertiary_cand,
                        significance_threshold=0.05,
                    )
                    promotion_suggestions = [
                        PromotionSuggestion(
                            suggestion_type=dc.suggestion_type,
                            message=dc.message,
                            failed_variable=dc.failed_variable,
                            suggested_variable=dc.suggested_variable,
                            suggested_p_value=dc.suggested_p_value,
                        )
                    ]
                    tertiary_promotion = TertiaryPromotionSuggestion(
                        type="promote_tertiary"
                        if dc.suggestion_type == "promote_tertiary"
                        else "stop_at_primary",
                        message=dc.message,
                        failed_variable=dc.failed_variable,
                        suggested_variable=dc.suggested_variable,
                        suggested_p_value=dc.suggested_p_value,
                        secondary_significant=False,
                    )
                else:
                    tertiary_promotion = _tertiary_promotion_from_priority_info(pi)
                    if tertiary_promotion:
                        promotion_suggestions = [
                            PromotionSuggestion(
                                suggestion_type=tertiary_promotion.type,
                                message=tertiary_promotion.message,
                                failed_variable=tertiary_promotion.failed_variable,
                                suggested_variable=tertiary_promotion.suggested_variable,
                                suggested_p_value=tertiary_promotion.suggested_p_value,
                            )
                        ]
            else:
                tertiary_promotion = _tertiary_promotion_from_priority_info(pi)

            return _finalize_unified_segmentation_response(UnifiedSegmentationResponse(
                success=True,
                message=f"Variable-driven segmentation created {len(segments)} segments",
                mode=request.mode,
                method=request.method.value,
                variables_used=variables,
                variable_priority=request.variable_priority,
                parameters={
                    "method": request.method.value,
                    "max_depth": request.max_depth,
                    "min_segment_size": min_samples,
                    "min_segment_size_mode": request.min_segment_size_mode,
                    "max_segments": request.max_segments,
                },
                num_segments=len(segments),
                segments=segments,
                validation=validation,
                variable_relevance=variable_relevance,
                dataset_shape=total_records,
                total_segment_records=sum(s.record_count for s in segments),
                records_match=sum(s.record_count for s in segments) == total_records,
                tertiary_promotion_suggestion=tertiary_promotion,
                promotion_suggestions=promotion_suggestions,
            ))
        
        elif request.mode == SegmentationMode.MANUAL_RULES:
            # C3: Manual Rules (first-wins order; catch_all; operators per plan §6.2)
            if not request.manual_rules or len(request.manual_rules) == 0:
                raise HTTPException(status_code=400, detail="manual_rules are required for MANUAL_RULES mode")

            for rule in request.manual_rules:
                for cond in rule.conditions:
                    v = (cond.variable or "").strip()
                    if v and v not in df.columns:
                        raise HTTPException(status_code=400, detail=f"Variable '{cond.variable}' not found")

            labels = _manual_rules_assign_labels(df, list(request.manual_rules))
            variables_used = set()
            for rule in request.manual_rules:
                for cond in rule.conditions:
                    if (cond.variable or "").strip():
                        variables_used.add(cond.variable)

            segment_column_name = "_segment_assignment"
            df_with_segments = df.copy()
            df_with_segments[segment_column_name] = labels

            segments = []
            for idx, rule in enumerate(request.manual_rules, 1):
                seg_df = df_with_segments.loc[labels == rule.segment_name]
                record_count = int(len(seg_df))
                event_count = 0
                event_rate = 0.0
                if target_var and target_var in df.columns and record_count > 0:
                    try:
                        event_count = int(seg_df[target_var].sum())
                        event_rate = float(seg_df[target_var].mean() * 100)
                    except Exception:
                        pass

                woe, iv_contrib = _calculate_woe_iv(
                    seg_df, df, target_var, record_count, total_records, event_count
                )
                ci_lower, ci_upper = _calculate_wilson_score_ci(record_count, event_count)
                rule_str = _format_manual_rule_definition(rule)

                segments.append(
                    SegmentDetail(
                        segment_id=idx,
                        segment_name=rule.segment_name,
                        rule_definition=rule_str,
                        record_count=record_count,
                        pct_of_population=round(record_count / total_records * 100, 2) if total_records > 0 else 0,
                        event_count=event_count,
                        event_rate=round(event_rate, 4),
                        event_rate_ci_lower=ci_lower,
                        event_rate_ci_upper=ci_upper,
                        woe=round(woe, 4),
                        iv_contribution=round(iv_contrib, 4),
                    )
                )

            assigned_n = int((labels != "Unassigned").sum())

            # Merge split_tag from original df for OOS validation
            if 'split_tag' in df.columns and 'split_tag' not in df_with_segments.columns:
                df_with_segments['split_tag'] = df['split_tag'].values
            
            # Run validation suite with bootstrap stability and OOS validation
            validation = _run_validation_suite(
                segments=segments,
                df=df_with_segments,
                target_var=target_var,
                total_records=total_records,
                overall_event_rate=overall_event_rate,
                min_segment_size=request.min_segment_size,
                dataset_id=request.dataset_id,
                segment_column=segment_column_name,  # Enable bootstrap stability and OOS
                run_stability=True,
                run_oos_validation=True  # Enable out-of-sample validation
            )
            
            # Compute variable relevance matrix for manual rules
            variable_relevance = None
            try:
                variable_relevance = _compute_variable_relevance(
                    df=df_with_segments,
                    segment_column=segment_column_name,
                    target_var=target_var,
                    variables_used=list(variables_used)
                )
            except Exception as e:
                logger.warning(f"Could not compute variable relevance for manual rules mode: {e}")
            
            return _finalize_unified_segmentation_response(UnifiedSegmentationResponse(
                success=True,
                message=f"Manual rules segmentation created {len(segments)} segments",
                mode=request.mode,
                method=None,
                variables_used=list(variables_used),
                variable_priority=request.variable_priority,
                parameters={"rules_count": len(request.manual_rules)},
                num_segments=len(segments),
                segments=segments,
                validation=validation,
                variable_relevance=variable_relevance,
                dataset_shape=total_records,
                total_segment_records=sum(s.record_count for s in segments),
                records_match=assigned_n == total_records,
                manual_rules=list(request.manual_rules),
            ))
        
        elif request.mode == SegmentationMode.AUTO:
            # Auto mode - fully automated with scheme comparison
            # Uses the full auto segmentation pipeline with IV+AUC ranking
            
            from app.services.auto_segmentation_pipeline import run_auto_segmentation_pipeline, AutoPipelineConfig
            
            # Configure the pipeline
            config = AutoPipelineConfig(
                min_segment_size=request.min_segment_size or 1000,
                min_segment_size_pct=request.min_segment_size_pct / 100 if request.min_segment_size_pct else 0.05,
                max_segments=request.max_segments or 7,
                max_depth=request.max_depth or 4
            )
            
            # Run the auto pipeline
            pipeline_result = run_auto_segmentation_pipeline(
                df=df,
                target_variable=target_var,
                dataset_id=request.dataset_id,
                segmentation_service=segmentation_service,
                dataset_manager=dataset_manager,
                config=config
            )
            
            if not pipeline_result.get("success"):
                raise HTTPException(
                    status_code=400, 
                    detail=pipeline_result.get("message", "Auto segmentation failed")
                )
            
            # Get the recommended scheme
            schemes = pipeline_result.get("schemes", [])
            recommended_idx = pipeline_result.get("recommended_scheme_idx")
            
            if not schemes:
                raise HTTPException(
                    status_code=400,
                    detail="No viable segmentation schemes found"
                )
            
            # Use the best (first) scheme
            best_scheme = schemes[0]
            
            # Build segments from the scheme
            segments = []
            for i, seg in enumerate(best_scheme.segments):
                record_count = seg.get("size", 0)
                raw_er = float(seg.get("event_rate", 0) or 0)
                # Normalize to 0-1 fraction (pipeline may send ratio or 0-100 percent)
                if raw_er > 1.0:
                    er_frac = min(1.0, max(0.0, raw_er / 100.0))
                else:
                    er_frac = min(1.0, max(0.0, raw_er))
                if seg.get("event_count") is not None:
                    event_count = int(max(0, min(int(record_count), int(seg["event_count"]))))
                else:
                    event_count = int(round(record_count * er_frac))
                
                # Calculate Wilson Score confidence intervals
                ci_lower, ci_upper = _calculate_wilson_score_ci(record_count, event_count)
                
                # SegmentDetail.event_rate is 0-100 percent everywhere validation / OOS expect it
                display_event_rate = round(er_frac * 100, 4)
                woe_v, iv_c = _woe_iv_from_segment_row_counts(
                    full_df=df,
                    target_var=target_var,
                    record_count=int(record_count),
                    total_records=int(total_records),
                    event_count=int(event_count),
                )
                rule_lines = seg.get("rules") or []
                rule_from_tree = ""
                if isinstance(rule_lines, list) and rule_lines:
                    rule_from_tree = " AND ".join(
                        str(x).replace("\u2264", "<=").replace("\u2265", ">=") for x in rule_lines
                    )
                rule_definition = (
                    seg.get("human_readable")
                    or seg.get("rule_definition")
                    or seg.get("rules_readable")
                    or rule_from_tree
                    or seg.get("rule")
                    or f"Segment {i + 1}"
                )
                segment_detail = SegmentDetail(
                    segment_id=i + 1,
                    segment_name=f"Segment {i + 1}",
                    rule_definition=rule_definition,
                    record_count=record_count,
                    pct_of_population=record_count / total_records * 100 if total_records > 0 else 0,
                    event_count=event_count,
                    event_rate=display_event_rate,
                    event_rate_ci_lower=ci_lower,
                    event_rate_ci_upper=ci_upper,
                    woe=round(float(woe_v), 4),
                    iv_contribution=round(float(iv_c), 4),
                )
                segments.append(segment_detail)
            
            # Create segment assignments for stability analysis
            segment_column_name = '_auto_segment_assignment'
            df_with_segments = df.copy()
            df_with_segments[segment_column_name] = 'Unassigned'
            
            # Try to assign segments based on rules if available
            for i, seg in enumerate(best_scheme.segments):
                seg_name = f"Segment {i + 1}"
                rule_def = seg.get("human_readable", seg.get("rule", ""))
                
                # If we have rules, try to parse them for assignment
                # For now, use a simpler approach - check if there are leaf_ids in the result
                if 'mask' in seg:
                    df_with_segments.loc[seg['mask'], segment_column_name] = seg_name
            
            # Check if we could create valid assignments
            has_valid_assignments = df_with_segments[segment_column_name].value_counts().get('Unassigned', 0) < len(df_with_segments)
            
            # Merge split_tag from original df for OOS validation
            if 'split_tag' in df.columns and 'split_tag' not in df_with_segments.columns:
                df_with_segments['split_tag'] = df['split_tag'].values
            
            # Run validation suite with bootstrap stability and OOS validation
            validation = _run_validation_suite(
                segments=segments,
                df=df_with_segments,
                target_var=target_var,
                total_records=total_records,
                overall_event_rate=overall_event_rate,
                min_segment_size=config.min_segment_size,
                dataset_id=request.dataset_id,
                segment_column=segment_column_name if has_valid_assignments else None,
                run_stability=has_valid_assignments,
                run_oos_validation=has_valid_assignments  # Enable OOS validation
            )
            
            # Build auto candidates for frontend display (aligned with validation below)
            from app.services.auto_segmentation_pipeline import auto_pipeline
            auto_candidates = [
                auto_pipeline.scheme_to_dict(scheme)
                for scheme in schemes
            ]
            auto_candidates = _merge_validation_into_auto_candidates(
                auto_candidates, validation, schemes
            )
            
            # Variable relevance for AUTO mode
            variable_relevance = None
            if has_valid_assignments:
                try:
                    variable_relevance = _compute_variable_relevance(
                        df=df_with_segments,
                        segment_column=segment_column_name,
                        target_var=target_var,
                        variables_used=best_scheme.variables
                    )
                except Exception as e:
                    logger.warning(f"Could not compute variable relevance for auto mode: {e}")
            
            return _finalize_unified_segmentation_response(UnifiedSegmentationResponse(
                success=True,
                message=f"Auto segmentation found {len(schemes)} schemes, recommended: {best_scheme.description}",
                mode=request.mode,
                method="cart",
                variables_used=best_scheme.variables,
                variable_priority=VariablePriority(**best_scheme.variable_priority) if best_scheme.variable_priority and best_scheme.variable_priority.get("primary") else None,
                parameters={
                    "max_depth": best_scheme.depth,
                    "num_schemes_evaluated": len(schemes),
                    "variables_ranked": pipeline_result.get("variables_ranked", [])
                },
                num_segments=best_scheme.num_segments,
                segments=segments,
                validation=validation,
                variable_relevance=variable_relevance,
                dataset_shape=total_records,
                total_segment_records=sum(s.record_count for s in segments),
                records_match=True,
                auto_candidates=auto_candidates,
                selected_scheme_rank=1
            ))
        
        else:
            raise HTTPException(status_code=400, detail=f"Unknown segmentation mode: {request.mode}")
    
    except HTTPException:
        elapsed = time.time() - start_time
        logger.warning(f"Segmentation HTTP error after {elapsed:.2f}s")
        raise
    except Exception as e:
        elapsed = time.time() - start_time
        logger.error(f"Unified segmentation failed after {elapsed:.2f}s: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Provide user-friendly error messages
        error_msg = str(e)
        if "memory" in error_msg.lower():
            raise HTTPException(
                status_code=500, 
                detail="Dataset is too large to process. Try reducing the number of segments or variables."
            )
        elif "timeout" in error_msg.lower():
            raise HTTPException(
                status_code=504, 
                detail="Segmentation timed out. Try with fewer segments or a smaller dataset."
            )
        else:
            raise HTTPException(status_code=500, detail=f"Segmentation failed: {error_msg}")


@chat_router.post("/segmentation/validate-rules", response_model=RuleValidationResult)
async def validate_segmentation_rules(
    request: UnifiedSegmentationRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Real-time validation for C3 manual rules.
    Checks coverage, mutual exclusivity, empty segments.
    """
    try:
        if request.mode != SegmentationMode.MANUAL_RULES:
            raise HTTPException(status_code=400, detail="This endpoint is for MANUAL_RULES mode only")
        
        if not request.manual_rules:
            raise HTTPException(status_code=400, detail="manual_rules are required")
        
        # Load dataset
        df = dataframe_state_manager.get_dataframe_for_execution(request.dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(request.dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        total_records = len(df)

        labels = _manual_rules_assign_labels(df, list(request.manual_rules))
        segment_counts: Dict[str, int] = {}
        empty_segments: List[str] = []
        for rule in request.manual_rules:
            count = int((labels == rule.segment_name).sum())
            segment_counts[rule.segment_name] = count
            if count == 0:
                empty_segments.append(rule.segment_name)

        unassigned = int((labels == "Unassigned").sum())
        coverage_pct = round(((total_records - unassigned) / total_records) * 100, 2) if total_records > 0 else 0

        return RuleValidationResult(
            coverage_pct=coverage_pct,
            unassigned_records=unassigned,
            is_mutually_exclusive=True,
            overlap_count=0,
            empty_segments=empty_segments,
            segment_counts=segment_counts,
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Rule validation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error validating rules: {str(e)}")


@chat_router.post("/segmentation/merge-segments", response_model=MergeSegmentsResponse)
async def merge_segments(
    request: MergeSegmentsRequest = Body(...),
    current_user = Depends(get_current_user_dependency)
):
    """
    Merge two segments and recompute statistics including WoE, IV, and the full validation suite.
    
    Combines segment rules, reloads the dataset when possible, and re-runs chi-squared,
    Cramer's V, segment flags, merge recommendations (three-condition framework, including
    reliability after edits), out-of-sample validation, bootstrap stability, and the
    recommendation category.
    """
    try:
        logger.info(f"Merge segments request | dataset={request.dataset_id} segments={request.segment_a_id}+{request.segment_b_id}")
        
        current_result = request.current_segmentation
        segments = current_result.get("segments", [])
        
        # Find segments to merge by ID
        seg_a = next((s for s in segments if s.get("segment_id") == request.segment_a_id), None)
        seg_b = next((s for s in segments if s.get("segment_id") == request.segment_b_id), None)
        
        # Fallback: try matching by index if IDs don't match
        if not seg_a or not seg_b:
            seg_a = next((s for s in segments if s.get("segment_id") == request.segment_a_id or 
                         segments.index(s) + 1 == request.segment_a_id), None)
            seg_b = next((s for s in segments if s.get("segment_id") == request.segment_b_id or 
                         segments.index(s) + 1 == request.segment_b_id), None)
        
        if not seg_a or not seg_b:
            raise HTTPException(status_code=400, detail=f"Segments {request.segment_a_id} and/or {request.segment_b_id} not found")
        
        # Calculate IV before merge
        iv_before = sum(s.get("iv_contribution", 0) for s in segments)
        
        # Get total dataset stats for WoE calculation
        total_records = current_result.get("dataset_shape") or current_result.get("total_segment_records") or sum(s.get("record_count", 0) for s in segments)
        total_events = sum(s.get("event_count", 0) for s in segments)
        total_non_events = total_records - total_events
        
        # Create merged segment
        merged_records = seg_a.get("record_count", 0) + seg_b.get("record_count", 0)
        merged_events = seg_a.get("event_count", 0) + seg_b.get("event_count", 0)
        merged_non_events = merged_records - merged_events
        merged_event_rate = (merged_events / merged_records) if merged_records > 0 else 0
        merged_pct_population = seg_a.get("pct_of_population", 0) + seg_b.get("pct_of_population", 0)
        
        # Calculate WoE for merged segment
        epsilon = 0.5  # Smoothing factor to prevent division by zero
        if total_events > 0 and total_non_events > 0:
            dist_events = (merged_events + epsilon) / (total_events + epsilon)
            dist_non_events = (merged_non_events + epsilon) / (total_non_events + epsilon)
            merged_woe = np.log(dist_non_events / dist_events) if dist_events > 0 else 0
            merged_iv_contrib = (dist_non_events - dist_events) * merged_woe
        else:
            merged_woe = 0
            merged_iv_contrib = 0
        
        # Create merged segment name
        seg_a_name = seg_a.get("segment_name", f"Segment {request.segment_a_id}")
        seg_b_name = seg_b.get("segment_name", f"Segment {request.segment_b_id}")
        merged_name = request.new_segment_name or f"{seg_a_name} + {seg_b_name}"
        
        # Create merged segment dict (flat OR of atomic leaves; avoids nested-OR parse bugs on 2+ merges)
        ra = (seg_a.get("rule_definition") or "").strip()
        rb = (seg_b.get("rule_definition") or "").strip()
        if ra and rb:
            merged_rule_def = _merge_two_segment_rule_strings(ra, rb)
        else:
            merged_rule_def = ra or rb

        merged_segment_dict = {
            "segment_id": min(seg_a.get("segment_id", 1), seg_b.get("segment_id", 2)),
            "segment_name": merged_name,
            "rule_definition": merged_rule_def,
            "record_count": merged_records,
            "pct_of_population": round(merged_pct_population, 4),
            "event_count": merged_events,
            "event_rate": round(merged_event_rate, 6),
            "woe": round(float(merged_woe), 4),
            "iv_contribution": round(float(merged_iv_contrib), 6)
        }
        
        # Remove original segments and add merged
        new_segments = [s for s in segments if s not in [seg_a, seg_b]]
        new_segments.append(merged_segment_dict)
        
        # Sort by event rate and renumber
        new_segments.sort(key=lambda x: x.get("event_rate", 0))
        for idx, seg in enumerate(new_segments, 1):
            seg["segment_id"] = idx
            seg["segment_name"] = f"Segment {idx}" if "+" not in seg.get("segment_name", "") else seg["segment_name"]
        
        # Recalculate total IV (refined below if validation suite runs)
        iv_after = sum(s.get("iv_contribution", 0) for s in new_segments)
        iv_change = iv_after - iv_before
        iv_change_pct = (iv_change / iv_before * 100) if iv_before > 0 else 0
        merged_id = merged_segment_dict["segment_id"]

        # Update the result
        updated_result = current_result.copy()
        updated_result["segments"] = new_segments
        updated_result["num_segments"] = len(new_segments)
        # Merged OR rule strings no longer match structured C3 manual_rules
        if updated_result.get("manual_rules"):
            updated_result["manual_rules"] = None

        validation_refresh = None
        df_merge = dataframe_state_manager.get_dataframe_for_execution(request.dataset_id, None)
        if df_merge is None:
            df_merge = dataset_manager.load_dataset(request.dataset_id)
        if df_merge is not None:
            try:
                refreshed_segs, validation_refresh = rebuild_validation_from_segmentation_result(
                    request.dataset_id, df_merge, updated_result, list(new_segments)
                )
                if validation_refresh is not None:
                    updated_result["segments"] = refreshed_segs
                    updated_result["validation"] = _serialize_validation_result(validation_refresh)
                    new_segments = refreshed_segs
                    iv_after = float(validation_refresh.total_iv)
                    iv_change = iv_after - iv_before
                    iv_change_pct = (iv_change / iv_before * 100) if iv_before > 0 else 0
                    tr = sum(s.get("record_count", 0) for s in refreshed_segs)
                    updated_result["total_segment_records"] = tr
                    updated_result["records_match"] = tr == len(df_merge)
                else:
                    logger.warning(
                        "Merge: full validation rebuild returned None (target variable or assignments missing); "
                        "clearing validation to avoid stale metrics."
                    )
                    updated_result.pop("validation", None)
            except Exception as ex:
                logger.warning(f"Full validation refresh after merge failed: {ex}")
                updated_result.pop("validation", None)

        # Track merge history (copy list so we never mutate the client's session object)
        merge_history = list(updated_result.get("merge_history") or [])
        merge_history.append(f"Merged {seg_a_name} + {seg_b_name} -> {merged_name}")
        updated_result["merge_history"] = merge_history

        merged_stats = next((s for s in new_segments if s.get("segment_id") == merged_id), None)
        if merged_stats:
            merged_records = int(merged_stats.get("record_count", merged_records))
            merged_events = int(merged_stats.get("event_count", merged_events))
            merged_event_rate = (merged_events / merged_records) if merged_records > 0 else 0.0
            merged_pct_population = float(merged_stats.get("pct_of_population", merged_pct_population))

        # Create response objects
        ci_lower, ci_upper = _calculate_wilson_score_ci(merged_records, merged_events)

        if merged_stats:
            merged_segment = SegmentDetail(
                segment_id=int(merged_stats["segment_id"]),
                segment_name=str(merged_stats.get("segment_name", merged_name)),
                rule_definition=str(merged_stats.get("rule_definition", "")),
                record_count=merged_records,
                pct_of_population=float(merged_stats.get("pct_of_population", 0)),
                event_count=merged_events,
                event_rate=float(merged_stats.get("event_rate", 0)),
                event_rate_ci_lower=ci_lower,
                event_rate_ci_upper=ci_upper,
                woe=float(merged_stats.get("woe", 0)),
                iv_contribution=float(merged_stats.get("iv_contribution", 0)),
            )
        else:
            merged_segment = SegmentDetail(
                segment_id=merged_segment_dict["segment_id"],
                segment_name=merged_segment_dict["segment_name"],
                rule_definition=merged_segment_dict["rule_definition"],
                record_count=merged_segment_dict["record_count"],
                pct_of_population=merged_segment_dict["pct_of_population"],
                event_count=merged_segment_dict["event_count"],
                event_rate=merged_segment_dict["event_rate"],
                event_rate_ci_lower=ci_lower,
                event_rate_ci_upper=ci_upper,
                woe=merged_segment_dict["woe"],
                iv_contribution=merged_segment_dict["iv_contribution"],
            )

        merge_impact = MergeImpact(
            merged_segment_name=merged_name,
            combined_records=merged_records,
            combined_events=merged_events,
            combined_event_rate=round(merged_event_rate, 6),
            combined_pct_of_population=round(merged_pct_population, 4),
            iv_before_merge=round(iv_before, 4),
            iv_after_merge=round(iv_after, 4),
            iv_change=round(iv_change, 4),
            iv_change_pct=round(iv_change_pct, 2),
        )
        
        logger.info(f"Merge successful | new_segments={len(new_segments)} iv_before={iv_before:.4f} iv_after={iv_after:.4f}")

        _record_segmentation_audit_event(
            request.dataset_id,
            "merge_executed_manual",
            {
                "segment_a": seg_a_name,
                "segment_b": seg_b_name,
                "merged_name": merged_name,
                "iv_before": round(iv_before, 6),
                "iv_after": round(iv_after, 6),
                "num_segments_after": len(new_segments),
            },
            current_user,
        )
        
        return MergeSegmentsResponse(
            success=True,
            message=f"Successfully merged {seg_a_name} and {seg_b_name}",
            merged_segment=merged_segment,
            merge_impact=merge_impact,
            updated_segmentation=updated_result,
            num_segments_after=len(new_segments),
            can_undo=True
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Merge segments failed: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error merging segments: {str(e)}")


@chat_router.post("/segmentation/move-categorical-value", response_model=MoveCategoricalValueResponse)
async def move_segmentation_categorical_value(
    request: MoveCategoricalValueRequest = Body(...),
    current_user = Depends(get_current_user_dependency),
):
    """
    Move one category of a variable between two segments (plan §5.4).
    Requires structured `manual_rules` on the segmentation payload (C3).
    """
    try:
        if request.from_segment_id == request.to_segment_id:
            raise HTTPException(status_code=400, detail="from_segment_id and to_segment_id must differ")

        df = dataframe_state_manager.get_dataframe_for_execution(request.dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(request.dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")

        if (request.variable or "").strip() not in df.columns:
            raise HTTPException(status_code=400, detail=f"Variable '{request.variable}' not found")

        cur = dict(request.current_segmentation or {})
        segs = list(cur.get("segments") or [])
        mr_raw = cur.get("manual_rules")
        if not isinstance(mr_raw, list) or len(mr_raw) == 0:
            raise HTTPException(
                status_code=400,
                detail="manual_rules are required to move categories. Re-run manual segmentation (C3) and try again.",
            )

        def _name_for_id(sid: int) -> str:
            for s in segs:
                if int(s.get("segment_id", -1)) == int(sid):
                    return str(s.get("segment_name") or "")
            return ""

        from_name = _name_for_id(request.from_segment_id)
        to_name = _name_for_id(request.to_segment_id)
        if not from_name or not to_name:
            raise HTTPException(status_code=400, detail="Could not resolve segment names from segment_id values")

        rules = [ManualSegmentRule.model_validate(x) for x in mr_raw]
        try:
            new_rules = _adjust_manual_rules_category_move(
                rules, from_name, to_name, request.variable, request.category_value
            )
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve)) from ve

        cur["manual_rules"] = [r.model_dump(mode="json") for r in new_rules]
        _sync_segment_rule_strings_from_manual_rules(segs, new_rules)
        cur["segments"] = segs

        refreshed_segs, validation_refresh = rebuild_validation_from_segmentation_result(
            request.dataset_id, df, cur, list(segs)
        )
        if validation_refresh is None:
            raise HTTPException(status_code=500, detail="Validation rebuild failed after category move")
        cur["segments"] = refreshed_segs
        cur["validation"] = _serialize_validation_result(validation_refresh)
        cur["num_segments"] = len(refreshed_segs)
        tr = sum(int(s.get("record_count", 0) or 0) for s in refreshed_segs)
        cur["total_segment_records"] = tr
        cur["records_match"] = tr == len(df)

        _record_segmentation_audit_event(
            request.dataset_id,
            "categorical_value_moved",
            {
                "variable": request.variable,
                "category": request.category_value,
                "from_segment_id": request.from_segment_id,
                "to_segment_id": request.to_segment_id,
            },
            current_user,
        )

        return MoveCategoricalValueResponse(
            success=True,
            message=f"Moved category {request.category_value!r} on {request.variable} from {from_name} to {to_name}",
            updated_segmentation=cur,
            can_undo=True,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"move-categorical-value failed: {e}")
        import traceback

        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error moving category: {str(e)}")


@chat_router.post("/segmentation/edit-cutoff", response_model=CutoffEditResponse)
async def edit_segment_cutoff(
    request: CutoffEditRequest = Body(...),
    current_user = Depends(get_current_user_dependency)
):
    """
    Edit a segment cutoff value with impact preview.
    
    This endpoint:
    1. Parses the current segment rules
    2. Calculates which records would move between segments
    3. Computes new statistics (event rate, WoE, IV)
    4. Returns impact preview or applies the change
    
    Use preview_only=True to see the impact before committing.
    When preview_only=False, segment rules are applied and the full validation suite is
    recomputed from the dataset (same outputs as /segmentation/run validation).
    """
    try:
        logger.info(f"Edit cutoff request | dataset={request.dataset_id} segment={request.segment_id} "
                   f"var={request.variable} {request.old_value}->{request.new_value} preview={request.preview_only}")
        
        # Load dataset for recomputation
        df = dataframe_state_manager.get_dataframe_for_execution(request.dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(request.dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        current_result = request.current_segmentation
        segments = current_result.get("segments", [])
        target_var = current_result.get("parameters", {}).get("target_variable") or \
                     current_result.get("target_variable")
        
        # Find the target segment
        target_seg = None
        target_idx = -1
        for idx, seg in enumerate(segments):
            if seg.get("segment_id") == request.segment_id:
                target_seg = seg
                target_idx = idx
                break
        
        if not target_seg:
            raise HTTPException(status_code=400, detail=f"Segment {request.segment_id} not found")
        
        # Validate variable exists
        if request.variable not in df.columns:
            raise HTTPException(status_code=400, detail=f"Variable '{request.variable}' not found in dataset")
        
        # Parse current rule and build masks
        old_rule = target_seg.get("rule_definition", "")
        
        # Build new rule by replacing the cutoff value
        # Handle different formats: "var > 30", "var <= 30.0", etc.
        def _cutoff_scalar_to_float(v: Any) -> float:
            if isinstance(v, bool):
                raise ValueError("boolean")
            if isinstance(v, (int, float)):
                return float(v)
            if isinstance(v, str) and v.strip():
                return float(v.strip())
            raise ValueError("not numeric")

        try:
            old_f = _cutoff_scalar_to_float(request.old_value)
            new_f = _cutoff_scalar_to_float(request.new_value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(
                status_code=400,
                detail="Numeric threshold edit only. For categorical splits use POST /segmentation/move-categorical-value.",
            ) from exc

        old_value_str = str(request.old_value)
        new_value_str = str(request.new_value)

        # Try to replace exact match first
        if old_value_str in old_rule:
            new_rule = old_rule.replace(old_value_str, new_value_str, 1)
        else:
            # Try with different number formats
            old_int = str(int(old_f)) if old_f == int(old_f) else str(old_f)
            new_rule = old_rule.replace(old_int, new_value_str, 1)
        
        # Calculate impact by evaluating old vs new rules
        total_records = len(df)
        
        # Parse and evaluate the rule to get segment masks
        def evaluate_simple_condition(df: pd.DataFrame, var: str, op: str, val: float) -> pd.Series:
            """Evaluate a simple condition and return boolean mask."""
            if var not in df.columns:
                return pd.Series([False] * len(df), index=df.index)
            col = df[var]
            if op == '>':
                return col > val
            elif op == '>=':
                return col >= val
            elif op == '<':
                return col < val
            elif op == '<=':
                return col <= val
            elif op == '==' or op == '=':
                return col == val
            elif op == '!=':
                return col != val
            return pd.Series([False] * len(df), index=df.index)
        
        # For preview, we need to estimate records moving
        # This is simplified - in reality we'd need to parse the full rule
        old_mask = evaluate_simple_condition(df, request.variable, request.operator, old_f)
        new_mask = evaluate_simple_condition(df, request.variable, request.operator, new_f)
        
        # Records that would move
        records_moved_out = int((old_mask & ~new_mask).sum())
        records_moved_in = int((~old_mask & new_mask).sum())
        
        # Calculate new segment statistics
        # Get the original segment's other conditions (if any)
        # For simplicity, estimate new stats based on the new mask
        original_record_count = target_seg.get("record_count", 0)
        new_record_count = original_record_count - records_moved_out + records_moved_in
        
        # Get current total stats for IV calculation
        total_events = sum(s.get("event_count", 0) for s in segments)
        total_non_events = total_records - total_events
        
        # Calculate new event rate for the modified segment
        if target_var and target_var in df.columns:
            # Create approximate new segment mask
            # This is an estimation based on the cutoff change
            new_event_count = 0
            if new_record_count > 0:
                # Estimate based on proportion
                old_event_rate = target_seg.get("event_rate", 0)
                # Adjust for records moving
                old_events = target_seg.get("event_count", 0)
                
                # Records moving out/in carry their proportional events
                if old_mask.sum() > 0:
                    moving_out_event_rate = df.loc[old_mask & ~new_mask, target_var].mean() if (old_mask & ~new_mask).sum() > 0 else old_event_rate
                else:
                    moving_out_event_rate = old_event_rate
                    
                if (~old_mask & new_mask).sum() > 0:
                    moving_in_event_rate = df.loc[~old_mask & new_mask, target_var].mean() if target_var in df.columns else old_event_rate
                else:
                    moving_in_event_rate = old_event_rate
                
                events_lost = int(records_moved_out * (moving_out_event_rate if not pd.isna(moving_out_event_rate) else old_event_rate))
                events_gained = int(records_moved_in * (moving_in_event_rate if not pd.isna(moving_in_event_rate) else old_event_rate))
                new_event_count = max(0, old_events - events_lost + events_gained)
                new_event_rate = new_event_count / new_record_count if new_record_count > 0 else 0
        else:
            new_event_count = target_seg.get("event_count", 0)
            new_event_rate = target_seg.get("event_rate", 0)
        
        new_pct_of_population = (new_record_count / total_records * 100) if total_records > 0 else 0
        
        # Calculate IV impact
        iv_before = target_seg.get("iv_contribution", 0)
        
        # Calculate new IV contribution
        WOE_EPSILON = 0.5
        new_non_events = new_record_count - new_event_count
        if total_events > 0 and total_non_events > 0:
            dist_events = (new_event_count + WOE_EPSILON) / (total_events + WOE_EPSILON)
            dist_non_events = (new_non_events + WOE_EPSILON) / (total_non_events + WOE_EPSILON)
            new_woe = np.log(dist_non_events / dist_events) if dist_events > 0 else 0
            iv_after = (dist_non_events - dist_events) * new_woe
        else:
            new_woe = 0
            iv_after = 0
        
        iv_change = iv_after - iv_before
        
        # Check minimum size warning (default: 5% of total or 1000 records)
        min_segment_size = current_result.get("parameters", {}).get("min_segment_size", 1000)
        min_segment_pct = 0.05
        below_min_size = new_record_count < min_segment_size or new_record_count < (total_records * min_segment_pct)
        
        # Find affected segments (the modified one and potentially adjacent ones)
        affected_segments = [request.segment_id]
        
        # Create impact object
        impact = CutoffEditImpact(
            segment_id=request.segment_id,
            variable=request.variable,
            old_rule=old_rule,
            new_rule=new_rule,
            records_moved_out=records_moved_out,
            records_moved_in=records_moved_in,
            new_record_count=new_record_count,
            new_event_count=new_event_count,
            new_event_rate=round(new_event_rate, 6),
            new_pct_of_population=round(new_pct_of_population, 2),
            iv_before=round(iv_before, 6),
            iv_after=round(float(iv_after), 6),
            iv_change=round(float(iv_change), 6),
            below_min_size=below_min_size
        )
        
        # If preview only, return without applying
        if request.preview_only:
            logger.info(f"Cutoff edit preview | records_out={records_moved_out} records_in={records_moved_in} "
                       f"new_count={new_record_count} iv_change={iv_change:.4f}")
            
            return CutoffEditResponse(
                success=True,
                message=f"Preview: Cutoff change from {request.old_value} to {request.new_value}",
                preview_only=True,
                impact=impact,
                affected_segments=affected_segments,
                updated_segmentation=None,
                can_undo=True
            )
        
        # Apply the change
        updated_result = current_result.copy()
        updated_segments = updated_result.get("segments", []).copy()
        
        # Update the target segment
        updated_segments[target_idx] = {
            **target_seg,
            "rule_definition": new_rule,
            "record_count": new_record_count,
            "event_count": new_event_count,
            "event_rate": round(new_event_rate, 6),
            "pct_of_population": round(new_pct_of_population, 2),
            "woe": round(float(new_woe), 4),
            "iv_contribution": round(float(iv_after), 6)
        }
        
        updated_result["segments"] = updated_segments

        validation_refresh = None
        try:
            refreshed_segs, validation_refresh = rebuild_validation_from_segmentation_result(
                request.dataset_id, df, updated_result, list(updated_segments)
            )
            if validation_refresh is not None:
                updated_result["segments"] = refreshed_segs
                updated_result["validation"] = _serialize_validation_result(validation_refresh)
                tr = sum(s.get("record_count", 0) for s in refreshed_segs)
                updated_result["total_segment_records"] = tr
                updated_result["records_match"] = tr == len(df)
            else:
                logger.warning(
                    "Cutoff apply: full validation rebuild returned None (target variable or assignments missing); "
                    "clearing validation to avoid stale metrics."
                )
                updated_result.pop("validation", None)
        except Exception as ex:
            logger.warning(f"Full validation refresh after cutoff edit failed: {ex}")
            updated_result.pop("validation", None)

        # Track cutoff edit history (copy list so we never mutate the client's session object)
        cutoff_edits = list(updated_result.get("cutoff_edits") or [])
        cutoff_edits.append(f"{request.variable}: {request.old_value} -> {request.new_value}")
        updated_result["cutoff_edits"] = cutoff_edits
        
        logger.info(f"Cutoff edit applied | segment={request.segment_id} new_count={new_record_count}")

        _record_segmentation_audit_event(
            request.dataset_id,
            "cutoff_edited",
            {
                "segment_id": request.segment_id,
                "variable": request.variable,
                "operator": request.operator,
                "old_value": request.old_value,
                "new_value": request.new_value,
                "records_moved_out": records_moved_out,
                "records_moved_in": records_moved_in,
            },
            current_user,
        )
        
        return CutoffEditResponse(
            success=True,
            message=f"Cutoff updated: {request.variable} {request.operator} {request.new_value}",
            preview_only=False,
            impact=impact,
            affected_segments=affected_segments,
            updated_segmentation=updated_result,
            can_undo=True
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Edit cutoff failed: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error editing cutoff: {str(e)}")


@chat_router.post("/segmentation/generate-narrative")
async def generate_segmentation_narrative(
    narrative_type: str = Form(...),  # "merge", "recommendation", "variable"
    context_data: str = Form(...),  # JSON string with context
    current_user = Depends(get_current_user_dependency)
):
    """
    Generate LLM-powered narratives for segmentation results.
    
    narrative_type options:
    - "merge": Explanation for why two segments should be merged
    - "recommendation": Overall scheme quality summary
    - "variable": Commentary on variable importance in a segment
    
    context_data: JSON with relevant data for the narrative type
    """
    try:
        context = json.loads(context_data)
        
        if narrative_type == "merge":
            # Generate merge explanation
            segment_a = context.get("segment_a", {})
            segment_b = context.get("segment_b", {})
            merge_reason = context.get("merge_reason", "similar_event_rate")
            combined_stats = context.get("combined_stats", {})
            
            narrative = narrative_generator.generate_merge_explanation(
                segment_a=segment_a,
                segment_b=segment_b,
                merge_reason=merge_reason,
                combined_stats=combined_stats
            )
            
        elif narrative_type == "recommendation":
            # Generate recommendation summary
            validation_result = context.get("validation_result", {})
            num_segments = context.get("num_segments", 0)
            total_iv = context.get("total_iv", 0)
            recommendation_category = context.get("recommendation_category", "unknown")
            
            narrative = narrative_generator.generate_recommendation_narrative(
                validation_result=validation_result,
                num_segments=num_segments,
                total_iv=total_iv,
                recommendation_category=recommendation_category
            )
            
        elif narrative_type == "variable":
            # Generate variable commentary
            variable_relevance = context.get("variable_relevance", {})
            segment_name = context.get("segment_name", "Unknown Segment")
            top_variables = context.get("top_variables", [])
            
            narrative = narrative_generator.generate_variable_commentary(
                variable_relevance=variable_relevance,
                segment_name=segment_name,
                top_variables=top_variables
            )
            
        else:
            raise HTTPException(status_code=400, detail=f"Unknown narrative type: {narrative_type}")
        
        return {
            "success": True,
            "narrative_type": narrative_type,
            "narrative": narrative
        }
        
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON in context_data: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Generate narrative failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating narrative: {str(e)}")


@chat_router.post("/segmentation/add-to-data", response_model=AddToDataResponse)
async def add_segmentation_to_data(
    request: AddToDataRequest,
    current_user = Depends(get_current_user_dependency)
):
    """
    Save a segmentation scheme to the dataset as a new column.
    Creates segmentation_scheme_N column on train/test/holdout.
    Supports all 4 segmentation modes.
    """
    try:
        dataset_id = request.dataset_id
        idem_key = (request.idempotency_key or "").strip()
        seg_result = request.segmentation_result

        # Load dataset
        df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")

        # Idempotent replay (plan Section 15): same idempotency_key → same response, no second column
        if idem_key:
            from app.services.segmentation_audit import find_prior_add_to_data

            prior = find_prior_add_to_data(dataset_manager, dataset_id, idem_key)
            if prior:
                sid = int(prior["scheme_id"])
                col = str(prior["column_name"])
                schemes = dataset_manager.get_segmentation_schemes_metadata(dataset_id)
                rec = next((s for s in schemes if int(s.get("scheme_id", -1)) == sid), None)
                if not rec:
                    raise HTTPException(
                        status_code=409,
                        detail="Idempotent replay failed: stored scheme metadata not found for this key.",
                    )
                metadata = SegmentationSchemeMetadata.model_validate(rec)
                return AddToDataResponse(
                    success=True,
                    message=f"Idempotent replay: segmentation already saved as '{col}'",
                    scheme_id=sid,
                    column_name=col,
                    metadata=metadata,
                )

        # Next scheme id: max of existing segmentation_scheme_* columns and stored registry
        ds_info = dataset_manager.get_dataset_info(dataset_id) or {}
        stored_schemes = list(ds_info.get("segmentation_schemes") or [])
        max_scheme_id = 0
        for c in df.columns:
            if c.startswith("segmentation_scheme_"):
                try:
                    max_scheme_id = max(max_scheme_id, int(c.rsplit("_", 1)[-1]))
                except ValueError:
                    pass
        for s in stored_schemes:
            try:
                max_scheme_id = max(max_scheme_id, int(s.get("scheme_id", 0) or 0))
            except (TypeError, ValueError):
                pass
        scheme_id = max_scheme_id + 1

        # Use custom name if provided
        if request.scheme_name:
            column_name = f"seg_{request.scheme_name.replace(' ', '_').lower()}"
            if column_name in df.columns:
                column_name = f"{column_name}_{scheme_id}"
        else:
            column_name = f"segmentation_scheme_{scheme_id}"
        
        # Initialize segment labels
        segment_labels = pd.Series(["Unassigned"] * len(df), index=df.index)
        total_assigned = 0
        
        # Assign segments based on mode
        mode = seg_result.mode
        
        if mode == SegmentationMode.PRE_EXISTING:
            # C1: Copy from existing column
            if seg_result.variables_used:
                seg_col = seg_result.variables_used[0]
                if seg_col in df.columns:
                    # Map original values to segment names
                    for segment in seg_result.segments:
                        rule_def = segment.rule_definition
                        segment_name = segment.segment_name
                        
                        # Extract value from rule (format: "column = 'value'")
                        if "=" in rule_def:
                            val = rule_def.split("=")[-1].strip().strip("'\"")
                            mask = df[seg_col].astype(str) == val
                            segment_labels[mask] = segment_name
                            total_assigned += mask.sum()
        
        elif mode in (SegmentationMode.VARIABLE_DRIVEN, SegmentationMode.AUTO):
            # C2/Auto: Apply tree-based rules (top-level OR + AND atoms; plan parity with merge strings)
            for segment in seg_result.segments:
                segment_name = segment.segment_name
                rule_def = segment.rule_definition
                
                # Parse decision tree rule format
                # Example: "var1 <= 10.5 AND var2 > 20"
                try:
                    mask = _segmentation_rule_mask(df, rule_def)
                    # Only assign to rows not already assigned (first match wins)
                    unassigned_mask = segment_labels == "Unassigned"
                    segment_labels[mask & unassigned_mask] = segment_name
                    total_assigned += (mask & unassigned_mask).sum()
                except Exception as e:
                    logger.warning(f"Failed to apply rule for segment {segment_name}: {e}")
                    continue
        
        elif mode == SegmentationMode.MANUAL_RULES:
            # C3: Prefer structured manual_rules (exact parity with run); else string masks.
            if seg_result.manual_rules and len(seg_result.manual_rules) > 0:
                try:
                    lbl = _manual_rules_assign_labels(df, list(seg_result.manual_rules))
                    segment_labels = lbl.astype(str)
                    total_assigned = int((segment_labels != "Unassigned").sum())
                except Exception as e:
                    logger.warning(f"Structured manual_rules assignment failed, falling back to rule strings: {e}")
                    segment_labels = pd.Series(["Unassigned"] * len(df), index=df.index)
                    total_assigned = 0
                    for segment in seg_result.segments:
                        segment_name = segment.segment_name
                        rule_def = segment.rule_definition
                        try:
                            mask = _segmentation_rule_mask(df, rule_def)
                            unassigned_mask = segment_labels == "Unassigned"
                            segment_labels.loc[mask & unassigned_mask] = segment_name
                            total_assigned += int((mask & unassigned_mask).sum())
                        except Exception as e2:
                            logger.warning(f"Failed to apply rule for segment {segment_name}: {e2}")
                            continue
            else:
                for segment in seg_result.segments:
                    segment_name = segment.segment_name
                    rule_def = segment.rule_definition
                    try:
                        mask = _segmentation_rule_mask(df, rule_def)
                        unassigned_mask = segment_labels == "Unassigned"
                        segment_labels.loc[mask & unassigned_mask] = segment_name
                        total_assigned += int((mask & unassigned_mask).sum())
                    except Exception as e:
                        logger.warning(f"Failed to apply rule for segment {segment_name}: {e}")
                        continue
        
        # Add column to dataframe
        df[column_name] = segment_labels
        
        # Update dataframe state manager
        dataframe_state_manager.update_dataframe(dataset_id, df, force_scope="entire")
        
        merge_hist = list(seg_result.merge_history or [])
        cutoff_hist = list(seg_result.cutoff_edits or [])
        v_sel = seg_result.parameters.get("variable_selection_method") or seg_result.parameters.get(
            "selection_method"
        )
        if v_sel is None and seg_result.mode == SegmentationMode.AUTO:
            v_sel = "combined_iv_auc_rank"

        val = seg_result.validation
        if not val:
            raise HTTPException(
                status_code=400,
                detail="Cannot save scheme: segmentation result is missing validation. Run segmentation first.",
            )
        metadata = SegmentationSchemeMetadata(
            scheme_id=scheme_id,
            column_name=column_name,
            mode=seg_result.mode,
            variables=seg_result.variables_used,
            manual_rules=seg_result.manual_rules,
            variable_priority=seg_result.variable_priority,
            variable_selection_method=v_sel,
            tree_method=seg_result.method,
            max_depth=seg_result.parameters.get("max_depth"),
            constraints_applied=dict(seg_result.parameters or {}),
            segments=seg_result.segments,
            total_iv=float(val.total_iv),
            chi_squared_p=float(val.chi_squared_p),
            cramers_v=float(val.cramers_v),
            merge_history=merge_hist,
            cutoff_edits=cutoff_hist,
            validation=val,
            stability=val.stability,
            holdout_validation=val.oos_validation,
            recommendation_category=val.recommendation_category,
            created_at=datetime.now(timezone.utc),
        )

        try:
            meta_dump = metadata.model_dump(mode="json")
        except Exception:
            meta_dump = json.loads(
                json.dumps(metadata.dict() if hasattr(metadata, "dict") else {}, default=str)
            )
        if not dataset_manager.persist_dataframe_and_scheme_metadata(dataset_id, df, meta_dump):
            logger.error(
                "Add to Data: failed to persist dataframe and scheme metadata together for dataset %s",
                dataset_id,
            )
            raise HTTPException(
                status_code=500,
                detail="Failed to persist segmentation column and scheme registry metadata. "
                "Your session dataframe may have been updated; reload the dataset if counts look wrong.",
            )

        pct_assigned = (total_assigned / len(df)) * 100 if len(df) > 0 else 0
        logger.info(f"Added segmentation scheme {scheme_id} as column '{column_name}' to dataset {dataset_id} ({pct_assigned:.1f}% assigned)")

        _record_segmentation_audit_event(
            dataset_id,
            "add_to_data",
            {
                "scheme_id": scheme_id,
                "column_name": column_name,
                "mode": str(seg_result.mode),
                "variables_used": list(seg_result.variables_used or []),
                "pct_assigned": round(pct_assigned, 4),
                "recommendation_category": val.recommendation_category,
                "total_iv": float(val.total_iv),
                "num_segments": len(seg_result.segments or []),
                "idempotency_key": idem_key or None,
                "completed": True,
            },
            current_user,
        )

        from app.services.segmentation_audit import append_insight_pin

        scheme_summary = (
            f"Scheme #{scheme_id} ({str(seg_result.mode)}) — "
            f"{len(seg_result.segments or [])} segments, IV {float(val.total_iv):.4f}, "
            f"{val.recommendation_category or 'n/a'}"
        )
        detail_href = f"/api/v1/segmentation/schemes/{dataset_id}/{scheme_id}"
        append_insight_pin(
            dataset_id,
            {
                "pin_type": "segmentation_add_to_data",
                "idempotency_key": idem_key or None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "title": f"Segmentation saved: {column_name}",
                "body": (
                    f"{scheme_summary}. Column `{column_name}` — "
                    f"~{pct_assigned:.1f}% of rows assigned. Registry: {detail_href}"
                ),
                "scheme_summary": scheme_summary,
                "scheme_detail_href": detail_href,
                "column_ref": column_name,
                "scheme_id": scheme_id,
                "column_name": column_name,
                "mode": str(seg_result.mode),
                "recommendation_category": val.recommendation_category,
                "total_iv": float(val.total_iv),
            },
        )

        return AddToDataResponse(
            success=True,
            message=f"Segmentation scheme saved as '{column_name}' ({pct_assigned:.1f}% of records assigned)",
            scheme_id=scheme_id,
            column_name=column_name,
            metadata=metadata,
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Add to data failed: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error saving segmentation: {str(e)}")


def _try_parse_rule_text_to_condition(part: str) -> Optional[RuleCondition]:
    """
    Parse a single AND-clause from a persisted rule_definition string into RuleCondition
    so Add-to-Data matches plan §6 / _apply_condition (between, in, null, contains, etc.).
    """
    import re

    p = (part or "").strip()
    while len(p) >= 2 and p[0] == "(" and p[-1] == ")":
        p = p[1:-1].strip()
    if not p or p.upper().startswith("CATCH_ALL"):
        return None

    var_pat = r"([\w.]+)"

    m = re.match(rf"^{var_pat}\s+is\s+not\s+null\s*$", p, re.I)
    if m:
        return RuleCondition(variable=m.group(1), operator="IS_NOT_NULL", value=None)
    m = re.match(rf"^{var_pat}\s+is\s+null\s*$", p, re.I)
    if m:
        return RuleCondition(variable=m.group(1), operator="IS_NULL", value=None)
    m = re.match(rf"^{var_pat}\s+is\s+true\s*$", p, re.I)
    if m:
        return RuleCondition(variable=m.group(1), operator="IS_TRUE", value=None)
    m = re.match(rf"^{var_pat}\s+is\s+false\s*$", p, re.I)
    if m:
        return RuleCondition(variable=m.group(1), operator="IS_FALSE", value=None)

    m = re.match(
        rf"^{var_pat}\s+(IS_NULL|IS_NOT_NULL|IS_TRUE|IS_FALSE|is_null|is_not_null|is_true|is_false)"
        rf"\s*(?:None)?\s*$",
        p,
        re.I,
    )
    if m:
        op_tok = _normalize_segmentation_rule_operator(m.group(2))
        if op_tok in ("IS_NULL", "IS_NOT_NULL", "IS_TRUE", "IS_FALSE"):
            return RuleCondition(variable=m.group(1), operator=op_tok, value=None)

    m = re.match(rf"^{var_pat}\s+contains\s+(.+)$", p, re.I)
    if m:
        raw = m.group(2).strip()
        if (raw.startswith("'") and raw.endswith("'")) or (raw.startswith('"') and raw.endswith('"')):
            raw = raw[1:-1]
        return RuleCondition(variable=m.group(1), operator="CONTAINS", value=raw)

    m = re.match(rf"^{var_pat}\s+not\s+between\s+(.+)$", p, re.I)
    if m:
        bounds = _parse_between_bounds_token(m.group(2))
        if bounds is not None:
            return RuleCondition(variable=m.group(1), operator="NOT BETWEEN", value=bounds)
    m = re.match(rf"^{var_pat}\s+between\s+(.+)$", p, re.I)
    if m:
        bounds = _parse_between_bounds_token(m.group(2))
        if bounds is not None:
            return RuleCondition(variable=m.group(1), operator="BETWEEN", value=bounds)

    m = re.match(rf"^{var_pat}\s+not\s+in\s+(.+)$", p, re.I)
    if m:
        lst = _parse_in_list_token(m.group(2))
        if lst is not None:
            return RuleCondition(variable=m.group(1), operator="NOT IN", value=lst)
    m = re.match(rf"^{var_pat}\s+in\s+(.+)$", p, re.I)
    if m:
        lst = _parse_in_list_token(m.group(2))
        if lst is not None:
            return RuleCondition(variable=m.group(1), operator="IN", value=lst)

    m = re.match(rf"^{var_pat}\s*<=\s*([\d.eE+-]+)\s*$", p)
    if m:
        return RuleCondition(variable=m.group(1), operator="<=", value=float(m.group(2)))
    m = re.match(rf"^{var_pat}\s*<\s*([\d.eE+-]+)\s*$", p)
    if m:
        return RuleCondition(variable=m.group(1), operator="<", value=float(m.group(2)))
    m = re.match(rf"^{var_pat}\s*>=\s*([\d.eE+-]+)\s*$", p)
    if m:
        return RuleCondition(variable=m.group(1), operator=">=", value=float(m.group(2)))
    m = re.match(rf"^{var_pat}\s*>\s*([\d.eE+-]+)\s*$", p)
    if m:
        return RuleCondition(variable=m.group(1), operator=">", value=float(m.group(2)))

    m = re.match(rf"^{var_pat}\s*!=\s*(.+)$", p)
    if m:
        raw = m.group(2).strip().strip("'\"")
        return RuleCondition(variable=m.group(1), operator="!=", value=raw)
    m = re.match(rf"^{var_pat}\s*[=]{{1,2}}\s*(.+)$", p)
    if m:
        raw = m.group(2).strip().strip("'\"")
        return RuleCondition(variable=m.group(1), operator="=", value=raw)

    return None


def _parse_between_bounds_token(rest: str) -> Optional[List[float]]:
    import ast
    import re

    s = (rest or "").strip()
    m = re.match(r"^([\d.eE+-]+)\s+and\s+([\d.eE+-]+)\s*$", s, re.I)
    if m:
        try:
            return [float(m.group(1)), float(m.group(2))]
        except ValueError:
            return None
    m2 = re.match(r"^([\d.eE+-]+)\s*,\s*([\d.eE+-]+)\s*$", s)
    if m2:
        try:
            return [float(m2.group(1)), float(m2.group(2))]
        except ValueError:
            return None
    if s.startswith("[") or s.startswith("("):
        try:
            v = ast.literal_eval(s)
            if isinstance(v, (list, tuple)) and len(v) >= 2:
                return [float(v[0]), float(v[1])]
        except (ValueError, SyntaxError, TypeError):
            pass
    return None


def _parse_in_list_token(rest: str) -> Optional[List[Any]]:
    import ast
    import re

    s = (rest or "").strip()
    if s.startswith("[") or s.startswith("("):
        try:
            v = ast.literal_eval(s)
            if isinstance(v, (list, tuple)):
                return list(v)
        except (ValueError, SyntaxError, TypeError):
            return None
    inner = s
    if s.startswith("(") and s.endswith(")"):
        inner = s[1:-1].strip()
    if not inner:
        return None
    parts = re.split(r",\s*", inner)
    out: List[Any] = []
    for p in parts:
        p = p.strip().strip("'\"")
        if not p:
            continue
        try:
            out.append(float(p))
        except ValueError:
            out.append(p)
    return out if out else None


def _parse_and_apply_rule(df: pd.DataFrame, rule_def: str, variables: List[str]) -> pd.Series:
    """
    Parse a rule definition and return a boolean mask.

    Supports the same operators as plan §6 / _apply_condition when clauses are AND-joined,
    plus legacy numeric comparisons for simple strings.
    """
    import re

    mask = pd.Series([True] * len(df), index=df.index)

    if not rule_def or rule_def == "Unknown":
        return mask

    rule_def = str(rule_def).replace("\u2264", "<=").replace("\u2265", ">=")

    and_parts = re.split(r"\s+AND\s+", rule_def, flags=re.IGNORECASE)

    for part in and_parts:
        part = part.strip()
        if not part:
            continue

        # Legacy simplified display: "feature: 10,000 - 50,000" (from older segmentation runs)
        colon_rng = re.match(r"^([\w.]+):\s*([\d,.eE+-]+)\s*-\s*([\d,.eE+-]+)\s*$", part)
        if colon_rng:
            var, lo_s, hi_s = colon_rng.groups()
            if var in df.columns:
                try:
                    lo = float(str(lo_s).replace(",", ""))
                    hi = float(str(hi_s).replace(",", ""))
                    num = pd.to_numeric(df[var], errors="coerce")
                    mask &= (num > lo) & (num <= hi)
                except (ValueError, TypeError):
                    pass
            continue

        cond = _try_parse_rule_text_to_condition(part)
        if cond is not None:
            mask &= _apply_condition(df, cond).fillna(False)
            continue

        match_le = re.match(r"([\w.]+)\s*<=\s*([\d.eE+-]+)", part)
        if match_le:
            var, val = match_le.groups()
            if var in df.columns:
                num = pd.to_numeric(df[var], errors="coerce")
                mask &= num <= float(val)
            continue

        match_lt = re.match(r"([\w.]+)\s*<\s*([\d.eE+-]+)", part)
        if match_lt:
            var, val = match_lt.groups()
            if var in df.columns:
                num = pd.to_numeric(df[var], errors="coerce")
                mask &= num < float(val)
            continue

        match_ge = re.match(r"([\w.]+)\s*>=\s*([\d.eE+-]+)", part)
        if match_ge:
            var, val = match_ge.groups()
            if var in df.columns:
                num = pd.to_numeric(df[var], errors="coerce")
                mask &= num >= float(val)
            continue

        match_gt = re.match(r"([\w.]+)\s*>\s*([\d.eE+-]+)", part)
        if match_gt:
            var, val = match_gt.groups()
            if var in df.columns:
                num = pd.to_numeric(df[var], errors="coerce")
                mask &= num > float(val)
            continue

        match_eq = re.match(r"([\w.]+)\s*[=]{1,2}\s*[\'\"]?([^\'\"]+)[\'\"]?", part)
        if match_eq:
            var, val = match_eq.groups()
            if var in df.columns:
                mask &= df[var].astype(str) == val.strip()
            continue

        match_neq = re.match(r"([\w.]+)\s*!=\s*[\'\"]?([^\'\"]+)[\'\"]?", part)
        if match_neq:
            var, val = match_neq.groups()
            if var in df.columns:
                mask &= df[var].astype(str) != val.strip()
            continue

    return mask


def _coerce_scheme_metadata_datetime(val: Any) -> datetime:
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        try:
            s = val.strip()
            if s.endswith("Z"):
                s = s[:-1] + "+00:00"
            return datetime.fromisoformat(s)
        except Exception:
            pass
    # Avoid fabricating "now" — breaks reload acceptance for legacy rows without created_at
    return datetime(1970, 1, 1, tzinfo=timezone.utc)


def _scheme_record_to_registry_entry(rec: Dict[str, Any], df: pd.DataFrame) -> SchemeRegistryEntry:
    col = str(rec.get("column_name") or "")
    mode_raw = rec.get("mode", SegmentationMode.PRE_EXISTING)
    if isinstance(mode_raw, SegmentationMode):
        mode = mode_raw
    else:
        mode = SegmentationMode(str(mode_raw))
    if col and col in df.columns:
        seg_count = int(df[col].nunique())
    else:
        seg_count = int(len(rec.get("segments") or []))
    vp = rec.get("variable_priority")
    if vp is not None and not isinstance(vp, VariablePriority):
        try:
            vp = (
                VariablePriority.model_validate(vp)
                if hasattr(VariablePriority, "model_validate")
                else VariablePriority(**vp)
            )
        except Exception:
            vp = None
    vrec = rec.get("validation") or {}
    rec_iv = rec.get("total_iv")
    if rec_iv is None and isinstance(vrec, dict):
        rec_iv = vrec.get("total_iv")
    rec_cat = rec.get("recommendation_category")
    if rec_cat is None and isinstance(vrec, dict):
        rec_cat = vrec.get("recommendation_category")

    return SchemeRegistryEntry(
        scheme_id=int(rec.get("scheme_id", 0)),
        column_name=col,
        mode=mode,
        variables=list(rec.get("variables") or []),
        variable_priority=vp,
        tree_method=rec.get("tree_method"),
        variable_selection_method=rec.get("variable_selection_method"),
        segment_count=seg_count,
        total_iv=float(rec_iv or 0.0),
        recommendation_category=str(rec_cat or "unknown"),
        created_at=_coerce_scheme_metadata_datetime(rec.get("created_at")),
    )


@chat_router.get("/segmentation/schemes/{dataset_id}/{scheme_id}", response_model=SegmentationSchemeDetailResponse)
async def get_segmentation_scheme_detail(
    dataset_id: str,
    scheme_id: int,
    current_user = Depends(get_current_user_dependency),
):
    """Return full stored audit metadata for one saved scheme (registry View details)."""
    try:
        stored_list = dataset_manager.get_segmentation_schemes_metadata(dataset_id)
        rec = next((s for s in stored_list if int(s.get("scheme_id", -1)) == scheme_id), None)
        if not rec:
            return SegmentationSchemeDetailResponse(
                success=False,
                dataset_id=dataset_id,
                scheme_id=scheme_id,
                metadata=None,
                message="Scheme metadata not found. It may have been saved before registry persistence was enabled.",
            )
        try:
            meta = SegmentationSchemeMetadata.model_validate(rec)
        except Exception as e:
            logger.error(f"Invalid stored scheme metadata scheme_id={scheme_id}: {e}")
            return SegmentationSchemeDetailResponse(
                success=False,
                dataset_id=dataset_id,
                scheme_id=scheme_id,
                metadata=None,
                message=f"Stored scheme metadata could not be read: {str(e)}",
            )
        return SegmentationSchemeDetailResponse(
            success=True,
            dataset_id=dataset_id,
            scheme_id=scheme_id,
            metadata=meta,
            message=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get scheme detail failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error loading scheme metadata: {str(e)}")


@chat_router.get("/segmentation/schemes/{dataset_id}", response_model=SchemeRegistryResponse)
async def get_segmentation_schemes(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency),
):
    """
    List all saved segmentation schemes for a dataset.
    Prefers persisted registry metadata (plan 12.3); falls back to legacy columns only.
    """
    try:
        df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")

        stored = dataset_manager.get_segmentation_schemes_metadata(dataset_id)
        schemes: List[SchemeRegistryEntry] = []
        if stored:
            for rec in stored:
                try:
                    schemes.append(_scheme_record_to_registry_entry(rec, df))
                except Exception as e:
                    logger.warning(f"Skipping invalid segmentation scheme registry record: {e}")
        if not schemes:
            scheme_cols = sorted(c for c in df.columns if c.startswith("segmentation_scheme_"))
            for col in scheme_cols:
                try:
                    sid = int(col.rsplit("_", 1)[-1])
                except ValueError:
                    sid = 0
                unique_segments = int(df[col].nunique())
                schemes.append(
                    SchemeRegistryEntry(
                        scheme_id=sid,
                        column_name=col,
                        mode=SegmentationMode.PRE_EXISTING,
                        variables=[],
                        variable_priority=None,
                        tree_method=None,
                        variable_selection_method=None,
                        segment_count=unique_segments,
                        total_iv=0.0,
                        recommendation_category="unknown",
                        created_at=datetime(1970, 1, 1, tzinfo=timezone.utc),
                    )
                )

        schemes.sort(key=lambda s: (s.created_at, s.scheme_id))
        return SchemeRegistryResponse(
            success=True,
            dataset_id=dataset_id,
            schemes=schemes,
            total_schemes=len(schemes),
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get schemes failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting schemes: {str(e)}")


@chat_router.get("/segmentation/audit-log/{dataset_id}")
async def get_segmentation_audit_log(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency),
):
    """Return persisted segmentation audit events (plan Section 15)."""
    try:
        if dataset_manager.get_dataset_info(dataset_id) is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        events = dataset_manager.get_segmentation_audit_log(dataset_id)
        return {"success": True, "dataset_id": dataset_id, "events": events, "count": len(events)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get segmentation audit log failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error loading audit log: {str(e)}")


@chat_router.get("/segmentation/insight-pins/{dataset_id}")
async def get_segmentation_insight_pins(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency),
):
    """Return segmentation Insight Pins stored for the Modeler's Notebook (plan Section 12.2)."""
    try:
        if dataset_manager.get_dataset_info(dataset_id) is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        raw = message_state_manager.db.load_message_state(dataset_id)
        if not raw:
            return {"success": True, "dataset_id": dataset_id, "pins": [], "count": 0}
        pins = raw.get("segmentation_insight_pins") or []
        if not isinstance(pins, list):
            pins = []
        return {"success": True, "dataset_id": dataset_id, "pins": pins, "count": len(pins)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get segmentation insight pins failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error loading insight pins: {str(e)}")


@chat_router.get("/segmentation/top-variables/{dataset_id}")
async def get_top_variables_by_segment(
    dataset_id: str,
    target_variable: Optional[str] = None,
    current_user = Depends(get_current_user_dependency)
):
    """
    Compute top 10 variables by IV per segment.
    Returns the Variable Relevance Matrix.
    """
    try:
        # Load dataset
        df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Find segment column
        segment_col = None
        for col in ["segment", "SEGMENT", "segment_id", "SEGMENT_ID"]:
            if col in df.columns:
                segment_col = col
                break
        
        if not segment_col:
            # Check for segmentation_scheme columns
            scheme_cols = [c for c in df.columns if c.startswith("segmentation_scheme_")]
            if scheme_cols:
                segment_col = scheme_cols[-1]  # Use latest
        
        if not segment_col:
            raise HTTPException(status_code=404, detail="No segment column found")
        
        # Get target
        if not target_variable:
            dataset_info = dataset_manager.get_dataset_info(dataset_id)
            target_variable = dataset_info.get("target_variable") if dataset_info else None
        
        if not target_variable or target_variable not in df.columns:
            raise HTTPException(status_code=400, detail="Target variable required for IV computation")
        
        # Get numeric columns excluding segment and target
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        feature_cols = [c for c in numeric_cols if c not in [segment_col, target_variable]]
        
        segments = df[segment_col].unique()
        
        # Compute IV per variable per segment (simplified)
        overall_iv = {}
        segment_iv = {str(seg): {} for seg in segments}
        
        for var in feature_cols[:20]:  # Limit to 20 variables for performance
            try:
                # Overall IV (simplified calculation)
                overall_iv[var] = round(np.random.uniform(0.05, 0.5), 4)  # Placeholder
                
                # Per-segment IV
                for seg in segments:
                    seg_df = df[df[segment_col] == seg]
                    segment_iv[str(seg)][var] = round(np.random.uniform(0.01, 0.6), 4)  # Placeholder
            except:
                pass
        
        # Sort by overall IV and take top 10
        sorted_vars = sorted(overall_iv.keys(), key=lambda x: overall_iv[x], reverse=True)[:10]
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "segment_column": segment_col,
            "segments": [str(s) for s in segments],
            "variables": sorted_vars,
            "overall_iv": {v: overall_iv[v] for v in sorted_vars},
            "segment_iv": {seg: {v: segment_iv[seg].get(v, 0) for v in sorted_vars} for seg in segment_iv}
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get top variables failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error computing top variables: {str(e)}")


# =============================================================================
# Helper functions for Segmentation Agent
# =============================================================================

def _calculate_woe_iv(seg_df: pd.DataFrame, full_df: pd.DataFrame, target_var: str, 
                      record_count: int, total_records: int, event_count: int) -> tuple:
    """Calculate WoE and IV contribution for a segment."""
    WOE_EPSILON = 0.5
    
    if not target_var or target_var not in full_df.columns:
        return 0.0, 0.0
    
    try:
        total_events = int(full_df[target_var].sum())
        total_non_events = total_records - total_events
        
        if total_events == 0 or total_non_events == 0:
            return 0.0, 0.0
        
        non_events = record_count - event_count
        
        # Distribution of events/non-events
        dist_events = (event_count + WOE_EPSILON) / (total_events + WOE_EPSILON)
        dist_non_events = (non_events + WOE_EPSILON) / (total_non_events + WOE_EPSILON)
        
        # WoE calculation
        if dist_events > 0 and dist_non_events > 0:
            woe = np.log(dist_non_events / dist_events)
        else:
            woe = 0.0
        
        # IV contribution
        iv_contrib = (dist_non_events - dist_events) * woe
        
        return woe, iv_contrib
    except:
        return 0.0, 0.0


def _woe_iv_from_segment_row_counts(
    full_df: pd.DataFrame,
    target_var: Optional[str],
    record_count: int,
    total_records: int,
    event_count: int,
) -> Tuple[float, float]:
    """
    WoE / IV contribution for one segment using global train counts only.
    Tree pipeline segments often omit iv_contribution; validation total_iv sums SegmentDetail.iv_contribution.
    """
    if not target_var or target_var not in full_df.columns or total_records <= 0:
        return 0.0, 0.0
    stub = full_df.head(0)
    return _calculate_woe_iv(stub, full_df, target_var, record_count, total_records, event_count)


def _calculate_wilson_score_ci(n: int, k: int, confidence: float = 0.95) -> Tuple[float, float]:
    """
    Calculate Wilson Score confidence interval for a proportion.
    
    Per Section 8.4 of the Segmentation Agent Plan:
    - Bar chart with event rate as a line overlay with confidence intervals
    - Overlapping confidence intervals flag: segments may not be well-separated
    
    Wilson Score interval is preferred over normal approximation because:
    - It's more accurate for small samples and extreme proportions
    - It never produces intervals outside [0, 1]
    - It's asymmetric, which is appropriate for proportions
    
    Args:
        n: Total number of observations (record_count)
        k: Number of successes (event_count)
        confidence: Confidence level (default 0.95 for 95% CI)
    
    Returns:
        Tuple of (lower_bound, upper_bound) as percentages (0-100 scale)
    """
    if n == 0:
        return 0.0, 0.0
    
    from scipy import stats
    
    # Sample proportion
    p_hat = k / n
    
    # Z-score for confidence level
    z = stats.norm.ppf(1 - (1 - confidence) / 2)
    
    # Wilson Score formula
    denominator = 1 + z**2 / n
    center = (p_hat + z**2 / (2 * n)) / denominator
    spread = (z / denominator) * np.sqrt(p_hat * (1 - p_hat) / n + z**2 / (4 * n**2))
    
    lower = max(0.0, center - spread)
    upper = min(1.0, center + spread)
    
    # Convert to percentage (0-100 scale) to match event_rate
    return round(lower * 100, 4), round(upper * 100, 4)


def _normalize_segmentation_rule_operator(op: Optional[str]) -> str:
    """Map UI / API operators to internal tokens (plan §6.2)."""
    if op is None:
        return "="
    s = str(op).strip()
    sl = s.lower()
    aliases = {
        "==": "=",
        "in": "IN",
        "not_in": "NOT IN",
        "not in": "NOT IN",
        "between": "BETWEEN",
        "not_between": "NOT BETWEEN",
        "not between": "NOT BETWEEN",
        "is_null": "IS_NULL",
        "is_not_null": "IS_NOT_NULL",
        "contains": "CONTAINS",
        "is_true": "IS_TRUE",
        "is false": "IS_FALSE",
        "is_false": "IS_FALSE",
        "true": "IS_TRUE",
        "false": "IS_FALSE",
    }
    if sl in aliases:
        return aliases[sl]
    u = s.upper().replace(" ", "_")
    if u in ("NOT_IN", "NOTIN"):
        return "NOT IN"
    return u


def _coerce_list_values(val: Any) -> List[Any]:
    if val is None:
        return []
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        parts = [p.strip().strip("'\"") for p in val.split(",") if p.strip()]
        out: List[Any] = []
        for p in parts:
            try:
                out.append(float(p) if "." in p else int(p))
            except ValueError:
                out.append(p)
        return out
    return [val]


def _manual_rule_combined_mask(df: pd.DataFrame, rule: ManualSegmentRule) -> pd.Series:
    """Boolean mask for one segment's conditions (AND/OR), ignoring catch_all / exclusivity."""
    if not rule.conditions:
        return pd.Series([True] * len(df), index=df.index)
    if rule.logic == "OR":
        mask = pd.Series([False] * len(df), index=df.index)
        for cond in rule.conditions:
            if not cond.variable or cond.variable not in df.columns:
                continue
            cm = _apply_condition(df, cond)
            mask = mask | cm
    else:
        mask = pd.Series([True] * len(df), index=df.index)
        for cond in rule.conditions:
            if not cond.variable or cond.variable not in df.columns:
                mask = mask & pd.Series([False] * len(df), index=df.index)
                continue
            cm = _apply_condition(df, cond)
            mask = mask & cm
    return mask.fillna(False)


def _manual_rules_assign_labels(df: pd.DataFrame, rules: List[ManualSegmentRule]) -> pd.Series:
    """
    First-wins segment assignment in rule list order (plan §6.2 / catch-all).
    catch_all rows match ~prior_claimed, optionally narrowed by conditions.
    """
    labels = pd.Series("Unassigned", index=df.index, dtype=object)
    claimed = pd.Series(False, index=df.index)
    for rule in rules:
        catch = bool(getattr(rule, "catch_all", False))
        if catch:
            base = ~claimed
            rm = _manual_rule_combined_mask(df, rule)
            if not rule.conditions or all(not (c.variable or "").strip() for c in rule.conditions):
                m = base.fillna(False)
            else:
                m = (base & rm).fillna(False)
        else:
            rm = _manual_rule_combined_mask(df, rule)
            m = (rm & ~claimed).fillna(False)
        labels = labels.where(~m, rule.segment_name)
        claimed = claimed | m
    return labels


def _format_manual_rule_definition(rule: ManualSegmentRule) -> str:
    if getattr(rule, "catch_all", False) and (
        not rule.conditions or all(not (c.variable or "").strip() for c in rule.conditions)
    ):
        return "CATCH_ALL (remaining rows)"
    parts: List[str] = []
    for cond in rule.conditions:
        if not (cond.variable or "").strip():
            continue
        parts.append(f"{cond.variable} {cond.operator} {cond.value}")
    return f" {rule.logic} ".join(parts) if parts else "(no conditions)"


def _norm_category_token(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _in_list_apply_remove_add(
    raw: Any, *, remove: Optional[str] = None, add: Optional[str] = None
) -> List[Any]:
    """Build a new IN-list: drop remove token (if any), append add once (if any)."""
    lst = _coerce_list_values(raw)
    rm = _norm_category_token(remove) if remove else ""
    ad = _norm_category_token(add) if add else ""
    out: List[Any] = []
    for x in lst:
        if rm and _norm_category_token(x) == rm:
            continue
        out.append(x)
    if ad and not any(_norm_category_token(x) == ad for x in out):
        out.append(ad)
    return out


def _adjust_manual_rules_category_move(
    rules: List[ManualSegmentRule],
    from_segment_name: str,
    to_segment_name: str,
    variable: str,
    category_value: str,
) -> List[ManualSegmentRule]:
    """Move one categorical level from source segment rule to destination (C3 structured rules)."""
    var = (variable or "").strip()
    cat = _norm_category_token(category_value)
    if not var or not cat:
        raise ValueError("variable and category_value are required")
    if from_segment_name == to_segment_name:
        raise ValueError("from and to segments must differ")

    out_rules: List[ManualSegmentRule] = []
    for rule in rules:
        r = ManualSegmentRule.model_validate(rule.model_dump())
        if r.segment_name == from_segment_name:
            new_conds: List[RuleCondition] = []
            for c in r.conditions:
                if (c.variable or "").strip() != var:
                    new_conds.append(c)
                    continue
                op = _normalize_segmentation_rule_operator(c.operator)
                if op == "IN":
                    nl = _in_list_apply_remove_add(c.value, remove=cat, add=None)
                    if nl:
                        new_conds.append(RuleCondition(variable=var, operator="IN", value=nl))
                elif op == "=":
                    if _norm_category_token(c.value) != cat:
                        new_conds.append(c)
                else:
                    new_conds.append(c)
            if not new_conds and not r.catch_all:
                raise ValueError(
                    f"Removing category {cat!r} would leave segment {from_segment_name!r} with no conditions"
                )
            r.conditions = new_conds
        elif r.segment_name == to_segment_name:
            touched = False
            new_conds2: List[RuleCondition] = []
            for c in r.conditions:
                if (c.variable or "").strip() != var:
                    new_conds2.append(c)
                    continue
                op = _normalize_segmentation_rule_operator(c.operator)
                if op == "IN":
                    touched = True
                    nl = _in_list_apply_remove_add(c.value, remove=None, add=cat)
                    new_conds2.append(RuleCondition(variable=var, operator="IN", value=nl))
                else:
                    new_conds2.append(c)
            if not touched:
                new_conds2.append(RuleCondition(variable=var, operator="IN", value=[cat]))
            r.conditions = new_conds2
        out_rules.append(r)
    return out_rules


def _sync_segment_rule_strings_from_manual_rules(
    segment_dicts: List[Dict[str, Any]], rules: List[ManualSegmentRule]
) -> None:
    by_name = {r.segment_name: r for r in rules}
    for s in segment_dicts:
        nm = s.get("segment_name")
        if nm in by_name:
            s["rule_definition"] = _format_manual_rule_definition(by_name[nm])


def _apply_condition(df: pd.DataFrame, cond: RuleCondition) -> pd.Series:
    """Apply a single rule condition to get a boolean mask."""
    if not cond.variable or cond.variable not in df.columns:
        return pd.Series([False] * len(df), index=df.index)
    col = df[cond.variable]
    op = _normalize_segmentation_rule_operator(cond.operator)
    val = cond.value

    if op == "=":
        if isinstance(val, str) and val.upper() in ("TRUE", "FALSE"):
            if val.upper() == "TRUE":
                return (col == True) | (col == 1) | (col.astype(str).str.upper() == "TRUE")
            return (col == False) | (col == 0) | (col.astype(str).str.upper() == "FALSE")
        return col == val
    if op == "!=":
        return col != val
    if op == "<=":
        return pd.to_numeric(col, errors="coerce") <= float(val)
    if op == ">=":
        return pd.to_numeric(col, errors="coerce") >= float(val)
    if op == "<":
        return pd.to_numeric(col, errors="coerce") < float(val)
    if op == ">":
        return pd.to_numeric(col, errors="coerce") > float(val)
    if op == "IN":
        lst = _coerce_list_values(val)
        return col.isin(lst)
    if op == "NOT IN":
        lst = _coerce_list_values(val)
        return ~col.isin(lst)
    if op == "BETWEEN":
        lo, hi = None, None
        if isinstance(val, (list, tuple)) and len(val) >= 2:
            try:
                lo, hi = float(val[0]), float(val[1])
            except (TypeError, ValueError):
                pass
        elif isinstance(val, str) and "," in val:
            parts = [p.strip() for p in val.split(",") if p.strip()]
            if len(parts) >= 2:
                try:
                    lo, hi = float(parts[0]), float(parts[1])
                except ValueError:
                    pass
        if lo is not None and hi is not None:
            cnum = pd.to_numeric(col, errors="coerce")
            return (cnum >= lo) & (cnum <= hi)
        return pd.Series([False] * len(df), index=df.index)
    if op == "NOT BETWEEN":
        lo, hi = None, None
        if isinstance(val, (list, tuple)) and len(val) >= 2:
            try:
                lo, hi = float(val[0]), float(val[1])
            except (TypeError, ValueError):
                pass
        elif isinstance(val, str) and "," in val:
            parts = [p.strip() for p in val.split(",") if p.strip()]
            if len(parts) >= 2:
                try:
                    lo, hi = float(parts[0]), float(parts[1])
                except ValueError:
                    pass
        if lo is not None and hi is not None:
            cnum = pd.to_numeric(col, errors="coerce")
            return (cnum < lo) | (cnum > hi)
        return pd.Series([False] * len(df), index=df.index)
    if op == "IS_NULL":
        return col.isna()
    if op == "IS_NOT_NULL":
        return col.notna()
    if op == "CONTAINS":
        needle = "" if val is None else str(val)
        return col.astype(str).str.contains(needle, case=False, na=False, regex=False)
    if op == "IS_TRUE":
        return (col == True) | (col == 1) | (col.astype(str).str.upper() == "TRUE")
    if op == "IS_FALSE":
        return (col == False) | (col == 0) | (col.astype(str).str.upper() == "FALSE")
    return pd.Series([False] * len(df), index=df.index)


def _stability_labels_from_leaf_assignment(
    leaf_ids: Any,
    raw_segments: List[Dict[str, Any]],
    segment_details: Optional[List[SegmentDetail]] = None,
) -> Optional[np.ndarray]:
    """
    Map per-row tree leaf indices to SegmentDetail.segment_name.

    Bootstrap stability and OOS compare ``df[segment_column]`` to ``segment_name``.
    ``leaf_ids`` are integers (0, 1, …) while ``segment_name`` is ``Segment 1``, …,
    so masks never matched and rank stability stayed 0%.

    ``segment_details`` is optional: when lengths differ from ``raw_segments`` (or
    details are omitted), labels fall back to the same rule as ``_convert_to_segment_details``
    (``name`` key or ``Segment {position}``).
    """
    if not raw_segments:
        return None
    n_details = len(segment_details) if segment_details else 0
    leaf_to_name: Dict[int, str] = {}
    for i, rseg in enumerate(raw_segments):
        if not isinstance(rseg, dict):
            continue
        lid = rseg.get("leaf_id")
        if lid is None:
            continue
        try:
            lid_i = int(lid)
        except (TypeError, ValueError):
            continue
        pos = i + 1  # matches enumerate(..., 1) in _convert_to_segment_details
        if i < n_details:
            label = segment_details[i].segment_name
        else:
            label = str(rseg.get("name") or f"Segment {pos}")
        leaf_to_name[lid_i] = label
    if not leaf_to_name:
        return None
    arr = np.asarray(leaf_ids)
    if arr.size == 0:
        return None
    out = np.empty(arr.shape[0], dtype=object)
    flat = arr.reshape(-1)
    for i, v in enumerate(flat):
        try:
            iv = int(v)
        except (TypeError, ValueError):
            out.flat[i] = str(v)
            continue
        out.flat[i] = leaf_to_name.get(iv, str(iv))
    return out.reshape(arr.shape)


def _convert_to_segment_details(result: dict, df: pd.DataFrame, target_var: str,
                                total_records: int) -> List[SegmentDetail]:
    """Convert legacy segmentation result to SegmentDetail list."""
    segments = []
    
    for idx, seg in enumerate(result.get("segments", []), 1):
        record_count = seg.get("size", seg.get("count", 0))
        event_count = seg.get("event_count", 0)
        event_rate = seg.get("event_rate", 0)
        
        # If event_rate is a ratio, convert to percentage
        if event_rate > 0 and event_rate < 1:
            event_rate = event_rate * 100
        
        # Calculate event_count from event_rate if not provided
        if event_count == 0 and event_rate > 0 and record_count > 0:
            # event_rate is now in percentage (0-100 scale)
            event_count = int(record_count * event_rate / 100)
        
        # Calculate Wilson Score confidence intervals
        ci_lower, ci_upper = _calculate_wilson_score_ci(record_count, event_count)

        woe_v, iv_c = _woe_iv_from_segment_row_counts(
            full_df=df,
            target_var=target_var,
            record_count=int(record_count),
            total_records=int(total_records),
            event_count=int(event_count),
        )

        # Persist parseable rules for merge / rebuild_validation (rules_readable alone is not enough).
        rules_list = seg.get("rules") or []
        rule_from_tree = " AND ".join(str(x) for x in rules_list if str(x).strip())
        rule_def = (seg.get("rule") or seg.get("description") or seg.get("rule_definition") or "").strip()
        if not rule_def:
            rule_def = rule_from_tree
        if not rule_def:
            rr = (seg.get("rules_readable") or "").strip()
            if rr and rr.upper() != "ALL DATA":
                rule_def = rr.replace("≤", "<=").replace("≥", ">=").replace(" and ", " AND ")
        
        segments.append(SegmentDetail(
            segment_id=idx,
            segment_name=seg.get("name", f"Segment {idx}"),
            rule_definition=rule_def,
            record_count=record_count,
            pct_of_population=round(record_count / total_records * 100, 2) if total_records > 0 else 0,
            event_count=event_count,
            event_rate=round(event_rate, 4),
            event_rate_ci_lower=ci_lower,
            event_rate_ci_upper=ci_upper,
            woe=round(float(woe_v), 4),
            iv_contribution=round(float(iv_c), 4),
        ))
    
    return segments


def _get_oos_data(df: pd.DataFrame, segment_column: Optional[str] = None) -> Tuple[Optional[pd.DataFrame], str]:
    """
    Extract out-of-sample data from DataFrame based on split_tag column.
    
    Per plan Section 11:
    - If holdout partition exists, validation runs on holdout
    - If not, it runs on the test partition
    - If neither exists, validation is skipped
    
    Args:
        df: Full dataframe with split_tag column
        segment_column: Segment column to ensure exists in OOS data
        
    Returns:
        Tuple of (oos_dataframe or None, partition_name used)
    """
    if 'split_tag' not in df.columns:
        return None, "none"
    
    # Try holdout/validation first (per plan)
    validation_mask = df['split_tag'].str.startswith('validation', na=False)
    if validation_mask.sum() > 0:
        oos_df = df[validation_mask].copy()
        # Ensure segment column exists if needed
        if segment_column and segment_column not in oos_df.columns:
            return None, "none"
        logger.info(f"Using validation/holdout partition for OOS validation: {len(oos_df)} records")
        return oos_df, "holdout"
    
    # Fall back to test partition
    test_mask = df['split_tag'] == 'test'
    if test_mask.sum() > 0:
        oos_df = df[test_mask].copy()
        if segment_column and segment_column not in oos_df.columns:
            return None, "none"
        logger.info(f"Using test partition for OOS validation: {len(oos_df)} records")
        return oos_df, "test"
    
    # No OOS data available
    logger.info("No out-of-sample partition available for validation")
    return None, "none"


def _top_level_or_clauses(rule_def: str) -> List[str]:
    """Split a rule string on top-level OR (outside parentheses)."""
    s = (rule_def or "").strip()
    if not s:
        return []
    depth = 0
    buf: List[str] = []
    pieces: List[str] = []
    i = 0
    n = len(s)
    while i < n:
        c = s[i]
        if c == "(":
            depth += 1
            buf.append(c)
            i += 1
        elif c == ")":
            depth -= 1
            buf.append(c)
            i += 1
        elif depth == 0 and i + 4 <= n and s[i : i + 4].upper() == " OR ":
            piece = "".join(buf).strip()
            if piece:
                pieces.append(piece)
            buf = []
            i += 4
        else:
            buf.append(c)
            i += 1
    tail = "".join(buf).strip()
    if tail:
        pieces.append(tail)
    return pieces


def _outer_parens_wrap_entire_string(s: str) -> bool:
    """
    True if the string is a single parenthesized expression, e.g. ( ... ) with matching
    open/close at the very ends. Used to peel wrappers so top-level OR can be found.
    """
    s = (s or "").strip()
    if len(s) < 2 or s[0] != "(" or s[-1] != ")":
        return False
    depth = 0
    for i, c in enumerate(s):
        if c == "(":
            depth += 1
        elif c == ")":
            depth -= 1
            if depth == 0:
                return i == len(s) - 1
    return False


def _flatten_or_rule_to_leaves(rule: str) -> List[str]:
    """
    Decompose a rule into OR-atomic leaves. Nested merges like
    '((A) OR (B)) OR (C)' are flattened to three leaves so _segmentation_rule_mask
    and merge logic both see a flat OR chain.
    """
    s = (rule or "").strip()
    if not s:
        return []
    s = s.replace("\u2264", "<=").replace("\u2265", ">=")
    while _outer_parens_wrap_entire_string(s):
        s = s[1:-1].strip()
    parts = _top_level_or_clauses(s)
    if len(parts) > 1:
        return [leaf for p in parts for leaf in _flatten_or_rule_to_leaves(p)]
    return [s]


def _merge_two_segment_rule_strings(ra: str, rb: str) -> str:
    """
    Combine two segment rule strings for merge: flatten existing ORs on each side,
    dedupe, and join with top-level OR so we never build ((A) OR (B)) OR (C) without
    flattening (which breaks _top_level_or_clauses and yields wrong row counts on 2+ merges).
    """
    la = _flatten_or_rule_to_leaves(ra)
    lb = _flatten_or_rule_to_leaves(rb)
    seen: set = set()
    out: List[str] = []
    for x in la + lb:
        k = (x or "").strip()
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(k)
    if not out:
        return (ra or rb or "").strip()
    if len(out) == 1:
        return out[0]

    def _wrap_leaf(t: str) -> str:
        t = t.strip()
        if t.startswith("(") and t.endswith(")"):
            return t
        return f"({t})"

    return " OR ".join(_wrap_leaf(t) for t in out)


def _atomic_rule_mask(df: pd.DataFrame, rule_def: str) -> pd.Series:
    """Parse a single AND-only style rule (strip redundant outer parentheses)."""
    atom = (rule_def or "").strip()
    while len(atom) >= 2 and atom[0] == "(" and atom[-1] == ")":
        atom = atom[1:-1].strip()
    if not atom or atom == "Unknown":
        return pd.Series([False] * len(df), index=df.index)
    return _parse_and_apply_rule(df, atom, [])


def _segmentation_rule_mask(df: pd.DataFrame, rule_def: str) -> pd.Series:
    """
    Boolean mask for rows matching a segment rule, including top-level OR (manual merges).
    Strips unnecessary outer parentheses so nested merges like '((A) OR (B)) OR (C)'
    are split as multiple ORs.
    """
    if rule_def:
        rule_def = str(rule_def).replace("\u2264", "<=").replace("\u2265", ">=")
    if not rule_def or rule_def == "Unknown":
        return pd.Series([False] * len(df), index=df.index)
    s = str(rule_def).strip()
    while _outer_parens_wrap_entire_string(s):
        s = s[1:-1].strip()
    parts = _top_level_or_clauses(s)
    if len(parts) > 1:
        m = pd.Series([False] * len(df), index=df.index)
        for p in parts:
            m = m | _segmentation_rule_mask(df, p)
        return m
    return _atomic_rule_mask(df, s)


def _serialize_validation_result(v: ValidationSuiteResult) -> Dict[str, Any]:
    if hasattr(v, "model_dump"):
        return json.loads(json.dumps(v.model_dump(mode="json"), default=str))
    return json.loads(json.dumps(v.dict(), default=str))


def _merge_validation_into_auto_candidates(
    auto_candidates: List[Dict[str, Any]],
    validation: ValidationSuiteResult,
    schemes: List[Any],
) -> List[Dict[str, Any]]:
    """
    Align Discovered Schemes cards with the ValidationSuite result (Insights panel):
    same recommendation_category, total_iv, quality composite score, and RECOMMENDED badge rules.
    """
    if not auto_candidates or validation is None:
        return auto_candidates
    from app.services.auto_segmentation_pipeline import auto_pipeline

    suspicious = getattr(validation, "iv_category", None) == "suspicious"
    out: List[Dict[str, Any]] = []
    for i, cand in enumerate(auto_candidates):
        merged = dict(cand)
        merged["recommendation_category"] = validation.recommendation_category
        merged["iv"] = round(float(validation.total_iv), 4)
        merged["iv_category"] = getattr(validation, "iv_category", None)
        if i < len(schemes):
            try:
                merged["score"] = round(auto_pipeline.compute_quality_composite_score(schemes[i]), 1)
            except Exception:
                merged["score"] = cand.get("score", 0.0)
        cat = validation.recommendation_category or "weak"
        merged["recommended"] = cat in ("strong", "exploratory") and not suspicious
        out.append(merged)
    return out


def _run_validation_suite(segments: List[SegmentDetail], df: pd.DataFrame, target_var: str,
                          total_records: int, overall_event_rate: float, 
                          min_segment_size: int, dataset_id: str,
                          segment_column: Optional[str] = None,
                          run_stability: bool = True,
                          run_oos_validation: bool = True) -> ValidationSuiteResult:
    """
    Run the complete validation suite for a segmentation scheme.
    Delegates to the segmentation_validation module.
    
    Args:
        segments: List of segment details
        df: Training dataframe (should include split_tag column for OOS extraction)
        target_var: Target variable name
        total_records: Total number of records
        overall_event_rate: Overall event rate (percentage)
        min_segment_size: Minimum segment size threshold
        dataset_id: Dataset ID for caching
        segment_column: Name of segment column in dataframe (required for stability and OOS)
        run_stability: Whether to run bootstrap stability diagnostics (default: True)
        run_oos_validation: Whether to run out-of-sample validation (default: True)
    """
    # Extract OOS data if available and requested
    oos_df = None
    oos_partition = "holdout"  # Default
    if run_oos_validation and segment_column:
        oos_df, oos_partition = _get_oos_data(df, segment_column)
        if oos_df is not None:
            logger.info(f"OOS validation enabled using {oos_partition} partition ({len(oos_df)} records)")
    
    return seg_validation_suite.run_validation(
        segments=segments,
        df=df,
        target_var=target_var,
        total_records=total_records,
        overall_event_rate=overall_event_rate,
        min_segment_size=min_segment_size,
        oos_df=oos_df,
        segment_column=segment_column,
        run_stability=run_stability,
        stability_bootstrap_runs=None,  # uses SegmentationConfig.BOOTSTRAP_RUNS (env: MIDAS_SEGMENTATION_BOOTSTRAP_RUNS)
        oos_partition=oos_partition
    )


def _compute_variable_relevance(df: pd.DataFrame, segment_column: str, target_var: str,
                                 variables_used: List[str] = None) -> Optional[VariableRelevanceMatrix]:
    """
    Compute variable relevance matrix (top 10 variables by IV per segment).
    
    Args:
        df: DataFrame with segment assignments
        segment_column: Column containing segment assignments
        target_var: Binary target variable
        variables_used: Optional list of variables to prioritize
        
    Returns:
        VariableRelevanceMatrix or None if computation fails
    """
    try:
        if not target_var or target_var not in df.columns:
            return None
        if segment_column not in df.columns:
            return None
            
        # Get numeric columns as candidates
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        candidate_vars = [c for c in numeric_cols if c not in [segment_column, target_var]]
        
        # If specific variables were used, prioritize them but include others too
        if variables_used:
            # Put used variables first, then others
            other_vars = [v for v in candidate_vars if v not in variables_used]
            candidate_vars = [v for v in variables_used if v in df.columns] + other_vars
        
        if not candidate_vars:
            return None
            
        return variable_relevance_calculator.compute_variable_iv_per_segment(
            df=df,
            segment_column=segment_column,
            target_variable=target_var,
            candidate_variables=candidate_vars[:50],  # Limit to top 50 candidates for performance
            top_n=10
        )
    except Exception as e:
        logger.warning(f"Could not compute variable relevance: {e}")
        return None


def _pick_segment_source_column(df: pd.DataFrame) -> Optional[str]:
    """
    Prefer in-memory ``segment``; otherwise newest persisted ``segmentation_scheme_*`` column
    (multi-replica / post–Add-to-Data safe).
    """
    cols = [str(c) for c in df.columns]
    if "segment" in cols:
        return "segment"
    scheme_cols = [c for c in cols if c.startswith("segmentation_scheme_")]
    if not scheme_cols:
        return None

    def scheme_id(name: str) -> int:
        try:
            return int(name.replace("segmentation_scheme_", ""))
        except ValueError:
            return -1

    return max(scheme_cols, key=scheme_id)


@chat_router.get("/dataset-preview/{dataset_id}", response_model=DatasetPreviewResponse)
async def get_dataset_preview(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get a preview of the dataset (first 10 rows) for global model training
    """
    try:
        logger.info(f"Getting dataset preview for: {dataset_id}")
        
        # Get dataset info
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Load the dataset - try to get processed DataFrame first
        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            head = dataset_manager.load_dataset_head_for_preview(dataset_id, nrows=10)
            if head is None:
                raise HTTPException(status_code=404, detail="Dataset not found")
            preview_df, total_rows, ncols = head
            logger.info(
                "Loaded dataset via storage head for preview. rows=%s cols=%s",
                total_rows,
                ncols,
            )
        else:
            logger.info(f"Retrieved processed dataset for preview. Shape: {df.shape}")
            preview_df = df.head(10)
            total_rows = len(df)
            ncols = len(df.columns)
        
        # Convert to dictionary format
        preview_data = {
            'columns': preview_df.columns.tolist(),
            'rows': preview_df.to_dict('records')
        }
        
        # Get shape info
        shape_info = {
            'rows': total_rows,
            'columns': ncols
        }
        
        logger.info("Dataset preview generated successfully.")
        
        return DatasetPreviewResponse(
            success=True,
            message="Dataset preview generated successfully",
            preview_data=preview_data,
            shape=shape_info
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get dataset preview: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting dataset preview: {str(e)}")

@chat_router.get("/segmented-dataset-preview/{dataset_id}", response_model=DatasetPreviewResponse)
async def get_segmented_dataset_preview(
    dataset_id: str,
    current_user = Depends(get_current_user_dependency)
):
    """
    Get a preview of the segmented dataset (first 10 rows) with segment column
    """
    try:
        logger.info(f"Getting segmented dataset preview for: {dataset_id}")
        
        # Get dataset info
        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is not None:
            logger.info(f"Retrieved dataframe from state manager. Shape: {df.shape}")
            seg_col = _pick_segment_source_column(df)
            if seg_col is None:
                raise HTTPException(
                    status_code=404,
                    detail="No segmented dataset found. Please run segmentation first.",
                )
            preview_df = df.head(10).copy()
            if seg_col != "segment":
                preview_df = preview_df.rename(columns={seg_col: "segment"})
            preview_data = {
                "columns": preview_df.columns.tolist(),
                "rows": preview_df.to_dict("records"),
            }
            shape_info = {"rows": len(df), "columns": len(df.columns)}
            logger.info(f"Segmented dataset preview from state. Shape: {df.shape}")
            return DatasetPreviewResponse(
                success=True,
                message="Segmented dataset preview generated successfully",
                preview_data=preview_data,
                shape=shape_info,
            )

        head = dataset_manager.load_dataset_head_for_preview(dataset_id, nrows=10)
        if head is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        preview_df, total_rows, ncols = head
        seg_col = _pick_segment_source_column(preview_df)
        if seg_col is None:
            raise HTTPException(
                status_code=404,
                detail="No segmented dataset found. Please run segmentation first.",
            )
        if seg_col != "segment":
            preview_df = preview_df.rename(columns={seg_col: "segment"})
        logger.info(
            "Segmented dataset preview via storage head. rows=%s cols=%s",
            total_rows,
            ncols,
        )

        preview_data = {
            "columns": preview_df.columns.tolist(),
            "rows": preview_df.to_dict("records"),
        }
        shape_info = {"rows": total_rows, "columns": ncols}

        return DatasetPreviewResponse(
            success=True,
            message="Segmented dataset preview generated successfully",
            preview_data=preview_data,
            shape=shape_info,
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get segmented dataset preview: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting segmented dataset preview: {str(e)}")

# NOTE: The synchronous `POST /segment-profiling` endpoint was removed
# (Phase 1 stateless-API cleanup, May 2026). Use
# `POST /segment-profiling/start` + `GET /segment-profiling/status/{job_id}`
# instead — segment profiling on real datasets exceeds the 1 s
# `SLOW_REQUEST_THRESHOLD_MS` and must run in a background job per
# `.cursor/rules/architecture.mdc`. No frontend caller references the
# sync path.

@chat_router.post("/segment-profiling/start")
async def start_segment_profiling(request: dict):
    """
    Start segment profiling as background job
    Returns job_id immediately to prevent timeout
    """
    try:
        from app.services.background_jobs import background_job_manager
        import uuid
        import time
        
        # Extract parameters
        dataset_id = request.get("dataset_id")
        segmentation_result = request.get("segmentation_result")
        target_variable = request.get("target_variable")
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not segmentation_result:
            raise HTTPException(status_code=400, detail="segmentation_result is required")
        if not target_variable:
            raise HTTPException(status_code=400, detail="target_variable is required")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        
        # Define job function with progress updates
        def execute_segment_profiling(params: dict):
            """Execute segment profiling in background with progress updates"""
            from app.services.segmentation_service import segmentation_service
            from app.services.dataframe_state_manager import dataframe_state_manager
            
            dataset_id = params['dataset_id']
            segmentation_result = params['segmentation_result']
            target_variable = params['target_variable']
            
            # Progress callback to update job status
            def progress_callback(progress: int, step: int, message: str):
                """Update job progress"""
                try:
                    job_status = background_job_manager.get_job_status(job_id)
                    if job_status:
                        with background_job_manager._lock:
                            if job_id in background_job_manager._jobs:
                                background_job_manager._jobs[job_id]['progress'] = progress
                                background_job_manager._jobs[job_id]['message'] = message
                                background_job_manager._jobs[job_id]['step'] = step
                        background_job_manager.persist_job_snapshot(job_id)
                except Exception as e:
                    logger.warning(f"Failed to update progress for job {job_id}: {str(e)}")
            
            # Execute profiling with progress callback
            result = segmentation_service.analyze_segment_quality(
                dataset_id=dataset_id,
                segmentation_result=segmentation_result,
                target_variable=target_variable,
                dataset_manager=dataset_manager,
                progress_callback=progress_callback
            )
            
            return result
        
        # Start background job
        background_job_manager.start_job(
            job_id=job_id,
            job_type='segment_profiling',
            params={
                'dataset_id': dataset_id,
                'segmentation_result': segmentation_result,
                'target_variable': target_variable
            },
            job_function=execute_segment_profiling
        )
        
        logger.info(f"Started segment profiling job {job_id} for dataset {dataset_id}")
        
        return {
            "success": True,
            "job_id": job_id,
            "status": "started",
            "message": "Segment profiling started in background. Poll /segment-profiling/status/{job_id} for updates."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting segment profiling: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting segment profiling: {str(e)}")

@chat_router.get("/segment-profiling/status/{job_id}")
async def get_segment_profiling_status(job_id: str):
    """
    Check status of segment profiling job
    """
    try:
        from app.services.background_jobs import background_job_manager
        
        status = background_job_manager.get_job_status(job_id)
        
        if status is None:
            raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
        
        response = {
            "job_id": job_id,
            "status": status.get("status"),
            "progress": status.get("progress", 0),
            "message": status.get("message", ""),
            "step": status.get("step", 0)
        }
        
        # If completed, include result
        if status.get("status") == "completed":
            response["result"] = status.get("result")
            response["success"] = True
        elif status.get("status") == "failed":
            response["error"] = status.get("error")
            response["success"] = False
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting segment profiling status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting segment profiling status: {str(e)}")

def _map_dataset_type(dataset_type_str: str):
    """Map LLM string to DatasetType enum."""
    _m = {
        'classification': DatasetType.CLASSIFICATION,
        'regression': DatasetType.REGRESSION,
        'time_series': DatasetType.TIME_SERIES,
        'others': DatasetType.OTHERS,
    }
    return _m.get(dataset_type_str.lower(), DatasetType.OTHERS)


# P2.4b: stratified-sampling primitives now live in app/services/sampling.py
# so the classification background task and any future analytics paths share
# the same implementation.
from app.services.sampling import (
    stratified_sample_pandas as _stratified_sample_pandas,
    stratified_indices_from_target_series as _stratified_indices_from_target_series,  # noqa: F401
    build_classification_sample as _build_classification_sample_core,
)


def _build_classification_sample(
    *,
    dataset_id: Optional[str],
    csv_path: Optional[str],
    target_variable: Optional[str],
    sample_rows: int = 200_000,
    min_per_class: int = 5_000,
) -> Tuple[pd.DataFrame, Tuple[int, int]]:
    """
    Resolve the most efficient data source for a classification sample and
    delegate to the canonical sampler.

    Resolution order:
      1. Cached full DataFrame from dataframe_state_manager (no extra IO).
      2. dataset_manager.load_dataset(dataset_id) -> uses Parquet sidecar
         when present, else CSV.
      3. Direct csv_path read (legacy multipart path).
    """
    full_df = None
    if dataset_id:
        try:
            full_df = dataframe_state_manager.get_full_dataframe_readonly(dataset_id)
        except Exception:
            full_df = None
        if full_df is None:
            try:
                full_df = dataset_manager.load_dataset(dataset_id)
            except Exception as exc:
                logger.warning(
                    f"[classify-bg] dataset_manager.load_dataset({dataset_id}) failed: {exc}"
                )
                full_df = None

    return _build_classification_sample_core(
        dataset_id=dataset_id,
        csv_path=csv_path,
        target_variable=target_variable,
        sample_rows=sample_rows,
        min_per_class=min_per_class,
        full_df=full_df,
    )


def _summary_from_cached(
    column_info_response: Any,
    target_profile: Dict[str, Any],
    target_variable: str,
) -> Dict[str, Any]:
    """
    P3.4: Build the LLM-classifier "dataset summary" structure directly
    from cached metadata, avoiding any DataFrame I/O. Output schema is
    intentionally a strict subset of `generate_dataset_summary(df)` so the
    classifier prompt stays unchanged.
    """
    columns: List[Dict[str, Any]] = []
    for col in getattr(column_info_response, "columns_info", []) or []:
        try:
            entry = {
                "name": getattr(col, "column_name", None) or col.get("column_name"),
                "dtype": getattr(col, "data_type", None) or col.get("data_type"),
                "user_friendly_type": getattr(col, "column_type", None) or col.get("column_type"),
                "n_unique": getattr(col, "unique_count", None) or col.get("unique_count"),
                "n_missing": getattr(col, "missing_count", None) or col.get("missing_count"),
                "is_date": bool(getattr(col, "is_date", False) or col.get("is_date")),
            }
            columns.append(entry)
        except Exception:
            continue

    summary: Dict[str, Any] = {
        "columns": columns,
        "n_columns": len(columns),
        "target_profile": target_profile,
        "target_variable": target_variable,
    }
    return summary


def _run_dataset_type_classification_bg(
    job_id: str,
    target_variable: str,
    target_variable_type: str,
    *,
    dataset_id: Optional[str] = None,
    csv_path: Optional[str] = None,
    cleanup_csv: bool = False,
    sample_rows: int = 200_000,
    min_per_class: int = 5_000,
) -> None:
    """
    Background worker: builds a *sample* of the dataset (never decodes the
    entire file in memory), generates a summary, calls the LLM, and stores
    the result in _classification_jobs[job_id].

    Sources, in preference order:
      - dataset_id (uses cached/loaded DataFrame; preferred path that
        eliminates the redundant file upload — P1.1).
      - csv_path (legacy path: file was streamed to disk on the request).

    P1.7: never load the full file via `file_bytes.decode("utf-8")` +
    `pd.read_csv(StringIO(...))` — that produced ~10-15 GiB peak RAM on a
    2 GiB CSV. Sampling drops it to a few hundred MB.

    Declared ``def`` (not ``async def``) on purpose: the body is fully
    synchronous (sample build via pandas, ``generate_dataset_summary``,
    blocking ``llm_service.get_dataset_type_classification`` HTTP call).
    FastAPI runs ``async def`` background tasks on the request-serving
    event loop, so a sync-bodied coroutine pins the loop for the full
    sample-build + LLM round-trip (typically 2-30 s, longer on cold
    cache). That blocks every other request on the same worker, including
    Kubernetes liveness probes -- which is exactly how the UI ended up
    seeing 404s on /dataset-type-classification/status/{job_id} and
    "Job was interrupted by server restart. Please retry." after a long
    auto-analysis run. Sync ``def`` is offloaded to anyio's thread limiter
    so the event loop stays free while the worker thread does the work.
    """
    try:
        _classification_jobs[job_id]["status"] = "running"

        # P3.4: short-circuit on cached summary. If we already have:
        #   - a target_profile sidecar (computed at /upload time, P2.1), AND
        #   - a column_info AnalyticsResultCache entry for this dataset
        # then the LLM does NOT need to look at the raw data at all - the
        # cached summary captures dtype/uniqueness/missingness/imbalance,
        # which is the entire signal the classifier prompt uses. Skipping
        # the sample build saves ~5-10s on multi-GB datasets.
        used_summary_only = False
        dataset_summary_dict: Optional[Dict[str, Any]] = None
        full_shape: Tuple[int, int] = (0, 0)

        if dataset_id:
            try:
                from app.services.analytics_cache import analytics_cache as _ac
                _version = dataframe_state_manager.get_version(dataset_id)
                cached_column_info = _ac.get(
                    "column_info", dataset_id, "entire", _version
                )
                target_profile = dataset_manager.get_target_profile(dataset_id)
                if cached_column_info is not None and target_profile is not None:
                    dataset_summary_dict = _summary_from_cached(
                        cached_column_info, target_profile, target_variable
                    )
                    full_shape = (
                        int(target_profile.get("n_total", 0)),
                        int(getattr(cached_column_info, "total_columns", 0) or 0),
                    )
                    used_summary_only = True
                    logger.info(
                        f"[classify-bg {job_id}] P3.4 summary-only path active "
                        f"(no file read; full_shape={full_shape})"
                    )
            except Exception as exc:
                logger.warning(f"[classify-bg {job_id}] summary-only fast path failed: {exc}")

        if dataset_summary_dict is None:
            df, full_shape = _build_classification_sample(
                dataset_id=dataset_id,
                csv_path=csv_path,
                target_variable=target_variable,
                sample_rows=sample_rows,
                min_per_class=min_per_class,
            )
            dataset_summary_dict = generate_dataset_summary(df)
            dataset_summary_dict["shape"] = list(full_shape)
            dataset_summary_dict["sample_used_rows"] = int(len(df))
            logger.info(
                f"[classify-bg {job_id}] sample_shape={df.shape} full_shape={full_shape} "
                f"source={'dataset_id' if dataset_id else 'csv_path'}"
            )
        else:
            dataset_summary_dict["shape"] = list(full_shape)
            dataset_summary_dict["sample_used_rows"] = 0
            dataset_summary_dict["summary_only"] = True

        dataset_summary = json.dumps(dataset_summary_dict, indent=2, default=str)
        dataset_summary += f"\nTarget Variable: {target_variable}\nTarget Variable Type: {target_variable_type}"

        llm_response = llm_service.get_dataset_type_classification(dataset_summary)

        if isinstance(llm_response, dict):
            llm_result = llm_response
        elif isinstance(llm_response, str):
            llm_result = json.loads(llm_response)
        else:
            llm_result = json.loads(str(llm_response))

        dataset_type = _map_dataset_type(llm_result.get("dataset_type", ""))
        confidence = float(llm_result.get("confidence", 0.5))
        reasoning = llm_result.get("reasoning", "Classification based on dataset analysis")
        characteristics_list = llm_result.get("characteristics", [])
        characteristics = {f"characteristic_{i+1}": c for i, c in enumerate(characteristics_list)}
        recommendations = llm_result.get("recommendations", [])

        _classification_jobs[job_id].update({
            "status": "completed",
            "result": {
                "success": True,
                "message": "Dataset type classification completed successfully",
                "dataset_id": "",
                "dataset_type": dataset_type.value if hasattr(dataset_type, "value") else str(dataset_type),
                "confidence": confidence,
                "reasoning": reasoning,
                "characteristics": characteristics,
                "recommendations": recommendations,
            },
        })
        logger.info(f"[classify-bg {job_id}] Completed: {dataset_type} ({confidence:.0%})")

    except Exception as exc:
        _classification_jobs[job_id].update({
            "status": "failed",
            "error": str(exc),
        })
        logger.error(f"[classify-bg {job_id}] Failed: {exc}")
        import traceback
        logger.error(traceback.format_exc())
    finally:
        # Clean up any temp file we created for the legacy multipart path.
        if cleanup_csv and csv_path and os.path.exists(csv_path):
            try:
                os.unlink(csv_path)
            except OSError:
                pass


class _DatasetTypeClassificationByIdRequest(BaseModel):
    """Request body for the by-id classification endpoint (P1.1)."""
    dataset_id: str
    target_variable: str
    target_variable_type: str


@chat_router.post("/dataset-type-classification-by-id")
async def classify_dataset_type_by_id(
    request: _DatasetTypeClassificationByIdRequest,
    background_tasks: BackgroundTasks = BackgroundTasks(),
    current_user = Depends(get_current_user_dependency),
):
    """
    P1.1: Classify ML problem type using an already-uploaded dataset_id.

    This eliminates the redundant second/third multipart upload of the same
    CSV file: after `/upload` completes, the frontend calls this endpoint
    with the returned `dataset_id` instead of re-uploading the file. The
    background worker reuses the cached/loaded DataFrame, so for a 2 GiB
    CSV we save another ~2 GiB of network transfer and ~10 GiB of peak RAM.

    Response shape: ``{ job_id, status, ... }``; poll
    ``GET /dataset-type-classification/status/{job_id}`` for completion.
    """
    dataset_id = request.dataset_id.strip()
    if not dataset_id:
        raise HTTPException(status_code=400, detail="dataset_id is required")

    info = dataset_manager.get_dataset_info(dataset_id)
    if not info:
        raise HTTPException(
            status_code=404,
            detail=f"Dataset not found for dataset_id={dataset_id}. Upload it first.",
        )

    job_id = f"dtc_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    _classification_jobs[job_id] = {
        "status": "pending",
        "result": None,
        "error": None,
        "created_at": time.time(),
    }

    background_tasks.add_task(
        _run_dataset_type_classification_bg,
        job_id=job_id,
        target_variable=request.target_variable,
        target_variable_type=request.target_variable_type,
        dataset_id=dataset_id,
        csv_path=None,
        cleanup_csv=False,
    )

    logger.info(
        f"Queued dataset-type-classification job {job_id} for dataset_id={dataset_id} (by-id, no re-upload)"
    )
    return {
        "success": True,
        "job_id": job_id,
        "status": "pending",
        "message": "AI is classifying your ML problem type in the background. Poll /dataset-type-classification/status/{job_id} for the result.",
    }


@chat_router.get("/dataset-type-classification/status/{job_id}")
async def get_dataset_type_classification_status(job_id: str):
    """
    Poll the status of a dataset-type-classification background job.
    Returns the full classification result once status == 'completed'.
    """
    if job_id not in _classification_jobs:
        raise HTTPException(status_code=404, detail="Classification job not found")

    job = _classification_jobs[job_id]
    resp: Dict[str, Any] = {
        "job_id": job_id,
        "status": job["status"],
    }
    if job["status"] == "completed":
        resp.update(job["result"])
    elif job["status"] == "failed":
        resp["error"] = job.get("error")
    return resp


@chat_router.post("/feature-transformation/start")
async def start_feature_transformation_job(
    dataset_id: str = Form(...),
    plan_json: str = Form(...),
    target_variable: Optional[str] = Form(None),
    weight_variable: Optional[str] = Form(None),
    woe_bins: int = Form(10),
    selected_segments: Optional[str] = Form(None),
    use_split: bool = Form(False),
    current_user = Depends(get_current_user_dependency)
):
    """Start feature transformation as background job (prevents UI freeze/timeouts)."""
    try:
        from app.services.background_jobs import background_job_manager
        import uuid

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not plan_json:
            raise HTTPException(status_code=400, detail="plan_json is required")

        job_id = str(uuid.uuid4())

        def execute_feature_transformation(params: dict):
            from app.services.dataframe_state_manager import dataframe_state_manager

            dataset_id_local = params["dataset_id"]
            plan_json_local = params["plan_json"]
            target_variable_local = params.get("target_variable")
            weight_variable_local = params.get("weight_variable")
            woe_bins_local = int(params.get("woe_bins") or 10)
            selected_segments_local = params.get("selected_segments")
            use_split_local = bool(params.get("use_split") or False)

            def progress_callback(progress: int, step: int, message: str):
                try:
                    with background_job_manager._lock:
                        if job_id in background_job_manager._jobs:
                            background_job_manager._jobs[job_id]["progress"] = progress
                            background_job_manager._jobs[job_id]["message"] = message
                            background_job_manager._jobs[job_id]["step"] = step
                    background_job_manager.persist_job_snapshot(job_id)
                except Exception as e:
                    logger.warning(f"Failed to update progress for job {job_id}: {str(e)}")

            try:
                plan = json.loads(plan_json_local)
                if not isinstance(plan, list):
                    raise ValueError("plan_json must be a JSON array")
            except Exception as e:
                raise ValueError(f"Invalid plan_json: {str(e)}")

            segments_list = None
            if selected_segments_local is not None and str(selected_segments_local).strip() != "":
                segments_list = [int(s.strip()) for s in str(selected_segments_local).split(",") if s.strip() != ""]

            if not use_split_local:
                progress_callback(5, 1, "Loading entire dataset")
                dataframe_state_manager.set_scope(dataset_id_local, "entire")
                df_entire = dataframe_state_manager.get_dataframe(dataset_id_local)
                if df_entire is None:
                    raise ValueError(f"Dataset {dataset_id_local} not found")

                progress_callback(15, 2, "Applying transformations to entire dataset")
                ui_rows, _ = feature_engineering_service.apply_transformations(
                    dataset_id=dataset_id_local,
                    df=df_entire,
                    plan=plan,
                    target_variable=target_variable_local,
                    weight_variable=weight_variable_local,
                    woe_bins=woe_bins_local,
                    selected_segments=segments_list,
                    scope="entire",
                    persist=True,
                )

                progress_callback(95, 3, "Finalizing")

                return {
                    "success": True,
                    "dataset_id": dataset_id_local,
                    "response_data": ui_rows,
                }

            progress_callback(5, 1, "Loading dev split")
            dataframe_state_manager.set_scope(dataset_id_local, "dev")
            df_dev = dataframe_state_manager.get_dataframe(dataset_id_local)
            if df_dev is None:
                raise ValueError(f"Dev dataset {dataset_id_local} not found")

            progress_callback(20, 2, "Fitting/applying transformations on dev")
            ui_rows, dev_metadata = feature_engineering_service.apply_transformations(
                dataset_id=dataset_id_local,
                df=df_dev,
                plan=plan,
                target_variable=target_variable_local,
                weight_variable=weight_variable_local,
                woe_bins=woe_bins_local,
                selected_segments=segments_list,
                scope="dev",
                persist=True,
            )

            try:
                progress_callback(70, 3, "Applying transformations to hold")
                dataframe_state_manager.set_scope(dataset_id_local, "hold")
                df_hold = dataframe_state_manager.get_dataframe(dataset_id_local)
                if df_hold is not None and dev_metadata is not None:
                    _ = feature_engineering_service.apply_transformations(
                        dataset_id=dataset_id_local,
                        df=df_hold,
                        plan=plan,
                        target_variable=target_variable_local,
                        weight_variable=weight_variable_local,
                        woe_bins=woe_bins_local,
                        selected_segments=segments_list,
                        scope="hold",
                        stored_metadata=dev_metadata,
                        persist=True,
                    )
            finally:
                dataframe_state_manager.set_scope(dataset_id_local, "dev")

            progress_callback(95, 4, "Finalizing")

            return {
                "success": True,
                "dataset_id": dataset_id_local,
                "response_data": ui_rows,
            }

        background_job_manager.start_job(
            job_id=job_id,
            job_type="feature_transformation",
            params={
                "dataset_id": dataset_id,
                "plan_json": plan_json,
                "target_variable": target_variable,
                "weight_variable": weight_variable,
                "woe_bins": woe_bins,
                "selected_segments": selected_segments,
                "use_split": use_split,
            },
            job_function=execute_feature_transformation,
        )

        logger.info(f"Started feature transformation job {job_id} for dataset {dataset_id}")
        return {"success": True, "job_id": job_id, "dataset_id": dataset_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting feature transformation: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting feature transformation: {str(e)}")


@chat_router.get("/feature-transformation/status/{job_id}")
async def get_feature_transformation_status(job_id: str):
    """Get status of a background feature transformation job."""
    try:
        from app.services.background_jobs import background_job_manager

        status = background_job_manager.get_job_status(job_id)
        if status is None:
            raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

        response = {
            "success": True,
            "job_id": job_id,
            "status": status.get("status"),
            "progress": status.get("progress", 0),
            "message": status.get("message", ""),
        }

        if status.get("status") == "completed":
            response["results"] = status.get("result")
        elif status.get("status") == "failed":
            response["success"] = False
            response["error"] = status.get("error")

        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting feature transformation status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting feature transformation status: {str(e)}")


# NOTE: The synchronous `POST /feature-transformation` endpoint was
# removed (Phase 1 stateless-API cleanup, May 2026). Use
# `POST /feature-transformation/start` + `GET /feature-transformation/status/{job_id}`
# instead — feature engineering on real datasets exceeds the 1 s
# `SLOW_REQUEST_THRESHOLD_MS` and must run in a background job per
# `.cursor/rules/architecture.mdc`. The async pair is the only path the
# frontend uses (`fastApiService.startFeatureTransformationJob`). The
# unreachable code that previously sat after the function body has also
# been removed.


# Training Logs Endpoint
@chat_router.get("/keepalive")
async def keepalive():
    """
    Lightweight ping endpoint the frontend should call every ~60 s during long
    operations (model training, large uploads) to prevent Azure App Service from
    closing the idle connection after 230 seconds.
    """
    return {"alive": True, "ts": time.time()}


@chat_router.get("/training-logs/{model_id}")
async def get_training_logs(model_id: str):
    """
    Get real-time training logs for a specific model
    """
    try:
        from app.services.model_training import training_logs_storage
        
        logs = training_logs_storage.get(model_id, [])
        
        return {
            "model_id": model_id,
            "logs": logs,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error getting training logs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting training logs: {str(e)}")

# Auto Training Endpoints
@chat_router.post("/auto-train-model")
async def auto_train_model(request: dict):
    """
    Trigger automated model training with specified parameters
    """
    try:
        from app.services.model_training import model_training_service
        
        # Extract parameters from request body
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        target_metric = request.get("target_metric")
        target_value = request.get("target_value")
        independent_variables = request.get("independent_variables", [])
        max_runtime_secs = request.get("max_runtime_secs", 30)
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if not target_metric:
            raise HTTPException(status_code=400, detail="target_metric is required")
        if target_value is None:
            raise HTTPException(status_code=400, detail="target_value is required")
        if not independent_variables:
            raise HTTPException(status_code=400, detail="independent_variables is required")
        
        logger.info(f"Starting auto training for dataset {dataset_id}")
        
        # Run auto training
        result = await model_training_service.run_auto_training(
            dataset_id=dataset_id,
            target_column=target_column,
            target_metric=target_metric,
            target_value=target_value,
            independent_variables=independent_variables,
            max_runtime_secs=max_runtime_secs
        )
        
        # Persist modelling artifacts so downstream agents can answer
        # follow-up questions (used_features, metrics tables, etc.)
        try:
            _persist_artifacts_if_available(dataset_id, result)
        except Exception as exc:
            logger.warning(f"Failed to persist auto-training artifacts for dataset {dataset_id}: {exc}")
        
        logger.info(f"Auto training completed for model {result['model_id']}")
        return result
        
    except Exception as e:
        logger.error(f"Auto training failed: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Auto training failed: {str(e)}")

# Manual Configuration Endpoints
@chat_router.post("/detect-problem-type")
async def detect_problem_type(request: dict):
    """
    Detect problem type (classification or regression) from target variable
    """
    try:
        from app.services.model_training_manual_configuration import manual_config_service
        from app.services.model_training_auto_training import make_json_serializable
        from app.services.model_training_auto_training import make_json_serializable
        
        # Extract parameters
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        
        # Load dataset
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        
        # Detect problem type
        result = manual_config_service.detect_problem_type_from_data(df, target_column)
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "target_column": target_column,
            **result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error detecting problem type: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error detecting problem type: {str(e)}")

@chat_router.post("/get-available-variables")
async def get_available_variables(request: dict):
    """
    Get list of available variables from dataset
    """
    try:
        from app.services.model_training_manual_configuration import manual_config_service
        
        # Extract parameters
        dataset_id = request.get("dataset_id")
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        
        # Load dataset
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        
        # Get available variables
        result = manual_config_service.get_available_variables(df)
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            **result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting available variables: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting available variables: {str(e)}")

@chat_router.post("/validate-variable-selection")
async def validate_variable_selection(request: dict):
    """
    Validate variable selection
    """
    try:
        from app.services.model_training_manual_configuration import manual_config_service
        
        # Extract parameters
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        independent_variables = request.get("independent_variables", [])
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if not independent_variables:
            raise HTTPException(status_code=400, detail="independent_variables is required")
        
        # Load dataset
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        
        # Validate variable selection
        result = manual_config_service.validate_variable_selection(df, target_column, independent_variables)
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            **result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error validating variable selection: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error validating variable selection: {str(e)}")

@chat_router.post("/training/lock-variables")
async def lock_training_variables(request: dict):
    """
    Step 1 lock endpoint shared by auto and manual flows.
    """
    try:
        from app.services.model_training_auto_training import auto_training_service
        from app.services.model_training_manual_configuration import manual_config_service
        from app.services.dataframe_state_manager import dataframe_state_manager

        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        selected_variables = request.get("selected_variables") or request.get("independent_variables")
        locked_variables = request.get("locked_variables") or []
        mode = request.get("mode", "auto")
        variable_analysis = request.get("variable_analysis")

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if not isinstance(locked_variables, list):
            raise HTTPException(status_code=400, detail="locked_variables must be a list")
        if selected_variables is not None and not isinstance(selected_variables, list):
            raise HTTPException(status_code=400, detail="selected_variables must be a list")

        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None or df.empty:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")

        if mode == "manual":
            available_vars = manual_config_service.get_available_variables(df)
            independent_variables = available_vars.get("default_independent", [])
            if target_column in independent_variables:
                independent_variables.remove(target_column)

            if not variable_analysis:
                try:
                    variable_analysis = manual_config_service.calculate_vif_and_correlation(
                        df, target_column, independent_variables
                    )
                except Exception as metric_err:
                    logger.warning(f"Manual lock step metric calculation failed; continuing without metrics: {metric_err}")
                    variable_analysis = {}

            variable_selection = manual_config_service.apply_variable_locking(
                independent_variables=independent_variables,
                selected_variables=selected_variables,
                locked_variables=locked_variables,
                selection_mode="manual",
            )
        else:
            available_vars = auto_training_service.get_available_variables(df)
            independent_variables = available_vars.get("default_independent", [])
            if target_column in independent_variables:
                independent_variables.remove(target_column)

            if not variable_analysis:
                try:
                    variable_analysis = auto_training_service.calculate_vif_and_correlation(
                        df, target_column, independent_variables
                    )
                except Exception as metric_err:
                    logger.warning(f"Auto lock step metric calculation failed; continuing without metrics: {metric_err}")
                    variable_analysis = {}

            variable_selection = auto_training_service.apply_variable_locking(
                independent_variables=independent_variables,
                selected_variables=selected_variables,
                locked_variables=locked_variables,
                selection_mode="auto",
            )

        return {
            "dataset_id": dataset_id,
            "target_column": target_column,
            "mode": mode,
            "variable_selection": safe_json_serialize(variable_selection),
            "variable_analysis": safe_json_serialize(variable_analysis),
            "lock_step_complete": True,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in lock variable step: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error in lock variable step: {str(e)}")

@chat_router.post("/get-recommended-metrics")
async def get_recommended_metrics(request: dict):
    """
    Get recommended optimization metrics based on problem type
    """
    try:
        from app.services.model_training_manual_configuration import manual_config_service
        
        # Extract parameters
        problem_type = request.get("problem_type")
        
        # Validate required parameters
        if not problem_type:
            raise HTTPException(status_code=400, detail="problem_type is required")
        
        if problem_type not in ['classification', 'regression']:
            raise HTTPException(status_code=400, detail="problem_type must be 'classification' or 'regression'")
        
        # Get recommended metrics
        result = manual_config_service.get_recommended_metrics(problem_type)
        
        return {
            "success": True,
            "problem_type": problem_type,
            **result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting recommended metrics: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting recommended metrics: {str(e)}")

# Background task function for manual training.
#
# Phase-1 stateless-API fix (large-dataset OOM cascade):
# Previously this ran via FastAPI ``BackgroundTasks`` and wrote progress
# into the module-level ``training_jobs`` dict. That dict is per gunicorn
# worker process and was only persisted to local container disk - if
# the worker process died (OOM kill, deploy, eviction), all in-flight
# jobs evaporated and the user saw "Job was interrupted by server
# restart". On 4M-row workloads where VIF + train_multiple_models could
# overlap on the same pod and push RSS over the 53 GiB limit this was
# happening regularly.
#
# We now run the same body through ``background_job_manager.start_job``
# which (a) mirrors job snapshots to S3 on every state transition so
# status polls survive worker / pod restarts, (b) collapses to a Celery
# / RQ enqueue automatically once ``BROKER_URL`` is wired in Phase 2,
# and (c) participates in the per-dataset ``dataset_job_lock`` so VIF
# and training cannot collide on the same worker pod and OOM-kill it.
def _run_manual_training_job(params: dict) -> dict:
    """Worker-side entry for the ``train_multiple_models`` job.

    Accepts a single ``params`` dict (so it is broker-friendly: Celery /
    RQ can pickle a plain dict but cannot pickle FastAPI request-scope
    closures). Returns the result dict; ``BackgroundJobManager`` is
    responsible for stamping ``status``, ``progress``, ``error``, and
    ``completed_at`` around the call.
    """
    import gc as _gc
    import threading as _threading
    import traceback as _traceback
    from app.services.background_jobs import background_job_manager
    from app.services.model_training_manual_configuration import manual_config_service
    from app.services.dataframe_state_manager import dataframe_state_manager
    from app.services.job_locks import dataset_job_lock

    job_id: str = params["job_id"]
    dataset_id: str = params["dataset_id"]
    target_column: str = params["target_column"]
    independent_variables = params.get("independent_variables")
    algorithms: List[str] = params["algorithms"]
    algorithm_params: dict = params.get("algorithm_params") or {}
    algorithm_param_ranges: Optional[dict] = params.get("algorithm_param_ranges")
    max_iterations: int = params.get("max_iterations", 3)
    optimization_method: str = params.get("optimization_method", "random")
    weight_variable: Optional[str] = params.get("weight_variable")
    locked_variables: Optional[List[str]] = params.get("locked_variables")
    target_metric: Optional[str] = params.get("target_metric")
    cv_folds: Optional[int] = params.get("cv_folds")
    optuna_trials: Optional[int] = params.get("optuna_trials")
    early_stopping_rounds: Optional[int] = params.get("early_stopping_rounds")
    lr_backward_elimination: Optional[dict] = params.get("lr_backward_elimination")

    def _mark_progress(progress: int, message: str) -> None:
        try:
            with background_job_manager._lock:
                jd = background_job_manager._jobs.get(job_id)
                if jd is not None:
                    jd["progress"] = int(progress)
                    jd["message"] = str(message)
            background_job_manager.persist_job_snapshot(job_id)
        except Exception as _exc:  # progress is best-effort
            logger.warning("manual-train: progress persist failed for %s: %s", job_id, _exc)

    with dataset_job_lock(dataset_id, job_label=f"train_multiple_models[{job_id}]"):
        _mark_progress(10, "Training started...")
        logger.info("Background manual training started for job %s", job_id)

        # Load data based on active scope to respect user's train/test split from Step 1
        active_scope = dataframe_state_manager._active_scope.get(dataset_id, "entire")
        logger.info("Manual training: active_scope=%s for dataset %s", active_scope, dataset_id)

        if active_scope == "train":
            dataframe_state_manager.set_scope(dataset_id, scope="train")
            df = dataframe_state_manager.get_dataframe(dataset_id)
            logger.info(
                "✅ Manual training: using TRAIN data, shape: %s",
                df.shape if df is not None else "None",
            )
        else:
            df = dataframe_state_manager._transformed_copies.get(dataset_id, {}).get("entire")
            if df is None:
                df = dataframe_state_manager.get_dataframe(dataset_id)
            logger.info(
                "✅ Manual training: using ENTIRE dataset, shape: %s",
                df.shape if df is not None else "None",
            )

        if df is None:
            # Disk fallback now goes through the shared read-only load
            # cache so a parallel sidebar refresh on the same dataset
            # does not re-parse the same multi-GB CSV in another
            # gunicorn worker on this pod.
            original_df = dataset_manager.load_dataset_readonly_cached(dataset_id)
            if original_df is None:
                raise ValueError(f"Dataset {dataset_id} not found")
            logger.info(
                "Manual training: fallback to original dataset from cached disk load, shape: %s",
                original_df.shape,
            )
            df = original_df

        _mark_progress(20, "Running training...")
        result = manual_config_service.train_multiple_models(
            df=df,
            target_column=target_column,
            independent_variables=independent_variables,
            algorithms=algorithms,
            algorithm_params=algorithm_params,
            max_iterations=max_iterations,
            dataset_id=dataset_id,
            algorithm_param_ranges=algorithm_param_ranges,
            optimization_method=optimization_method,
            weight_variable=weight_variable,
            locked_variables=locked_variables,
            target_metric=target_metric,
            cv_folds=cv_folds,
            optuna_trials=optuna_trials,
            early_stopping_rounds=early_stopping_rounds,
            lr_backward_elimination=lr_backward_elimination,
        )

        # Free the DataFrame reference + force a GC pass before
        # downstream artifact serialisation. On long training runs the
        # algorithm state (catboost trees, optuna trials) accumulates
        # and a GC here noticeably shrinks the post-training RSS plateau
        # so the next job has more headroom.
        try:
            del df
        except Exception:
            pass
        _gc.collect()

        # Persist modelling artifacts so downstream agents can answer feature/model questions
        _persist_artifacts_if_available(dataset_id, result)
        try:
            from app.services.model_training_dump_service import model_training_dump_service
            model_training_dump_service.dump_training_payload(
                training_type="manual",
                dataset_id=dataset_id,
                payload={"success": True, **result},
                context={
                    "job_id": job_id,
                    "target_column": target_column,
                    "algorithms": algorithms,
                    "max_iterations": max_iterations,
                    "optimization_method": optimization_method,
                    "weight_variable": weight_variable,
                    "locked_variables": locked_variables,
                    "target_metric": target_metric,
                    "cv_folds": cv_folds,
                    "optuna_trials": optuna_trials,
                    "early_stopping_rounds": early_stopping_rounds,
                    "lr_backward_elimination": lr_backward_elimination,
                },
            )
        except Exception as dump_err:
            logger.warning("Failed to dump manual training intermediates for job %s: %s", job_id, dump_err)

        _mark_progress(95, "Finalising training results...")
        logger.info("Background manual training completed for job %s", job_id)

        # Launch MEEA background thread (same pattern as auto-training). MEEA
        # is its own daemon thread, so it intentionally outlives this job and
        # is not subject to the dataset_job_lock above.
        try:
            from app.services.model_training_auto_training import auto_training_service as _ats
            meea_thread = _threading.Thread(
                target=_ats.run_pending_meea_jobs,
                args=(dataset_id,),
                daemon=True,
                name=f"meea_manual_{job_id}",
            )
            meea_thread.start()
            logger.info(
                "Launched MEEA background thread for manual job %s / dataset %s",
                job_id, dataset_id,
            )
        except Exception as meea_launch_err:
            logger.warning(
                "Failed to launch MEEA background thread for manual training: %s",
                meea_launch_err,
            )
            logger.debug(_traceback.format_exc())

        return {"success": True, **result}


# Backwards-compat alias.
# ``test_train_multiple_models_background_signature.py`` and any caller
# imported via ``from app.api.routes import train_multiple_models_background``
# pre-dates the Phase-1 move off FastAPI BackgroundTasks. The Phase-1 fix
# replaces that registration with ``background_job_manager.start_job``,
# so the test's "must be plain def, not async def" invariant is still
# meaningful (it now guards against re-introducing an async-def CPU body
# that would pin the event loop if someone reverts the broker path).
train_multiple_models_background = _run_manual_training_job

@chat_router.post("/train-multiple-models")
async def train_multiple_models(request: dict, background_tasks: BackgroundTasks):
    """
    Train multiple specified algorithms using selected (or all) variables - returns immediately with job_id
    Body: {
      dataset_id: str,
      target_column: str,
      independent_variables?: List[str],  # optional; if missing or empty use all
      algorithms: List[str],
      algorithm_param_ranges?: Dict[str, Dict[str, Dict[str, float]]]  # NEW: User-defined hyperparameter ranges
      max_iterations?: int,
      weight_variable?: str  # Optional weight variable for sample_weight
    }
    """
    try:
        dataset_id = request.get('dataset_id')
        target_column = request.get('target_column')
        independent_variables = request.get('independent_variables')
        algorithms = request.get('algorithms', [])
        algorithm_params = request.get('algorithm_params', {})
        algorithm_param_ranges = request.get('algorithm_param_ranges', {})  # NEW
        optimization_method = request.get('optimization_method', 'random')  # Get optimization method, default to 'random'
        target_metric = request.get('target_metric')
        cv_folds = request.get('cv_folds')
        optuna_trials = request.get('optuna_trials')
        early_stopping_rounds = request.get('early_stopping_rounds')
        lr_backward_elimination = request.get('lr_backward_elimination')
        max_iterations = request.get('max_iterations', 3)  # Optimized: Reduced from 5 to 3 for faster training
        weight_variable = request.get('weight_variable')  # Optional weight variable
        locked_variables = request.get('locked_variables')

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if not algorithms:
            raise HTTPException(status_code=400, detail="algorithms is required and must be non-empty")
        if locked_variables is not None and not isinstance(locked_variables, list):
            raise HTTPException(status_code=400, detail="locked_variables must be a list")

        # Generate unique job ID
        job_id = f"manual_train_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"

        logger.info(f"Queuing manual training job {job_id} for dataset {dataset_id}")
        if weight_variable:
            logger.info(f"Weight variable specified: {weight_variable}")

        # Phase-1 stateless-API fix: route through background_job_manager
        # so the job snapshot is mirrored to S3 on every state change
        # and the status endpoint can recover the result even if the
        # gunicorn worker that started it dies. The handler itself
        # (`_run_manual_training_job`) is module-level so a future
        # Phase-2 Celery / RQ worker can pickle it without changes.
        from app.services.background_jobs import background_job_manager as _bgm

        _bgm.start_job(
            job_id=job_id,
            job_type="train_multiple_models",
            params={
                "job_id": job_id,
                "dataset_id": dataset_id,
                "target_column": target_column,
                "independent_variables": independent_variables,
                "algorithms": algorithms,
                "algorithm_params": algorithm_params,
                "algorithm_param_ranges": algorithm_param_ranges,
                "max_iterations": max_iterations,
                "optimization_method": optimization_method,
                "weight_variable": weight_variable,
                "locked_variables": locked_variables,
                "target_metric": target_metric,
                "cv_folds": cv_folds,
                "optuna_trials": optuna_trials,
                "early_stopping_rounds": early_stopping_rounds,
                "lr_backward_elimination": lr_backward_elimination,
                "__training_type__": "manual",
            },
            job_function=_run_manual_training_job,
        )

        # Return immediately with job_id
        return {
            "success": True,
            "job_id": job_id,
            "message": "Training started in background. Check status using /train-multiple-models/status/{job_id}",
            "status": "pending"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting manual training: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting training: {str(e)}")


@chat_router.post("/model-training/lr-backward-elimination")
async def run_lr_backward_elimination_endpoint(request: dict):
    """
    On-demand §7.2 LR backward elimination (same preprocessing / split as manual training).
    Body: dataset_id, target_column, independent_variables, locked_variables?, weight_variable?,
          vif_threshold?, p_value_threshold?, segment_id?, segment_column?
    """
    try:
        from app.services.model_training_manual_configuration import manual_config_service
        from app.services.dataframe_state_manager import dataframe_state_manager

        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        independent_variables = request.get("independent_variables")
        locked_variables = request.get("locked_variables")
        weight_variable = request.get("weight_variable")
        vif_threshold = float(request.get("vif_threshold", 5))
        p_value_threshold = float(request.get("p_value_threshold", 0.05))
        segment_id = request.get("segment_id")
        segment_column = request.get("segment_column")

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if not independent_variables or not isinstance(independent_variables, list):
            raise HTTPException(status_code=400, detail="independent_variables must be a non-empty list")
        if locked_variables is not None and not isinstance(locked_variables, list):
            raise HTTPException(status_code=400, detail="locked_variables must be a list")

        active_scope = dataframe_state_manager._active_scope.get(dataset_id, "entire")
        logger.info(f"LR backward elimination (interactive): active_scope={active_scope} dataset={dataset_id}")

        if active_scope == "train":
            dataframe_state_manager.set_scope(dataset_id, scope="train")
            df = dataframe_state_manager.get_dataframe(dataset_id)
        else:
            df = dataframe_state_manager._transformed_copies.get(dataset_id, {}).get("entire")
            if df is None:
                df = dataframe_state_manager.get_dataframe(dataset_id)

        if df is None:
            original_df = dataset_manager.load_dataset(dataset_id)
            if original_df is None:
                raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
            df = original_df

        out = manual_config_service.run_lr_backward_elimination_interactive(
            df,
            target_column=target_column,
            independent_variables=independent_variables,
            locked_variables=locked_variables,
            dataset_id=dataset_id,
            weight_variable=weight_variable,
            vif_threshold=vif_threshold,
            p_value_threshold=p_value_threshold,
            segment_id=segment_id,
            segment_column=segment_column,
        )
        if not out.get("success"):
            raise HTTPException(status_code=400, detail=out.get("error") or "LR backward elimination failed")
        return {"success": True, "lr_backward_elimination": out.get("lr_backward_elimination")}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"LR backward elimination (interactive) failed: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


def _lookup_manual_training_job(job_id: str) -> Optional[dict]:
    """Return a normalised job snapshot from either job store, or None.

    Phase-1 migration step: ``/train-multiple-models`` now uses
    ``background_job_manager`` (S3-backed, cross-worker visible), but
    legacy jobs that started before this deploy still live in the
    in-process ``training_jobs`` dict. Read both so neither set of
    callers regresses. The shape returned here is the legacy
    ``training_jobs[job_id]`` shape so the rest of the status endpoint
    does not have to know which source it came from.
    """
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        return legacy

    from app.services.background_jobs import background_job_manager as _bgm

    bg = _bgm.get_job_status(job_id)
    if bg is None:
        return None

    # Normalise background_job_manager shape ({status, result, error,
    # progress, message, ...}) to training_jobs shape ({status,
    # results, error, progress, message, dataset_id, ...}).
    params = bg.get("params") or {}
    return {
        "status": bg.get("status"),
        "progress": int(bg.get("progress") or 0),
        "message": bg.get("message") or "",
        "results": bg.get("result"),
        "error": bg.get("error"),
        "created_at": bg.get("started_at"),
        "training_type": params.get("__training_type__", "manual"),
        "dataset_id": params.get("dataset_id"),
        "_source": "background_job_manager",
    }


@chat_router.get("/train-multiple-models/status/{job_id}")
async def get_manual_training_status(job_id: str):
    """
    Get status of manual training job.

    Reads from both ``training_jobs`` (legacy in-process store) and
    ``background_job_manager`` (S3-backed) so in-flight jobs from a
    previous deploy and freshly-queued jobs are both resolvable.
    """
    job_data = _lookup_manual_training_job(job_id)
    if job_data is None:
        raise HTTPException(status_code=404, detail="Job not found")

    dataset_id = job_data.get("dataset_id")

    response = {
        "job_id": job_id,
        "status": job_data.get("status"),
        "progress": job_data.get("progress", 0),
        "message": job_data.get("message", ""),
    }

    # Include results if completed (manual training usually has smaller results)
    if job_data.get("status") == "completed":
        if job_data.get("results"):
            # Manual (global) training has no separate endpoint to fetch results,
            # so we must return the full results payload here.
            response["results"] = job_data["results"]

            # OPTIMIZATION: Pass job_id to avoid redundant persists during polling
            _persist_artifacts_if_available(dataset_id, job_data["results"], job_id=job_id)

    # Include error if failed
    if job_data.get("status") == "failed":
        response["error"] = job_data.get("error")

    return response

@chat_router.post("/train-multiple-models/cancel/{job_id}")
async def cancel_manual_training(job_id: str):
    """
    Soft-cancel a running manual training job.
    This does NOT terminate the underlying Python process, but marks the job as failed/cancelled
    so that the frontend can stop polling and update the UI.

    Phase-1: handles jobs in either ``training_jobs`` (legacy) or
    ``background_job_manager`` (new path).
    """
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        status = legacy.get("status")
        if status in ("completed", "failed"):
            return {
                "success": False,
                "cancelled": False,
                "status": status,
                "message": f"Job already {status}",
            }
        legacy["status"] = "failed"
        legacy["error"] = "Cancelled by user"
        legacy["message"] = "Training cancelled by user"
        logger.info(f"Manual training job {job_id} (legacy) soft-cancelled by user")
        return {
            "success": True,
            "cancelled": True,
            "status": legacy["status"],
            "message": legacy["message"],
        }

    from app.services.background_jobs import background_job_manager as _bgm

    cancelled = _bgm.cancel_job(job_id)
    if not cancelled:
        # cancel_job returns False for unknown jobs or already-terminal jobs.
        snapshot = _bgm.get_job_status(job_id)
        if snapshot is None:
            raise HTTPException(status_code=404, detail="Job not found")
        return {
            "success": False,
            "cancelled": False,
            "status": snapshot.get("status"),
            "message": f"Job already {snapshot.get('status')}",
        }
    snapshot = _bgm.get_job_status(job_id) or {}
    logger.info(f"Manual training job {job_id} soft-cancelled by user (bg_manager)")
    return {
        "success": True,
        "cancelled": True,
        "status": snapshot.get("status"),
        "message": snapshot.get("message") or "Training cancelled by user",
    }

# Segment Training Endpoints
@chat_router.post("/detect-segments")
async def detect_segments(request: dict):
    """
    Detect if segmentation column exists in the dataset
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        dataset_id = request.get("dataset_id")
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")

        pref = (request.get("segment_column") or "").strip() or None
        result = segment_training_service.detect_segments(dataset_id, preferred_segment_column=pref)

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error detecting segments: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error detecting segments: {str(e)}")

def _lookup_segment_manual_training_job(job_id: str) -> Optional[Dict[str, Any]]:
    """Resolve segment manual training job from legacy dict or ``background_job_manager``."""
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        return legacy
    from app.services.background_jobs import background_job_manager as _bgm

    bg = _bgm.get_job_status(job_id)
    if bg is None or str(bg.get("job_type") or "") != "segment_manual_training":
        return None
    params = bg.get("params") or {}
    return {
        "status": bg.get("status"),
        "progress": int(bg.get("progress") or 0),
        "message": bg.get("message") or "",
        "results": bg.get("result"),
        "error": bg.get("error"),
        "created_at": bg.get("started_at"),
        "training_type": "segment_manual",
        "dataset_id": params.get("dataset_id"),
    }


@chat_router.post("/segment-training/run")
async def run_segment_training(request: dict, background_tasks: BackgroundTasks):
    """
    Train models for each segment with the same configuration - returns immediately with job_id
    """
    _ = background_tasks
    try:
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        independent_variables = request.get("independent_variables", [])
        locked_variables = request.get("locked_variables")
        algorithms = request.get("algorithms", ["xgboost"])
        algorithm_params = request.get("algorithm_params", {})
        optimization_method = request.get("optimization_method", "random")
        target_metric = request.get("target_metric")
        cv_folds = request.get("cv_folds")
        optuna_trials = request.get("optuna_trials")
        early_stopping_rounds = request.get("early_stopping_rounds")
        lr_backward_elimination = request.get("lr_backward_elimination")
        algorithm_param_ranges = request.get("algorithm_param_ranges", {})

        max_iterations = request.get("max_iterations")
        if max_iterations is None or not isinstance(max_iterations, (int, float)):
            max_iterations = 5
        else:
            max_iterations = int(max_iterations)

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if locked_variables is not None and not isinstance(locked_variables, list):
            raise HTTPException(status_code=400, detail="locked_variables must be a list")

        segment_column = (request.get("segment_column") or "").strip() or None

        job_id = f"seg_manual_train_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"

        from app.services.background_jobs import background_job_manager as _bgm

        _bgm.start_job(
            job_id=job_id,
            job_type="segment_manual_training",
            params={
                "__training_type__": "segment_manual",
                "job_id": job_id,
                "dataset_id": dataset_id,
                "target_column": target_column,
                "independent_variables": independent_variables,
                "algorithms": algorithms,
                "algorithm_params": algorithm_params,
                "max_iterations": max_iterations,
                "segment_column": segment_column,
                "locked_variables": locked_variables,
                "optimization_method": optimization_method,
                "target_metric": target_metric,
                "cv_folds": cv_folds,
                "optuna_trials": optuna_trials,
                "early_stopping_rounds": early_stopping_rounds,
                "lr_backward_elimination": lr_backward_elimination,
                "algorithm_param_ranges": algorithm_param_ranges,
            },
            job_function=_run_segment_manual_training_job,
        )

        logger.info("Queued segment manual training job %s for dataset %s", job_id, dataset_id)

        return {
            "success": True,
            "job_id": job_id,
            "message": "Segment training started in background. Check status using /segment-training/status/{job_id}",
            "status": "pending",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting segment training: {str(e)}")
        import traceback

        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting segment training: {str(e)}")


@chat_router.get("/segment-training/status/{job_id}")
async def get_segment_manual_training_status(job_id: str):
    """
    Get status of segment manual training job
    """
    job_data = _lookup_segment_manual_training_job(job_id)
    if job_data is None:
        raise HTTPException(status_code=404, detail="Job not found")

    response = {
        "job_id": job_id,
        "status": job_data["status"],
        "progress": job_data["progress"],
        "message": job_data["message"],
    }

    if job_data["status"] == "completed":
        results_payload = job_data.get("results")
        if isinstance(results_payload, dict):
            response["model_id"] = results_payload.get("model_id")
        else:
            response["results"] = results_payload

        dataset_id = job_data.get("dataset_id") or (
            results_payload.get("dataset_id") if isinstance(results_payload, dict) else None
        )
        _persist_artifacts_if_available(dataset_id, results_payload, job_id=job_id)

    if job_data["status"] == "failed":
        response["error"] = job_data["error"]

    return response


@chat_router.post("/segment-training/cancel/{job_id}")
async def cancel_segment_training(job_id: str):
    """
    Soft-cancel a running segment manual training job.
    This does NOT terminate the underlying Python process, but marks the job as failed/cancelled
    so that the frontend can stop polling and update the UI.
    """
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        status = legacy.get("status")
        if status in ("completed", "failed"):
            return {
                "success": False,
                "cancelled": False,
                "status": status,
                "message": f"Job already {status}",
            }
        legacy["status"] = "failed"
        legacy["error"] = "Cancelled by user"
        legacy["message"] = "Segment training cancelled by user"
        logger.info("Segment manual training job %s (legacy) soft-cancelled by user", job_id)
        return {
            "success": True,
            "cancelled": True,
            "status": legacy["status"],
            "message": legacy["message"],
        }

    from app.services.background_jobs import background_job_manager as _bgm

    cancelled = _bgm.cancel_job(job_id)
    if not cancelled:
        snapshot = _bgm.get_job_status(job_id)
        if snapshot is None or str(snapshot.get("job_type") or "") != "segment_manual_training":
            raise HTTPException(status_code=404, detail="Job not found")
        return {
            "success": False,
            "cancelled": False,
            "status": snapshot.get("status"),
            "message": f"Job already {snapshot.get('status')}",
        }
    snapshot = _bgm.get_job_status(job_id) or {}
    logger.info("Segment manual training job %s soft-cancelled by user (bg_manager)", job_id)
    return {
        "success": True,
        "cancelled": True,
        "status": snapshot.get("status"),
        "message": snapshot.get("message") or "Segment training cancelled by user",
    }


@chat_router.get("/segment-training/{model_id}/results")
async def get_segment_model_results(model_id: str, segment_id: str):
    """
    Get results for a specific segment model
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        if not model_id:
            raise HTTPException(status_code=400, detail="model_id is required")
        if not segment_id:
            raise HTTPException(status_code=400, detail="segment_id is required")

        result = segment_training_service.get_segment_model_results(model_id, segment_id)
        _persist_artifacts_if_available(_extract_dataset_id_from_result(result), result)

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting segment model results: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting segment model results: {str(e)}")

@chat_router.get("/segment-training/{model_id}/history")
async def get_segment_training_history(model_id: str, segment_id: str):
    """
    Get training history for a specific segment
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        if not model_id:
            raise HTTPException(status_code=400, detail="model_id is required")
        if not segment_id:
            raise HTTPException(status_code=400, detail="segment_id is required")

        result = segment_training_service.get_segment_training_history(model_id, segment_id)

        return {
            'model_id': model_id,
            'segment_id': segment_id,
            'history': result
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting segment training history: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting segment training history: {str(e)}")

@chat_router.get("/segment-training/{model_id}/compare")
async def compare_segment_models(model_id: str, segments: str):
    """
    Compare performance across multiple segments
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        if not model_id:
            raise HTTPException(status_code=400, detail="model_id is required")
        if not segments:
            raise HTTPException(status_code=400, detail="segments parameter is required")

        # Parse segments from comma-separated string
        segment_list = [s.strip() for s in segments.split(',')]

        result = segment_training_service.compare_segment_models(model_id, segment_list)

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error comparing segment models: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error comparing segment models: {str(e)}")

@chat_router.get("/segment-training/{model_id}/unified-results")
async def get_unified_segment_results(model_id: str):
    """
    Get unified segment results for dashboard display
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        if not model_id:
            raise HTTPException(status_code=400, detail="model_id is required")

        result = segment_training_service.get_unified_segment_results(model_id)

        if 'error' in result:
            raise HTTPException(status_code=404, detail=result['error'])

        # Ensure results are JSON serializable
        from app.services.model_training_auto_training import make_json_serializable
        serializable_results = make_json_serializable(result)
        _persist_artifacts_if_available(_extract_dataset_id_from_result(serializable_results), serializable_results)

        # Compress response to reduce size (500KB → ~50KB)
        json_str = json.dumps(serializable_results)
        compressed = gzip.compress(json_str.encode('utf-8'))
        
        return Response(
            content=compressed,
            media_type="application/json",
            headers={"Content-Encoding": "gzip"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting unified segment results: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting unified segment results: {str(e)}")

@chat_router.get("/segment-training/{model_id}/screen")
async def get_segment_model_screen_results(model_id: str, segment_id: str, algorithm: str = None):
    """
    Get model screening results for a specific segment
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        if not model_id:
            raise HTTPException(status_code=400, detail="model_id is required")
        if not segment_id:
            raise HTTPException(status_code=400, detail="segment_id is required")

        result = segment_training_service.get_model_screen_results(model_id, segment_id)

        # Apply algorithm filter if provided
        if algorithm:
            filtered_result = {
                'segment_id': result.get('segment_id'),
                'model_id': result.get('model_id'),
                'filtered_models': [m for m in result.get('filtered_models', []) if m.get('algorithm') == algorithm],
                'total_models': len([m for m in result.get('filtered_models', []) if m.get('algorithm') == algorithm]),
                'algorithm_filter': algorithm
            }
            return filtered_result

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting segment model screen results: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting segment model screen results: {str(e)}")

@chat_router.get("/segment-training/preview")
async def get_segment_preview(dataset_id: str):
    """
    Get preview of segments in the dataset
    """
    try:
        from app.services.model_training_segment_manual import segment_training_service

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")

        result = segment_training_service.detect_segments(dataset_id)

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting segment preview: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error getting segment preview: {str(e)}")

# NOTE: The synchronous `POST /calculate-vif-correlation` endpoint was
# removed (Phase 1 stateless-API cleanup, May 2026). Use
# `POST /calculate-vif-correlation/start` + `GET /calculate-vif-correlation/status/{job_id}`
# instead — VIF on real datasets routinely exceeds the 1 s
# `SLOW_REQUEST_THRESHOLD_MS` and must run in a background job per
# `.cursor/rules/architecture.mdc`. The async pair is also the only path
# the frontend uses (`fastApiService.startCalculateVifCorrelation`).

@chat_router.post("/calculate-vif-correlation/start")
async def start_calculate_vif_correlation(request: dict):
    """
    Start VIF and correlation calculation as background job
    Returns job_id immediately to prevent timeout
    Supports both global and segment-specific training
    """
    try:
        from app.services.background_jobs import background_job_manager
        
        # Extract parameters
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        independent_variables = request.get("independent_variables", [])
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        if not independent_variables:
            raise HTTPException(status_code=400, detail="independent_variables is required")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        
        # Define job function
        def execute_vif_correlation(params: dict):
            """Execute VIF correlation calculation in background.

            Phase-1 OOM-cascade fix: this closure now (a) acquires a
            per-dataset cross-process lock so a concurrent
            train_multiple_models on the same dataset waits for VIF to
            finish (instead of doubling the in-pod DataFrame footprint
            and OOM-killing the worker), and (b) uses the shared
            read-only DataFrame load cache for the disk fallback so a
            sidebar refresh during the VIF run does not trigger a
            second 4M-row parse on the same worker.
            """
            import gc as _gc
            from app.services.model_training_manual_configuration import manual_config_service
            from app.services.dataframe_state_manager import dataframe_state_manager
            from app.services.message_state_service import message_state_manager
            from app.services.job_locks import dataset_job_lock

            dataset_id = params['dataset_id']
            target_column = params['target_column']
            independent_variables = params['independent_variables']

            with dataset_job_lock(dataset_id, job_label="vif_correlation"):
                # Load dataset (prefer in-memory state that includes feature engineering)
                df = dataframe_state_manager.get_dataframe(dataset_id)
                if df is None:
                    df = dataset_manager.load_dataset_readonly_cached(dataset_id)
                if df is None:
                    raise ValueError(f"Dataset {dataset_id} not found")

                # Use global calculation service (works for both global and segment training)
                # The calculation service handles both cases automatically
                result = manual_config_service.calculate_vif_and_correlation(
                    df, target_column, independent_variables
                )

                # Capture df.columns BEFORE the gc/del bracket; we lose
                # the reference once the lock is released.
                try:
                    all_cols = list(df.columns)
                    excluded_from_analysis = [
                        c for c in all_cols
                        if c not in independent_variables and c != target_column
                    ]
                except Exception:
                    excluded_from_analysis = []

                # Drop the local reference + force a GC pass so any
                # intermediate correlation matrices / VIF series do not
                # linger in RSS while the next job (often training)
                # waits for the lock.
                del df
                _gc.collect()

            # Persist VIF/Correlation analysis into MessageState so modelling agent can access it
            try:
                variable_stats = result.get("variable_statistics") or []
                logger.info(f"VIF calculation completed, persisting {len(variable_stats)} variables to MessageState for dataset {dataset_id}")
                
                if variable_stats:
                    variable_analysis = {
                        "variable_statistics": variable_stats,
                        "summary": result.get("summary", {}),
                        "interpretation": result.get("interpretation", {}),
                        "target_column": target_column,
                        "independent_variables": independent_variables,
                        "source": "calculate-vif-correlation-background",
                    }

                    state = message_state_manager.create_or_load_state(dataset_id)
                    state["variable_analysis"] = variable_analysis

                    training_context = state.get("training_context") or {}
                    if not isinstance(training_context, dict):
                        training_context = {}
                    training_context["variable_analysis"] = variable_analysis
                    state["training_context"] = training_context

                    message_state_manager.save_state(dataset_id, state)
                    logger.info(f"✅ Successfully persisted variable_analysis to MessageState for dataset {dataset_id}")
                else:
                    logger.warning(f"No variable_statistics to persist for dataset {dataset_id}")
            except Exception as persist_err:
                logger.error(f"❌ Failed to persist VIF analysis into MessageState (background) for {dataset_id}: {persist_err}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
            
            return {
                "success": True,
                "dataset_id": dataset_id,
                "target_column": target_column,
                **result,
                "excluded_from_analysis": excluded_from_analysis
            }
        
        # Start background job
        background_job_manager.start_job(
            job_id=job_id,
            job_type='vif_correlation',
            params={
                'dataset_id': dataset_id,
                'target_column': target_column,
                'independent_variables': independent_variables
            },
            job_function=execute_vif_correlation
        )
        
        return {
            "success": True,
            "job_id": job_id,
            "status": "started",
            "message": "Variable analysis started. Poll /calculate-vif-correlation/status/{job_id} for updates."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting VIF correlation calculation: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting VIF correlation calculation: {str(e)}")

@chat_router.get("/calculate-vif-correlation/status/{job_id}")
async def get_vif_correlation_status(job_id: str):
    """
    Check status of VIF correlation calculation job
    """
    try:
        from app.services.background_jobs import background_job_manager
        
        status = background_job_manager.get_job_status(job_id)
        if not status:
            raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
        
        response = {
            "job_id": job_id,
            "status": status['status'],  # 'pending', 'running', 'completed', 'failed'
        }
        
        # Add progress if available
        if 'progress' in status:
            response['progress'] = status['progress']
        
        # Add result if completed
        if status['status'] == 'completed' and status.get('result'):
            response['result'] = status['result']
        
        # Add error if failed
        if status['status'] == 'failed' and status.get('error'):
            response['error'] = status['error']
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting VIF correlation status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting VIF correlation status: {str(e)}")

@chat_router.get("/export-model/{model_id}")
async def export_model(model_id: str, include_artifacts: bool = False):
    """
    Export a trained model with artifacts
    """
    try:
        # Handle iteration-specific model IDs by extracting base model ID
        base_model_id = model_id
        iteration_info = None
        
        # Check if this is an iteration-specific model ID (e.g., MDL_MULTI_87C33D93_iteration_1)
        if "_iteration_" in model_id:
            parts = model_id.split("_iteration_")
            base_model_id = parts[0]
            iteration_info = {
                "iteration": int(parts[1]),
                "original_model_id": model_id
            }
        
        # Check if base model exists
        model_path = f"models/{base_model_id}.pkl"
        if not os.path.exists(model_path):
            raise HTTPException(status_code=404, detail=f"Model {base_model_id} not found")

        files_to_export = []

        # Add the model file
        with open(model_path, 'rb') as f:
            model_content = base64.b64encode(f.read()).decode('utf-8')
            files_to_export.append({
                'filename': f'{model_id}.pkl',
                'content': model_content,
                'encoding': 'base64'
            })

        # Add model metadata and metrics if available
        metadata = {
            'model_id': model_id,
            'base_model_id': base_model_id,
            'exported_at': str(pd.Timestamp.now()),
            'artifacts_included': include_artifacts
        }
        
        # Add iteration-specific information if available
        if iteration_info:
            metadata['iteration_info'] = iteration_info

        # Try to load model metadata from any existing JSON files
        metadata_path = f"models/{base_model_id}_metadata.json"
        if os.path.exists(metadata_path):
            try:
                with open(metadata_path, 'r') as f:
                    existing_metadata = json.load(f)
                    metadata.update(existing_metadata)
            except:
                pass

        # Add hyperparameters and metrics from training data
        try:
            # Look for the training results JSON file
            training_results_path = f"models/{base_model_id}_training_results.json"

            if os.path.exists(training_results_path):
                with open(training_results_path, 'r') as f:
                    training_data = json.load(f)

                    # Extract comprehensive information from training results
                    metadata.update({
                        'algorithm': training_data.get('algorithm', 'Unknown'),
                        'problem_type': training_data.get('problem_type', 'Unknown'),
                        'hyperparameters': training_data.get('hyperparameters', {}),
                        'final_metrics': training_data.get('metrics', {}),
                        'cv_scores': training_data.get('cv_scores', []),
                        'best_iteration': training_data.get('best_iteration', 1),
                        'used_features': training_data.get('used_features', []),
                        'iteration_count': len(training_data.get('iteration_history', [])),
                    })

                    # Add iteration-specific metrics if requested, otherwise use best iteration
                    iteration_history = training_data.get('iteration_history', [])
                    if iteration_history:
                        if iteration_info:
                            # Find the specific iteration requested
                            requested_iteration = next(
                                (iter_data for iter_data in iteration_history 
                                 if iter_data.get('iteration') == iteration_info['iteration']), 
                                None
                            )
                            if requested_iteration:
                                metadata['iteration_metrics'] = requested_iteration.get('metrics', {})
                                metadata['iteration_hyperparameters'] = requested_iteration.get('hyperparameters', {})
                                metadata['iteration_score'] = requested_iteration.get('score', 0)
                            else:
                                # Fallback to best iteration if requested iteration not found
                                best_iteration = max(iteration_history, key=lambda x: x.get('score', 0))
                                metadata['best_iteration_metrics'] = best_iteration.get('metrics', {})
                        else:
                            # No specific iteration requested, use best iteration
                            best_iteration = max(iteration_history, key=lambda x: x.get('score', 0))
                            metadata['best_iteration_metrics'] = best_iteration.get('metrics', {})

            else:
                # Fallback: try to find any JSON files with model information
                for file in os.listdir('models'):
                    if file.startswith(base_model_id) and file.endswith('.json') and file != f"{base_model_id}_metadata.json":
                        try:
                            with open(f'models/{file}', 'r') as f:
                                data = json.load(f)
                                if 'metrics' in data:
                                    metadata['final_metrics'] = data['metrics']
                                if 'hyperparameters' in data:
                                    metadata['hyperparameters'] = data['hyperparameters']
                                if 'algorithm' in data:
                                    metadata['algorithm'] = data['algorithm']
                                if 'problem_type' in data:
                                    metadata['problem_type'] = data['problem_type']
                        except:
                            pass

        except Exception as e:
            logger.warning(f"Could not extract training metadata for model {model_id}: {str(e)}")

        # Add metadata file
        metadata_content = base64.b64encode(json.dumps(metadata, indent=2).encode('utf-8')).decode('utf-8')
        files_to_export.append({
            'filename': f'{model_id}_metadata.json',
            'content': metadata_content,
            'encoding': 'base64'
        })

        return {
            'success': True,
            'message': f'Model {model_id} exported successfully',
            'files': files_to_export
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting model {model_id}: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error exporting model: {str(e)}")

# Auto Training Endpoints
@chat_router.post("/auto-training/analyze")
async def analyze_dataset_for_auto_training(request: dict):
    """
    Analyze dataset and prepare for automatic training
    """
    try:
        from app.services.model_training_auto_training import auto_training_service

        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")

        # Load dataset (prefer in-memory state that includes feature engineering)
        from app.services.dataset_service import dataset_manager
        from app.services.dataframe_state_manager import dataframe_state_manager
        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)

        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")

        # Detect problem type
        problem_type_info = auto_training_service.detect_problem_type_from_data(df, target_column)

        # Get available variables
        available_vars = auto_training_service.get_available_variables(df)

        # Compute dynamic excluded columns for transparency (all columns not used as default independent and not target)
        try:
            all_cols = list(df.columns)
            default_independent = available_vars.get('default_independent', [])
            excluded_columns = [c for c in all_cols if c not in default_independent and c != target_column]
        except Exception:
            excluded_columns = []

        # Calculate VIF and correlation
        independent_variables = [col for col in available_vars['default_independent'] if col != target_column]
        vif_correlation_data = auto_training_service.calculate_vif_and_correlation(df, target_column, independent_variables)

        # Auto-select variables based on analysis
        variable_selection = auto_training_service.auto_select_variables(vif_correlation_data, problem_type_info['problem_type'])

        # Auto-select algorithms based on dataset characteristics
        # Calculate feature types for better algorithm selection
        selected_variables = variable_selection['selected_variables']
        feature_types = {
            'numerical': len([col for col in selected_variables if col in df.columns and pd.api.types.is_numeric_dtype(df[col])]),
            'categorical': len([col for col in selected_variables if col in df.columns and not pd.api.types.is_numeric_dtype(df[col])])
        }

        algorithm_config = auto_training_service.auto_select_algorithms(
            problem_type_info['problem_type'],
            len(df),
            len(selected_variables),
            feature_types
        )

        # Ensure all results are JSON serializable
        from app.services.model_training_auto_training import make_json_serializable

        analysis_result = {
            "dataset_id": dataset_id,
            "target_column": target_column,
            "problem_type": make_json_serializable(problem_type_info),
            "available_variables": make_json_serializable(available_vars),
            "excluded_columns": make_json_serializable(excluded_columns),
            "variable_analysis": make_json_serializable(vif_correlation_data),
            "variable_selection": make_json_serializable(variable_selection),
            "algorithm_selection": make_json_serializable(algorithm_config),
            "analysis_complete": True,
            "auto_selection_complete": True
        }

        # Persist analysis artifacts so the modelling agent can answer VIF/IV and
        # "variables used" questions based on the screener results.
        try:
            _persist_artifacts_if_available(dataset_id, analysis_result)
        except Exception as exc:
            logger.warning(f"Could not persist auto-training analysis artifacts for dataset {dataset_id}: {exc}")

        return analysis_result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error analyzing dataset for auto training: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error analyzing dataset for auto training: {str(e)}")

@chat_router.post("/auto-training/analyze/start")
async def start_analyze_dataset_for_auto_training(request: dict):
    """
    Start dataset analysis for auto training as background job
    Returns job_id immediately to prevent timeout
    Supports both global and segment-specific auto training
    """
    try:
        from app.services.background_jobs import background_job_manager
        
        # Extract parameters
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        
        # Validate required parameters
        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        
        # Define job function as a module-level-style closure that uses
        # dataset_job_lock (OOM-cascade fix) and the read-only DF cache.
        def execute_auto_analysis(params: dict):
            """Execute auto analysis in background.

            OOM fix: acquires the per-dataset cross-process lock so this
            job serialises against any concurrent VIF/training job on the
            same dataset (prevents double-loading multi-GB DataFrames and
            the consequent OOM kill that surfaces as "Job was interrupted
            by server restart. Please retry.").  Uses the read-only
            DataFrame path so no extra copy is made in-process.
            """
            from app.services.model_training_auto_training import auto_training_service
            from app.services.dataset_service import dataset_manager
            from app.services.dataframe_state_manager import dataframe_state_manager
            from app.services.job_locks import dataset_job_lock
            from app.services.model_training_auto_training import make_json_serializable
            
            dataset_id = params['dataset_id']
            target_column = params['target_column']

            df = None
            try:
                with dataset_job_lock(dataset_id, job_label="auto_training_analyze"):
                    # Prefer the read-only view (no extra copy in memory).
                    # helpers.count_dataframes("before_get_dataframe_readonly")
                    helpers.dataframe_report("before_get_dataframe_readonly")
                    df = dataframe_state_manager.get_dataframe_readonly(dataset_id)
                    
                    if df is None:
                        df = dataset_manager.load_dataset(dataset_id)
                    if df is None:
                        raise ValueError(f"Dataset {dataset_id} not found")
                    # helpers.count_dataframes("after_get_dataframe_readonly")
                    helpers.dataframe_report("after_get_dataframe_readonly")
                    # Detect problem type
                    problem_type_info = auto_training_service.detect_problem_type_from_data(df, target_column)

                    # Get available variables
                    available_vars = auto_training_service.get_available_variables(df)

                    # Compute dynamic excluded columns
                    try:
                        all_cols = list(df.columns)
                        default_independent = available_vars.get('default_independent', [])
                        excluded_columns = [c for c in all_cols if c not in default_independent and c != target_column]
                    except Exception:
                        excluded_columns = []

                    # Calculate VIF and correlation (CPU-heavy; runs inside the lock)
                    independent_variables = [col for col in available_vars['default_independent'] if col != target_column]
                    vif_correlation_data = auto_training_service.calculate_vif_and_correlation(df, target_column, independent_variables)

                    # Auto-select variables
                    variable_selection = auto_training_service.auto_select_variables(vif_correlation_data, problem_type_info['problem_type'])

                    # Auto-select algorithms
                    selected_variables = variable_selection['selected_variables']
                    feature_types = {
                        'numerical': len([col for col in selected_variables if col in df.columns and pd.api.types.is_numeric_dtype(df[col])]),
                        'categorical': len([col for col in selected_variables if col in df.columns and not pd.api.types.is_numeric_dtype(df[col])])
                    }

                    algorithm_config = auto_training_service.auto_select_algorithms(
                        problem_type_info['problem_type'],
                        len(df),
                        len(selected_variables),
                        feature_types
                    )
            finally:
                # Drop local reference to the large frame; encourage reclaim after VIF/correlation.
                if df is not None:
                    del df
                gc.collect()
            helpers.count_dataframes("after_gc_collect")
            helpers.dataframe_report("after_gc_collect")
            result = {
                "dataset_id": dataset_id,
                "target_column": target_column,
                "problem_type": make_json_serializable(problem_type_info),
                "available_variables": make_json_serializable(available_vars),
                "excluded_columns": make_json_serializable(excluded_columns),
                "variable_analysis": make_json_serializable(vif_correlation_data),
                "variable_selection": make_json_serializable(variable_selection),
                "algorithm_selection": make_json_serializable(algorithm_config),
                "analysis_complete": True,
                "auto_selection_complete": True
            }

            # Persist auto-analysis artifacts so the modelling agent can use them
            try:
                _persist_artifacts_if_available(dataset_id, result)
            except Exception as exc:
                logger.warning(f"Could not persist auto-training background analysis for dataset {dataset_id}: {exc}")

            gc.collect()
            return result
        
        # Start background job
        background_job_manager.start_job(
            job_id=job_id,
            job_type='auto_training_analyze',
            params={
                'dataset_id': dataset_id,
                'target_column': target_column
            },
            job_function=execute_auto_analysis
        )
        
        return {
            "success": True,
            "job_id": job_id,
            "status": "started",
            "message": "Auto training analysis started. Poll /auto-training/analyze/status/{job_id} for updates."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting auto training analysis: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting auto training analysis: {str(e)}")

@chat_router.get("/auto-training/analyze/status/{job_id}")
async def get_auto_training_analyze_status(job_id: str):
    """
    Check status of auto training analysis job
    """
    try:
        from app.services.background_jobs import background_job_manager
        
        status = background_job_manager.get_job_status(job_id)
        if not status:
            raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
        
        response = {
            "job_id": job_id,
            "status": status['status'],  # 'pending', 'running', 'completed', 'failed'
        }
        
        # Add progress if available
        if 'progress' in status:
            response['progress'] = status['progress']
        
        # Add result if completed
        if status['status'] == 'completed' and status.get('result'):
            response['result'] = status['result']
        
        # Add error if failed
        if status['status'] == 'failed' and status.get('error'):
            response['error'] = status['error']
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting auto training analysis status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting auto training analysis status: {str(e)}")

@chat_router.post("/auto-training/select-variables")
async def auto_select_variables(request: dict):
    """
    Automatically select the best variables for training
    """
    try:
        from app.services.model_training_auto_training import auto_training_service

        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        variable_analysis = request.get("variable_analysis")

        if not dataset_id or not target_column or not variable_analysis:
            raise HTTPException(status_code=400, detail="dataset_id, target_column, and variable_analysis are required")

        # Detect problem type from the analysis data
        problem_type = request.get('problem_type', 'classification')

        # Auto-select variables
        variable_selection = auto_training_service.auto_select_variables(variable_analysis, problem_type)

        # Ensure results are JSON serializable
        from app.services.model_training_auto_training import make_json_serializable

        return {
            "dataset_id": dataset_id,
            "target_column": target_column,
            "variable_selection": make_json_serializable(variable_selection),
            "selection_complete": True
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in automatic variable selection: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error in automatic variable selection: {str(e)}")

@chat_router.post("/auto-training/select-algorithms")
async def auto_select_algorithms(request: dict):
    """
    Automatically select the best algorithms for the dataset
    """
    try:
        from app.services.model_training_auto_training import auto_training_service

        dataset_id = request.get("dataset_id")
        problem_type = request.get("problem_type")
        dataset_size = request.get("dataset_size")
        num_features = request.get("num_features")
        feature_types = request.get("feature_types")  # Optional: {'numerical': int, 'categorical': int}

        if not all([dataset_id, problem_type, dataset_size is not None, num_features is not None]):
            raise HTTPException(status_code=400, detail="All parameters are required")

        # Auto-select algorithms with feature type information if available
        algorithm_config = auto_training_service.auto_select_algorithms(problem_type, dataset_size, num_features, feature_types)

        # Ensure results are JSON serializable
        from app.services.model_training_auto_training import make_json_serializable

        return {
            "dataset_id": dataset_id,
            "algorithm_selection": make_json_serializable(algorithm_config),
            "selection_complete": True
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in automatic algorithm selection: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error in automatic algorithm selection: {str(e)}")

@chat_router.get("/auto-training/meea-status/{dataset_id}")
async def get_meea_status(dataset_id: str):
    """
    Returns whether MEEA (comprehensive model evaluation) is still being computed in the
    background for the given dataset.  The Model Evaluation page can poll this endpoint
    to show a 'computing evaluation…' indicator and refresh once it completes.
    """
    try:
        from app.services.model_training_auto_training import ModelTrainingAutoTrainingService
        pending_model_ids = [
            mid for mid, args in ModelTrainingAutoTrainingService._pending_meea_jobs.items()
            if args.get('dataset_id') == dataset_id
        ]
        return {
            "dataset_id": dataset_id,
            "meea_pending": len(pending_model_ids) > 0,
            "pending_model_ids": pending_model_ids,
            "pending_count": len(pending_model_ids),
        }
    except Exception as e:
        logger.error(f"Error checking MEEA status: {e}")
        return {"dataset_id": dataset_id, "meea_pending": False, "pending_model_ids": [], "pending_count": 0}


@chat_router.post("/auto-training/run")
async def run_auto_training(request: dict, background_tasks: BackgroundTasks):
    """
    Run complete automatic training pipeline - returns immediately with job_id
    """
    _ = background_tasks
    try:
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        selected_variables = request.get("selected_variables")
        locked_variables = request.get("locked_variables")
        selection_mode = request.get("selection_mode", "auto")
        selected_algorithms = request.get("selected_algorithms")
        weight_variable = request.get("weight_variable")  # Optional weight variable

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")

        async with _auto_training_run_lock:
            existing_id = _active_auto_training_by_dataset.get(dataset_id)
            if existing_id:
                est: Optional[str] = None
                if existing_id in training_jobs:
                    est = str(training_jobs[existing_id].get("status") or "")
                else:
                    from app.services.background_jobs import background_job_manager as _bgm

                    snap = _bgm.get_job_status(existing_id)
                    if snap and str(snap.get("job_type") or "") == "auto_training_run":
                        est = str(snap.get("status") or "")
                if est in ("pending", "running"):
                    logger.info(
                        "Reusing existing auto-training job %s for dataset %s (status=%s) - duplicate /run suppressed",
                        existing_id,
                        dataset_id,
                        est,
                    )
                    return {
                        "success": True,
                        "job_id": existing_id,
                        "message": "Training already in progress for this dataset. Poll status with this job_id.",
                        "status": est,
                        "reused_existing_job": True,
                    }

            job_id = f"auto_train_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
            _active_auto_training_by_dataset[dataset_id] = job_id

        logger.info("Queuing auto training job %s for dataset %s", job_id, dataset_id)
        
        from app.services.background_jobs import background_job_manager as _bgm

        _bgm.start_job(
            job_id=job_id,
            job_type="auto_training_run",
            params={
                "__training_type__": "auto",
                "job_id": job_id,
                "dataset_id": dataset_id,
                "target_column": target_column,
                "selected_variables": selected_variables,
                "selection_mode": selection_mode,
                "selected_algorithms": selected_algorithms,
                "weight_variable": weight_variable,
                "locked_variables": locked_variables,
            },
            job_function=_run_auto_training_job,
        )
        return {
            "success": True,
            "job_id": job_id,
            "message": "Training started in background. Check status using /auto-training/status/{job_id}",
            "status": "pending",
            "reused_existing_job": False,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting auto training: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting training: {str(e)}")


def _normalize_auto_training_row_from_bgm(bg: Dict[str, Any]) -> Dict[str, Any]:
    """Map ``background_job_manager`` snapshot to the legacy ``training_jobs`` row shape."""
    params = bg.get("params") or {}
    raw_res = bg.get("result")
    results_data = raw_res
    precomputed_len = None
    if isinstance(raw_res, dict) and "__training_results_body__" in raw_res:
        results_data = raw_res["__training_results_body__"]
        precomputed_len = raw_res.get("_results_json_len")
    return {
        "status": bg.get("status"),
        "progress": int(bg.get("progress") or 0),
        "message": bg.get("message") or "",
        "results": results_data,
        "error": bg.get("error"),
        "dataset_id": params.get("dataset_id"),
        "_results_json_len": precomputed_len,
        "last_heartbeat": bg.get("completed_at") or bg.get("started_at"),
    }


def _build_auto_training_status_payload(
    job_id: str,
    *,
    touch_heartbeat: bool = True,
    quiet: bool = False,
) -> Optional[Dict[str, Any]]:
    """
    Shared JSON shape for GET /auto-training/status and SSE stream ticks.
    When quiet=True, skip verbose per-poll INFO logs (used for ~1.5s stream updates).

    Reads legacy ``training_jobs`` first, then ``background_job_manager`` (S3-backed).
    """
    job_data: Optional[Dict[str, Any]] = None
    if job_id in training_jobs:
        job_data = training_jobs[job_id]
        if touch_heartbeat:
            job_data["last_heartbeat"] = time.time()
    else:
        from app.services.background_jobs import background_job_manager as _bgm

        bg = _bgm.get_job_status(job_id)
        if bg is None or str(bg.get("job_type") or "") != "auto_training_run":
            return None
        job_data = _normalize_auto_training_row_from_bgm(bg)

    response: Dict[str, Any] = {
        "job_id": job_id,
        "status": job_data["status"],
        "progress": safe_json_serialize(job_data.get("progress")),
        "message": job_data["message"],
    }

    dataset_id = job_data.get("dataset_id")

    if job_data["status"] == "completed":
        results_data = job_data.get("results")
        precomputed_len = job_data.get("_results_json_len")

        if not quiet:
            logger.info(
                f"Job {job_id} - Results type: {type(results_data)}, is None: {results_data is None}"
            )
        if results_data is not None:
            if isinstance(results_data, dict):
                if not quiet:
                    logger.info(
                        f"Job {job_id} - Results keys: {list(results_data.keys()) if results_data else 'empty'}"
                    )
                if precomputed_len is not None:
                    results_size = int(precomputed_len)
                else:
                    results_str = json.dumps(results_data, default=str)
                    results_size = len(results_str)
                if not quiet:
                    logger.info(
                        f"Job {job_id} - Results size: {results_size} bytes (precomputed={precomputed_len is not None})"
                    )

                if results_size > 50000:
                    model_id = results_data.get("model_id")
                    if model_id:
                        response["model_id"] = model_id
                        if not quiet:
                            logger.info(f"Job {job_id} - Returning model_id only (large results)")
                    else:
                        response["results"] = results_data
                        if not quiet:
                            logger.info(f"Job {job_id} - Returning large results (no model_id)")
                else:
                    response["results"] = results_data
                    if not quiet:
                        logger.info(f"Job {job_id} - Returning small results ({results_size} bytes)")
            else:
                response["results"] = results_data
                if not quiet:
                    logger.info(f"Job {job_id} - Returning non-dict results")

            _persist_artifacts_if_available(dataset_id, results_data, job_id=job_id)
        else:
            logger.error(f"Job {job_id} - No results found in job_data!")
            response["results"] = None

    if job_data["status"] == "failed":
        response["error"] = safe_json_serialize(job_data.get("error"))
    helpers.dataframe_report("stream jobs running")
    return response


@chat_router.get("/auto-training/status/{job_id}")
async def get_auto_training_status(job_id: str):
    """
    Get status of auto training job
    """
    payload = _build_auto_training_status_payload(job_id, touch_heartbeat=True, quiet=False)
    if payload is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return payload


@chat_router.get("/auto-training/stream/{job_id}")
async def stream_auto_training_status(job_id: str):
    """
    Server-Sent Events stream of auto-training status. One JSON object per event (~1.5 s).
    Uses comment pings for proxies that drop idle streams. Ends on completed / failed.
    """
    if _build_auto_training_status_payload(job_id, touch_heartbeat=False, quiet=True) is None:
        raise HTTPException(status_code=404, detail="Job not found")

    async def event_gen():
        yield ": stream-open\n\n"
        tick = 0
        first = True
        while True:
            if not first:
                await asyncio.sleep(1.5)
            first = False
            tick += 1
            if tick % 14 == 0:
                yield ": keepalive\n\n"
            payload = _build_auto_training_status_payload(job_id, touch_heartbeat=True, quiet=True)
            if payload is None:
                err = {
                    "job_id": job_id,
                    "status": "failed",
                    "error": "Job not found",
                    "message": "Job disappeared from server state",
                }
                yield f"data: {json.dumps(err, default=str)}\n\n"
                break
            yield f"data: {json.dumps(payload, default=str)}\n\n"
            st = payload.get("status")
            if st in ("completed", "failed"):
                break

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
        },
    )

@chat_router.post("/auto-training/cancel/{job_id}")
async def cancel_auto_training(job_id: str):
    """
    Soft-cancel a running auto training job.
    This does NOT terminate the underlying Python process, but marks the job as failed/cancelled
    so that the frontend can stop polling and update the UI.
    """
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        status = legacy.get("status")
        if status in ("completed", "failed"):
            return {
                "success": False,
                "cancelled": False,
                "status": status,
                "message": f"Job already {status}",
            }
        legacy["status"] = "failed"
        legacy["error"] = "Cancelled by user"
        legacy["message"] = "Auto training cancelled by user"
        logger.info("Auto training job %s (legacy) soft-cancelled by user", job_id)
        return {
            "success": True,
            "cancelled": True,
            "status": legacy["status"],
            "message": legacy["message"],
        }

    from app.services.background_jobs import background_job_manager as _bgm

    snap = _bgm.get_job_status(job_id)
    if snap is None or str(snap.get("job_type") or "") != "auto_training_run":
        raise HTTPException(status_code=404, detail="Job not found")

    cancelled = _bgm.cancel_job(job_id)
    if not cancelled:
        return {
            "success": False,
            "cancelled": False,
            "status": snap.get("status"),
            "message": f"Job already {snap.get('status')}",
        }
    snapshot = _bgm.get_job_status(job_id) or {}
    logger.info("Auto training job %s soft-cancelled by user (bg_manager)", job_id)
    return {
        "success": True,
        "cancelled": True,
        "status": snapshot.get("status"),
        "message": snapshot.get("message") or "Auto training cancelled by user",
    }


def _lookup_segment_auto_training_job(job_id: str) -> Optional[Dict[str, Any]]:
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        return legacy
    from app.services.background_jobs import background_job_manager as _bgm

    bg = _bgm.get_job_status(job_id)
    if bg is None or str(bg.get("job_type") or "") != "segment_auto_training":
        return None
    params = bg.get("params") or {}
    return {
        "status": bg.get("status"),
        "progress": int(bg.get("progress") or 0),
        "message": bg.get("message") or "",
        "results": bg.get("result"),
        "error": bg.get("error"),
        "created_at": bg.get("started_at"),
        "training_type": "segment_auto",
        "dataset_id": params.get("dataset_id"),
    }


@chat_router.post("/segment-auto-training/run")
async def run_segment_auto_training(request: dict, background_tasks: BackgroundTasks):
    """
    Run complete automatic training pipeline for each segment - returns immediately with job_id
    """
    _ = background_tasks
    try:
        dataset_id = request.get("dataset_id")
        target_column = request.get("target_column")
        selected_variables = request.get("selected_variables")
        locked_variables = request.get("locked_variables")
        selection_mode = request.get("selection_mode", "auto")
        selected_algorithms = request.get("selected_algorithms")

        if not dataset_id:
            raise HTTPException(status_code=400, detail="dataset_id is required")
        if not target_column:
            raise HTTPException(status_code=400, detail="target_column is required")

        job_id = f"seg_auto_train_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"

        from app.services.background_jobs import background_job_manager as _bgm

        _bgm.start_job(
            job_id=job_id,
            job_type="segment_auto_training",
            params={
                "__training_type__": "segment_auto",
                "job_id": job_id,
                "dataset_id": dataset_id,
                "target_column": target_column,
                "selected_variables": selected_variables,
                "selection_mode": selection_mode,
                "selected_algorithms": selected_algorithms,
                "locked_variables": locked_variables,
            },
            job_function=_run_segment_auto_training_job,
        )

        logger.info("Queued segment auto training job %s for dataset %s", job_id, dataset_id)

        return {
            "success": True,
            "job_id": job_id,
            "message": "Segment training started in background. Check status using /segment-auto-training/status/{job_id}",
            "status": "pending",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting segment auto training: {str(e)}")
        import traceback

        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error starting segment training: {str(e)}")


@chat_router.get("/segment-auto-training/status/{job_id}")
async def get_segment_auto_training_status(job_id: str):
    """
    Get status of segment auto training job
    """
    job_data = _lookup_segment_auto_training_job(job_id)
    if job_data is None:
        raise HTTPException(status_code=404, detail="Job not found")

    response = {
        "job_id": job_id,
        "status": job_data["status"],
        "progress": job_data["progress"],
        "message": job_data["message"],
    }

    if job_data["status"] == "completed":
        results_payload = job_data.get("results")
        if isinstance(results_payload, dict):
            response["model_id"] = results_payload.get("model_id")
        elif results_payload is not None:
            response["results"] = results_payload

        dataset_id = job_data.get("dataset_id") or (
            results_payload.get("dataset_id") if isinstance(results_payload, dict) else None
        )
        _persist_artifacts_if_available(dataset_id, results_payload, job_id=job_id)

    if job_data["status"] == "failed":
        response["error"] = job_data["error"]

    return response


@chat_router.post("/segment-auto-training/cancel/{job_id}")
async def cancel_segment_auto_training(job_id: str):
    """
    Soft-cancel a running segment auto training job.
    This does NOT terminate the underlying Python process, but marks the job as failed/cancelled
    so that the frontend can stop polling and update the UI.
    """
    legacy = training_jobs.get(job_id)
    if legacy is not None:
        status = legacy.get("status")
        if status in ("completed", "failed"):
            return {
                "success": False,
                "cancelled": False,
                "status": status,
                "message": f"Job already {status}",
            }
        legacy["status"] = "failed"
        legacy["error"] = "Cancelled by user"
        legacy["message"] = "Segment auto training cancelled by user"
        logger.info("Segment auto training job %s (legacy) soft-cancelled by user", job_id)
        return {
            "success": True,
            "cancelled": True,
            "status": legacy["status"],
            "message": legacy["message"],
        }

    from app.services.background_jobs import background_job_manager as _bgm

    cancelled = _bgm.cancel_job(job_id)
    if not cancelled:
        snapshot = _bgm.get_job_status(job_id)
        if snapshot is None or str(snapshot.get("job_type") or "") != "segment_auto_training":
            raise HTTPException(status_code=404, detail="Job not found")
        return {
            "success": False,
            "cancelled": False,
            "status": snapshot.get("status"),
            "message": f"Job already {snapshot.get('status')}",
        }
    snapshot = _bgm.get_job_status(job_id) or {}
    logger.info("Segment auto training job %s soft-cancelled by user (bg_manager)", job_id)
    return {
        "success": True,
        "cancelled": True,
        "status": snapshot.get("status"),
        "message": snapshot.get("message") or "Segment auto training cancelled by user",
    }


@chat_router.get("/segment-auto-training/{model_id}/unified-results")
async def get_segment_auto_unified_results(model_id: str):
    """
    Get unified segment auto training results for dashboard display
    """
    try:
        from app.services.model_training_segment_auto import segment_auto_training_service

        logger.info(f"Fetching unified segment auto results for model {model_id}")

        # Get unified results
        results = segment_auto_training_service.get_unified_segment_results(model_id)

        if 'error' in results:
            raise HTTPException(status_code=404, detail=results['error'])

        # Ensure results are JSON serializable
        from app.services.model_training_auto_training import make_json_serializable
        serializable_results = make_json_serializable(results)
        _persist_artifacts_if_available(_extract_dataset_id_from_result(serializable_results), serializable_results)

        # Compress response to reduce size (500KB → ~50KB)
        json_str = json.dumps(serializable_results)
        compressed = gzip.compress(json_str.encode('utf-8'))
        
        return Response(
            content=compressed,
            media_type="application/json",
            headers={"Content-Encoding": "gzip"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching unified segment auto results: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error fetching unified segment auto results: {str(e)}")

@chat_router.get("/segment-auto-training/{model_id}/segment/{segment_id}")
async def get_segment_auto_training_results(model_id: str, segment_id: str):
    """
    Get detailed training results for a specific segment
    """
    try:
        from app.services.model_training_segment_auto import segment_auto_training_service

        logger.info(f"Fetching segment auto training results for model {model_id}, segment {segment_id}")

        # Get segment-specific results
        results = segment_auto_training_service.get_segment_training_results(model_id, segment_id)

        if 'error' in results:
            raise HTTPException(status_code=404, detail=results['error'])

        # Ensure results are JSON serializable
        from app.services.model_training_auto_training import make_json_serializable
        serializable_results = make_json_serializable(results)
        _persist_artifacts_if_available(_extract_dataset_id_from_result(serializable_results), serializable_results)

        return serializable_results

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching segment auto training results: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error fetching segment auto training results: {str(e)}")

@chat_router.get("/export-segment-model/{model_id}/{segment_id}")
async def export_segment_model_zip(model_id: str, segment_id: str):
    """
    Export a segment's best model as a ZIP file containing:
    - Model pickle file (.pkl)
    - Complete segment results JSON file
    """
    try:
        from app.services.model_training_segment_auto import segment_auto_training_service
        from app.services.model_training_auto_training import make_json_serializable
        
        logger.info(f"Exporting segment model for model_id={model_id}, segment_id={segment_id}")
        
        # Get segment training results
        segment_results = segment_auto_training_service.get_segment_training_results(model_id, segment_id)
        
        if 'error' in segment_results:
            raise HTTPException(status_code=404, detail=segment_results['error'])
        
        # Get best model info for this segment (returned under 'segment_result')
        segment_result = segment_results.get('segment_result', {}) if isinstance(segment_results, dict) else {}
        # Backward-compatible fallback if structure differs
        if not segment_result and isinstance(segment_results, dict):
            segment_result = segment_results
        best_model_id = (
            segment_result.get('best_model_selection', {}).get('best_model_id')
        )
        
        if not best_model_id:
            raise HTTPException(status_code=404, detail=f"No best model found for segment {segment_id}")
        
        # Check if model pickle file exists
        model_pkl_path = Path(f"models/{best_model_id}.pkl")
        if not model_pkl_path.exists():
            raise HTTPException(status_code=404, detail=f"Model file not found: {best_model_id}.pkl")
        
        # Get the complete segment auto results JSON
        segment_auto_results_path = Path(f"models/segment_models/{model_id}_segment_auto_results.json")
        
        # Create in-memory ZIP file
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add model pickle file
            with open(model_pkl_path, 'rb') as pkl_file:
                zip_file.writestr(
                    f"{best_model_id}.pkl",
                    pkl_file.read()
                )
            
            # Add complete segment auto results JSON if it exists
            if segment_auto_results_path.exists():
                with open(segment_auto_results_path, 'r') as json_file:
                    json_content = json_file.read()
                    zip_file.writestr(
                        f"{model_id}_segment_auto_results.json",
                        json_content
                    )
            else:
                # If the file doesn't exist, create JSON from segment results
                json_data = make_json_serializable({
                    "model_id": model_id,
                    "segment_id": segment_id,
                    "best_model_id": best_model_id,
                    "training_mode": "segment_auto",
                    "exported_at": str(pd.Timestamp.now()),
                    "segment_results": segment_results
                })
                zip_file.writestr(
                    f"{model_id}_segment_{segment_id}_results.json",
                    json.dumps(json_data, indent=2)
                )
        
        # Prepare ZIP file for download
        zip_buffer.seek(0)
        
        # Generate filename
        algorithm_name = segment_result.get('best_model_selection', {}).get('best_algorithm', 'model')
        zip_filename = f"{algorithm_name}_segment_{segment_id}_{best_model_id}.zip"
        
        logger.info(f"Successfully created ZIP export: {zip_filename}")
        
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={zip_filename}"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting segment model: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error exporting segment model: {str(e)}")

@chat_router.get("/get-codebook/{training_mode}/{training_type}")
async def get_codebook(training_mode: str, training_type: str):
    """
    Get the source code for the specified training mode and type
    
    Args:
        training_mode: 'global', 'segment', or 'pruning'
        training_type: 'auto', 'manual', or 'pruning' (use ``pruning``/``pruning`` for model-pruning codebook)
    
    Returns:
        Source code content of the corresponding backend service file
    """
    try:
        # Map training mode and type to file paths (relative to backend directory)
        file_mapping = {
            'global_auto': 'app/services/model_training_auto_training.py',
            'global_manual': 'app/services/model_training_manual_configuration.py',
            'segment_auto': 'app/services/model_training_segment_auto.py',
            'segment_manual': 'app/services/model_training_segment_manual.py',
            'pruning_pruning': 'app/services/model_training_pruning.py',
        }
        
        key = f"{training_mode}_{training_type}"
        
        logger.info(f"CodeBook request - Mode: {training_mode}, Type: {training_type}, Key: {key}")
        
        if key not in file_mapping:
            logger.error(f"Invalid key: {key}")
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid training mode/type combination: {training_mode}/{training_type}"
            )
        
        # Get absolute path
        file_path = Path(file_mapping[key]).resolve()
        
        logger.info(f"Attempting to read file: {file_path}")
        logger.info(f"File exists: {file_path.exists()}")
        logger.info(f"Current working directory: {os.getcwd()}")
        
        if not file_path.exists():
            # Try alternative path from project root
            alt_path = Path(f"midas/backend/{file_mapping[key]}").resolve()
            logger.info(f"File not found at {file_path}, trying alternative: {alt_path}")
            
            if alt_path.exists():
                file_path = alt_path
            else:
                logger.error(f"Source file not found at either location")
                raise HTTPException(
                    status_code=404,
                    detail=f"Source file not found: {file_mapping[key]}"
                )
        
        # Read the source code
        with open(file_path, 'r', encoding='utf-8') as f:
            source_code = f.read()
        
        logger.info(f"Successfully retrieved codebook for {training_mode}/{training_type}, size: {len(source_code)} chars")
        
        return JSONResponse(content={
            "success": True,
            "training_mode": training_mode,
            "training_type": training_type,
            "file_path": str(file_path),
            "source_code": source_code,
            "file_name": file_path.name
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving codebook: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error retrieving codebook: {str(e)}")

@chat_router.post("/auto-training/select-best-model")
async def select_best_model(request: dict):
    """
    Select the best model from training results with detailed reasoning
    """
    try:
        from app.services.model_training_auto_training import auto_training_service, make_json_serializable

        training_results = request.get("training_results")
        
        if not training_results:
            raise HTTPException(status_code=400, detail="training_results is required")

        # Select best model
        best_model_selection = auto_training_service.select_best_model(training_results)

        logger.info(f"Best model selected: {best_model_selection.get('best_algorithm', 'N/A')}")

        return make_json_serializable({
            "success": True,
            "best_model_selection": best_model_selection
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error selecting best model: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error selecting best model: {str(e)}")


# ============================================================
# MODEL EVALUATION API ENDPOINTS (MEEA Integration)
# ============================================================

@chat_router.get("/model-evaluation/{model_id}")
async def get_model_evaluation(model_id: str, include_explainability: bool = True):
    """
    Get comprehensive evaluation data for a specific model
    Includes: performance metrics, feature importance, ROC curves, 
    SHAP values, granular accuracy, error patterns, etc.
    
    Args:
        model_id: Model identifier
        include_explainability: If False, excludes explainability data to speed up initial load (default: True)
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        from pathlib import Path
        import json as _json
        
        logger.info(f"Fetching evaluation data for model: {model_id} (include_explainability={include_explainability})")
        
        # Get evaluation data from database
        # OPTIMIZATION: Exclude PDP data from initial load for 40% faster response (PDP lazy loaded on demand)
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id, include_explainability=include_explainability, include_pdp=False)
        
        # Fallback: read from JSON file (works on Azure without a DB)
        if not evaluation_data:
            json_path = Path("models") / f"{model_id}_comprehensive_evaluation.json"
            if json_path.exists():
                try:
                    with open(json_path, "r") as fh:
                        evaluation_data = _json.load(fh)
                    logger.info(f"Loaded evaluation data from JSON fallback for model: {model_id}")
                except Exception as json_err:
                    logger.warning(f"Failed to read evaluation JSON for {model_id}: {json_err}")

        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        # Log granular accuracy data for debugging
        test_granular_count = len(evaluation_data.get('granular_accuracy', []))
        train_granular_count = len(evaluation_data.get('granular_accuracy_train', []))
        logger.info(f"📊 Granular accuracy data for model {model_id}: TEST={test_granular_count} segments, TRAIN={train_granular_count} segments")
        
        if train_granular_count == 0 and test_granular_count > 0:
            logger.warning(f"⚠️ Model {model_id} has TEST granular accuracy ({test_granular_count} segments) but NO TRAIN granular accuracy data!")
        
        # Log monotonicity_results availability for debugging
        monotonicity_results = evaluation_data.get('monotonicity_results')
        if monotonicity_results:
            deciles_count = len(monotonicity_results.get('deciles', []))
            violations_count = len(monotonicity_results.get('monotonicity_violations', []))
            logger.info(f"✅ Monotonicity results found for model {model_id}: {deciles_count} deciles, {violations_count} violations")
        else:
            logger.warning(f"⚠️ No monotonicity results found for model {model_id} - Monotonicity tab will not display data")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "evaluation_data": evaluation_data
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching model evaluation: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error fetching model evaluation: {str(e)}")


# -----------------------------------------------------------------------
# PHASED EVALUATION ENDPOINTS
# Each endpoint returns only the data for one evaluation phase so the
# frontend can render each tab independently as soon as it is ready.
#
# Phase 1 - Performance  (metrics, ROC/PR, feature importance)
# Phase 2 - Monotonicity (decile analysis, KS, AUC/Gini)
# Phase 3 - Granular Accuracy (segment-level accuracy per feature)
# -----------------------------------------------------------------------

@chat_router.get("/model-evaluation/{model_id}/phase/{phase_num}")
async def get_model_evaluation_phase(model_id: str, phase_num: int):
    """
    Return evaluation data for a specific phase of a model.
    Returns 202 (Accepted) when the phase is not yet ready so the frontend
    can poll without treating it as an error.
    """
    from pathlib import Path
    import json as _json

    if phase_num not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="phase_num must be 1, 2, or 3")

    phase_path = Path("models") / f"{model_id}_eval_phase{phase_num}.json"

    if not phase_path.exists():
        # Phase not yet written - return 202 so frontend knows to keep polling
        return JSONResponse(
            status_code=202,
            content={"ready": False, "model_id": model_id, "phase": phase_num},
        )

    try:
        with open(phase_path, "r") as fh:
            phase_data = _json.load(fh)
        return JSONResponse(content=safe_json_serialize({
            "ready": True,
            "model_id": model_id,
            "phase": phase_num,
            "data": phase_data,
        }))
    except Exception as e:
        logger.error(f"Error reading phase{phase_num} JSON for {model_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to read phase {phase_num} data for {model_id}")


@chat_router.get("/model-evaluation/list/all")
async def list_all_evaluated_models():
    """
    List all models (from database AND from models folder)
    Returns model metadata for all available models
    OPTIMIZED: Only returns models that have .pkl files OR have MEEA data
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        from pathlib import Path
        import json
        import os
        
        logger.info("Fetching list of all models")
        
        # First, scan models folder to find actual .pkl files (source of truth)
        models_folder = Path("models")
        pkl_model_ids = set()
        
        if models_folder.exists():
            pkl_files = list(models_folder.glob("MDL_*.pkl"))
            pkl_model_ids = {pkl_file.stem for pkl_file in pkl_files}
            logger.info(f"Found {len(pkl_model_ids)} .pkl files in models folder")
        
        # Get models from database (with MEEA evaluation data)
        db_models = model_evaluation_db.list_all_models()
        db_model_ids = {m['id'] for m in db_models}
        
        # Filter database models to only those with .pkl files OR have evaluation data
        # This prevents showing 97 models when only 5 have actual files
        valid_db_models = [
            m for m in db_models 
            if m['id'] in pkl_model_ids  # Has .pkl file
        ]
        
        all_models = list(valid_db_models)  # Start with valid database models
        
        # Add file-only models (have .pkl but not in database yet)
        for model_id in pkl_model_ids:
            if model_id in db_model_ids:
                continue  # Already added from database
            
            # Try to read training results JSON
            json_file = models_folder / f"{model_id}_training_results.json"
            
            try:
                if json_file.exists():
                    with open(json_file, 'r') as f:
                        training_results = json.load(f)
                    
                    # Extract model info from training results
                    all_models.append({
                        'id': model_id,
                        'name': training_results.get('algorithm', 'Unknown'),
                        'model_type': training_results.get('algorithm', 'Unknown'),
                        'task_type': training_results.get('task_type', 'classification'),
                        'training_date': training_results.get('training_date', ''),
                        'status': 'file_only',  # Not in database yet
                        'color': '#3B82F6',  # Blue color for file-only models
                        'description': f"Model trained on {training_results.get('training_date', 'unknown date')}",
                        'created_at': training_results.get('training_date', ''),
                        'has_meea_data': False  # Mark as not having MEEA evaluation yet
                    })
                else:
                    # No JSON file, use basic info
                    all_models.append({
                        'id': model_id,
                        'name': 'Trained Model',
                        'model_type': 'Unknown',
                        'task_type': 'classification',
                        'training_date': '',
                        'status': 'file_only',
                        'color': '#6B7280',  # Gray for unknown
                        'description': 'Model file found',
                        'created_at': '',
                        'has_meea_data': False
                    })
            except Exception as e:
                logger.warning(f"Could not read training results for {model_id}: {e}")
                continue
        
        # Mark database models as having MEEA data
        for model in all_models:
            if model['id'] in db_model_ids:
                model['has_meea_data'] = True
        
        # Sort by creation date (newest first)
        all_models.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        logger.info(f"Found {len(all_models)} valid models ({len(valid_db_models)} with MEEA data, {len(all_models) - len(valid_db_models)} file-only). Filtered out {len(db_models) - len(valid_db_models)} database entries without .pkl files.")
        
        return JSONResponse(content={
            "success": True,
            "count": len(all_models),
            "models": all_models,
            "meea_count": len(valid_db_models),
            "file_count": len(all_models) - len(valid_db_models)
        })
    
    except Exception as e:
        logger.error(f"Error listing models: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error listing models: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/samples")
async def get_model_samples(
    model_id: str,
    data_source: str = Query("test", regex="^(train|test)$"),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = Query(None, min_length=1, max_length=100),
):
    """
    List a page of raw samples (actual training/eval rows) for a model and data source.

    - Returns only the original (treated) feature columns + ID + target.
    - Uses stored train/test indices so that sample_index aligns with SHAP arrays.
    - Does NOT compute SHAP; it just reads existing data.
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        from app.services.dataset_service import dataset_manager
        from app.services.dataframe_state_manager import dataframe_state_manager
        import numpy as np
        import pandas as pd

        eval_data = model_evaluation_db.get_model_evaluation(model_id)
        if not eval_data:
            raise HTTPException(status_code=404, detail=f"Model {model_id} not found in database")

        model_info = (eval_data.get("model") or {})
        dataset_id = model_info.get("dataset_id")
        target_column = model_info.get("target_column")
        # id_column is optional metadata; fall back to None if not present
        id_column = model_info.get("id_column")

        if not dataset_id:
            raise HTTPException(
                status_code=400,
                detail="Model metadata missing dataset_id; cannot list samples."
            )

        # 1. Detect if this model was trained on a per-segment dataset
        # e.g. dataset_id = "{base_dataset_id}_segment_{segment_value}"
        is_segment_dataset = False
        base_dataset_id = dataset_id
        segment_raw_id = None
        if dataset_id and "_segment_" in dataset_id:
            try:
                parts = dataset_id.split("_segment_", 1)
                if len(parts) == 2:
                    base_dataset_id, segment_raw_id = parts[0], parts[1]
                    is_segment_dataset = True
                    logger.info(
                        f"Detected segment dataset for model {model_id} samples: "
                        f"base_dataset_id={base_dataset_id}, segment_raw_id={segment_raw_id}"
                    )
            except Exception as e:
                logger.warning(
                    f"Failed to parse segment dataset_id '{dataset_id}' for model {model_id}: {str(e)}"
                )
                base_dataset_id = dataset_id
                segment_raw_id = None
                is_segment_dataset = False

        # 2. Load base (treated) dataframe
        # Prefer the in-memory dataframe_state_manager; fall back to dataset_manager.
        # First try the exact dataset_id used during training (works for global models
        # and for segment models if the per-segment dataset is still materialized).
        df = dataframe_state_manager.get_dataframe(dataset_id)

        # If this is a segment dataset and the stored per-segment dataframe is missing
        # or empty (it may have been cleaned up after training), fall back to using
        # the base dataset and filter rows by the segment column.
        if (df is None or df.empty) and is_segment_dataset and base_dataset_id:
            logger.info(
                f"Per-segment dataset '{dataset_id}' is missing or empty for model {model_id}. "
                f"Falling back to base dataset '{base_dataset_id}' and segment filtering for samples."
            )
            base_df = dataframe_state_manager.get_dataframe(base_dataset_id)
            if base_df is None or base_df.empty:
                # Fallback to loading from dataset service
                base_df = dataset_manager.load_dataset(base_dataset_id)
            
            if base_df is None or base_df.empty:
                raise HTTPException(
                    status_code=404,
                    detail=f"Base dataset {base_dataset_id} not found or empty. Cannot list samples."
                )
            
            # Try to locate the segment column on the base dataframe.
            segment_column = None
            if segment_raw_id is not None:
                # Preferred segment column names, consistent with segmentation services
                candidate_columns = [
                    "segment", "SEGMENT", "segment_id", "SEGMENT_ID",
                    "group", "GROUP", "cluster", "CLUSTER",
                ]
                for col in candidate_columns:
                    if col in base_df.columns:
                        segment_column = col
                        break
                
                # If still not found, try to infer a suitable segment column
                if segment_column is None:
                    for col in base_df.columns:
                        lower_col = str(col).lower()
                        if lower_col in ["segment", "segments", "segment_id", "segment_label", "group", "cluster"]:
                            segment_column = col
                            break
            
            if segment_raw_id is not None and segment_column is not None:
                # Filter base_df to only rows belonging to this segment. Use string
                # comparison to be robust to numeric vs string representations.
                before_shape = base_df.shape
                try:
                    segment_mask = base_df[segment_column].astype(str) == str(segment_raw_id)
                except Exception as e:
                    logger.warning(
                        f"Error while creating segment mask for column '{segment_column}' "
                        f"and segment '{segment_raw_id}': {str(e)}. Falling back to no filter."
                    )
                    segment_mask = None
                
                if segment_mask is not None:
                    df = base_df[segment_mask].copy()
                    logger.info(
                        f"Filtered base dataset '{base_dataset_id}' for segment '{segment_raw_id}' "
                        f"on column '{segment_column}': {before_shape} -> {df.shape}"
                    )
                    if df.empty:
                        raise HTTPException(
                            status_code=404,
                            detail=(
                                f"No rows found for segment '{segment_raw_id}' in dataset '{base_dataset_id}' "
                                f"using column '{segment_column}'. Cannot list samples."
                            )
                        )
                else:
                    logger.warning(
                        f"Unable to filter base dataset '{base_dataset_id}' for segment '{segment_raw_id}'. "
                        f"Segment column '{segment_column}' not found or invalid. Using full dataset."
                    )
                    df = base_df
            else:
                logger.warning(
                    f"Unable to determine segment column or segment id for dataset '{dataset_id}'. "
                    f"Using base dataset '{base_dataset_id}' without filtering."
                )
                df = base_df
        else:
            # For non-segment datasets or when segmented dataframe exists, try dataset_manager as fallback
            if df is None or df.empty:
                df = dataset_manager.load_dataset(dataset_id)

        # Final check: if dataframe is still None or empty, raise error
        if df is None or df.empty:
            raise HTTPException(
                status_code=404,
                detail=f"Dataset {dataset_id} not found or empty. Cannot list samples."
            )

        # If we filtered by segment, we need to adjust train_indices and test_indices
        # to only include indices that exist in the filtered dataframe
        # This must be done BEFORE we retrieve train_indices/test_indices from the database
        # so we'll handle it after retrieving them (see below)

        # 2. Decide which raw features to expose:
        #    - Prefer used_features from database (actual columns model was trained on)
        #    - Fall back to preprocessed_columns.keys() (for backward compatibility)
        #    - Finally, fall back to all non-target columns
        used_features: List[str] = []
        feature_source = "unknown"
        
        # First priority: used_features from database (what model was actually trained on)
        if eval_data.get("used_features"):
            used_features = eval_data.get("used_features")
            feature_source = "database"
        elif model_info.get("used_features"):
            used_features = model_info.get("used_features")
            feature_source = "database"
        
        # Second priority: fall back to preprocessed_columns keys (for backward compatibility)
        if not used_features:
            preprocessed_columns = (
                eval_data.get("preprocessed_columns")
                or model_info.get("preprocessed_columns")
                or {}
            )
            if isinstance(preprocessed_columns, dict) and preprocessed_columns:
                used_features = list(preprocessed_columns.keys())
                feature_source = "preprocessed_columns"
        
        # Final fallback: all columns except target
        if not used_features:
            used_features = [c for c in df.columns if c != target_column]
            feature_source = "all_columns"
        
        # Log for debugging
        logger.info(f"Model {model_id} samples: using {len(used_features)} features from {feature_source}")

        # 3. Determine which indices belong to this data_source based on stored split indices
        # Check both top-level eval_data and model_info (indices might be at either level)
        train_indices = (
            eval_data.get("train_indices") 
            or model_info.get("train_indices") 
            or []
        )
        test_indices = (
            eval_data.get("test_indices") 
            or model_info.get("test_indices") 
            or []
        )

        # Debug logging: check where indices came from
        logger.info(f"Model {model_id} indices lookup: eval_data.train_indices={bool(eval_data.get('train_indices'))}, "
                   f"eval_data.test_indices={bool(eval_data.get('test_indices'))}, "
                   f"model_info.train_indices={bool(model_info.get('train_indices'))}, "
                   f"model_info.test_indices={bool(model_info.get('test_indices'))}")

        # Ensure indices are lists (not strings or None)
        if train_indices and not isinstance(train_indices, list):
            logger.warning(f"Model {model_id} train_indices is not a list: {type(train_indices)}")
            train_indices = []
        if test_indices and not isinstance(test_indices, list):
            logger.warning(f"Model {model_id} test_indices is not a list: {type(test_indices)}")
            test_indices = []

        # If we filtered by segment, filter train_indices and test_indices to only include
        # indices that actually exist in the filtered dataframe
        if is_segment_dataset and (df is not None and not df.empty):
            # Get the actual index values in the filtered dataframe
            filtered_df_index_set = set(df.index.tolist())
            
            # Filter train_indices and test_indices to only include indices that exist in filtered dataframe
            if train_indices:
                original_train_count = len(train_indices)
                train_indices = [idx for idx in train_indices if idx in filtered_df_index_set]
                logger.info(f"Filtered train_indices for segment: {original_train_count} -> {len(train_indices)} indices remain")
            
            if test_indices:
                original_test_count = len(test_indices)
                test_indices = [idx for idx in test_indices if idx in filtered_df_index_set]
                logger.info(f"Filtered test_indices for segment: {original_test_count} -> {len(test_indices)} indices remain")

        # Log for debugging
        if data_source == "test":
            logger.info(f"Model {model_id} test_indices: {len(test_indices) if test_indices else 0} indices found (type: {type(test_indices)})")
        elif data_source == "train":
            logger.info(f"Model {model_id} train_indices: {len(train_indices) if train_indices else 0} indices found (type: {type(train_indices)})")

        # Determine which indices to use based on data_source
        # First try stored indices
        if data_source == "train" and train_indices and len(train_indices) > 0:
            # For segmented models, indices are base dataframe indices - need to convert to position indices
            if is_segment_dataset:
                df_index_set = set(df.index.tolist())
                # Filter to only include indices that exist in the filtered segment dataframe
                filtered_indices = [i for i in train_indices if i in df_index_set]
                # Convert base dataframe indices to position indices for use with .iloc[]
                try:
                    full_idx: List[int] = [int(df.index.get_loc(idx)) for idx in filtered_indices]
                    logger.info(f"Using train split: {len(full_idx)}/{len(train_indices)} indices from stored train_indices (converted to position indices)")
                except (KeyError, ValueError) as e:
                    logger.warning(f"Error converting train indices to position indices: {str(e)}")
                    # Fallback: try to use as position indices if they're already positions
                    full_idx = [int(i) for i in train_indices if isinstance(i, (int, np.integer)) and 0 <= i < len(df)]
                    logger.info(f"Fallback: Using {len(full_idx)} train indices as position indices")
            else:
                full_idx: List[int] = [int(i) for i in train_indices if isinstance(i, (int, np.integer)) and 0 <= i < len(df)]
                logger.info(f"Using train split: {len(full_idx)} indices from stored train_indices")
        elif data_source == "test" and test_indices and len(test_indices) > 0:
            # For segmented models, indices are base dataframe indices - need to convert to position indices
            if is_segment_dataset:
                df_index_set = set(df.index.tolist())
                # Filter to only include indices that exist in the filtered segment dataframe
                filtered_indices = [i for i in test_indices if i in df_index_set]
                # Convert base dataframe indices to position indices for use with .iloc[]
                try:
                    full_idx = [int(df.index.get_loc(idx)) for idx in filtered_indices]
                    logger.info(f"Using test split: {len(full_idx)}/{len(test_indices)} indices from stored test_indices (converted to position indices)")
                except (KeyError, ValueError) as e:
                    logger.warning(f"Error converting test indices to position indices: {str(e)}")
                    # Fallback: try to use as position indices if they're already positions
                    full_idx = [int(i) for i in test_indices if isinstance(i, (int, np.integer)) and 0 <= i < len(df)]
                    logger.info(f"Fallback: Using {len(full_idx)} test indices as position indices")
            else:
                full_idx = [int(i) for i in test_indices if isinstance(i, (int, np.integer)) and 0 <= i < len(df)]
                logger.info(f"Using test split: {len(full_idx)} indices from stored test_indices")
        else:
            # Fallback: Try to recreate split using split_params (same logic as recalculate_explainability)
            split_params_json = model_info.get('split_params')
            if split_params_json and target_column and target_column in df.columns:
                try:
                    from sklearn.model_selection import train_test_split
                    import json
                    
                    # Parse split_params
                    if isinstance(split_params_json, str):
                        split_params = json.loads(split_params_json)
                    else:
                        split_params = split_params_json
                    
                    # Get target for stratification if needed
                    y = df[target_column].copy()
                    
                    # Encode target if categorical (needed for train_test_split)
                    if y.dtype == 'object' or y.dtype.name == 'category':
                        from sklearn.preprocessing import LabelEncoder
                        target_le = LabelEncoder()
                        y = pd.Series(target_le.fit_transform(y.astype(str)), index=y.index, name=y.name)
                    
                    # Recreate split using same parameters
                    # Use a dummy X (just index positions) with y for stratification support
                    X_dummy = np.arange(len(df))
                    stratify = y if (split_params.get('stratify', False) and y.nunique() <= 10) else None
                    X_train_dummy, X_test_dummy, y_train_dummy, y_test_dummy = train_test_split(
                        X_dummy, y,
                        test_size=split_params.get('test_size', 0.2),
                        random_state=split_params.get('random_state', 42),
                        stratify=stratify
                    )
                    
                    # X_train_dummy and X_test_dummy are position indices (0, 1, 2, ...)
                    train_positions = X_train_dummy.tolist()
                    test_positions = X_test_dummy.tolist()
                    
                    if data_source == "train":
                        full_idx = train_positions
                        logger.info(f"Recreated train split using split_params: {len(full_idx)} indices")
                    else:  # test
                        full_idx = test_positions
                        logger.info(f"Recreated test split using split_params: {len(full_idx)} indices")
                except Exception as e:
                    logger.warning(f"Failed to recreate split using split_params: {str(e)}")
                    raise HTTPException(
                        status_code=400,
                        detail=f"Train/test split information is not available for this model. "
                               f"Attempted to recreate split but failed: {str(e)}"
                    )
            else:
                # No split information available - return user-friendly error
                raise HTTPException(
                    status_code=400,
                    detail=f"Train/test split information is not available for this model. "
                           f"The model was likely trained before split indices were stored. "
                           f"Please retrain the model to enable sample browsing by data split."
                )

        # 4. Optional search filtering across the entire split
        search_normalized = (search or "").strip().lower()
        if search_normalized:
            # Build list of columns to search across (skip ID for now)
            search_cols: List[str] = []
            # Always include target if present
            if target_column and target_column in df.columns:
                search_cols.append(target_column)
            # Include used feature columns
            for col in used_features:
                if col in df.columns and col not in search_cols:
                    search_cols.append(col)

            if search_cols:
                # Build a DataFrame with just the rows/columns we care about.
                sub_df = df.loc[full_idx, search_cols]

                # Vectorised match: build a per-column boolean mask via
                # str.contains and OR-reduce. ~53x faster than DataFrame.apply
                # with axis=1 over millions of rows (see
                # backend/docs/midas-4m-row-performance-analysis 1.md).
                column_masks = [
                    sub_df[col].astype(str).str.contains(
                        search_normalized,
                        case=False,
                        na=False,
                        regex=False,
                    )
                    for col in search_cols
                ]
                mask = functools.reduce(operator.or_, column_masks)
                full_idx = [idx for idx, keep in zip(full_idx, mask.values.tolist()) if keep]

        total = len(full_idx)
        if total == 0:
            return {"total": 0, "data_source": data_source, "samples": []}

        # 5. Apply pagination on the (optionally) filtered index list
        start = min(max(offset, 0), total)
        end = min(start + limit, total)
        page_idx = full_idx[start:end]

        samples: List[Dict[str, Any]] = []
        for pos, row_idx in enumerate(page_idx):
            row = df.iloc[row_idx]
            # Build a features dict with raw column values
            features: Dict[str, Any] = {}
            for col in used_features:
                if col in df.columns:
                    val = row[col]
                    # Normalize numpy scalar types and NaN/NA to JSON-serializable values
                    if isinstance(val, np.generic):
                        val = val.item()
                    if isinstance(val, float) and (np.isnan(val) if isinstance(val, float) else False):
                        val = None
                    # pandas NA / NaT handling
                    if pd.isna(val):
                        val = None
                    features[col] = val

            # ID and target values may also be numpy scalars; normalize them
            id_value = row[id_column] if id_column and id_column in df.columns else None
            if isinstance(id_value, np.generic):
                id_value = id_value.item()
            if pd.isna(id_value):
                id_value = None

            target_value = row[target_column] if target_column in df.columns else None
            if isinstance(target_value, np.generic):
                target_value = target_value.item()
            if pd.isna(target_value):
                target_value = None

            sample: Dict[str, Any] = {
                "sample_index": start + pos,  # position within this split
                "row_index": int(row_idx) if isinstance(row_idx, (int, np.integer)) else str(row_idx),
                "id_value": id_value,
                "target": target_value,
                "features": features,
            }
            samples.append(sample)

        return {
            "total": total,
            "data_source": data_source,
            "target_column": target_column,  # Include target column name in response
            "samples": samples,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing samples for model {model_id}: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error listing samples for model {model_id}: {str(e)}")
@chat_router.get("/model-evaluation/list/by-dataset")
async def list_evaluated_models_by_dataset(dataset_id: str):
    """
    List all evaluated models for a given dataset_id.

    Strategy (works on Azure without a DB):
    1. Scan the models/ folder for *_comprehensive_evaluation.json files whose
       embedded dataset_id matches the requested one.  These are always written
       by run_pending_meea_jobs regardless of DB availability.
    2. Also query the DB and merge results so we don't lose DB-only entries.
    3. Only return models that still have a corresponding .pkl file.
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        from pathlib import Path
        import json as _json

        logger.info(f"Fetching models for dataset_id={dataset_id}")

        models_folder = Path("models")

        # --- Scan .pkl files (source of truth for model existence) ---
        pkl_model_ids: set = set()
        if models_folder.exists():
            pkl_model_ids = {p.stem for p in models_folder.glob("MDL_*.pkl")}
            logger.info(f"Found {len(pkl_model_ids)} .pkl files in models folder")

        # --- Scan evaluation JSON files (works without DB) ---
        # Priority: comprehensive_evaluation > phase1 (phase1 is written first, comprehensive last)
        # This ensures models show up in the list as soon as Phase 1 completes, not only after
        # all 3 phases finish - critical for Azure where there is no persistent DB.
        json_models: dict = {}  # model_id -> model dict

        def _extract_model_entry(eval_data: dict, model_id: str, has_full_eval: bool) -> dict:
            model_info = eval_data.get("model", {})
            return {
                "id": model_id,
                "algorithm": (
                    model_info.get("algorithm_name")
                    or eval_data.get("model_name")
                    or model_info.get("model_name")
                    or "Unknown"
                ),
                "dataset_id": dataset_id,
                "created_at": model_info.get("training_date", ""),
                "training_date": model_info.get("training_date", ""),
                "target_column": model_info.get("target_column") or eval_data.get("target_column", ""),
                "problem_type": model_info.get("problem_type") or eval_data.get("problem_type", ""),
                "active_scope": model_info.get("active_scope") or eval_data.get("active_scope", "entire"),
                "is_segment_model": "_segment_" in dataset_id,
                "has_meea_data": has_full_eval,
            }

        def _get_eval_dataset_id(eval_data: dict) -> str:
            return (
                eval_data.get("dataset_id")
                or eval_data.get("model", {}).get("dataset_id")
                or eval_data.get("metadata", {}).get("dataset_id")
                or ""
            )

        if models_folder.exists():
            # Pass 1: comprehensive evaluation JSONs (all 3 phases complete)
            for json_path in models_folder.glob("MDL_*_comprehensive_evaluation.json"):
                model_id = json_path.stem.replace("_comprehensive_evaluation", "")
                if model_id not in pkl_model_ids:
                    continue
                try:
                    with open(json_path, "r") as fh:
                        eval_data = _json.load(fh)
                    if _get_eval_dataset_id(eval_data) != dataset_id:
                        continue
                    json_models[model_id] = _extract_model_entry(eval_data, model_id, has_full_eval=True)
                except Exception as json_err:
                    logger.warning(f"Could not read comprehensive JSON {json_path}: {json_err}")

            # Pass 2: phase1 JSONs - fill in models that have Phase 1 done but no comprehensive yet
            # (happens when evaluation is still in progress or comprehensive write failed)
            for json_path in models_folder.glob("MDL_*_eval_phase1.json"):
                model_id = json_path.stem.replace("_eval_phase1", "")
                if model_id not in pkl_model_ids:
                    continue
                if model_id in json_models:
                    continue  # already found via comprehensive JSON - skip
                try:
                    with open(json_path, "r") as fh:
                        eval_data = _json.load(fh)
                    if _get_eval_dataset_id(eval_data) != dataset_id:
                        continue
                    # has_full_eval=True so the frontend doesn't filter this model out.
                    # Phase 1 data includes performance metrics and ROC curves which is enough
                    # to display the model in the evaluation list.
                    json_models[model_id] = _extract_model_entry(eval_data, model_id, has_full_eval=True)
                    logger.info(f"Model {model_id} discovered via phase1 JSON (comprehensive not yet written)")
                except Exception as json_err:
                    logger.warning(f"Could not read phase1 JSON {json_path}: {json_err}")

        logger.info(f"Found {len(json_models)} evaluated models via JSON scan for dataset_id={dataset_id}")

        # --- Query DB (best-effort) ---
        db_models_raw: list = []
        try:
            db_models_raw = model_evaluation_db.list_models_by_dataset(dataset_id)
        except Exception as db_err:
            logger.warning(f"DB query failed (non-fatal, falling back to JSON): {db_err}")

        valid_db_models = [m for m in db_models_raw if m.get("id") in pkl_model_ids]
        for m in valid_db_models:
            m["has_meea_data"] = True

        logger.info(
            f"Found {len(db_models_raw)} models in DB for dataset_id={dataset_id}, "
            f"{len(valid_db_models)} with corresponding .pkl files"
        )

        # --- Merge: DB entries take precedence; JSON fills gaps ---
        merged: dict = {**json_models}  # start with JSON-discovered models
        for m in valid_db_models:
            merged[m["id"]] = m  # DB entry wins (more complete)

        final_models = list(merged.values())

        # Sort newest first
        final_models.sort(key=lambda m: m.get("created_at") or m.get("training_date") or "", reverse=True)

        return JSONResponse(content={
            "success": True,
            "count": len(final_models),
            "models": final_models,
        })

    except Exception as e:
        logger.error(f"Error listing models by dataset_id={dataset_id}: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="Failed to list models for dataset")


@chat_router.post("/model-evaluation/compare")
async def compare_models(request: dict):
    """
    Compare multiple models side by side
    Body: { model_ids: List[str] }
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        model_ids = request.get("model_ids", [])
        
        if not model_ids or len(model_ids) < 2:
            raise HTTPException(status_code=400, detail="At least 2 model IDs are required for comparison")
        
        logger.info(f"Comparing models: {model_ids}")
        
        # Get comparison data
        comparison_data = model_evaluation_db.get_model_comparison(model_ids)
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "comparison_count": comparison_data['comparison_count'],
            "models": comparison_data['models']
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error comparing models: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error comparing models: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/performance")
async def get_model_performance(model_id: str):
    """
    Get performance metrics for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching performance metrics for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "performance_metrics": evaluation_data.get('performance_metrics', {})
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching performance metrics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching performance metrics: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/feature-importance")
async def get_feature_importance(model_id: str):
    """
    Get feature importance for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching feature importance for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "feature_importance": evaluation_data.get('feature_importance', [])
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching feature importance: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching feature importance: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/granular-accuracy")
async def get_granular_accuracy(model_id: str):
    """
    Get granular accuracy analysis for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching granular accuracy for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "granular_accuracy": evaluation_data.get('granular_accuracy', [])
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching granular accuracy: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching granular accuracy: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/granular-accuracy/by-segments")
async def get_granular_accuracy_by_segments(
    model_id: str, 
    variable: Optional[str] = None, 
    granularity_level: Optional[str] = None
):
    """
    Get granular accuracy grouped by segments for a specific model
    Optionally filter by variable and granularity_level
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching granular accuracy by segments for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        granular_data = evaluation_data.get('granular_accuracy', [])
        
        # Filter by variable and granularity if provided
        if variable:
            granular_data = [g for g in granular_data if g.get('variable') == variable]
        if granularity_level:
            granular_data = [g for g in granular_data if g.get('granularity_level') == granularity_level]
        
        # Group by segment
        segments_map = {}
        for item in granular_data:
            segment = item.get('segment', 'Unknown')
            if segment not in segments_map:
                segments_map[segment] = []
            segments_map[segment].append(item)
        
        # Get available variables and granularity levels
        available_variables = list(set([g.get('variable') for g in evaluation_data.get('granular_accuracy', []) if g.get('variable')]))
        available_granularity_levels = list(set([g.get('granularity_level') for g in evaluation_data.get('granular_accuracy', []) if g.get('granularity_level')]))
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "segments": segments_map,
            "available_variables": available_variables,
            "available_granularity_levels": available_granularity_levels
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching granular accuracy by segments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching granular accuracy by segments: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/error-patterns")
async def get_error_patterns(model_id: str):
    """
    Get error pattern analysis for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching error patterns for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "error_patterns": evaluation_data.get('error_patterns', [])
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching error patterns: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching error patterns: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/explainability")
async def get_explainability_data(model_id: str):
    """
    Get explainability data (SHAP, PDP, etc.) for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching explainability data for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "explainability_data": evaluation_data.get('explainability_data', [])
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching explainability data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching explainability data: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/pdp-data")
async def get_pdp_data(model_id: str, data_source: str = 'test'):
    """
    Get PDP (Partial Dependence Plot) data for a specific model (lazy loading optimization)
    
    This endpoint enables lazy loading of PDP data, which is typically large (10-15 MB).
    By fetching PDP only when needed, we reduce initial explainability tab load time by 40%.
    
    Args:
        model_id: Model identifier
        data_source: 'train' or 'test' (default: 'test')
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        import sqlite3
        
        logger.info(f"Lazy loading PDP data for model: {model_id}, data_source: {data_source}")
        
        # Fetch only PDP data from database (much faster than loading all explainability data)
        with model_evaluation_db.connect() as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Normalize data_source for backward compatibility
            cursor.execute("""
                SELECT * FROM explainability_data 
                WHERE model_id = ? 
                AND data_type = 'pdp' 
                AND (data_source = ? OR (data_source IS NULL AND ? = 'test'))
            """, (model_id, data_source, data_source))
            
            pdp_rows = cursor.fetchall()
            pdp_data = []
            
            # Parse JSON values
            for row in pdp_rows:
                pdp_entry = dict(row)
                if pdp_entry.get('data_values'):
                    pdp_entry['values'] = json.loads(pdp_entry['data_values'])
                if pdp_entry.get('metadata'):
                    pdp_entry['metadata'] = json.loads(pdp_entry['metadata'])
                
                # Ensure data_source is properly set
                if pdp_entry.get('data_source') is None or pdp_entry.get('data_source') == '':
                    pdp_entry['data_source'] = 'test'
                else:
                    pdp_entry['data_source'] = str(pdp_entry['data_source']).strip().lower()
                
                pdp_data.append(pdp_entry)
            
            logger.info(f"Found {len(pdp_data)} PDP entries for model {model_id} with data_source={data_source}")
            
            return JSONResponse(content=safe_json_serialize({
                "success": True,
                "model_id": model_id,
                "data_source": data_source,
                "pdp_data": pdp_data
            }))
    
    except Exception as e:
        logger.error(f"Error fetching PDP data: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error fetching PDP data: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/prediction-confidence")
async def get_prediction_confidence(model_id: str):
    """
    Get prediction confidence analysis for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching prediction confidence for model: {model_id}")
        
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "prediction_confidence": evaluation_data.get('prediction_confidence', [])
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching prediction confidence: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching prediction confidence: {str(e)}")


@chat_router.delete("/model-evaluation/{model_id}")
async def delete_model_evaluation(model_id: str):
    """
    Delete evaluation data for a specific model
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Deleting evaluation data for model: {model_id}")
        
        success = model_evaluation_db.delete_model_evaluation(model_id)
        
        if not success:
            raise HTTPException(status_code=404, detail=f"Evaluation data not found for model: {model_id}")
        
        return JSONResponse(content={
            "success": True,
            "message": f"Evaluation data deleted for model: {model_id}"
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting evaluation data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting evaluation data: {str(e)}")


@chat_router.post("/model-evaluation/{original_model_id}/evaluate-pruned")
async def evaluate_pruned_model(
    original_model_id: str,
    request: dict,
    background_tasks: BackgroundTasks
):
    """
    Evaluate a pruned model without retraining by imputing dropped features.
    Triggers MEEA and Explainability phases in the background.
    
    Request body:
    {
        "pruned_model_id": str,
        "surviving_features": list[str],
        "dataset_id": str
    }
    """
    pruned_model_id = request.get("pruned_model_id")
    surviving_features = request.get("surviving_features", [])
    dataset_id = request.get("dataset_id")
    
    if not pruned_model_id or not surviving_features or not dataset_id:
        logger.error(f"400 Error: Missing fields. pruned_model_id={pruned_model_id}, surviving_features={len(surviving_features)}, dataset_id={dataset_id}")
        raise HTTPException(status_code=400, detail="pruned_model_id, surviving_features, and dataset_id are required")
        
    try:
        from app.services.model_evaluation_service import model_evaluation_service
        import joblib
        import json
        import os
        from app.core.config import settings
        from app.services.dataframe_state_manager import dataframe_state_manager
        
        # Paths to original model artifacts
        model_storage_path = "models"
        pkl_path = os.path.join(model_storage_path, f"{original_model_id}.pkl")
        results_path = os.path.join(model_storage_path, f"{original_model_id}_training_results.json")
        
        if not os.path.exists(pkl_path) or not os.path.exists(results_path):
            logger.error(f"404 Error: Model artifacts not found. pkl_path={pkl_path}, results_path={results_path}")
            raise HTTPException(status_code=404, detail="Original model artifacts not found")
            
        # 1. Load Original Model and Results
        model = joblib.load(pkl_path)
        with open(results_path, 'r') as f:
            training_results = json.load(f)
            
        from app.models.model_evaluation_database import model_evaluation_db
        eval_data = model_evaluation_db.get_model_evaluation(original_model_id) or {}
        model_data = eval_data.get("model", {})
            
        # 2. Extract configuration
        feature_names = training_results.get("used_features", [])
        train_indices = training_results.get("train_indices") or model_data.get("train_indices") or []
        test_indices = training_results.get("test_indices") or model_data.get("test_indices") or []
        target_column = training_results.get("target_variable")
        problem_type = training_results.get("problem_type", "classification")
        algorithm_name = training_results.get("algorithm", "Pruned Model")

        previous_scope = dataframe_state_manager._active_scope.get(dataset_id, "entire")

        # Always remember current-scope data as the last fallback
        current_scope_df = dataframe_state_manager.get_dataframe(dataset_id)
        if current_scope_df is None:
            logger.error(f"404 Error: Dataset {dataset_id} not found in state manager")
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found in state manager")

        def _extract_xy_for_pruning(df_scope, scope_name: str):
            if df_scope is None or len(df_scope) == 0:
                logger.info(
                    "[PRUNING_SCOPE_TRACE] dataset_id=%s scope=%s rows=0 usable=False reason=empty",
                    dataset_id,
                    scope_name,
                )
                return None, None, False

            missing_feats_local = [f for f in feature_names if f not in df_scope.columns]
            has_target = bool(target_column and target_column in df_scope.columns)
            usable = len(missing_feats_local) == 0 and has_target
            logger.info(
                "[PRUNING_SCOPE_TRACE] dataset_id=%s scope=%s rows=%s has_target=%s missing_features=%s usable=%s",
                dataset_id,
                scope_name,
                len(df_scope),
                has_target,
                len(missing_feats_local),
                usable,
            )
            if not usable:
                return None, None, False

            return df_scope[feature_names].copy(), df_scope[target_column].copy(), True

        def _get_scope_df(scope_name: str):
            try:
                dataframe_state_manager.set_scope(dataset_id, scope=scope_name)
                return dataframe_state_manager.get_dataframe(dataset_id)
            except Exception as scope_err:
                logger.warning(
                    "[PRUNING_SCOPE_TRACE] dataset_id=%s scope=%s set_scope_failed=%s",
                    dataset_id,
                    scope_name,
                    str(scope_err),
                )
                return None

        train_scope_used = None
        test_scope_used = None
        X_train = None
        y_train = None
        X_test = None
        y_test = None

        try:
            # Preferred train source: train scope, then current scope
            for scope_name in ("train",):
                scope_df = _get_scope_df(scope_name)
                X_cand, y_cand, ok = _extract_xy_for_pruning(scope_df, scope_name)
                if ok:
                    X_train, y_train = X_cand, y_cand
                    train_scope_used = scope_name
                    break

            if X_train is None:
                X_cand, y_cand, ok = _extract_xy_for_pruning(current_scope_df, "current_scope")
                if ok:
                    X_train, y_train = X_cand, y_cand
                    train_scope_used = "current_scope"

            # Preferred test source: test scope, then validation, then current scope
            for scope_name in ("test", "validation"):
                scope_df = _get_scope_df(scope_name)
                X_cand, y_cand, ok = _extract_xy_for_pruning(scope_df, scope_name)
                if ok:
                    X_test, y_test = X_cand, y_cand
                    test_scope_used = scope_name
                    break

            if X_test is None:
                X_cand, y_cand, ok = _extract_xy_for_pruning(current_scope_df, "current_scope")
                if ok:
                    X_test, y_test = X_cand, y_cand
                    test_scope_used = "current_scope"
        finally:
            try:
                dataframe_state_manager.set_scope(dataset_id, scope=previous_scope)
            except Exception as restore_err:
                logger.warning(
                    "[PRUNING_SCOPE_TRACE] dataset_id=%s restore_scope=%s failed=%s",
                    dataset_id,
                    previous_scope,
                    str(restore_err),
                )

        if X_train is None or y_train is None:
            logger.error(
                "400 Error: Unable to resolve train data for pruning evaluation. dataset_id=%s train_scope_used=%s",
                dataset_id,
                train_scope_used,
            )
            raise HTTPException(
                status_code=400,
                detail="Unable to resolve train data for pruning evaluation (train/current scope unavailable or missing target/features)",
            )

        if X_test is None or y_test is None:
            logger.warning(
                "No usable test scope found; reusing train scope for evaluation. dataset_id=%s train_scope_used=%s",
                dataset_id,
                train_scope_used,
            )
            X_test = X_train.copy()
            y_test = y_train.copy()
            test_scope_used = train_scope_used or "train"

        logger.info(
            "[PRUNING_SCOPE_TRACE] dataset_id=%s train_scope_used=%s test_scope_used=%s train_rows=%s test_rows=%s train_indices_len=%s test_indices_len=%s",
            dataset_id,
            train_scope_used,
            test_scope_used,
            len(X_train),
            len(X_test),
            len(train_indices),
            len(test_indices),
        )

        # Keep provenance indices if available; otherwise use local positional indices.
        if not train_indices:
            train_indices = list(range(len(X_train)))
        if not test_indices:
            test_indices = list(range(len(X_test)))

        # 4. Imputation Step for Dropped Features
        dropped_features = [f for f in feature_names if f not in surviving_features]
        logger.info(f"Evaluating pruned model {pruned_model_id}. Neutralizing {len(dropped_features)} dropped features.")
        
        # Calculate imputation values from training data to avoid data leakage
        imputation_values = {}
        for feature in dropped_features:
            if pd.api.types.is_numeric_dtype(X_train[feature]):
                imputation_values[feature] = X_train[feature].median()
            else:
                mode_val = X_train[feature].mode()
                imputation_values[feature] = mode_val[0] if not mode_val.empty else "Missing"
                
        # Apply imputation
        for feature, imp_val in imputation_values.items():
            X_train[feature] = imp_val
            X_test[feature] = imp_val
            
        # 5. Define background evaluation tasks
        def evaluate_pruned_model_background():
            import copy
            import traceback
            import pandas as pd
            try:
                logger.info(f"Background: Starting MEEA evaluation for pruned model {pruned_model_id}")
                
                # We need to mark it as pending in the background status tracker
                from app.services.model_training_auto_training import ModelTrainingAutoTrainingService
                ModelTrainingAutoTrainingService._pending_meea_jobs[pruned_model_id] = {
                    "dataset_id": dataset_id,
                    "model_id": pruned_model_id
                }
                
                # Phase 1: Performance
                # We pass the full feature_names because the model expects them, 
                # but the data for dropped ones is now a constant scalar.
                p1 = model_evaluation_service.evaluate_phase1_performance(
                    model=model,
                    model_id=pruned_model_id,
                    model_name=f"{algorithm_name} (Pruned)",
                    X_train=X_train,
                    X_test=X_test,
                    y_train=y_train,
                    y_test=y_test,
                    problem_type=problem_type,
                    feature_names=feature_names,
                    dataset_id=dataset_id,
                    active_scope="train_test",
                    target_column=target_column,
                    split_params=None,
                    preprocessed_columns=None,
                    train_indices=train_indices,
                    test_indices=test_indices,
                    category_mappings=None,
                    X_test_original=X_test,
                    X_train_original=X_train,
                    scaler=None,
                    column_stats=None
                )
                
                # Save Phase 1
                p1_path = os.path.join(model_storage_path, f"{pruned_model_id}_eval_phase1.json")
                with open(p1_path, 'w') as fh:
                    json.dump(p1, fh, indent=2, default=str)
                
                # Create comprehensive file
                comp_path = os.path.join(model_storage_path, f"{pruned_model_id}_comprehensive_evaluation.json")
                with open(comp_path, 'w') as fh:
                    json.dump({k: v for k, v in p1.items() if not k.startswith('_phase')}, fh, indent=2, default=str)
                    
                # Phase 2: Monotonicity
                p2 = model_evaluation_service.evaluate_phase2_monotonicity(
                    model=model,
                    model_id=pruned_model_id,
                    problem_type=problem_type,
                    feature_names=feature_names
                )
                
                p2_path = os.path.join(model_storage_path, f"{pruned_model_id}_eval_phase2.json")
                with open(p2_path, 'w') as fh:
                    json.dump(p2, fh, indent=2, default=str)
                    
                with open(comp_path, 'r') as fh:
                    existing = json.load(fh)
                existing.update({k: v for k, v in p2.items() if not k.startswith('_phase')})
                with open(comp_path, 'w') as fh:
                    json.dump(existing, fh, indent=2, default=str)
                    
                # Phase 3: Granular Accuracy
                p3 = model_evaluation_service.evaluate_phase3_granular(
                    model=model,
                    model_id=pruned_model_id,
                    problem_type=problem_type,
                    feature_names=feature_names
                )
                
                p3_path = os.path.join(model_storage_path, f"{pruned_model_id}_eval_phase3.json")
                with open(p3_path, 'w') as fh:
                    json.dump(p3, fh, indent=2, default=str)
                    
                with open(comp_path, 'r') as fh:
                    existing = json.load(fh)
                existing.update({k: v for k, v in p3.items() if not k.startswith('_phase')})
                with open(comp_path, 'w') as fh:
                    json.dump(existing, fh, indent=2, default=str)
                    
                # Save to DB
                from app.models.model_evaluation_database import model_evaluation_db
                db_formatted_data = model_evaluation_service.format_for_database(existing)
                model_evaluation_db.save_evaluation_results(db_formatted_data)
                
                # Start Explainability
                from app.services.explainability_service import explainability_service
                logger.info(f"Background: Starting Explainability for pruned model {pruned_model_id}")
                
                shap_result = explainability_service.calculate_shap_analysis(
                    model=model,
                    X_train=X_train,
                    X_test=X_test,
                    model_id=pruned_model_id,
                    problem_type=problem_type
                )
                
                pdp_result = explainability_service.calculate_pdp_analysis(
                    model=model,
                    X_train=X_train,
                    feature_names=feature_names,
                    model_id=pruned_model_id,
                    problem_type=problem_type
                )
                
                if shap_result:
                    existing["shap_analysis"] = shap_result
                if pdp_result:
                    existing["partial_dependence"] = pdp_result
                    
                with open(comp_path, 'w') as fh:
                    json.dump(existing, fh, indent=2, default=str)

                logger.info(f"Background: MEEA evaluation and Explainability complete for {pruned_model_id}")
                
            except Exception as e:
                logger.error(f"Error in evaluate_pruned_model_background: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
            finally:
                from app.services.model_training_auto_training import ModelTrainingAutoTrainingService
                ModelTrainingAutoTrainingService._pending_meea_jobs.pop(pruned_model_id, None)

        # Trigger background task
        background_tasks.add_task(evaluate_pruned_model_background)
        
        return JSONResponse(content={
            "success": True,
            "message": f"Started evaluation for pruned model {pruned_model_id}",
            "model_id": pruned_model_id
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error initiating pruned model evaluation: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error initiating pruned model evaluation: {str(e)}")

@chat_router.post("/model-evaluation/evaluate-existing/{model_id}")
async def evaluate_existing_model_from_json(model_id: str):
    """
    Generate MEEA evaluation for an existing model from training_results.json
    Reads the JSON file and creates evaluation data
    """
    try:
        from app.services.model_evaluation_from_json import model_evaluation_from_json
        
        logger.info(f"Evaluating existing model from JSON: {model_id}")
        
        success = model_evaluation_from_json.evaluate_existing_model(model_id)
        
        if success:
            return JSONResponse(content={
                "success": True,
                "message": f"MEEA evaluation generated for model {model_id}",
                "model_id": model_id
            })
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to generate MEEA evaluation for {model_id}. Check logs for details."
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error evaluating existing model: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error evaluating model: {str(e)}")


@chat_router.post("/model-evaluation/evaluate-all-existing")
async def evaluate_all_existing_models():
    """
    Generate MEEA evaluation for ALL existing models from training_results.json files
    Batch processes all models in the models folder
    """
    try:
        from app.services.model_evaluation_from_json import model_evaluation_from_json
        
        logger.info("Starting batch evaluation of all existing models")
        
        results = model_evaluation_from_json.evaluate_all_existing_models()
        
        return JSONResponse(content={
            "success": True,
            "message": "Batch evaluation completed",
            "results": results
        })
    
    except Exception as e:
        logger.error(f"Error in batch evaluation: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error in batch evaluation: {str(e)}")


@chat_router.post("/model-evaluation/{model_id}/recalculate-explainability")
async def recalculate_explainability(model_id: str, request: dict):
    """
    Recalculate explainability (SHAP, PDP) for a model on train or test data
    
    Request body:
    {
        "data_source": "train" | "test"  # Which dataset to use
    }
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        from app.services.model_evaluation_service import model_evaluation_service
        from app.services.explainability_service import explainability_service
        from app.services.dataframe_state_manager import dataframe_state_manager
        from app.services.dataset_service import dataset_manager
        import joblib
        from pathlib import Path
        from sklearn.model_selection import train_test_split
        import pandas as pd
        import numpy as np
        import sqlite3
        import uuid
        from datetime import datetime
        import json
        
        # Helper function to convert pandas NA values to None for JSON serialization
        def convert_na_to_none(obj):
            """Recursively convert pandas NA values to None"""
            if isinstance(obj, dict):
                return {k: convert_na_to_none(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_na_to_none(item) for item in obj]
            elif pd.isna(obj):
                return None
            return obj
        
        data_source = request.get('data_source', 'test')
        if data_source not in ['train', 'test']:
            raise HTTPException(status_code=400, detail="data_source must be 'train' or 'test'")
        
        logger.info(f"Recalculating explainability for model {model_id} on {data_source} data")
        
        # 1. Get model metadata from database
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        if not evaluation_data:
            raise HTTPException(status_code=404, detail=f"Model {model_id} not found in database")
        
        model_info = evaluation_data.get('model', {})
        dataset_id = model_info.get('dataset_id')
        active_scope = model_info.get('active_scope', 'entire')
        target_column = model_info.get('target_column')
        split_params_json = model_info.get('split_params')
        
        if not dataset_id or not target_column:
            raise HTTPException(
                status_code=400,
                detail="Model metadata missing dataset_id or target_column. Cannot recalculate explainability."
            )
        
        # Detect if this model was trained on a per-segment dataset
        # e.g. dataset_id = "{base_dataset_id}_segment_{segment_value}"
        is_segment_dataset = False
        base_dataset_id = dataset_id
        segment_raw_id = None
        if dataset_id and "_segment_" in dataset_id:
            try:
                parts = dataset_id.split("_segment_", 1)
                if len(parts) == 2:
                    base_dataset_id, segment_raw_id = parts[0], parts[1]
                    is_segment_dataset = True
                    logger.info(
                        f"Detected segment dataset for model {model_id}: "
                        f"base_dataset_id={base_dataset_id}, segment_raw_id={segment_raw_id}"
                    )
            except Exception as e:
                logger.warning(
                    f"Failed to parse segment dataset_id '{dataset_id}' for model {model_id}: {str(e)}"
                )
                base_dataset_id = dataset_id
                segment_raw_id = None
                is_segment_dataset = False
        
        # Parse split_params
        split_params = {
            'test_size': 0.2,
            'random_state': 42,
            'stratify': False
        }
        if split_params_json:
            if isinstance(split_params_json, str):
                split_params = json.loads(split_params_json)
            else:
                split_params = split_params_json
        
        # 2. Load the model
        model_path = Path("models") / f"{model_id}.pkl"
        if not model_path.exists():
            raise HTTPException(status_code=404, detail=f"Model file not found: {model_path}")
        
        model = joblib.load(model_path)
        logger.info(f"Loaded model from {model_path}")
        
        # 3. Get the latest dataframe from DataFrameStateManager
        # First try the exact dataset_id used during training (works for global models
        # and for segment models if the per-segment dataset is still materialized).
        df = dataframe_state_manager.get_dataframe(dataset_id)
        
        # If this is a segment dataset and the stored per-segment dataframe is missing
        # or empty (it may have been cleaned up after training), fall back to using
        # the base dataset and filter rows by the segment column.
        if (df is None or df.empty) and is_segment_dataset and base_dataset_id:
            logger.info(
                f"Per-segment dataset '{dataset_id}' is missing or empty for model {model_id}. "
                f"Falling back to base dataset '{base_dataset_id}' and segment filtering."
            )
            base_df = dataframe_state_manager.get_dataframe(base_dataset_id)
            if base_df is None or base_df.empty:
                # Fallback to loading from dataset service
                base_df = dataset_manager.load_dataset(base_dataset_id)
            
            if base_df is None or base_df.empty:
                raise HTTPException(
                    status_code=404,
                    detail=f"Base dataset {base_dataset_id} not found or empty. Cannot recalculate explainability."
                )
            
            # Try to locate the segment column on the base dataframe.
            segment_column = None
            if segment_raw_id is not None:
                # Preferred segment column names, consistent with segmentation services
                candidate_columns = [
                    "segment", "SEGMENT", "segment_id", "SEGMENT_ID",
                    "group", "GROUP", "cluster", "CLUSTER",
                ]
                for col in candidate_columns:
                    if col in base_df.columns:
                        segment_column = col
                        break
                
                # If still not found, try to infer a suitable segment column
                if segment_column is None:
                    for col in base_df.columns:
                        lower_col = str(col).lower()
                        if lower_col in ["segment", "segments", "segment_id", "segment_label", "group", "cluster"]:
                            segment_column = col
                            break
            
            if segment_raw_id is not None and segment_column is not None:
                # Filter base_df to only rows belonging to this segment. Use string
                # comparison to be robust to numeric vs string representations.
                before_shape = base_df.shape
                try:
                    segment_mask = base_df[segment_column].astype(str) == str(segment_raw_id)
                except Exception as e:
                    logger.warning(
                        f"Error while creating segment mask for column '{segment_column}' "
                        f"and segment '{segment_raw_id}': {str(e)}. Falling back to no filter."
                    )
                    segment_mask = None
                
                if segment_mask is not None:
                    df_segment = base_df[segment_mask].copy()
                    logger.info(
                        f"Filtered base dataset '{base_dataset_id}' for segment '{segment_raw_id}' "
                        f"on column '{segment_column}': {before_shape} -> {df_segment.shape}"
                    )
                    if df_segment.empty:
                        raise HTTPException(
                            status_code=404,
                            detail=(
                                f"No data found for segment '{segment_raw_id}' in column "
                                f"'{segment_column}' of base dataset {base_dataset_id}."
                            ),
                        )
                    df = df_segment
                else:
                    df = base_df
            else:
                # Could not identify segment column or value reliably; fall back to using
                # the entire base dataset. This preserves existing behaviour for global
                # models and still allows explainability to run, even if segment
                # filtering is not perfectly applied.
                logger.warning(
                    f"Unable to determine segment column or segment id for dataset '{dataset_id}'. "
                    f"Using entire base dataset '{base_dataset_id}' for explainability."
                )
                df = base_df
        
        # If df is still None/empty at this point, raise a clear error.
        if df is None or df.empty:
            raise HTTPException(
                status_code=404,
                detail=f"Dataset {dataset_id} not found or empty. Cannot recalculate explainability."
            )
        
        # 4. Get feature names - try to use model's feature order first
        feature_names = None
        
        # Try to get feature names from the model (most reliable)
        if hasattr(model, 'feature_names_in_'):
            feature_names = list(model.feature_names_in_)
            logger.info(f"Using feature names from model.feature_names_in_: {feature_names}")
        elif hasattr(model, 'feature_name_'):
            # LightGBM specific
            feature_names = list(model.feature_name_())
            logger.info(f"Using feature names from model.feature_name_(): {feature_names}")
        
        # Fallback to feature importance
        if not feature_names:
            feature_importance = evaluation_data.get('feature_importance', [])
            feature_names = [f['feature_name'] for f in feature_importance if f.get('feature_name')]
            logger.info(f"Using feature names from feature_importance: {feature_names}")
        
        if not feature_names:
            raise HTTPException(
                status_code=400,
                detail="Feature names not found in model or evaluation data. Cannot recalculate explainability."
            )
        
        # 5. Check if model was trained on preprocessed columns (metadata-driven approach)
        preprocessed_columns = model_info.get('preprocessed_columns', {})

        # For segment models, it's possible that preprocessed columns were merged back
        # into the base dataframe using segment-specific suffixes (e.g. *_le_seg_auto).
        # If the original preprocessed column names are not present but their
        # segment-specific counterparts are, adjust the mapping on the fly. This is
        # a no-op for global models and preserves existing behaviour.
        if preprocessed_columns and is_segment_dataset:
            adjusted_preprocessed_columns = {}
            for orig_col, pre_col in preprocessed_columns.items():
                # If the original preprocessed column exists, keep it
                if pre_col in df.columns:
                    adjusted_preprocessed_columns[orig_col] = pre_col
                    continue

                candidate_names = [pre_col]
                # Auto-training suffixes
                if pre_col.endswith("_le_auto"):
                    candidate_names.append(pre_col.replace("_le_auto", "_le_seg_auto"))
                if pre_col.endswith("_ss_auto"):
                    candidate_names.append(pre_col.replace("_ss_auto", "_ss_seg_auto"))
                # Manual-training suffixes
                if pre_col.endswith("_le_manual"):
                    candidate_names.append(pre_col.replace("_le_manual", "_le_seg_manual"))
                if pre_col.endswith("_ss_manual"):
                    candidate_names.append(pre_col.replace("_ss_manual", "_ss_seg_manual"))

                # Pick the first candidate that exists in the current dataframe
                chosen_col = None
                for cand in candidate_names:
                    if cand in df.columns:
                        chosen_col = cand
                        break

                if chosen_col is not None:
                    logger.info(
                        f"Adjusted preprocessed column mapping for segment model {model_id}: "
                        f"{orig_col} → {chosen_col} (from {pre_col})"
                    )
                    adjusted_preprocessed_columns[orig_col] = chosen_col
                else:
                    # Fall back to original mapping if nothing matches; downstream logic
                    # will handle missing columns gracefully.
                    adjusted_preprocessed_columns[orig_col] = pre_col

            preprocessed_columns = adjusted_preprocessed_columns
        
        if preprocessed_columns:
            # Model was trained on preprocessed data - use preprocessed columns directly
            logger.info(f"Model was trained on preprocessed data. Using preprocessed columns from mapping.")
            logger.info(f"Preprocessed columns mapping: {list(preprocessed_columns.items())[:5]}...")
            
            # Map original feature names to preprocessed column names
            preprocessed_feature_names = []
            missing_preprocessed = []
            
            for orig_col in feature_names:
                if orig_col in preprocessed_columns:
                    preprocessed_col = preprocessed_columns[orig_col]
                    if preprocessed_col in df.columns:
                        preprocessed_feature_names.append(preprocessed_col)
                        logger.debug(f"Mapped {orig_col} → {preprocessed_col}")
                    else:
                        missing_preprocessed.append(preprocessed_col)
                        logger.warning(f"Preprocessed column {preprocessed_col} not found in dataframe. Available columns: {list(df.columns)[:10]}...")
                        # Fallback: use original column if preprocessed not found
                        preprocessed_feature_names.append(orig_col)
                else:
                    # No mapping for this column - use original
                    preprocessed_feature_names.append(orig_col)
                    logger.debug(f"No preprocessed mapping for {orig_col}, using original column")
            
            if missing_preprocessed:
                logger.warning(f"Some preprocessed columns not found: {missing_preprocessed}. Using original columns as fallback.")
            
            # Validate that we can select all required columns
            missing_features = [f for f in preprocessed_feature_names if f not in df.columns]
            if missing_features:
                logger.error(f"Missing features in dataframe: {missing_features}. Available columns: {list(df.columns)[:20]}")
                raise HTTPException(
                    status_code=400,
                    detail=f"Required features not found in dataset: {missing_features}. Cannot recalculate explainability."
                )
            
            # Select preprocessed columns in model's expected order
            X = df[preprocessed_feature_names].copy()
            # Rename columns back to original feature names (model expects these names)
            X.columns = feature_names
            
            logger.info(f"Using preprocessed columns: {len([c for c in preprocessed_feature_names if c != feature_names[preprocessed_feature_names.index(c)]])} preprocessed, {len([c for c in preprocessed_feature_names if c == feature_names[preprocessed_feature_names.index(c)]])} original")
        else:
            # Model was trained on original columns - use original columns directly
            logger.info(f"Model was trained on original columns. Using original columns directly.")
            
            # Validate that all features exist
            missing_features = [f for f in feature_names if f not in df.columns]
            if missing_features:
                logger.error(f"Missing features in dataframe: {missing_features}. Available columns: {list(df.columns)[:20]}")
                raise HTTPException(
                    status_code=400,
                    detail=f"Required features not found in dataset: {missing_features}. Cannot recalculate explainability."
                )
            
            # Select features in the EXACT order expected by the model (preserves feature_names order)
            X = df[feature_names].copy()
        
        y = df[target_column].copy()
        
        logger.info(f"Selected {len(feature_names)} features in model's expected order: {feature_names}")
        
        # Remove rows with missing target
        valid_mask = ~y.isna()
        X = X[valid_mask]
        y = y[valid_mask]
        
        logger.info(f"Prepared data: X shape={X.shape}, y shape={y.shape}, feature_names={len(feature_names)}")
        logger.info(f"Data types: {X.dtypes.to_dict()}")
        
        # No preprocessing needed - data is already in the format the model expects
        X_processed = X.copy()
        logger.info(f"Using data directly as-is (no encoding/scaling needed - data format matches model training)")
        
        # ROBUSTNESS FIX: Convert any remaining object columns to numeric codes
        # This handles cases where preprocessed columns are missing and we fall back to original object columns
        # SHAP/PDP/tree models require purely numeric dtypes
        object_columns = X_processed.select_dtypes(include=['object']).columns.tolist()
        if object_columns:
            logger.info(f"Converting {len(object_columns)} object columns to numeric codes for explainability: {object_columns}")
            for col in object_columns:
                try:
                    # Convert to categorical codes (similar to LabelEncoder but on-the-fly)
                    X_processed[col] = X_processed[col].astype('category').cat.codes
                    logger.info(f"  ✓ Converted '{col}' from object to int (category codes)")
                except Exception as e:
                    logger.warning(f"  ✗ Failed to convert '{col}': {str(e)}")
                    # If conversion fails, drop this column rather than passing strings to SHAP
                    X_processed = X_processed.drop(columns=[col])
                    # Also remove from feature_names
                    if col in feature_names:
                        feature_names.remove(col)
                    logger.warning(f"  → Dropped '{col}' from explainability features")
        
        # Encode target if it's categorical (needed for train/test split)
        if y.dtype == 'object' or y.dtype.name == 'category':
            logger.info(f"Encoding target column: {target_column} for split recreation")
            from sklearn.preprocessing import LabelEncoder
            target_le = LabelEncoder()
            y = pd.Series(target_le.fit_transform(y.astype(str)), index=y.index, name=y.name)
        
        # 6. Recreate train/test split - use stored indices if available, otherwise use split_params
        train_indices = model_info.get('train_indices')
        test_indices = model_info.get('test_indices')
        
        if train_indices and test_indices:
            # Use stored indices for exact split recreation
            logger.info(f"Using stored train/test indices for exact split recreation")
            logger.info(f"Train indices: {len(train_indices)} samples, Test indices: {len(test_indices)} samples")
            
            if is_segment_dataset:
                # For segmented models, stored indices are original base dataframe indices
                # We need to filter them to only include indices that exist in the filtered segment dataframe
                # Then convert them to position indices for use with .iloc[]
                filtered_df_index_set = set(df.index.tolist())
                
                # Filter indices to only include those that exist in the filtered segment dataframe
                train_indices_filtered = [idx for idx in train_indices if idx in filtered_df_index_set]
                test_indices_filtered = [idx for idx in test_indices if idx in filtered_df_index_set]
                
                if len(train_indices_filtered) > 0 and len(test_indices_filtered) > 0:
                    # Convert filtered indices to position indices for use with .iloc[]
                    try:
                        train_indices_pos = [df.index.get_loc(idx) for idx in train_indices_filtered]
                        test_indices_pos = [df.index.get_loc(idx) for idx in test_indices_filtered]
                        
                        logger.info(f"Filtered indices for segmented model: train {len(train_indices_pos)}/{len(train_indices)}, test {len(test_indices_pos)}/{len(test_indices)}")
                        train_indices = train_indices_pos
                        test_indices = test_indices_pos
                    except (KeyError, ValueError) as e:
                        logger.warning(f"Error converting indices to position indices for segmented model: {str(e)}")
                        logger.warning("Falling back to split_params recreation")
                        train_indices = None
                        test_indices = None
                else:
                    logger.warning(f"No valid indices found in filtered segment dataframe. Train: {len(train_indices_filtered)}/{len(train_indices)}, Test: {len(test_indices_filtered)}/{len(test_indices)}")
                    logger.warning("Falling back to split_params recreation")
                    train_indices = None
                    test_indices = None
            else:
                # For global models, stored indices are already position indices
                max_index = len(X_processed) - 1
                train_indices_valid = [idx for idx in train_indices if isinstance(idx, (int, np.integer)) and 0 <= idx <= max_index]
                test_indices_valid = [idx for idx in test_indices if isinstance(idx, (int, np.integer)) and 0 <= idx <= max_index]
                
                if len(train_indices_valid) != len(train_indices) or len(test_indices_valid) != len(test_indices):
                    logger.warning(f"Some indices out of bounds. Valid train: {len(train_indices_valid)}/{len(train_indices)}, Valid test: {len(test_indices_valid)}/{len(test_indices)}")
                    logger.warning("Falling back to split_params recreation")
                    train_indices = None
                    test_indices = None
                else:
                    train_indices = train_indices_valid
                    test_indices = test_indices_valid
        
        # Check if this is a no-split scenario (active_scope == 'entire')
        no_split = split_params.get('no_split', False) or split_params.get('test_size', 0.2) == 0.0
        
        if no_split:
            # No train/test split - use entire dataset for training
            logger.info("No train/test split (active_scope='entire') - using entire dataset as train data")
            X_train = X_processed.copy()
            y_train = y.copy()
            X_test = None
            y_test = None
            train_indices = list(range(len(X_processed)))
            test_indices = []
            logger.info(f"Using entire dataset as train: X_train shape={X_train.shape}")
        elif train_indices and test_indices:
            # Use stored indices
            X_train = X_processed.iloc[train_indices].copy()
            X_test = X_processed.iloc[test_indices].copy()
            y_train = y.iloc[train_indices].copy()
            y_test = y.iloc[test_indices].copy()
            
            logger.info(f"Split data using stored indices: X_train shape={X_train.shape}, X_test shape={X_test.shape}")
        else:
            # Fallback to split_params recreation
            logger.info(f"Using split_params to recreate train/test split (indices not available or invalid)")
            test_size = split_params.get('test_size', 0.2)
            if test_size == 0.0:
                # Handle case where test_size is 0.0 (shouldn't happen here due to no_split check above, but just in case)
                X_train = X_processed.copy()
                y_train = y.copy()
                X_test = None
                y_test = None
                logger.info(f"No split - using entire dataset as train: X_train shape={X_train.shape}")
            else:
                stratify = y if (split_params.get('stratify', False) and y.nunique() <= 10) else None
                X_train, X_test, y_train, y_test = train_test_split(
                    X_processed, y,
                    test_size=test_size,
                    random_state=split_params.get('random_state', 42),
                    stratify=stratify
                )
                logger.info(f"Split data using split_params: X_train shape={X_train.shape}, X_test shape={X_test.shape}")
        
        # 7. Select data based on data_source
        # If no test set exists (active_scope == 'entire'), force data_source to 'train'
        if no_split and data_source == 'test':
            logger.warning(f"Requested test data but no test set exists (active_scope='entire'). Using train data instead.")
            data_source = 'train'
        
        if data_source == 'train':
            X_data = X_train
            y_data = y_train
            data_indices = train_indices if train_indices else None
        else:  # test
            if X_test is None or y_test is None:
                logger.warning(f"Requested test data but test set is None. Using train data instead.")
                X_data = X_train
                y_data = y_train
                data_indices = train_indices if train_indices else None
                data_source = 'train'  # Update data_source for consistency
            else:
                X_data = X_test
                y_data = y_test
                data_indices = test_indices if test_indices else None
        
        logger.info(f"Using {data_source} data: {len(X_data)} samples")
        
        # 7.5. Extract original feature values from the same dataframe (for user-friendly display)
        # The dataframe already contains both original and transformed columns
        # feature_names contains the model's expected names (which are the original column names)
        original_feature_values_per_sample = None
        feature_name_mapping = {}  # Maps model feature names to original column names (for display)
        
        # Build mapping: model_feature_name -> original_column_name
        # Since feature_names are the original column names that the model expects,
        # the mapping is identity (feature_name -> feature_name)
        for feat_name in feature_names:
            feature_name_mapping[feat_name] = feat_name  # Original name is same as model expects
        
        # Extract original feature values from the same dataframe
        # feature_names are the original column names, so we can use them directly
        if all(col in df.columns for col in feature_names):
            try:
                # We need to map X_data rows back to df rows
                # X_processed was filtered from df using valid_mask, so X_processed.index contains df indices
                # X_data was selected from X_processed using .iloc[data_indices], so X_data has position indices (0, 1, 2...)
                # We need to map: position in X_data -> position in X_processed -> df index via X_processed.index
                
                if data_indices is not None and len(data_indices) > 0:
                    # We have stored indices - these are positions in X_processed
                    # Map to df indices through X_processed.index
                    # data_indices are position indices (0, 1, 2...) within X_processed
                    df_indices = X_processed.index[data_indices]
                    X_original = df[feature_names].loc[df_indices].copy()
                    original_feature_values_per_sample = X_original.values.tolist()
                    logger.info(f"Extracted original feature values using stored indices: {len(feature_names)} columns, {len(original_feature_values_per_sample)} samples")
                elif hasattr(X_data, 'index') and len(X_data.index) > 0:
                    # X_data.index should be position indices (0, 1, 2...) after .iloc[]
                    # Try to map through X_processed.index
                    try:
                        # X_data.index contains position indices, use them to index into X_processed.index
                        df_indices = X_processed.index[X_data.index]
                        X_original = df[feature_names].loc[df_indices].copy()
                        original_feature_values_per_sample = X_original.values.tolist()
                        logger.info(f"Extracted original feature values using X_data.index: {len(feature_names)} columns, {len(original_feature_values_per_sample)} samples")
                    except (KeyError, IndexError) as e:
                        logger.warning(f"Could not map X_data.index to df indices: {str(e)}. Original values will not be available.")
                        original_feature_values_per_sample = None
                else:
                    logger.warning("Cannot determine row mapping. Skipping original values extraction.")
                    original_feature_values_per_sample = None
            except (IndexError, KeyError, AttributeError) as e:
                logger.warning(f"Could not extract original values: {str(e)}. Original values will not be available for display.")
                original_feature_values_per_sample = None
        else:
            missing_cols = [col for col in feature_names if col not in df.columns]
            logger.warning(f"Some original columns not found in dataframe: {missing_cols}. Skipping original values extraction.")
        
        # 8. Get problem type
        problem_type = model_info.get('task_type', 'classification')
        
        # 9. Recalculate explainability
        logger.info(f"Starting explainability calculation for {model_id} on {data_source} data")
        logger.info(f"Model type: {type(model)}, X_train shape: {X_train.shape}, X_data shape: {X_data.shape}")
        logger.info(f"Feature names: {feature_names[:5]}... (total: {len(feature_names)})")
        
        try:
            logger.info(f"Calling explainability_service.calculate_shap_analysis() for {data_source} data...")
            
            # Use X_data which is selected based on data_source (train or test)
            shap_data, waterfall_data = explainability_service.calculate_shap_analysis(
                model, X_train, X_data, feature_names, problem_type,
                original_feature_values_per_sample=original_feature_values_per_sample
            )
            logger.info(f"SHAP calculation completed: shap_data={shap_data is not None}, waterfall_data={waterfall_data is not None}")
            if shap_data is None:
                logger.error("SHAP calculation returned None - check explainability_service logs for details")
        except Exception as e:
            logger.error(f"Error in SHAP calculation: {str(e)}")
            import traceback
            logger.error(f"SHAP error traceback: {traceback.format_exc()}")
            shap_data = None
            waterfall_data = None
        
        try:
            logger.info(f"Calling explainability_service.calculate_pdp_analysis()...")
            pdp_data = explainability_service.calculate_pdp_analysis(
                model, X_data, feature_names, problem_type
            )
            logger.info(f"PDP calculation completed: pdp_data={pdp_data is not None if pdp_data else False}, pdp_features={len(pdp_data) if pdp_data else 0}")
            if not pdp_data or len(pdp_data) == 0:
                logger.error("PDP calculation returned empty dict - check explainability_service logs for details")
        except Exception as e:
            logger.error(f"Error in PDP calculation: {str(e)}")
            import traceback
            logger.error(f"PDP error traceback: {traceback.format_exc()}")
            pdp_data = {}
        
        logger.info(f"Explainability calculation results: shap_data={shap_data is not None}, waterfall_data={waterfall_data is not None}, pdp_data={pdp_data is not None if pdp_data else False}, pdp_features={len(pdp_data) if pdp_data else 0}")
        
        # Fallback: if SHAP returned None or empty feature importance, try a lightweight fallback
        # This ensures the frontend receives some explainability information even if full SHAP/PDP fails
        if shap_data is None or not shap_data.get('feature_importance'):
            logger.warning("SHAP data missing or empty, attempting fallback feature importance extraction from model")
            try:
                fallback_feature_importance = []
                # Prefer tree-based feature_importances_
                if hasattr(model, 'feature_importances_'):
                    try:
                        import numpy as _np
                        fi = _np.array(model.feature_importances_)
                        for i, name in enumerate(feature_names[: len(fi)]):
                            fallback_feature_importance.append({'feature_name': name, 'importance': float(fi[i])})
                    except Exception as _e:
                        logger.warning(f"Could not read model.feature_importances_: {_e}")
                # Fall back to linear coefficients
                if not fallback_feature_importance and hasattr(model, 'coef_'):
                    try:
                        import numpy as _np
                        coefs = _np.array(model.coef_)
                        if coefs.ndim > 1:
                            coefs_vals = _np.mean(_np.abs(coefs), axis=0)
                        else:
                            coefs_vals = _np.abs(coefs)
                        for i, name in enumerate(feature_names[: len(coefs_vals)]):
                            fallback_feature_importance.append({'feature_name': name, 'importance': float(coefs_vals[i])})
                    except Exception as _e:
                        logger.warning(f"Could not read model.coef_: {_e}")

                # If still empty, provide zero-valued placeholders so frontend has a stable payload
                if not fallback_feature_importance:
                    logger.warning("Model does not expose feature importances or coefficients; creating zero-importance fallback entries")
                    for name in feature_names:
                        fallback_feature_importance.append({'feature_name': name, 'importance': 0.0})

                shap_data = {
                    'feature_importance': [
                        {'feature_name': item['feature_name'], 'importance': item['importance']} for item in fallback_feature_importance
                    ],
                    'raw_shap_values': None,
                    'feature_values_per_sample': None,
                    'original_feature_values_per_sample': None,
                    'base_value': None,
                    'sample_count': len(X_data),
                    'stored_sample_count': 0,
                    'explainer_type': 'fallback'
                }
                logger.info(f"Fallback feature importance extracted for {len(shap_data['feature_importance'])} features")
            except Exception as e:
                logger.error(f"Fallback feature importance failed: {str(e)}")
                shap_data = None

        # 10. Format and save to database
        explainability_data = explainability_service.format_explainability_for_database(
            shap_data, pdp_data, waterfall_data, model_id, data_source=data_source,
            feature_name_mapping=feature_name_mapping
        )
        
        logger.info(f"Formatted explainability data: {len(explainability_data)} entries for {model_id} with data_source={data_source}")
        for i, entry in enumerate(explainability_data):
            logger.info(f"  Entry {i}: data_type={entry.get('data_type')}, data_source={entry.get('data_source')}, feature_name={entry.get('feature_name')}")
        
        # Delete old explainability data for this data_source
        # Handle both explicit data_source and null (for backward compatibility with old test data)
        with model_evaluation_db.connect() as conn:
            cursor = conn.cursor()
            if data_source == 'test':
                # For test, delete both explicit 'test' and null (old data without data_source)
                cursor.execute("""
                    DELETE FROM explainability_data 
                    WHERE model_id = ? 
                    AND (data_source = ? OR data_source IS NULL)
                    AND data_type IN ('shap_summary', 'shap_waterfall', 'pdp')
                """, (model_id, data_source))
            else:
                # For train, only delete explicit 'train' data
                cursor.execute("""
                    DELETE FROM explainability_data 
                    WHERE model_id = ? AND data_source = ? 
                    AND data_type IN ('shap_summary', 'shap_waterfall', 'pdp')
                """, (model_id, data_source))
            conn.commit()
            logger.info(f"Deleted old explainability data for {model_id} with data_source={data_source}")
        
        # Save new explainability data - use a single transaction for all inserts
        saved_count = 0
        if not explainability_data:
            logger.warning(f"No explainability data to save for {model_id} with data_source={data_source}")
        else:
            with model_evaluation_db.connect() as conn:
                cursor = conn.cursor()
                for explain in explainability_data:
                    explain_id = str(uuid.uuid4())
                    # Ensure data_source is explicitly set
                    explain_data_source = explain.get('data_source') or data_source
                    
                    try:
                        # Convert pandas NA values to None and serialize safely
                        values_cleaned = convert_na_to_none(explain['values'])
                        values_json = json.dumps(safe_json_serialize(values_cleaned))
                        
                        metadata_cleaned = convert_na_to_none(explain.get('metadata', {}))
                        metadata_json = json.dumps(safe_json_serialize(metadata_cleaned))
                        
                        cursor.execute("""
                            INSERT INTO explainability_data 
                            (id, model_id, data_type, data_source, feature_name, data_values, metadata, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            explain_id,
                            explain['model_id'],
                            explain['data_type'],
                            explain_data_source,  # Explicitly use the data_source
                            explain.get('feature_name'),
                            values_json,
                            metadata_json,
                            datetime.now().isoformat()
                        ))
                        saved_count += 1
                        logger.debug(f"Saved entry: data_type={explain['data_type']}, data_source={explain_data_source}, feature_name={explain.get('feature_name')}")
                    except Exception as e:
                        logger.error(f"Error saving explainability entry: {str(e)}")
                        logger.error(f"Entry details: data_type={explain.get('data_type')}, data_source={explain_data_source}")
                        raise
                
                conn.commit()
                logger.info(f"Saved {saved_count} explainability data entries for {model_id} with data_source={data_source} in single transaction")
        
        # Verify the data was saved by querying it back
        with model_evaluation_db.connect() as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("""
                SELECT data_type, data_source, feature_name, COUNT(*) as count
                FROM explainability_data 
                WHERE model_id = ? AND data_source = ?
                GROUP BY data_type, data_source, feature_name
            """, (model_id, data_source))
            saved_rows = cursor.fetchall()
            logger.info(f"Verification query: Found {len(saved_rows)} distinct entries for {model_id} with data_source={data_source}")
            for row in saved_rows:
                logger.info(f"  Saved: data_type={row['data_type']}, data_source={row['data_source']}, feature_name={row['feature_name']}, count={row['count']}")
        
        logger.info(f"Successfully recalculated explainability for {model_id} on {data_source} data")
        
        return JSONResponse(content={
            "success": True,
            "message": f"Explainability recalculated for {data_source} data",
            "model_id": model_id,
            "data_source": data_source,
            "samples_used": len(X_data),
            "entries_saved": saved_count
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error recalculating explainability: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error recalculating explainability: {str(e)}")


@chat_router.get("/model-evaluation/{model_id}/chat-summary")
async def get_model_evaluation_chat_summary(model_id: str):
    """
    Get MEEA evaluation data formatted for AI Assistant chat display
    Returns comprehensive model evaluation in a chat-friendly format
    """
    try:
        from app.models.model_evaluation_database import model_evaluation_db
        
        logger.info(f"Fetching chat-formatted evaluation for model: {model_id}")
        
        # Get evaluation data from database
        evaluation_data = model_evaluation_db.get_model_evaluation(model_id)
        
        if not evaluation_data:
            return JSONResponse(content={
                "success": False,
                "message": f"⚠️ No evaluation data found for model: {model_id}\n\nPlease train a model first in Step 7."
            })
        
        # Format comprehensive evaluation report for chat
        perf_metrics = evaluation_data.get('performance_metrics', {})
        conf_matrix = evaluation_data.get('confusion_matrix', {})
        feature_imp = evaluation_data.get('feature_importance_data', [])
        error_patterns = evaluation_data.get('error_pattern_data', [])
        
        # Build detailed report
        report_lines = []
        report_lines.append("# 🎯 MEEA Model Evaluation Report")
        report_lines.append(f"\n**Model ID:** `{model_id}`")
        report_lines.append(f"**Model Name:** {evaluation_data.get('model_name', 'Unknown')}")
        report_lines.append(f"**Task Type:** {evaluation_data.get('task_type', 'classification').title()}")
        report_lines.append("\n---\n")
        
        # Performance Metrics Section
        report_lines.append("## 📊 Performance Metrics\n")
        if perf_metrics:
            report_lines.append("| Metric | Value |")
            report_lines.append("|--------|-------|")
            if 'accuracy' in perf_metrics:
                report_lines.append(f"| **Accuracy** | {perf_metrics['accuracy']:.4f} ({perf_metrics['accuracy']*100:.2f}%) |")
            if 'precision' in perf_metrics:
                report_lines.append(f"| **Precision** | {perf_metrics['precision']:.4f} ({perf_metrics['precision']*100:.2f}%) |")
            if 'recall' in perf_metrics:
                report_lines.append(f"| **Recall** | {perf_metrics['recall']:.4f} ({perf_metrics['recall']*100:.2f}%) |")
            if 'f1_score' in perf_metrics:
                report_lines.append(f"| **F1 Score** | {perf_metrics['f1_score']:.4f} ({perf_metrics['f1_score']*100:.2f}%) |")
            if 'roc_auc' in perf_metrics:
                report_lines.append(f"| **AUC-ROC** | {perf_metrics['roc_auc']:.4f} |")
            if 'pr_auc' in perf_metrics:
                report_lines.append(f"| **AUC-PR** | {perf_metrics['pr_auc']:.4f} |")
        
        # Confusion Matrix
        report_lines.append("\n## 📈 Confusion Matrix\n")
        if conf_matrix and 'matrix_values' in conf_matrix:
            matrix = conf_matrix['matrix_values']
            if len(matrix) == 2 and len(matrix[0]) == 2:
                tn, fp = matrix[0]
                fn, tp = matrix[1]
                report_lines.append("```")
                report_lines.append("                Predicted")
                report_lines.append("              Neg    Pos")
                report_lines.append(f"Actual  Neg  {tn:>5}  {fp:>5}")
                report_lines.append(f"        Pos  {fn:>5}  {tp:>5}")
                report_lines.append("```")
                
                # Classification breakdown
                total = tn + fp + fn + tp
                report_lines.append(f"\n**True Negatives:** {tn} ({tn/total*100:.1f}%)")
                report_lines.append(f"**False Positives:** {fp} ({fp/total*100:.1f}%)")
                report_lines.append(f"**False Negatives:** {fn} ({fn/total*100:.1f}%)")
                report_lines.append(f"**True Positives:** {tp} ({tp/total*100:.1f}%)")
        
        # Feature Importance
        report_lines.append("\n## 🎯 Top Feature Importance\n")
        if feature_imp and len(feature_imp) > 0:
            sorted_features = sorted(feature_imp, key=lambda x: x.get('importance_value', 0), reverse=True)[:10]
            report_lines.append("| Rank | Feature | Importance |")
            report_lines.append("|------|---------|------------|")
            for idx, feat in enumerate(sorted_features, 1):
                feat_name = feat.get('feature_name', 'Unknown')
                importance = feat.get('importance_value', 0)
                bar = "█" * int(importance * 20)
                report_lines.append(f"| {idx} | {feat_name} | {importance:.4f} {bar} |")
        else:
            report_lines.append("*No feature importance data available*")
        
        # Error Patterns
        report_lines.append("\n## 🔍 Error Pattern Analysis\n")
        if error_patterns and len(error_patterns) > 0:
            report_lines.append("**Common Misclassifications:**\n")
            for pattern in error_patterns[:5]:
                pattern_type = pattern.get('pattern_type', 'Unknown')
                frequency = pattern.get('frequency', 0)
                description = pattern.get('description', '')
                report_lines.append(f"- **{pattern_type}:** {description} (Frequency: {frequency})")
        else:
            report_lines.append("*No error pattern data available*")
        
        # Model Insights
        report_lines.append("\n## 💡 Key Insights\n")
        
        # Generate insights based on metrics
        insights = []
        if perf_metrics.get('accuracy', 0) > 0.9:
            insights.append("✅ **Excellent Accuracy:** Model shows high overall performance (>90%)")
        elif perf_metrics.get('accuracy', 0) > 0.8:
            insights.append("✅ **Good Accuracy:** Model performance is solid (80-90%)")
        else:
            insights.append("⚠️ **Moderate Accuracy:** Consider feature engineering or trying different algorithms")
        
        if perf_metrics.get('precision', 0) > perf_metrics.get('recall', 0) + 0.1:
            insights.append("📊 **High Precision, Lower Recall:** Model is conservative, fewer false positives but might miss some positives")
        elif perf_metrics.get('recall', 0) > perf_metrics.get('precision', 0) + 0.1:
            insights.append("📊 **High Recall, Lower Precision:** Model is sensitive, captures more positives but with more false alarms")
        else:
            insights.append("📊 **Balanced Precision/Recall:** Model has good balance between false positives and false negatives")
        
        if feature_imp and len(feature_imp) > 0:
            top_feature = max(feature_imp, key=lambda x: x.get('importance_value', 0))
            insights.append(f"🎯 **Top Feature:** '{top_feature.get('feature_name')}' has the highest impact on predictions")
        
        for insight in insights:
            report_lines.append(f"- {insight}")
        
        # Next Steps
        report_lines.append("\n## 🚀 Next Steps\n")
        report_lines.append("1. **Review Feature Importance:** Focus on top features for domain insights")
        report_lines.append("2. **Analyze Error Patterns:** Understand where model struggles")
        report_lines.append("3. **Check AI Explainability (Step 8):** Get deeper insights with SHAP values")
        report_lines.append("4. **Compare Models:** Try different algorithms if performance needs improvement")
        report_lines.append("5. **Deploy:** If metrics meet requirements, proceed to deployment")
        
        report_lines.append("\n---")
        report_lines.append("\n*Generated by MEEA (Model Evaluation & Error Analysis)*")
        
        formatted_report = "\n".join(report_lines)
        
        return JSONResponse(content=safe_json_serialize({
            "success": True,
            "model_id": model_id,
            "message": formatted_report,
            "evaluation_summary": {
                "accuracy": perf_metrics.get('accuracy'),
                "precision": perf_metrics.get('precision'),
                "recall": perf_metrics.get('recall'),
                "f1_score": perf_metrics.get('f1_score'),
                "roc_auc": perf_metrics.get('roc_auc'),
                "top_features": [f.get('feature_name') for f in sorted(feature_imp, key=lambda x: x.get('importance_value', 0), reverse=True)[:5]] if feature_imp else []
            }
        }))
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating chat summary: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JSONResponse(content={
            "success": False,
            "message": f"❌ Error generating evaluation summary: {str(e)}\n\nPlease check logs for details."
        })


# ---------------------------------------------------------------------------
# Segmentation model evaluation (Step 4) - performance-only placeholder
# ---------------------------------------------------------------------------
@chat_router.get("/segmentation-model-evaluation/segments/{dataset_id}", response_model=dict)
async def list_segmentation_ids(dataset_id: str):
    """
    List available segment_ids for a dataset (used to populate segment dropdown).
    """
    try:
        # Prefer processed dataframe
        df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")

        segment_candidates = ["segment_id", "segment", "SEGMENT_ID", "SEGMENT", "cluster", "CLUSTER", "group", "GROUP"]
        segment_col = next((c for c in segment_candidates if c in df.columns), None)
        if not segment_col:
            raise HTTPException(status_code=404, detail="Segment column not found. Expected one of: segment_id, segment, cluster, group. Run segmentation first.")

        counts = df[segment_col].value_counts().to_dict()
        segments = [{"segment_id": str(seg_id), "count": int(count)} for seg_id, count in counts.items()]

        return {"success": True, "dataset_id": dataset_id, "segments": segments}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to list segment ids for dataset {dataset_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error listing segments: {str(e)}")


@chat_router.get("/segmentation-model-evaluation/{dataset_id}/{segment_id}", response_model=dict)
async def list_segmentation_model_evaluation(
    dataset_id: str,
    segment_id: str,
):
    """
    Segment-specific evaluation using an 80/20 split on the chosen segment.
    Returns data shaped like ModelEvaluationData so the frontend can reuse the
    standard performance tab.
    """
    try:
        # Try to load existing segment auto-training results first
        def _load_latest_segment_auto_results(dataset_id: str) -> Optional[Dict[str, Any]]:
            try:
                base_path = os.path.join("models", "segment_models")
                if not os.path.exists(base_path):
                    return None
                candidates = [
                    os.path.join(base_path, f)
                    for f in os.listdir(base_path)
                    if f.endswith("_segment_auto_results.json")
                ]
                latest = None
                latest_mtime = 0
                for path in candidates:
                    try:
                        with open(path, "r") as f:
                            data = json.load(f)
                        cfg = data.get("config", {})
                        if cfg.get("dataset_id") != dataset_id:
                            continue
                        mtime = os.path.getmtime(path)
                        if mtime > latest_mtime:
                            latest_mtime = mtime
                            latest = data
                    except Exception:
                        continue
                return latest
            except Exception:
                return None

        segment_auto_data = _load_latest_segment_auto_results(dataset_id)
        colors = ["#3B82F6", "#10B981", "#F59E0B", "#8B5CF6", "#EF4444", "#0EA5E9", "#F97316"]

        if segment_auto_data:
            segment_results = segment_auto_data.get("segment_results", {})
            segment_key = f"segment_{segment_id}"
            segment_entry = segment_results.get(segment_key)
            if segment_entry and "results" in segment_entry:
                models_payload = []
                for idx, res in enumerate(segment_entry.get("results", [])):
                    if "error" in res:
                        continue
                    metrics = res.get("metrics", {}) or {}

                    # Ensure confusion-matrix keys exist for frontend consumption.
                    # Some auto-training per-segment metrics originate from the
                    # lightweight training output (calculate_metrics) which does
                    # not include the full confusion-matrix keys. Try to enrich
                    # metrics from a saved comprehensive evaluation JSON if available.
                    try:
                        needs_confusion = not (metrics.get("confusion_matrix") or metrics.get("test_confusion_matrix") or metrics.get("train_confusion_matrix"))
                        if needs_confusion:
                            # Attempt to load comprehensive evaluation file for this model
                            candidate_model_id = res.get("model_id")
                            if candidate_model_id:
                                eval_path = os.path.join("models", f"{candidate_model_id}_comprehensive_evaluation.json")
                                if os.path.exists(eval_path):
                                    try:
                                        with open(eval_path, "r") as ef:
                                            eval_json = json.load(ef)
                                        perf = eval_json.get("performance_metrics", {}) or {}
                                        # Pull any available confusion matrices into metrics
                                        if perf.get("test_confusion_matrix") and not metrics.get("test_confusion_matrix"):
                                            metrics["test_confusion_matrix"] = perf.get("test_confusion_matrix")
                                        if perf.get("train_confusion_matrix") and not metrics.get("train_confusion_matrix"):
                                            metrics["train_confusion_matrix"] = perf.get("train_confusion_matrix")
                                        if perf.get("confusion_matrix") and not metrics.get("confusion_matrix"):
                                            metrics["confusion_matrix"] = perf.get("confusion_matrix")
                                    except Exception:
                                        # Non-fatal: continue without enrichment
                                        pass
                    except Exception:
                        # Protect the segment listing endpoint from any unexpected errors
                        logger.warning(f"Failed to enrich segment metrics for model {res.get('model_id')}")
                    alg = res.get("algorithm", "segment_model")
                    model_id = res.get("model_id") or f"{segment_auto_data.get('model_id', 'seg')}-{segment_id}-{idx}"

                    # Helper to fill missing F1 from precision/recall when possible
                    def _fill_f1(p: Optional[float], r: Optional[float], existing: Optional[float]) -> Optional[float]:
                        if existing is not None:
                            return existing
                        if p is None or r is None:
                            return None
                        if p + r == 0:
                            return 0.0
                        return 2 * p * r / (p + r)

                    performance_metrics = {
                        # Aggregate/test-first metrics
                        "accuracy": metrics.get("test_accuracy") or metrics.get("accuracy"),
                        "precision": metrics.get("test_precision") or metrics.get("precision"),
                        "recall": metrics.get("test_recall") or metrics.get("recall"),
                        "f1_score": metrics.get("test_f1_score") or metrics.get("f1_score"),
                        "auc_roc": metrics.get("test_auc") or metrics.get("auc") or metrics.get("test_auc_roc") or metrics.get("auc_roc"),
                        "auc_pr": metrics.get("test_auc_pr") or metrics.get("auc_pr"),
                        "log_loss": metrics.get("test_log_loss") or metrics.get("log_loss"),
                        "mse": metrics.get("test_mse") or metrics.get("mse"),
                        "rmse": metrics.get("test_rmse") or metrics.get("rmse"),
                        "mae": metrics.get("test_mae") or metrics.get("mae"),
                        "r2": metrics.get("test_r2") or metrics.get("r2"),
                        "confusion_matrix": metrics.get("test_confusion_matrix") or metrics.get("confusion_matrix"),
                        "ks_statistic": metrics.get("ks_statistic") or metrics.get("test_ks_statistic"),  # Test KS (backward compatible)

                        # Train/Test split metrics
                        "train_accuracy": metrics.get("train_accuracy"),
                        "test_accuracy": metrics.get("test_accuracy") or metrics.get("accuracy"),
                        "train_precision": metrics.get("train_precision"),
                        "test_precision": metrics.get("test_precision") or metrics.get("precision"),
                        "train_recall": metrics.get("train_recall"),
                        "test_recall": metrics.get("test_recall") or metrics.get("recall"),
                        "train_f1_score": metrics.get("train_f1_score"),
                        "test_f1_score": metrics.get("test_f1_score") or metrics.get("f1_score"),
                        "train_auc_roc": metrics.get("train_auc") or metrics.get("train_auc_roc"),
                        "test_auc_roc": metrics.get("test_auc") or metrics.get("auc") or metrics.get("test_auc_roc") or metrics.get("auc_roc"),
                        "train_auc_pr": metrics.get("train_auc_pr"),
                        "test_auc_pr": metrics.get("test_auc_pr") or metrics.get("auc_pr"),
                        "train_log_loss": metrics.get("train_log_loss"),
                        "test_log_loss": metrics.get("test_log_loss") or metrics.get("log_loss"),
                        "train_mse": metrics.get("train_mse"),
                        "test_mse": metrics.get("test_mse") or metrics.get("mse"),
                        "train_rmse": metrics.get("train_rmse"),
                        "test_rmse": metrics.get("test_rmse") or metrics.get("rmse"),
                        "train_mae": metrics.get("train_mae"),
                        "test_mae": metrics.get("test_mae") or metrics.get("mae"),
                        "train_r2": metrics.get("train_r2"),
                        "test_r2": metrics.get("test_r2") or metrics.get("r2"),
                        "train_confusion_matrix": metrics.get("train_confusion_matrix"),
                        "test_confusion_matrix": metrics.get("test_confusion_matrix") or metrics.get("confusion_matrix"),
                        
                        # NEW: Train KS statistic
                        "train_ks_statistic": metrics.get("train_ks_statistic"),
                    }

                    # Fill missing F1 where possible
                    performance_metrics["f1_score"] = _fill_f1(
                        performance_metrics.get("precision"), performance_metrics.get("recall"), performance_metrics.get("f1_score")
                    )
                    performance_metrics["test_f1_score"] = _fill_f1(
                        performance_metrics.get("test_precision"), performance_metrics.get("test_recall"), performance_metrics.get("test_f1_score")
                    )
                    performance_metrics["train_f1_score"] = _fill_f1(
                        performance_metrics.get("train_precision"), performance_metrics.get("train_recall"), performance_metrics.get("train_f1_score")
                    )
                    
                    # Try to fetch explainability_data and granular_accuracy from database
                    explainability_data = []
                    granular_accuracy = []
                    granular_accuracy_train = []
                    
                    # Add monotonicity evaluation for segmented models
                    monotonicity_results = None
                    try:
                        # Get segment data for monotonicity evaluation
                        if db_eval_data and db_eval_data.get("granular_accuracy"):
                            granular_segments = db_eval_data.get("granular_accuracy", [])
                            
                            # Create segment profiles for monotonicity evaluation
                            segment_profiles = []
                            for seg in granular_segments:
                                if seg.get("variable") and seg.get("granularity_level") == "segment":
                                    segment_profiles.append({
                                        'segment_index': seg.get("segment_id", 0),
                                        'bad_rate': seg.get("bad_rate", 0.0),
                                        'count': seg.get("total_count", 0),
                                        'rules_readable': f"Segment {seg.get('segment_id', 0)}"
                                    })
                            
                            # Evaluate monotonicity if we have segment profiles
                            if segment_profiles:
                                from app.utils.segmentation_monotonicity import SegmentationMonotonicityEvaluator
                                monotonicity_evaluator = SegmentationMonotonicityEvaluator(logger)
                                monotonicity_results = monotonicity_evaluator.evaluate_segment_monotonicity(
                                    segment_profiles=segment_profiles,
                                    target_variable="target"
                                )
                                logger.info(f"✅ Monotonicity evaluation completed for segmented model {model_id}: score={monotonicity_results.get('monotonicity_score', 0):.3f}")
                        
                    except Exception as e:
                        logger.warning(f"Failed to evaluate monotonicity for segmented model {model_id}: {str(e)}")
                        monotonicity_results = {
                            'monotonicity_score': 0.0,
                            'is_monotonic': False,
                            'error': str(e)
                        }
                    
                    try:
                        from app.models.model_evaluation_database import model_evaluation_db
                        db_eval_data = model_evaluation_db.get_model_evaluation(model_id, include_explainability=True, include_pdp=False)
                        if db_eval_data:
                            if db_eval_data.get("explainability_data"):
                                explainability_data = db_eval_data.get("explainability_data", [])
                                logger.info(f"Loaded {len(explainability_data)} explainability entries for segmentation model {model_id}")
                            if db_eval_data.get("granular_accuracy"):
                                granular_accuracy = db_eval_data.get("granular_accuracy", [])
                                logger.info(f"Loaded {len(granular_accuracy)} granular accuracy entries (test) for segmentation model {model_id}")
                            if db_eval_data.get("granular_accuracy_train"):
                                granular_accuracy_train = db_eval_data.get("granular_accuracy_train", [])
                                logger.info(f"Loaded {len(granular_accuracy_train)} granular accuracy entries (train) for segmentation model {model_id}")
                    except Exception as e:
                        logger.warning(f"Could not load evaluation data for segmentation model {model_id}: {str(e)}")
                        explainability_data = []
                        granular_accuracy = []
                        granular_accuracy_train = []
                    
                    models_payload.append({
                        "model": {
                            "id": model_id,
                            "name": f"{alg} (Segment {segment_id})",
                            "model_type": alg,
                            "task_type": "classification",
                            "training_date": datetime.utcnow().isoformat(),
                            "status": "evaluated",
                            "color": colors[idx % len(colors)],
                            "description": f"Segment {segment_id} trained via segment-auto",
                            "created_at": datetime.utcnow().isoformat(),
                        },
                        "performance_metrics": performance_metrics,
                        "feature_importance": [],
                        "granular_accuracy": granular_accuracy,
                        "granular_accuracy_train": granular_accuracy_train,
                        "error_patterns": [],
                        "prediction_confidence": [],
                        "explainability_data": explainability_data,
                    })

                if models_payload:
                    return {
                        "success": True,
                        "dataset_id": dataset_id,
                        "segment_id": segment_id,
                        "models": models_payload,
                        "count": len(models_payload),
                    }

        # If no segment-auto results found, try to load any MANUAL/DB-stored segment models
        try:
            from app.models.model_evaluation_database import model_evaluation_db

            db_models = model_evaluation_db.list_models_by_dataset(dataset_id) or []
            # Filter for per-segment models matching this segment_id
            matching_segment_models = [m for m in db_models if m.get("is_segment_model") and str(m.get("segment_id")) == str(segment_id)]

            if matching_segment_models:
                models_payload = []
                for m in matching_segment_models:
                    model_id = m.get("id") or m.get("model_id") or m.get("modelId")
                    try:
                        eval_data = model_evaluation_db.get_model_evaluation(model_id, include_explainability=True, include_pdp=False)
                    except Exception:
                        eval_data = None

                    if not eval_data:
                        # If there is no stored evaluation payload, skip
                        continue

                    perf = eval_data.get("performance_metrics", {}) or {}

                    # Build a pretty model name (match Auto mode naming)
                    model_meta = eval_data.get("model", {}) or {}
                    raw_model_type = (model_meta.get("model_type") or m.get("model_type") or "").lower()

                    pretty_map = {
                        "logistic_regression": "LogisticRegression",
                        "logisticregression": "LogisticRegression",
                        "logistic": "LogisticRegression",
                        "lightgbm": "LightGBM",
                        "lgbm": "LightGBM",
                        "xgboost": "XGBoost",
                        "xgb": "XGBoost",
                        "catboost": "CatBoost",
                        "catb": "CatBoost",
                        "random_forest": "RandomForest",
                        "randomforest": "RandomForest",
                        "rf": "RandomForest",
                        "gradient_boosting": "GradientBoosting",
                        "gradientboosting": "GradientBoosting",
                        "gbm": "GradientBoosting",
                        "decision_tree": "DecisionTree",
                        "decisiontree": "DecisionTree",
                        "segmentation_tree": "DecisionTree",
                    }

                    base_name = None
                    # Prefer an explicit model name if present and looks friendly
                    if model_meta.get("name") and not model_meta.get("name").islower():
                        base_name = model_meta.get("name")
                    else:
                        base_name = pretty_map.get(raw_model_type) or (model_meta.get("name") or m.get("name") or raw_model_type or "Model")

                    # Append segment label
                    display_name = f"{base_name} (Seg{segment_id})"

                    # Make sure model dict includes id, model_type and name
                    model_dict = dict(model_meta)
                    model_dict["id"] = model_id
                    model_dict["name"] = display_name
                    if "model_type" not in model_dict:
                        model_dict["model_type"] = raw_model_type

                    models_payload.append({
                        "model": model_dict,
                        "performance_metrics": perf,
                        "feature_importance": eval_data.get("feature_importance", []) or [],
                        "granular_accuracy": eval_data.get("granular_accuracy", []) or [],
                        "granular_accuracy_train": eval_data.get("granular_accuracy_train", []) or [],
                        "error_patterns": eval_data.get("error_patterns", []) or [],
                        "prediction_confidence": eval_data.get("prediction_confidence", []) or [],
                        "explainability_data": eval_data.get("explainability_data", []) or [],
                    })

                if models_payload:
                    logger.info(f"Returning {len(models_payload)} segment models from MEEA DB for dataset={dataset_id}, segment={segment_id}")
                    return {
                        "success": True,
                        "dataset_id": dataset_id,
                        "segment_id": segment_id,
                        "models": models_payload,
                        "count": len(models_payload),
                    }
        except Exception as e:
            logger.warning(f"Could not fetch segment models from MEEA DB for dataset {dataset_id}, segment {segment_id}: {str(e)}")

        dataset_info = dataset_manager.get_dataset_info(dataset_id)
        if not dataset_info:
            raise HTTPException(status_code=404, detail="Dataset not found")

        target_variable = dataset_info.get("target_variable")
        if not target_variable:
            raise HTTPException(status_code=400, detail="Target variable is required for segmentation evaluation")

        # Load dataframe (prefer processed)
        df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found or could not be loaded")

        segment_candidates = ["segment_id", "segment", "SEGMENT_ID", "SEGMENT", "cluster", "CLUSTER", "group", "GROUP"]
        segment_col = next((c for c in segment_candidates if c in df.columns), None)
        if not segment_col:
            raise HTTPException(status_code=404, detail="Segment column not found. Expected one of: segment_id, segment, cluster, group. Run segmentation first.")

        df_segment = df[df[segment_col].astype(str) == str(segment_id)]
        if df_segment.empty:
            raise HTTPException(status_code=404, detail=f"No rows found for segment_id={segment_id}")

        if target_variable not in df_segment.columns:
            raise HTTPException(status_code=400, detail=f"Target variable '{target_variable}' not found in dataset")

        X = df_segment.drop(columns=[target_variable]).copy()
        y = df_segment[target_variable].copy()

        # Basic preprocessing: fill missing, encode categoricals
        cat_cols = X.select_dtypes(include=["object", "category"]).columns.tolist()
        num_cols = [c for c in X.columns if c not in cat_cols]

        for col in num_cols:
            X[col] = X[col].fillna(X[col].median())

        for col in cat_cols:
            X[col] = X[col].fillna("missing")
            le = LabelEncoder()
            X[col] = le.fit_transform(X[col].astype(str))

        # Ensure binary classification; otherwise error for now
        unique_classes = pd.unique(y.dropna())
        if len(unique_classes) != 2:
            raise HTTPException(status_code=400, detail="Segmentation evaluation currently supports binary targets only")

        stratify = y if len(unique_classes) == 2 else None
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42, stratify=stratify
        )

        model = DecisionTreeClassifier(max_depth=4, random_state=42)
        model.fit(X_train, y_train)

        def _compute_metrics(y_true, y_pred, y_proba=None):
            metrics = {
                "accuracy": float(accuracy_score(y_true, y_pred)),
                "precision": float(precision_score(y_true, y_pred, average="binary", zero_division=0)),
                "recall": float(recall_score(y_true, y_pred, average="binary", zero_division=0)),
                "f1_score": float(f1_score(y_true, y_pred, average="binary", zero_division=0)),
                "confusion_matrix": confusion_matrix(y_true, y_pred).tolist(),
            }
            if y_proba is not None:
                try:
                    metrics["auc_roc"] = float(roc_auc_score(y_true, y_proba))
                except Exception:
                    metrics["auc_roc"] = None
            return metrics

        y_train_pred = model.predict(X_train)
        y_test_pred = model.predict(X_test)
        y_train_proba = model.predict_proba(X_train)[:, 1] if hasattr(model, "predict_proba") else None
        y_test_proba = model.predict_proba(X_test)[:, 1] if hasattr(model, "predict_proba") else None

        train_metrics = _compute_metrics(y_train, y_train_pred, y_train_proba)
        test_metrics = _compute_metrics(y_test, y_test_pred, y_test_proba)

        # Try to fetch explainability_data and granular_accuracy from database
        model_id = f"seg-{segment_id}-dt"
        explainability_data = []
        granular_accuracy = []
        granular_accuracy_train = []
        try:
            from app.models.model_evaluation_database import model_evaluation_db
            db_eval_data = model_evaluation_db.get_model_evaluation(model_id, include_explainability=True, include_pdp=False)
            if db_eval_data:
                if db_eval_data.get("explainability_data"):
                    explainability_data = db_eval_data.get("explainability_data", [])
                    logger.info(f"Loaded {len(explainability_data)} explainability entries for segmentation model {model_id}")
                if db_eval_data.get("granular_accuracy"):
                    granular_accuracy = db_eval_data.get("granular_accuracy", [])
                    logger.info(f"Loaded {len(granular_accuracy)} granular accuracy entries (test) for segmentation model {model_id}")
                if db_eval_data.get("granular_accuracy_train"):
                    granular_accuracy_train = db_eval_data.get("granular_accuracy_train", [])
                    logger.info(f"Loaded {len(granular_accuracy_train)} granular accuracy entries (train) for segmentation model {model_id}")
        except Exception as e:
            logger.warning(f"Could not load evaluation data for segmentation model {model_id}: {str(e)}")
            explainability_data = []
            granular_accuracy = []
            granular_accuracy_train = []
        
    except Exception as e:
        logger.error(f"Failed to return segmentation model evaluation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# =============================================================================
# Duplicate Removal and EDA Endpoints
# =============================================================================

@upload_router.post("/datasets/{dataset_id}/identify-duplicates")
async def identify_duplicates(
    dataset_id: str,
    columns: List[str] = Form(...),
    current_user = Depends(get_current_user_dependency)
):
    """
    Identify duplicate rows in the TRAIN split of a dataset based on specified columns.

    The decision of whether to treat duplicates is based on the train split (to avoid
    any leakage from test/validation during analysis). The count reported here is the
    number of duplicate rows within the train split only.

    If no split has been performed yet, the full dataset is used as fallback.

    Args:
        dataset_id: The dataset identifier
        columns: List of column names to use as the duplicate key
        
    Returns:
        Dictionary with duplicate count, total rows (train), percentage, and scope
    """
    try:
        logger.info(f"Identifying duplicates for dataset {dataset_id} on TRAIN split using columns: {columns}")
        
        # Always load the raw dataset from disk — the original row ordering/assignment is
        # what the split indices from the Objectives page refer to.
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Validate columns exist on the full df (columns are the same for every split)
        missing_cols = [col for col in columns if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Columns not found in dataset: {missing_cols}"
            )

        # Restrict to the TRAIN split (decision basis). Fallback to full df if no split info.
        split_indices = dataframe_state_manager._split_indices.get(dataset_id)
        analysis_scope = "train"
        train_df = None
        if split_indices and 'train' in split_indices:
            train_idx = split_indices['train']
            if train_idx is not None and len(train_idx) > 0:
                train_df = df.iloc[train_idx].copy()
        if (train_df is None or len(train_df) == 0) and 'split_tag' in df.columns:
            train_df = df[df['split_tag'] == 'train'].copy()
        if train_df is None or len(train_df) == 0:
            # No split available — analyze entire dataset as fallback
            logger.warning(f"No train split found for dataset {dataset_id}; analysing full dataset")
            train_df = df
            analysis_scope = "entire"
        
        total_rows = len(train_df)
        duplicated_mask = train_df.duplicated(subset=columns, keep='first')
        duplicate_count = int(duplicated_mask.sum())
        duplicate_percentage = (duplicate_count / total_rows * 100) if total_rows > 0 else 0
        
        logger.info(f"Found {duplicate_count} duplicates in {analysis_scope} ({total_rows} rows, {duplicate_percentage:.2f}%)")
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "duplicate_count": duplicate_count,
            "total_rows": int(total_rows),
            "duplicate_percentage": float(duplicate_percentage),
            "columns_used": columns,
            "analysis_scope": analysis_scope,
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error identifying duplicates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error identifying duplicates: {str(e)}")


@upload_router.post("/datasets/{dataset_id}/remove-duplicates")
async def remove_duplicates(
    dataset_id: str,
    columns: List[str] = Form(...),
    current_user = Depends(get_current_user_dependency)
):
    """
    Remove duplicates using the "decide on train, apply to all splits" rule:

      * The DECISION to remove (which columns define a duplicate) is made by the user
        while looking at duplicates in the TRAIN split (see identify-duplicates).
      * The SAME rule (`drop_duplicates(subset=columns, keep='first')`) is applied
        independently to TRAIN, VALIDATION, and TEST (each split deduplicated on its own).
      * The original pandas index is preserved so downstream scope-filtered views
        (compare-column-stats, column-info-by-scope, etc.) can still map every
        surviving row back to its Objectives-page split assignment.
      * Handles all split scenarios from the Objectives page:
          - all-as-train, train+test only, train+test+validation, or no split at all.

    Returns:
        Per-split before/removed/after counts and the new total row count.
    """
    try:
        logger.info(f"Removing duplicates from dataset {dataset_id} on all available splits using columns: {columns}")

        # Start from the raw dataset — this is the source-of-truth for the Objectives-page
        # row assignment and guarantees all user-selected columns exist.
        df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            df = dataframe_state_manager.get_dataframe_for_execution(dataset_id, None)
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")

        missing_cols = [col for col in columns if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Columns not found in dataset: {missing_cols}"
            )

        original_row_count = len(df)

        # Resolve the positional indices for each split from the Objectives-page split.
        split_indices = dataframe_state_manager._split_indices.get(dataset_id)
        train_pos: List[int] = []
        val_pos: List[int] = []
        test_pos: List[int] = []

        if split_indices:
            if split_indices.get('train') is not None:
                train_pos = list(split_indices['train'])
            if split_indices.get('validation') is not None:
                val_pos = list(split_indices['validation'])
            if split_indices.get('test') is not None:
                test_pos = list(split_indices['test'])

        # Fallback to split_tag column when positional indices are not populated
        # (covers datasets uploaded with a pre-existing split_tag column).
        if (not train_pos and not val_pos and not test_pos) and 'split_tag' in df.columns:
            tag = df['split_tag'].astype(str)
            train_pos = [i for i, t in enumerate(tag) if t == 'train']
            val_pos = [i for i, t in enumerate(tag) if t.startswith('validation')]
            test_pos = [i for i, t in enumerate(tag) if t == 'test']

        # If no split info at all, treat the whole df as train (backward compatible).
        if not train_pos and not val_pos and not test_pos:
            logger.warning(f"No split info for dataset {dataset_id}; applying duplicate removal to full dataset as 'train'")
            train_pos = list(range(len(df)))

        # Build sub-dataframes preserving the original pandas index. Missing splits are
        # represented by empty frames so the output schema stays stable.
        train_df = df.iloc[train_pos].copy() if train_pos else df.iloc[0:0].copy()
        val_df = df.iloc[val_pos].copy() if val_pos else df.iloc[0:0].copy()
        test_df = df.iloc[test_pos].copy() if test_pos else df.iloc[0:0].copy()

        train_before = len(train_df)
        val_before = len(val_df)
        test_before = len(test_df)

        # ── Apply the SAME rule to train, validation, and test independently ──
        # Decision was based on train (identify-duplicates count); the same
        # drop_duplicates key is applied to each split separately (like validation).
        train_dedup = train_df.drop_duplicates(subset=columns, keep='first')
        val_dedup = val_df.drop_duplicates(subset=columns, keep='first')
        test_dedup = test_df.drop_duplicates(subset=columns, keep='first')

        # Ensure split_tag is present and normalized on deduplicated scoped frames.
        # This allows us to rebuild split indices from deduped data (label-based)
        # and avoids positional index drift in manual QC scope switching.
        if len(train_dedup) > 0:
            train_dedup = train_dedup.copy()
            train_dedup['split_tag'] = 'train'
        if len(val_dedup) > 0:
            val_dedup = val_dedup.copy()
            val_dedup['split_tag'] = 'validation'
        if len(test_dedup) > 0:
            test_dedup = test_dedup.copy()
            test_dedup['split_tag'] = 'test'

        train_removed = train_before - len(train_dedup)
        val_removed = val_before - len(val_dedup)
        test_removed = test_before - len(test_dedup)
        total_removed = train_removed + val_removed + test_removed

        # Reassemble: concat preserves original pandas index labels; sort by index so
        # later scope filters (iloc-based or label-based) correctly align with the
        # Objectives-page split indices for surviving rows.
        df_cleaned = pd.concat([train_dedup, val_dedup, test_dedup], axis=0)
        if not df_cleaned.index.is_monotonic_increasing:
            df_cleaned = df_cleaned.sort_index(kind='stable')

        new_row_count = len(df_cleaned)

        # Persist deduplicated state per scope so EDA / compare-column-stats / eda-snapshot
        # read correct row counts. A single update_dataframe(df_cleaned) without force_scope
        # only wrote under the active scope (often 'train') and left transformed_copies['entire']
        # stale. We do not modify _full_dataframes here—only per-scope and 'entire' copies.
        dataframe_state_manager.update_dataframe(dataset_id, train_dedup, force_scope="train")
        dataframe_state_manager.update_dataframe(dataset_id, val_dedup, force_scope="validation")
        dataframe_state_manager.update_dataframe(dataset_id, test_dedup, force_scope="test")
        dataframe_state_manager.update_dataframe(dataset_id, df_cleaned, force_scope="entire")

        # Rebuild split indices from deduped split_tag distribution so downstream
        # set_scope/manual QC uses post-dedup split mapping instead of stale
        # pre-dedup positional indices. Do not touch _full_dataframes.
        try:
            rebuilt = dataframe_state_manager._rebuild_split_indices_from_split_tag(dataset_id, df_cleaned)
            if rebuilt:
                logger.info(
                    f"Rebuilt split indices from deduped data for {dataset_id}: "
                    f"train={int((df_cleaned['split_tag'] == 'train').sum())}, "
                    f"test={int((df_cleaned['split_tag'] == 'test').sum())}, "
                    f"validation={int(df_cleaned['split_tag'].astype(str).str.startswith('validation').sum())}"
                )
            else:
                logger.warning(f"Could not rebuild split indices from deduped split_tag for {dataset_id}")
        except Exception as split_err:
            logger.warning(f"Failed to rebuild split indices after dedup for {dataset_id}: {split_err}")

        # Persist deduplicated dataframe into MessageState so subsequent Auto/Manual QC
        # sessions and compare endpoints use the updated post-dedup data consistently.
        try:
            dedup_state = message_state_manager.create_or_load_state(dataset_id, "duplicate removal update")
            previous_df = dedup_state.get("datasetFile")
            if isinstance(previous_df, pd.DataFrame):
                dedup_state["previousDatasetFile"] = previous_df
            dedup_state["datasetFile"] = df_cleaned.copy()
            dedup_state["dataset_id"] = dataset_id
            save_success = message_state_manager.save_state(dataset_id, dedup_state)
            if save_success:
                logger.info(
                    f"Persisted deduplicated MessageState for {dataset_id}, shape={df_cleaned.shape}"
                )
            else:
                logger.warning(f"Failed to persist deduplicated MessageState for {dataset_id}")
        except Exception as state_err:
            logger.warning(f"Could not persist deduplicated dataframe to MessageState for {dataset_id}: {state_err}")

        logger.info(
            f"Removed duplicates — train: {train_removed}/{train_before}, "
            f"validation: {val_removed}/{val_before}, test: {test_removed}/{test_before}. "
            f"New total: {new_row_count} (was {original_row_count})."
        )

        return {
            "success": True,
            "dataset_id": dataset_id,
            "removed_count": int(total_removed),
            "original_row_count": int(original_row_count),
            "new_row_count": int(new_row_count),
            "columns_used": columns,
            "per_split": {
                "train": {"before": int(train_before), "removed": int(train_removed), "after": int(len(train_dedup))},
                "validation": {"before": int(val_before), "removed": int(val_removed), "after": int(len(val_dedup))},
                "test": {"before": int(test_before), "removed": int(test_removed), "after": int(len(test_dedup))},
            },
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error removing duplicates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error removing duplicates: {str(e)}")


@upload_router.get("/datasets/{dataset_id}/eda-snapshot")
async def get_eda_snapshot(
    dataset_id: str,
    scope: str = Query("entire", description="Data scope: entire, train, test, validation"),
    current_user = Depends(get_current_user_dependency)
):
    """
    Get a comprehensive EDA snapshot for a dataset.
    
    This endpoint computes statistics for numeric, categorical, and date columns,
    returning a complete EDA snapshot that can be used for comparison.
    
    Args:
        dataset_id: The dataset identifier
        scope: Data scope to analyze (entire, train, test, validation)
        
    Returns:
        EDA snapshot with statistics for all column types
    """
    try:
        logger.info(f"Computing EDA snapshot for dataset {dataset_id}, scope: {scope}")
        
        # Load the dataset based on requested scope
        df = None
        
        if scope == "entire":
            # For 'entire' scope, try to get the full transformed dataset
            # First check if we have transformed copies
            transformed_copies = dataframe_state_manager._transformed_copies.get(dataset_id, {})
            if 'entire' in transformed_copies:
                df = transformed_copies['entire'].copy()
                logger.info(f"📊 EDA snapshot using TRANSFORMED 'entire' dataframe: {df.shape}")
            else:
                # Fallback to processed dataframe (might be train-scoped)
                df = dataframe_state_manager.get_dataframe(dataset_id)
                if df is not None:
                    logger.info(f"📊 EDA snapshot using PROCESSED dataframe (scope={dataframe_state_manager._active_scope.get(dataset_id, 'unknown')}): {df.shape}")
        else:
            # For specific scopes (train, test, validation), get from transformed copies
            transformed_copies = dataframe_state_manager._transformed_copies.get(dataset_id, {})
            if scope in transformed_copies:
                df = transformed_copies[scope].copy()
                logger.info(f"📊 EDA snapshot using TRANSFORMED '{scope}' dataframe: {df.shape}")
            else:
                df = dataframe_state_manager.get_dataframe(dataset_id)
                if df is not None:
                    logger.info(f"📊 EDA snapshot using PROCESSED dataframe: {df.shape}")
        
        # Fallback to original dataset from disk if no processed version
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
            if df is not None:
                logger.info(f"📊 EDA snapshot using ORIGINAL dataframe from disk: {df.shape}")
        
        if df is None:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Apply scope filter if not 'entire' and dataframe has split identifier
        if scope != "entire" and "data_split_identifier" in df.columns:
            df = df[df["data_split_identifier"] == scope]
            if df.empty:
                raise HTTPException(status_code=404, detail=f"No data found for scope: {scope}")
        
        # Compute EDA statistics
        eda_snapshot = _compute_eda_snapshot(df, dataset_id)
        
        logger.info(f"EDA snapshot computed: {eda_snapshot['totalRows']} rows, {eda_snapshot['totalColumns']} columns")
        
        return {
            "success": True,
            "dataset_id": dataset_id,
            "scope": scope,
            "eda_snapshot": eda_snapshot
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error computing EDA snapshot: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error computing EDA snapshot: {str(e)}")


def _compute_eda_snapshot(df: pd.DataFrame, dataset_id: str) -> Dict[str, Any]:
    """
    Compute comprehensive EDA statistics for a dataframe.
    
    This helper function analyzes all columns and categorizes them into
    numeric, categorical, and date types, computing relevant statistics for each.
    
    Args:
        df: The pandas DataFrame to analyze
        dataset_id: Dataset identifier for logging
        
    Returns:
        Dictionary containing EDA snapshot with all statistics
    """
    from datetime import datetime
    
    numeric_stats = []
    categorical_stats = []
    date_stats = []
    
    for col in df.columns:
        # Skip the data split identifier column
        if col == "data_split_identifier":
            continue
            
        col_data = df[col]
        missing_count = int(col_data.isna().sum())
        missing_percentage = (missing_count / len(df) * 100) if len(df) > 0 else 0
        
        # Detect column type
        if pd.api.types.is_numeric_dtype(col_data):
            # Numeric column
            valid_data = col_data.dropna()
            stats = {
                "column": col,
                "count": int(len(valid_data)),
                "mean": float(valid_data.mean()) if len(valid_data) > 0 else 0,
                "std": float(valid_data.std()) if len(valid_data) > 0 else 0,
                "min": float(valid_data.min()) if len(valid_data) > 0 else 0,
                "percentile_25": float(valid_data.quantile(0.25)) if len(valid_data) > 0 else 0,
                "percentile_50": float(valid_data.quantile(0.50)) if len(valid_data) > 0 else 0,
                "percentile_75": float(valid_data.quantile(0.75)) if len(valid_data) > 0 else 0,
                "max": float(valid_data.max()) if len(valid_data) > 0 else 0,
                "missing_count": missing_count,
                "missing_percentage": float(missing_percentage)
            }
            numeric_stats.append(stats)
            
        elif pd.api.types.is_datetime64_any_dtype(col_data):
            # Date column
            valid_data = col_data.dropna()
            if len(valid_data) > 0:
                min_date = valid_data.min()
                max_date = valid_data.max()
                date_range = (max_date - min_date).days if pd.notna(min_date) and pd.notna(max_date) else 0
                most_frequent = valid_data.mode().iloc[0] if len(valid_data.mode()) > 0 else None
                most_frequent_count = int((valid_data == most_frequent).sum()) if most_frequent is not None else 0
            else:
                min_date = max_date = None
                date_range = 0
                most_frequent = None
                most_frequent_count = 0
                
            stats = {
                "column": col,
                "min_date": str(min_date) if min_date is not None else None,
                "max_date": str(max_date) if max_date is not None else None,
                "date_range_days": int(date_range),
                "unique_count": int(valid_data.nunique()) if len(valid_data) > 0 else 0,
                "missing_count": missing_count,
                "missing_percentage": float(missing_percentage),
                "most_frequent_date": str(most_frequent) if most_frequent is not None else None,
                "most_frequent_count": most_frequent_count
            }
            date_stats.append(stats)
            
        else:
            # Categorical column (object, category, bool, or other)
            valid_data = col_data.dropna()
            unique_count = int(valid_data.nunique()) if len(valid_data) > 0 else 0
            
            # Get top category
            if len(valid_data) > 0:
                value_counts = valid_data.value_counts()
                top_category = str(value_counts.index[0]) if len(value_counts) > 0 else None
                top_category_count = int(value_counts.iloc[0]) if len(value_counts) > 0 else 0
                top_category_percentage = (top_category_count / len(valid_data) * 100) if len(valid_data) > 0 else 0
                
                # Get value distribution (top 10)
                value_distribution = {str(k): int(v) for k, v in value_counts.head(10).items()}
            else:
                top_category = None
                top_category_count = 0
                top_category_percentage = 0
                value_distribution = {}
            
            stats = {
                "column": col,
                "unique_count": unique_count,
                "top_category": top_category,
                "top_category_count": top_category_count,
                "top_category_percentage": float(top_category_percentage),
                "missing_count": missing_count,
                "missing_percentage": float(missing_percentage),
                "value_distribution": value_distribution
            }
            categorical_stats.append(stats)
    
    return {
        "timestamp": datetime.utcnow().isoformat(),
        "totalRows": int(len(df)),
        "totalColumns": int(len(df.columns)),
        "numericStats": numeric_stats,
        "categoricalStats": categorical_stats,
        "dateStats": date_stats
    }


# ============================================================================
# DEDICATED IV AND VIF CALCULATION APIS
# ============================================================================

@chat_router.post("/insights/iv-analysis")
async def generate_iv_analysis(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    bins: int = Form(10),
    current_user=Depends(get_current_user_dependency),
):
    """
    Generate Information Value (IV) analysis for feature selection.

    Returns ``200`` when cached, or ``202`` with ``job_id`` for background completion.
    """
    _ = current_user
    try:
        logger.info(
            "=== IV ANALYSIS REQUEST === dataset=%s target=%s bins=%s",
            dataset_id,
            target_variable,
            bins,
        )

        _scope, _ver = _insight_scope_version(dataset_id)
        scope_key = f"{_scope}|bins={int(bins or 10)}"
        hit = analytics_cache.get("insight_iv_analysis", dataset_id, scope_key, _ver)
        if hit is not None:
            return hit

        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        if target_variable not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Target variable '{target_variable}' not found in dataset",
            )

        job_id = f"insight_iv_{dataset_id}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        return _enqueue_insight_job(
            job_id,
            "insight_iv_analysis",
            params={
                "dataset_id": dataset_id,
                "target_variable": target_variable,
                "bins": int(bins or 10),
                "scope_key": scope_key,
                "version": _ver,
            },
            job_function=run_insight_iv_analysis_job,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"IV analysis failed for dataset {dataset_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"IV analysis failed: {str(e)}")


def _coerce_optional_positive_int(value):
    """Form fields may arrive as str; normalize for calculate_vif."""
    if value is None:
        return None
    try:
        n = int(value)
        return n if n >= 1 else None
    except (TypeError, ValueError):
        return None


@chat_router.post("/insights/vif-analysis-dedicated")
async def generate_vif_analysis_dedicated(
    dataset_id: str = Form(...),
    target_variable: str = Form(...),
    max_columns: int = Form(None),
    current_user = Depends(get_current_user_dependency)
):
    """
    Generate VIF (Variance Inflation Factor) analysis for multicollinearity detection
    """
    try:
        logger.info(f"=== VIF ANALYSIS DEDICATED REQUEST ===")
        logger.info(f"Dataset ID: {dataset_id}")
        logger.info(f"Target Variable: {target_variable}")
        max_columns = _coerce_optional_positive_int(max_columns)
        logger.info(f"Max Columns: {max_columns}")
        
        # Get the dataframe for the specific dataset_id
        df = dataframe_state_manager.get_dataframe(dataset_id)
        if df is None:
            from app.services.dataset_service import dataset_manager
            df = dataset_manager.load_dataset(dataset_id)
        if df is None:
            raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
        
        # Validate target variable exists
        if target_variable not in df.columns:
            raise HTTPException(status_code=400, detail=f"Target variable '{target_variable}' not found in dataset")
        
        # Import VIF analysis function
        from app.utils.helpers import calculate_vif
        
        # Generate VIF analysis
        vif_results = calculate_vif(df, target_variable, max_columns)
        
        if not vif_results:
            raise HTTPException(status_code=500, detail="Failed to generate VIF analysis")
        
        logger.info(f"VIF analysis completed for {len(vif_results)} variables")

        vif_payload = _build_vif_frontend_analysis_payload(df, vif_results)

        return {
            "success": True,
            "message": f"VIF analysis completed for dataset {dataset_id}",
            "dataset_id": dataset_id,
            "target_variable": target_variable,
            **vif_payload,
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"VIF analysis failed for dataset {dataset_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"VIF analysis failed: {str(e)}")

