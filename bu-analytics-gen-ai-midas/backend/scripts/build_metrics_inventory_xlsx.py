"""
Generator for observability metrics inventory: Excel (two sheets) + two CSV files.
Run: python scripts/build_metrics_inventory_xlsx.py
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)), 90))
        ws.column_dimensions[letter].width = max(12, max_len + 2)


def main() -> None:
    out = Path(__file__).resolve().parent.parent / "docs" / "metrics_observability_inventory.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    app = wb.active
    app.title = "Application"

    app_rows: list[tuple[str, str, str]] = [
        (
            "JSON log envelope (all structured lines)",
            "When LOG_FORMAT=json, each log line is one JSON object with timestamp (UTC ISO Z), level, service (LOG_SERVICE_NAME), environment (LOG_ENVIRONMENT), logger name, message string, and caller {file,line,function}.",
            "CloudWatch Logs Insights compatible; consistent parsing and filtering.",
        ),
        (
            "request_id",
            "HTTP correlation id from X-Request-ID header or generated UUID; stored in contextvars and echoed on responses.",
            "Tie all logs for one HTTP request together across handlers and LLM calls.",
        ),
        (
            "dataset_id",
            "Optional dataset correlation from contextvars when the app sets it.",
            "Link logs to a specific dataset without logging raw data.",
        ),
        (
            "user_id",
            "Numeric user id from authenticated session (UserInDB.id) after SessionValidationMiddleware succeeds.",
            "Audit and troubleshooting per user without logging PII like email in every line.",
        ),
        (
            "tenant_id",
            "Opaque tenant/org id from X-Tenant-ID header (trimmed, max 128 chars) when present.",
            "Multi-tenant filtering in logs when you adopt tenant headers.",
        ),
        (
            "trace_id / span_id",
            "Distributed tracing ids from W3C traceparent (or X-Trace-Id / X-Span-Id fallback).",
            "Correlate with upstream gateways and future OpenTelemetry wiring.",
        ),
        (
            "event (http_request)",
            "Structured HTTP access log: method, path, route/operation (FastAPI template when available), status_code, duration_ms, outcome (success|redirect|client_error|server_error|unknown), log_category=http.",
            "Latency and error mix by route; SRE dashboards and SLIs.",
        ),
        (
            "client_ip_hash",
            "SHA-256 of client IP when LOG_CLIENT_IP=true (not raw IP).",
            "Rough client attribution with lower PII exposure.",
        ),
        (
            "event (llm_call)",
            "Per litellm call: usage (chat|knowledge_graph|embedding), model, provider, duration_ms, success, outcome, message_count, prompt_chars, finish_reason, response_chars, token fields when API returns them, optional prompt/input hashes if LOG_PROMPT_HASH=true, error_type on failure, log_category=llm, integration=litellm.",
            "LLM cost/latency/reliability observability without logging prompts or bodies.",
        ),
        (
            "event (health_check)",
            "Ops probe: vector_store_initialized, documents_count, log_category=ops, outcome=success.",
            "Liveness/readiness signal and vector store sanity.",
        ),
        (
            "event (auth_session_rejected)",
            "Bearer present but session/token invalid or expired: auth_failure_reason, path, method, log_category=security, outcome=failure.",
            "Detect auth spikes and credential abuse patterns.",
        ),
        (
            "event (rate_limit_exceeded)",
            "HTTP 429 path: rate_limit_bucket, identity_kind (user|token_hash|ip), identity_hash, rate_limit_max, retry_after_seconds, log_category=security.",
            "Capacity and abuse monitoring for rate limits.",
        ),
        (
            "event (dependency_call)",
            "Optional helper log_dependency_event: dependency name (e.g. postgres), dependency_operation, duration_ms, success, outcome, error_type.",
            "Downstream latency and failure rates per dependency.",
        ),
        (
            "event (agent_prompt_init)",
            "Agent prompt construction: summary_info_chars, optional summary_info_sha256 if LOG_PROMPT_HASH=true, log_category=agent.",
            "Confirm agent context size without logging full summaries.",
        ),
        (
            "event (plan_context_for_code_gen)",
            "Plan sizing for code gen: latest_plan_chars, processed_plan_chars, plan section flags, log_category=agent.",
            "Debug plan-handling without logging plan content.",
        ),
        (
            "error object (exceptions)",
            "On logged exceptions: error.type, error.message; optional stackTrace when LOG_JSON_STACK_TRACE=true.",
            "Root-cause analysis while keeping stacks optional in prod.",
        ),
        (
            "LOG_LEVEL / LOG_FILE / ENABLE_CONSOLE_LOGGING",
            "Control verbosity, file path (empty disables file logging), and stdout logging.",
            "Operational tuning for dev vs containers.",
        ),
        (
            "LOG_SENSITIVE_DEBUG / LOG_PROMPT_HASH",
            "Toggles for previews/hashes of sensitive content in select logs.",
            "Strict prod privacy vs deeper support debugging.",
        ),
    ]

    app.append(["Metric", "Definition", "Purpose"])
    _style_header(app, 1, 3)
    for row in app_rows:
        app.append(list(row))
    for r in range(2, app.max_row + 1):
        for c in range(1, 4):
            app.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(app, 3)

    infra = wb.create_sheet("Infrastructure")
    infra_rows: list[tuple[str, str, str, str]] = [
        (
            "EKS control plane logs (api, audit, authenticator, scheduler, controller)",
            "EKS (AWS / CloudWatch Logs)",
            "Log group: /aws/eks/<cluster>/cluster — API audit and auth to the Kubernetes API server.",
            "Security forensics, API misuse, and compliance for cluster operations.",
        ),
        (
            "Container Insights — cluster/node/pod metrics",
            "EKS + CloudWatch Agent / Observability addon (Amazon CloudWatch)",
            "CPU, memory, network, filesystem, pod restarts; namespace/workload breakdowns.",
            "Capacity planning, saturation alerts, and crash-loop detection.",
        ),
        (
            "Application container stdout (MIDAS JSON)",
            "EKS (Pod logs → Fluent Bit / CloudWatch Logs)",
            "Ship stdout/stderr JSON from midas pods to a dedicated log group (e.g. /midas/backend).",
            "Same business/LLM/http logs as local; central in CloudWatch.",
        ),
        (
            "Ingress / ALB access logs",
            "AWS ALB (ELB) + S3 or CloudWatch",
            "HTTP host, path, target status, latency, TLS; optional to CloudWatch.",
            "Edge latency and 4xx/5xx before traffic reaches pods.",
        ),
        (
            "RDS PostgreSQL",
            "RDS + CloudWatch Logs (postgresql log)",
            "Slow queries, errors, connections; log group e.g. /aws/rds/instance/<id>/postgresql.",
            "Database performance and query failures.",
        ),
        (
            "RDS OS metrics",
            "RDS + CloudWatch (AWS/RDS metrics)",
            "CPUUtilization, FreeStorageSpace, DatabaseConnections, ReadLatency, WriteLatency.",
            "DB health, storage exhaustion, connection limits.",
        ),
        (
            "ElastiCache Redis",
            "ElastiCache + CloudWatch (AWS/ElastiCache)",
            "CPU, memory, evictions, connections, replication lag (if cluster mode).",
            "Cache saturation and session/rate-limit store reliability.",
        ),
        (
            "S3 bucket metrics",
            "S3 + CloudWatch (AWS/S3)",
            "Bucket size, request counts, 4xx/5xx, replication (if used).",
            "Upload volume and error rates for uploads.",
        ),
        (
            "Bedrock invocation metrics",
            "Amazon Bedrock + CloudWatch (AWS/Bedrock)",
            "Invocations, invocation errors, latency, throttles by model/region.",
            "Infra-side LLM usage and quota limits vs app-level llm_call logs.",
        ),
        (
            "EC2 (EKS worker nodes)",
            "EC2 + CloudWatch (AWS/EC2) + Container Insights node metrics",
            "CPU, disk, network, status checks; node NotReady correlates with pod issues.",
            "Node failure and capacity for the Kubernetes data plane.",
        ),
        (
            "VPC Flow Logs",
            "VPC / Transit Gateway + CloudWatch Logs or S3",
            "Accepted/rejected traffic flows between subnets, LB, and NAT.",
            "Network segmentation and anomaly detection.",
        ),
        (
            "AWS WAF (if deployed)",
            "WAF + CloudWatch (AWS/WAFV2)",
            "Blocked requests, rate-based rules, geo matches.",
            "Edge security and bot mitigation.",
        ),
        (
            "Secrets Manager / API call failures",
            "CloudTrail + CloudWatch Logs (optional)",
            "GetSecretValue, failures, unusual access patterns.",
            "Audit credential access and secret rotation issues.",
        ),
    ]

    infra.append(["Metric", "Source", "Definition", "Purpose"])
    _style_header(infra, 1, 4)
    for row in infra_rows:
        infra.append(list(row))
    for r in range(2, infra.max_row + 1):
        for c in range(1, 5):
            infra.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(infra, 4)

    wb.save(out)
    print(f"Wrote {out}")

    csv_app = out.parent / "metrics_application.csv"
    with csv_app.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Metric", "Definition", "Purpose"])
        w.writerows(app_rows)
    print(f"Wrote {csv_app}")

    csv_infra = out.parent / "metrics_infrastructure.csv"
    with csv_infra.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Metric", "Source", "Definition", "Purpose"])
        w.writerows(infra_rows)
    print(f"Wrote {csv_infra}")


if __name__ == "__main__":
    main()
